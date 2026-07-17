"""
app/api/v1/endpoints/doi_soat_danh_gia.py
=========================================
ĐỐI SOÁT HOÀN THÀNH ĐÁNH GIÁ THÁNG (dành cho TCCB / CCT).

Mục tiêu: TCCB tự phục vụ, KHÔNG cần chạy SQL tay để biết mỗi kỳ ai CHƯA hoàn tất
kê khai / duyệt (đây chính là việc đối soát "hệ thống vs phụ lục giấy" lặp lại hàng
tháng). Endpoint CHỈ ĐỌC — không ghi gì vào KPI production.

Cách xác định đơn vị của công chức "tại tháng" tái dùng NGUYÊN hàm biểu quyết đa số
3 nguồn `_don_vi_tai_thang_expr` của báo cáo xếp loại → "sự thật" khớp đúng báo cáo,
không sinh cảnh báo "ma".

Nhóm triệu chứng (đúng danh sách đối soát T4/T5 đã kiểm chứng):
  1. cc_chua_ke_khai_cv     — Công chức thường chưa kê khai công việc trong tháng.
  2. hd111_chua_ke_khai     — HĐ 111 chưa kê khai VB714 (và cũng chưa có form lãnh đạo).
  3. hd111_cho_duyet        — HĐ 111 đã kê khai VB714 nhưng CHƯA được duyệt (NHAP/CHO_DUYET).
  4. tcc_chua_ke_khai       — Chưa kê khai/hoàn tất tiêu chí chung (không có bản, hoặc còn NHAP/bị trả).
  5. tcc_cho_duyet          — Tiêu chí chung đã gửi nhưng đang CHỜ DUYỆT (kèm người cần duyệt).
  6. ban_ghi_0_diem         — Đã có trong báo cáo nhưng điểm tổng = 0 (thai sản/nghỉ/nhầm → đơn vị xác nhận).
"""

from __future__ import annotations

import io
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func

from app.api.deps import DatabaseDep, ActiveUserDep
from app.models.user_org import CongChuc, DonVi, VaiTro, CapBacVaiTro
from app.models.kpi_assessment import DanhGiaThang
from app.models.kpi_submission import KeKhaiCongViec
from app.models.leader_kpi import KeKhaiLanhDao
from app.models.hdld import HdldDanhGia
from app.models.bao_cao_xep_loai import BaoCaoXepLoai, ChiTietXepLoai
from app.schemas.common import success_response, error_response

# Tái dùng NGUYÊN biểu quyết đa số 3 nguồn của báo cáo xếp loại (giữ "sự thật" khớp báo cáo)
from app.api.v1.endpoints.bao_cao_xep_loai import _don_vi_tai_thang_expr

router = APIRouter()


# =============================================================================
# METADATA CÁC NHÓM (để FE render nhất quán, không hardcode nhãn ở FE)
# =============================================================================

NHOM_META = [
    {
        "key": "cc_chua_ke_khai_cv",
        "ten": "Công chức chưa kê khai công việc",
        "mo_ta": "Công chức thường chưa nộp kê khai công việc trong tháng → KPI = 0.",
        "muc_do": "cao",
        "nguoi_xu_ly": "Chính công chức",
    },
    {
        "key": "hd111_chua_ke_khai",
        "ten": "HĐ 111 chưa kê khai VB714",
        "mo_ta": "Hợp đồng 111 chưa nộp form VB714 (và cũng chưa có kê khai lãnh đạo) → KPI = 0.",
        "muc_do": "cao",
        "nguoi_xu_ly": "Chính HĐ 111",
    },
    {
        "key": "hd111_cho_duyet",
        "ten": "HĐ 111 — VB714 chờ duyệt",
        "mo_ta": "Đã kê khai VB714 nhưng chưa được Đội trưởng duyệt (còn nháp/chờ duyệt) → KPI = 0.",
        "muc_do": "cao",
        "nguoi_xu_ly": "Đội trưởng đơn vị",
    },
    {
        "key": "tcc_chua_ke_khai",
        "ten": "Chưa kê khai tiêu chí chung",
        "mo_ta": "Chưa có bản đánh giá tiêu chí chung, hoặc còn nháp/bị trả lại → điểm chung = 0.",
        "muc_do": "cao",
        "nguoi_xu_ly": "Chính công chức",
    },
    {
        "key": "tcc_cho_duyet",
        "ten": "Tiêu chí chung — chờ duyệt",
        "mo_ta": "Đã chấm tiêu chí chung nhưng đang kẹt ở bước duyệt (Phó ĐT/ĐT).",
        "muc_do": "trung_binh",
        "nguoi_xu_ly": "Phó Đội trưởng / Đội trưởng",
    },
    {
        "key": "diem_bat_thuong",
        "ten": "Điểm bất thường trên báo cáo (thành phần = 0)",
        "mo_ta": (
            "Lưới an toàn: mọi công chức ĐÃ lên báo cáo nhưng có thành phần điểm = 0 "
            "(KPI = 0, hoặc tiêu chí chung = 0, hoặc điểm tổng = 0). Quét đúng triệu chứng "
            "đã kiểm chứng nên KHÔNG bỏ sót. Nguyên nhân mỗi người MỘT KHÁC — xem cột "
            "\"Chi tiết\" (vd HĐ111 VB714 chờ duyệt, lãnh đạo chưa có bản tiêu chí chung, "
            "TCC còn nháp, thai sản/nghỉ, hoặc đã kê khai nhưng KPI=0). Phần lớn trùng với "
            "các nhóm trên; giá trị riêng là bắt ca \"đã kê khai nhưng KPI=0\"."
        ),
        "muc_do": "trung_binh",
        "nguoi_xu_ly": "Đơn vị xác nhận / xem lại",
    },
]


# =============================================================================
# HELPER
# =============================================================================

def _ev(x) -> Optional[str]:
    """Trả về giá trị chuỗi của enum (hoặc chính chuỗi/None)."""
    return x.value if hasattr(x, "value") else x


def _check_quyen_doi_soat(user: CongChuc) -> bool:
    """
    Ai được xem đối soát toàn Chi cục: TCCB, CCT/PCCT, Super Admin, hoặc user có
    cờ can_view_all_units (đọc toàn đơn vị).
    """
    if getattr(user, "is_system_admin", False):
        return True
    if getattr(user, "can_view_all_units", False):
        return True
    if not user.vai_tro:
        return False
    return user.vai_tro.cap_bac in (
        CapBacVaiTro.TCCB,
        CapBacVaiTro.CHI_CUC_TRUONG,
        CapBacVaiTro.PHO_CHI_CUC_TRUONG,
    )


async def _thu_thap_doi_soat(
    db: DatabaseDep,
    thang: int,
    nam: int,
    don_vi_id: Optional[UUID] = None,
) -> dict:
    """
    Thu thập dữ liệu đối soát cho tháng/năm. Trả về dict:
      { "nhom": {key: [item,...]}, "meta": NHOM_META, "tong_hop": {...}, "thang", "nam" }
    Mỗi item: ma_cc, ho_ten, chuc_vu, vai_tro, don_vi_id, don_vi_ten, chi_tiet(str), nguoi_xu_ly(str|None)
    """
    dv_expr = _don_vi_tai_thang_expr(thang, nam)

    # --- 1. Roster: công chức đang hoạt động + đơn vị "tại tháng" (biểu quyết) ---
    #     Loại vai trò không thuộc diện đánh giá KPI: Super Admin, TCCB.
    roster_stmt = (
        select(
            CongChuc.id,
            CongChuc.ma_cc,
            CongChuc.ho_ten,
            CongChuc.chuc_vu,
            CongChuc.is_lanh_dao,
            VaiTro.ma_vai_tro,
            VaiTro.cap_bac,
            dv_expr.label("dv_id"),
            DonVi.ten_don_vi.label("dv_ten"),
        )
        .select_from(CongChuc)
        .outerjoin(VaiTro, VaiTro.id == CongChuc.vai_tro_id)
        .outerjoin(DonVi, DonVi.id == dv_expr)
        .where(
            CongChuc.is_active == True,  # noqa: E712
            CongChuc.is_deleted == False,  # noqa: E712
        )
    )
    if don_vi_id is not None:
        roster_stmt = roster_stmt.where(dv_expr == don_vi_id)

    roster = (await db.execute(roster_stmt)).all()

    # Bỏ các vai trò không đánh giá (TCCB / Super Admin)
    KHONG_DANH_GIA = {"TCCB", "SUPER_ADMIN"}
    roster = [r for r in roster if (r.ma_vai_tro or "") not in KHONG_DANH_GIA]

    cc_ids = [r.id for r in roster]
    if not cc_ids:
        return {
            "thang": thang, "nam": nam,
            "meta": NHOM_META,
            "nhom": {m["key"]: [] for m in NHOM_META},
            "tong_hop": {m["key"]: 0 for m in NHOM_META},
            "tong_so_ca": 0,
        }

    # --- 2. Các tín hiệu trạng thái (mỗi bảng 1 query, gom theo cong_chuc_id) ---
    # 2a. Đã kê khai công việc (CC thường)
    kk_cv_rows = await db.execute(
        select(KeKhaiCongViec.cong_chuc_id)
        .where(
            KeKhaiCongViec.cong_chuc_id.in_(cc_ids),
            KeKhaiCongViec.thang == thang,
            KeKhaiCongViec.nam == nam,
            KeKhaiCongViec.is_deleted == False,  # noqa: E712
        )
        .distinct()
    )
    set_kk_cv = {r[0] for r in kk_cv_rows}

    # 2b. Đánh giá tháng (tiêu chí chung): trạng thái + người duyệt
    dgt_rows = await db.execute(
        select(
            DanhGiaThang.cong_chuc_id,
            DanhGiaThang.trang_thai_tc,
            DanhGiaThang.nguoi_phe_duyet_tc_cap1_id,
            DanhGiaThang.nguoi_phe_duyet_tc_cap2_id,
        ).where(
            DanhGiaThang.cong_chuc_id.in_(cc_ids),
            DanhGiaThang.thang == thang,
            DanhGiaThang.nam == nam,
            DanhGiaThang.is_deleted == False,  # noqa: E712
        )
    )
    map_dgt = {r.cong_chuc_id: r for r in dgt_rows}

    # 2c. HĐ 111 — VB714 (hdld_danh_gia): trạng thái
    hdld_rows = await db.execute(
        select(HdldDanhGia.cong_chuc_id, HdldDanhGia.trang_thai).where(
            HdldDanhGia.cong_chuc_id.in_(cc_ids),
            HdldDanhGia.thang == thang,
            HdldDanhGia.nam == nam,
        )
    )
    map_hdld = {r.cong_chuc_id: _ev(r.trang_thai) for r in hdld_rows}

    # 2d. Kê khai theo form lãnh đạo (fallback cho HĐ 111)
    kkld_rows = await db.execute(
        select(KeKhaiLanhDao.cong_chuc_id).where(
            KeKhaiLanhDao.cong_chuc_id.in_(cc_ids),
            KeKhaiLanhDao.thang == thang,
            KeKhaiLanhDao.nam == nam,
            KeKhaiLanhDao.is_deleted == False,  # noqa: E712
        ).distinct()
    )
    set_kkld = {r[0] for r in kkld_rows}

    # 2e. Chi tiết xếp loại đã có trong báo cáo (điểm) — join báo cáo để lọc tháng
    ct_rows = await db.execute(
        select(
            ChiTietXepLoai.cong_chuc_id,
            ChiTietXepLoai.diem_kpi,
            ChiTietXepLoai.diem_tieu_chi_chung,
            ChiTietXepLoai.diem_tong,
        )
        .select_from(ChiTietXepLoai)
        .join(BaoCaoXepLoai, BaoCaoXepLoai.id == ChiTietXepLoai.bao_cao_id)
        .where(
            ChiTietXepLoai.cong_chuc_id.in_(cc_ids),
            BaoCaoXepLoai.thang == thang,
            BaoCaoXepLoai.nam == nam,
            BaoCaoXepLoai.is_deleted == False,  # noqa: E712
        )
    )
    map_ct = {r.cong_chuc_id: r for r in ct_rows}

    # --- 3. Người duyệt: tra tên/mã để hiển thị "ai cần xử lý" ---
    nguoi_duyet_ids = set()
    for r in map_dgt.values():
        nid = r.nguoi_phe_duyet_tc_cap2_id or r.nguoi_phe_duyet_tc_cap1_id
        if nid:
            nguoi_duyet_ids.add(nid)
    map_nguoi = {}
    if nguoi_duyet_ids:
        nd_rows = await db.execute(
            select(CongChuc.id, CongChuc.ma_cc, CongChuc.ho_ten).where(
                CongChuc.id.in_(nguoi_duyet_ids)
            )
        )
        map_nguoi = {r.id: f"{r.ho_ten} ({r.ma_cc})" for r in nd_rows}

    # --- 4. Phân loại từng người ---
    nhom = {m["key"]: [] for m in NHOM_META}

    def _base(r) -> dict:
        return {
            "cong_chuc_id": str(r.id),
            "ma_cc": r.ma_cc,
            "ho_ten": r.ho_ten,
            "chuc_vu": r.chuc_vu,
            "vai_tro": r.ma_vai_tro,
            "don_vi_id": str(r.dv_id) if r.dv_id else None,
            "don_vi_ten": r.dv_ten,
        }

    LANH_DAO_CAP = {
        CapBacVaiTro.CHI_CUC_TRUONG,
        CapBacVaiTro.PHO_CHI_CUC_TRUONG,
        CapBacVaiTro.TRUONG_DON_VI,
        CapBacVaiTro.PHO_DON_VI,
        CapBacVaiTro.QUAN_LY_DON_VI,
    }

    for r in roster:
        is_hd111 = r.ma_vai_tro == "HD_111"
        is_ld = bool(r.is_lanh_dao) or (r.cap_bac in LANH_DAO_CAP)

        # 4.1 — KPI kê khai
        if is_hd111:
            st = map_hdld.get(r.id)  # None | NHAP | CHO_DUYET | DA_DUYET
            if st is None and r.id not in set_kkld:
                item = _base(r)
                item["chi_tiet"] = "Chưa kê khai VB714"
                item["nguoi_xu_ly"] = None
                nhom["hd111_chua_ke_khai"].append(item)
            elif st is not None and st != "DA_DUYET":
                item = _base(r)
                nhan = "Nháp, chưa gửi" if st == "NHAP" else "Đã gửi, chờ Đội trưởng duyệt"
                item["chi_tiet"] = f"VB714: {nhan}"
                item["nguoi_xu_ly"] = "Đội trưởng đơn vị"
                nhom["hd111_cho_duyet"].append(item)
        elif not is_ld:
            # Công chức thường → cần kê khai công việc
            if r.id not in set_kk_cv:
                item = _base(r)
                item["chi_tiet"] = "Chưa kê khai công việc"
                item["nguoi_xu_ly"] = None
                nhom["cc_chua_ke_khai_cv"].append(item)
        # Lãnh đạo: KPI tính tự động từ SP cấp dưới (V2) → không cờ ở đây.

        # 4.2 — Tiêu chí chung
        dg = map_dgt.get(r.id)
        if dg is None:
            item = _base(r)
            item["chi_tiet"] = "Chưa có bản đánh giá tiêu chí chung"
            item["nguoi_xu_ly"] = None
            nhom["tcc_chua_ke_khai"].append(item)
        else:
            tt = _ev(dg.trang_thai_tc)
            if tt in ("NHAP", "TU_CHOI", None):
                item = _base(r)
                nhan = "Nháp, chưa gửi" if tt in ("NHAP", None) else "Bị trả lại"
                item["chi_tiet"] = f"Tiêu chí chung: {nhan}"
                item["nguoi_xu_ly"] = None
                nhom["tcc_chua_ke_khai"].append(item)
            elif tt in ("CHO_PHE_DUYET", "CHO_CAP2"):
                item = _base(r)
                cap = "Phó ĐT (cấp 1)" if tt == "CHO_PHE_DUYET" else "Đội trưởng (cấp 2)"
                item["chi_tiet"] = f"Tiêu chí chung: chờ {cap} duyệt"
                nid = dg.nguoi_phe_duyet_tc_cap2_id or dg.nguoi_phe_duyet_tc_cap1_id
                item["nguoi_xu_ly"] = map_nguoi.get(nid) if nid else None
                nhom["tcc_cho_duyet"].append(item)
            # DA_PHE_DUYET → hoàn tất

        # 4.3 — Điểm bất thường trên báo cáo: lưới an toàn quét ĐÚNG triệu chứng đã kiểm
        #        chứng (KPI=0 HOẶC chung=0 HOẶC tổng=0). Bắt được cả ca "đã kê khai nhưng
        #        KPI=0" (vd 0042: kê khai 9 CV, KPI=0, chung=19.5, tổng≠0) mà các nhóm
        #        "chưa kê khai" ở trên bỏ sót.
        ct = map_ct.get(r.id)
        if ct is not None:
            kpi = float(ct.diem_kpi or 0)
            chung = float(ct.diem_tieu_chi_chung or 0)
            tong = float(ct.diem_tong or 0)
            if kpi == 0 or chung == 0 or tong == 0:
                # Chẩn đoán nguyên nhân THỰC từng người từ tín hiệu nguồn (không đoán chung).
                ly_do = []
                if kpi == 0:
                    if is_hd111:
                        st = map_hdld.get(r.id)
                        if st is None and r.id not in set_kkld:
                            ly_do.append("KPI=0: HĐ111 chưa kê khai VB714")
                        elif st == "NHAP":
                            ly_do.append("KPI=0: VB714 còn nháp, chưa gửi")
                        elif st == "CHO_DUYET":
                            ly_do.append("KPI=0: VB714 chờ Đội trưởng duyệt")
                        else:
                            ly_do.append("KPI=0: HĐ111 (VB714)")
                    elif is_ld:
                        ly_do.append("KPI=0: lãnh đạo (KPI tính tự động từ SP cấp dưới)")
                    elif r.id in set_kk_cv:
                        ly_do.append("KPI=0: đã kê khai nhưng SP chưa đạt/chưa duyệt")
                    else:
                        ly_do.append("KPI=0: chưa kê khai công việc")
                if chung == 0:
                    if dg is None:
                        if is_ld:
                            ly_do.append("chung=0: lãnh đạo chưa có bản tiêu chí chung")
                        else:
                            ly_do.append("chung=0: chưa có bản tiêu chí chung")
                    else:
                        ttv = _ev(dg.trang_thai_tc)
                        if ttv in ("NHAP", None):
                            ly_do.append("chung=0: tiêu chí chung còn nháp, chưa gửi")
                        elif ttv == "TU_CHOI":
                            ly_do.append("chung=0: tiêu chí chung bị trả lại")
                        else:
                            ly_do.append("chung=0: tiêu chí chung chưa được cộng")
                if not ly_do:
                    ly_do.append("điểm tổng = 0")
                item = _base(r)
                item["chi_tiet"] = (
                    "; ".join(ly_do)
                    + f" (KPI={kpi:.1f}, chung={chung:.1f}, tổng={tong:.1f})"
                )
                item["nguoi_xu_ly"] = "Đơn vị xác nhận / xem lại"
                nhom["diem_bat_thuong"].append(item)

    # Sắp xếp mỗi nhóm theo đơn vị rồi mã CC cho dễ đọc
    for k in nhom:
        nhom[k].sort(key=lambda x: ((x["don_vi_ten"] or ""), (x["ma_cc"] or "")))

    tong_hop = {m["key"]: len(nhom[m["key"]]) for m in NHOM_META}
    return {
        "thang": thang,
        "nam": nam,
        "meta": NHOM_META,
        "nhom": nhom,
        "tong_hop": tong_hop,
        "tong_so_ca": sum(tong_hop.values()),
    }


# =============================================================================
# ENDPOINTS
# =============================================================================

@router.get("/thang/{thang}/nam/{nam}")
async def get_doi_soat(
    db: DatabaseDep,
    current_user: ActiveUserDep,
    thang: int,
    nam: int,
    don_vi_id: Optional[UUID] = Query(None, description="Lọc theo 1 đơn vị (mặc định: toàn Chi cục)"),
) -> dict:
    """
    Đối soát hoàn thành đánh giá tháng — trả về các nhóm công chức CHƯA hoàn tất.
    Quyền: TCCB, CCT/PCCT, Super Admin, hoặc user có can_view_all_units.
    """
    if not _check_quyen_doi_soat(current_user):
        raise HTTPException(status_code=403, detail=error_response(
            code="PERM_003",
            message="Chỉ TCCB / Lãnh đạo Chi cục mới được xem đối soát toàn Chi cục",
        ))
    if not (1 <= thang <= 12):
        raise HTTPException(status_code=400, detail=error_response(
            code="VALIDATION_ERROR", message="Tháng không hợp lệ"))

    data = await _thu_thap_doi_soat(db, thang, nam, don_vi_id)
    return success_response(
        data=data,
        message=f"Đối soát tháng {thang}/{nam}: {data['tong_so_ca']} trường hợp cần xử lý",
    )


@router.get("/thang/{thang}/nam/{nam}/export")
async def export_doi_soat(
    db: DatabaseDep,
    current_user: ActiveUserDep,
    thang: int,
    nam: int,
    don_vi_id: Optional[UUID] = Query(None),
) -> StreamingResponse:
    """Xuất Excel danh sách đối soát (một sheet phẳng, dễ gửi đơn vị)."""
    if not _check_quyen_doi_soat(current_user):
        raise HTTPException(status_code=403, detail=error_response(
            code="PERM_003",
            message="Chỉ TCCB / Lãnh đạo Chi cục mới được xuất đối soát",
        ))

    data = await _thu_thap_doi_soat(db, thang, nam, don_vi_id)

    # Lazy import để không nặng khi chỉ gọi API JSON
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = f"Doi soat T{thang}-{nam}"

    tieu_de = [
        "STT", "Nhóm vấn đề", "Mã CC", "Họ tên", "Chức vụ",
        "Đơn vị", "Vai trò", "Chi tiết", "Người cần xử lý",
    ]
    ws.append([f"ĐỐI SOÁT HOÀN THÀNH ĐÁNH GIÁ THÁNG {thang}/{nam}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(tieu_de))
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])
    ws.append(tieu_de)
    header_row = ws.max_row
    fill = PatternFill("solid", fgColor="D9E1F2")
    for c in range(1, len(tieu_de) + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ten_nhom = {m["key"]: m["ten"] for m in NHOM_META}
    stt = 0
    for m in NHOM_META:
        for it in data["nhom"][m["key"]]:
            stt += 1
            ws.append([
                stt,
                ten_nhom[m["key"]],
                it["ma_cc"],
                it["ho_ten"],
                it.get("chuc_vu") or "",
                it.get("don_vi_ten") or "",
                it.get("vai_tro") or "",
                it.get("chi_tiet") or "",
                it.get("nguoi_xu_ly") or "",
            ])

    widths = [5, 30, 12, 26, 20, 26, 10, 34, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"DoiSoat_T{thang:02d}_{nam}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
