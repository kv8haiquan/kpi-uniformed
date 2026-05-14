"""
app/api/v1/endpoints/export_bao_cao.py
======================================
API Endpoints xuất báo cáo DOCX/PDF cho hệ thống KPI.

Endpoints:
1. GET /export/ca-nhan/thang/{thang}/nam/{nam}          - Xuất Mẫu 01 + 02 (cá nhân)
2. GET /export/don-vi/thang/{thang}/nam/{nam}            - Xuất Mẫu 03 (đơn vị)
3. GET /export/tong-hop/thang/{thang}/nam/{nam}          - Xuất Mẫu 04 toàn Chi cục (CCT/PCCT)
4. GET /export/don-vi-tong-hop/thang/{thang}/nam/{nam}   - Xuất Mẫu 03 tất cả đơn vị
5. GET /export/mau05-doi-moi/thang/{thang}/nam/{nam}     - Xuất Mẫu 05 CC có thành tích đổi mới
6. GET /export/bao-cao-tong-hop/thang/{thang}/nam/{nam}  - Xuất 5 báo cáo thống kê Excel (ZIP)

Output: DOCX hoặc PDF (query param ?format=docx|pdf) hoặc ZIP (endpoint 6)

Phiên bản: 1.2.0 (15/03/2026) - Thêm endpoint xuất 5 báo cáo Excel ZIP
Phiên bản: 1.1.0 (11/02/2026) - Thêm Mẫu 04 cho tổng hợp toàn Chi cục
Phiên bản: 1.0.0 (02/02/2026)
"""

import io
import logging
import subprocess
import tempfile
import traceback
import calendar
from datetime import datetime, date
from decimal import Decimal
from pathlib import Path
from typing import Optional, List
from uuid import UUID

logger = logging.getLogger(__name__)

from fastapi import APIRouter, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.deps import DatabaseDep, ActiveUserDep, is_qldv
from app.models.user_org import CongChuc, DonVi, VaiTro, CapBacVaiTro
from app.models.kpi_submission import KeKhaiCongViec, TrangThaiKeKhai
from app.models.leader_kpi import KeKhaiLanhDao, TrangThaiKeKhaiLD, TrangThaiHoanThanh
from app.models.kpi_assessment import DanhGiaThang, TieuChiChung, TieuChiChungDanhGia
from app.models.bao_cao_xep_loai import (
    BaoCaoXepLoai, ChiTietXepLoai, TrangThaiBaoCao
)
from app.models.bao_cao_xep_loai_quy import (
    BaoCaoXepLoaiQuy, ChiTietXepLoaiQuy,
)
from app.schemas.common import error_response


router = APIRouter()


# =============================================================================
# HELPER: SORT CC THEO CHỨC VỤ
# =============================================================================

# Thứ tự sort theo chức vụ (cấp bậc vai trò)
SORT_ORDER_CAP_BAC = {
    "SUPER_ADMIN": 0,
    "CHI_CUC_TRUONG": 1,
    "PHO_CHI_CUC_TRUONG": 2,
    "TRUONG_DON_VI": 3,
    "QUAN_LY_DON_VI": 4,
    "PHO_DON_VI": 5,
    "CONG_CHUC": 6,
    "TCCB": 7,
}


def get_sort_key_chuc_vu(cc: CongChuc) -> tuple:
    """
    Tạo sort key cho công chức theo chức vụ (vai trò) → họ tên.

    FIX Issue #2 (27/02/2026): Sort báo cáo theo thứ tự chức vụ thay vì mã CC.

    Thứ tự:
    1. Cấp bậc vai trò (CCT → PCCT → TDV → QLDV → PDV → CC → TCCB)
    2. Họ tên (A-Z)

    Args:
        cc: CongChuc object (must have vai_tro loaded)

    Returns:
        tuple (order_index, ho_ten)
    """
    cap_bac = cc.vai_tro.cap_bac.value if cc.vai_tro and cc.vai_tro.cap_bac else "CONG_CHUC"
    order_index = SORT_ORDER_CAP_BAC.get(cap_bac, 99)
    return (order_index, cc.ho_ten or "")


# =============================================================================
# HELPER: CONVERT DOCX -> PDF VIA LIBREOFFICE
# =============================================================================

def convert_docx_to_pdf(docx_bytes: bytes) -> bytes:
    """
    Convert DOCX bytes sang PDF bytes sử dụng LibreOffice headless.
    
    Raises:
        HTTPException nếu LibreOffice không khả dụng hoặc convert lỗi.
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        docx_path = Path(tmpdir) / "report.docx"
        docx_path.write_bytes(docx_bytes)
        
        try:
            result = subprocess.run(
                [
                    "libreoffice", "--headless", "--convert-to", "pdf",
                    "--outdir", tmpdir, str(docx_path)
                ],
                capture_output=True, text=True, timeout=60
            )
        except FileNotFoundError:
            raise HTTPException(
                status_code=500,
                detail=error_response(
                    code="SYS_010",
                    message="LibreOffice chưa được cài đặt trên server. Vui lòng xuất DOCX."
                )
            )
        except subprocess.TimeoutExpired:
            raise HTTPException(
                status_code=500,
                detail=error_response(
                    code="SYS_011",
                    message="Chuyển đổi PDF quá thời gian cho phép."
                )
            )
        
        pdf_path = Path(tmpdir) / "report.pdf"
        if not pdf_path.exists():
            raise HTTPException(
                status_code=500,
                detail=error_response(
                    code="SYS_012",
                    message="Không thể tạo file PDF. Vui lòng xuất DOCX."
                )
            )
        
        return pdf_path.read_bytes()


# =============================================================================
# HELPER: STREAMING RESPONSE
# =============================================================================

def make_file_response(
    file_bytes: bytes,
    filename: str,
    fmt: str = "docx"
) -> StreamingResponse:
    """Tạo StreamingResponse cho file download."""
    if fmt == "pdf":
        media_type = "application/pdf"
        filename = filename.replace(".docx", ".pdf")
    else:
        media_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    
    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        }
    )


# =============================================================================
# HELPER: KIỂM TRA QUYỀN
# =============================================================================

def _get_cap_bac(user: CongChuc) -> Optional[str]:
    """Lấy cấp bậc vai trò."""
    if not user.vai_tro:
        return None
    return user.vai_tro.cap_bac


def _is_lanh_dao_chi_cuc(user: CongChuc) -> bool:
    """CCT hoặc PCCT hoặc user có flag can_view_all_units."""
    # v1.1.0: User có flag can_view_all_units
    if getattr(user, 'can_view_all_units', False):
        return True
    cap_bac = _get_cap_bac(user)
    return cap_bac in [CapBacVaiTro.CHI_CUC_TRUONG, CapBacVaiTro.PHO_CHI_CUC_TRUONG]


def _is_lanh_dao_don_vi(user: CongChuc) -> bool:
    """ĐT hoặc Phó ĐT hoặc QLDV (có quyền export)."""
    cap_bac = _get_cap_bac(user)
    return cap_bac in [CapBacVaiTro.TRUONG_DON_VI, CapBacVaiTro.PHO_DON_VI] or is_qldv(user)


# =============================================================================
# HELPER: LẤY DỮ LIỆU
# =============================================================================

async def _get_danh_gia_thang(
    db: AsyncSession, cong_chuc_id: UUID, thang: int, nam: int
) -> Optional[DanhGiaThang]:
    """Lấy đánh giá tháng (tiêu chí chung) của 1 CC."""
    stmt = (
        select(DanhGiaThang)
        .options(
            selectinload(DanhGiaThang.tieu_chi_chungs)
            .selectinload(TieuChiChungDanhGia.tieu_chi)
        )
        .where(
            DanhGiaThang.cong_chuc_id == cong_chuc_id,
            DanhGiaThang.thang == thang,
            DanhGiaThang.nam == nam,
            DanhGiaThang.is_deleted == False,
        )
    )
    result = await db.execute(stmt)
    return result.scalar_one_or_none()


async def _get_ke_khai_list(
    db: AsyncSession, cong_chuc_id: UUID, thang: int, nam: int
) -> list:
    """Lấy danh sách kê khai đã duyệt của 1 CC."""
    stmt = (
        select(KeKhaiCongViec)
        .options(
            selectinload(KeKhaiCongViec.danh_muc_sp),
            selectinload(KeKhaiCongViec.cap_do),
        )
        .where(
            KeKhaiCongViec.cong_chuc_id == cong_chuc_id,
            KeKhaiCongViec.thang == thang,
            KeKhaiCongViec.nam == nam,
            KeKhaiCongViec.trang_thai == TrangThaiKeKhai.DA_PHE_DUYET,
            KeKhaiCongViec.is_deleted == False,
        )
        .order_by(KeKhaiCongViec.created_at)
    )
    result = await db.execute(stmt)
    return list(result.scalars().all())


async def _get_ke_khai_lanh_dao_list(
    db: AsyncSession, cong_chuc_id: UUID, thang: int, nam: int
) -> list:
    """Lấy danh sách kê khai LĐ đã duyệt của 1 CC.

    Phase 3 (29/04/2026): HĐ 111 dùng cùng form ke_khai_lanh_dao như Lãnh đạo.
    Export Mẫu 02 cần đọc bảng ke_khai_lanh_dao thay vì ke_khai_cong_viec.
    """
    stmt = (
        select(KeKhaiLanhDao)
        .where(
            KeKhaiLanhDao.cong_chuc_id == cong_chuc_id,
            KeKhaiLanhDao.thang == thang,
            KeKhaiLanhDao.nam == nam,
            KeKhaiLanhDao.trang_thai == TrangThaiKeKhaiLD.DA_PHE_DUYET.value,
            KeKhaiLanhDao.is_deleted == False,
        )
        .order_by(KeKhaiLanhDao.ngay_thuc_hien, KeKhaiLanhDao.created_at)
    )
    result = await db.execute(stmt)
    return list(result.scalars().all())


async def _get_tieu_chi_chung_list(db: AsyncSession) -> list:
    """Lấy danh sách tiêu chí chung (master data)."""
    stmt = (
        select(TieuChiChung)
        .where(TieuChiChung.is_active == True)
        .order_by(TieuChiChung.thu_tu)
    )
    result = await db.execute(stmt)
    return list(result.scalars().all())


async def _get_bao_cao_don_vi(
    db: AsyncSession, don_vi_id: UUID, thang: int, nam: int
) -> Optional[BaoCaoXepLoai]:
    """Lấy báo cáo xếp loại đơn vị kèm chi tiết."""
    stmt = (
        select(BaoCaoXepLoai)
        .options(
            selectinload(BaoCaoXepLoai.don_vi),
            selectinload(BaoCaoXepLoai.nguoi_lap),
            selectinload(BaoCaoXepLoai.nguoi_phe_duyet),
            selectinload(BaoCaoXepLoai.chi_tiets).selectinload(ChiTietXepLoai.cong_chuc).selectinload(CongChuc.vai_tro),
        )
        .where(
            BaoCaoXepLoai.don_vi_id == don_vi_id,
            BaoCaoXepLoai.thang == thang,
            BaoCaoXepLoai.nam == nam,
            BaoCaoXepLoai.is_deleted == False,
        )
    )
    result = await db.execute(stmt)
    return result.scalar_one_or_none()


async def _get_all_bao_cao(
    db: AsyncSession, thang: int, nam: int
) -> list:
    """Lấy tất cả báo cáo xếp loại toàn Chi cục."""
    stmt = (
        select(BaoCaoXepLoai)
        .options(
            selectinload(BaoCaoXepLoai.don_vi),
            selectinload(BaoCaoXepLoai.nguoi_lap),
            selectinload(BaoCaoXepLoai.chi_tiets).selectinload(ChiTietXepLoai.cong_chuc).selectinload(CongChuc.vai_tro),
        )
        .where(
            BaoCaoXepLoai.thang == thang,
            BaoCaoXepLoai.nam == nam,
            BaoCaoXepLoai.is_deleted == False,
        )
        .order_by(BaoCaoXepLoai.don_vi_id)
    )
    result = await db.execute(stmt)
    return list(result.scalars().all())


async def _get_bao_cao_don_vi_quy(
    db: AsyncSession, don_vi_id: UUID, quy: int, nam: int
) -> Optional[BaoCaoXepLoaiQuy]:
    """Lấy báo cáo xếp loại QUÝ của đơn vị kèm chi tiết."""
    stmt = (
        select(BaoCaoXepLoaiQuy)
        .options(
            selectinload(BaoCaoXepLoaiQuy.don_vi),
            selectinload(BaoCaoXepLoaiQuy.nguoi_lap),
            selectinload(BaoCaoXepLoaiQuy.nguoi_phe_duyet),
            selectinload(BaoCaoXepLoaiQuy.chi_tiets)
            .selectinload(ChiTietXepLoaiQuy.cong_chuc)
            .selectinload(CongChuc.vai_tro),
        )
        .where(
            BaoCaoXepLoaiQuy.don_vi_id == don_vi_id,
            BaoCaoXepLoaiQuy.quy == quy,
            BaoCaoXepLoaiQuy.nam == nam,
            BaoCaoXepLoaiQuy.is_deleted == False,
        )
    )
    result = await db.execute(stmt)
    return result.scalar_one_or_none()


async def _get_all_bao_cao_quy(
    db: AsyncSession, quy: int, nam: int
) -> list:
    """Lấy tất cả báo cáo xếp loại QUÝ toàn Chi cục."""
    stmt = (
        select(BaoCaoXepLoaiQuy)
        .options(
            selectinload(BaoCaoXepLoaiQuy.don_vi),
            selectinload(BaoCaoXepLoaiQuy.nguoi_lap),
            selectinload(BaoCaoXepLoaiQuy.chi_tiets)
            .selectinload(ChiTietXepLoaiQuy.cong_chuc)
            .selectinload(CongChuc.vai_tro),
        )
        .where(
            BaoCaoXepLoaiQuy.quy == quy,
            BaoCaoXepLoaiQuy.nam == nam,
            BaoCaoXepLoaiQuy.is_deleted == False,
        )
        .order_by(BaoCaoXepLoaiQuy.don_vi_id)
    )
    result = await db.execute(stmt)
    return list(result.scalars().all())


# =============================================================================
# HELPER: BUILD DATA DICTS CHO DOCX GENERATOR
# =============================================================================

def _build_mau01_data(
    user: CongChuc,
    thang: int, nam: int,
    danh_gia: Optional[DanhGiaThang],
    chi_tiet_xep_loai: Optional[ChiTietXepLoai],
) -> dict:
    """Build data dict cho Mẫu 01 - Phiếu theo dõi đánh giá."""
    # Tiêu chí chung chi tiết
    tieu_chi_items = []
    if danh_gia and hasattr(danh_gia, 'tieu_chi_chungs') and danh_gia.tieu_chi_chungs:
        for tc_dg in danh_gia.tieu_chi_chungs:
            tc = tc_dg.tieu_chi  # TieuChiChung master record
            tieu_chi_items.append({
                "ten": tc.ten_tieu_chi if tc else "",
                "diem_toi_da": float(tc.diem_toi_da) if tc else 0,
                "diem_tu_cham": float(tc_dg.diem_tu_cham) if tc_dg.diem_tu_cham else 0,
                "diem_lanh_dao": float(tc_dg.diem_phe_duyet) if tc_dg.diem_phe_duyet else 0,
            })
    
    diem_tcc = float(chi_tiet_xep_loai.diem_tieu_chi_chung) if chi_tiet_xep_loai else 0
    diem_kpi = float(chi_tiet_xep_loai.diem_kpi) if chi_tiet_xep_loai else 0
    diem_tong = float(chi_tiet_xep_loai.diem_tong) if chi_tiet_xep_loai else 0
    
    return {
        "ho_ten": user.ho_ten,
        "chuc_vu": user.chuc_vu or "",
        "don_vi": user.don_vi.ten_don_vi if user.don_vi else "",
        "thang": thang,
        "nam": nam,
        "tieu_chi_items": tieu_chi_items,
        "diem_tieu_chi_chung": diem_tcc,
        "diem_kpi": diem_kpi,
        "diem_tong": diem_tong,
        "xep_loai": chi_tiet_xep_loai.xep_loai_cuoi_cung if chi_tiet_xep_loai else "E",
    }


def _build_mau02_data(
    user: CongChuc,
    thang: int, nam: int,
    ke_khais: list,
    chi_tiet_xep_loai: Optional[ChiTietXepLoai],
) -> dict:
    """Build data dict cho Mẫu 02 - Bảng kê công việc cá nhân."""
    cong_viec_items = []
    tong_sp_quy_doi = Decimal("0")
    tong_sp_cl = Decimal("0")
    tong_sp_td = Decimal("0")
    
    for kk in ke_khais:
        sp_qd = kk.so_sp_goc_quy_doi or Decimal("0")
        sp_cl = kk.so_sp_chat_luong or Decimal("0")
        sp_td = kk.so_sp_tien_do or Decimal("0")
        
        tong_sp_quy_doi += sp_qd
        tong_sp_cl += sp_cl
        tong_sp_td += sp_td
        
        cong_viec_items.append({
            "ten_cong_viec": kk.danh_muc_sp.ten_cong_viec if kk.danh_muc_sp else "",
            "cap_do": kk.cap_do.ma_cap_do if kk.cap_do else "",
            "so_luong": kk.so_luong,
            "sp_quy_doi": float(sp_qd),
            "tu_dg_tien_do": kk.tu_danh_gia_tien_do or 0,
            "tu_dg_chat_luong": kk.tu_danh_gia_chat_luong or 0,
            "so_loi_tien_do": kk.so_loi_tien_do or 0,
            "so_loi_chat_luong": kk.so_loi_chat_luong or 0,
            "sp_chat_luong": float(sp_cl),
            "sp_tien_do": float(sp_td),
        })
    
    so_ngay_lv = float(chi_tiet_xep_loai.so_ngay_lam_viec or 0) if chi_tiet_xep_loai else 0
    so_ngay_nghi = float(chi_tiet_xep_loai.so_ngay_nghi or 0) if chi_tiet_xep_loai else 0
    diem_tcc = float(chi_tiet_xep_loai.diem_tieu_chi_chung) if chi_tiet_xep_loai else 0
    diem_kpi = float(chi_tiet_xep_loai.diem_kpi) if chi_tiet_xep_loai else 0
    diem_tong = float(chi_tiet_xep_loai.diem_tong) if chi_tiet_xep_loai else 0
    
    return {
        "ho_ten": user.ho_ten,
        "chuc_vu": user.chuc_vu or "",
        "don_vi": user.don_vi.ten_don_vi if user.don_vi else "",
        "thang": thang,
        "nam": nam,
        "so_ngay_lam_viec": so_ngay_lv,
        "so_ngay_nghi": so_ngay_nghi,
        "cong_viec_items": cong_viec_items,
        "tong_sp_quy_doi": float(tong_sp_quy_doi),
        "tong_sp_chat_luong": float(tong_sp_cl),
        "tong_sp_tien_do": float(tong_sp_td),
        "target_sp": so_ngay_lv * 96,
        "diem_tieu_chi_chung": diem_tcc,
        "diem_kpi": diem_kpi,
        "diem_tong": diem_tong,
        "xep_loai": chi_tiet_xep_loai.xep_loai_cuoi_cung if chi_tiet_xep_loai else "E",
    }


def _build_mau02_data_lanh_dao_style(
    user: CongChuc,
    thang: int, nam: int,
    ke_khais: list,
    chi_tiet_xep_loai: Optional[ChiTietXepLoai],
) -> dict:
    """Build data dict cho Mẫu 02 — kê khai theo form Lãnh đạo (LĐ + HĐ 111).

    Khác với CC thường:
    - Mỗi CV không có cap_do / SP quy đổi theo hệ số.
    - Mỗi CV "đếm" theo số lượng (so_luong, mặc định 1).
    - SP CL = so_luong × max(0, 1 - so_loi_cl × 0.25); SP TĐ tương tự.
    - target_sp = tổng số CV được giao (không phải ngày × 96).
    """
    cong_viec_items = []
    tong_sp_quy_doi = 0.0  # Tổng CV hoàn thành
    tong_sp_cl = 0.0       # Tổng điểm chất lượng
    tong_sp_td = 0.0       # Tổng điểm tiến độ
    tong_so_luong = 0      # Tổng CV được giao

    for kk in ke_khais:
        # Trang_thai_hoan_thanh có thể là enum hoặc string
        tt_ht = kk.trang_thai_hoan_thanh
        if hasattr(tt_ht, "value"):
            tt_ht = tt_ht.value
        is_done = tt_ht == TrangThaiHoanThanh.DA_HOAN_THANH.value

        so_luong = kk.so_luong or 1
        loi_cl = kk.so_loi_chat_luong or 0
        loi_td = kk.so_loi_tien_do or 0

        # Mỗi CV hoàn thành = 1 đơn vị; tính điểm CL/TĐ theo công thức leader.
        sp_qd = float(so_luong) if is_done else 0.0
        sp_cl = float(so_luong) * max(0.0, 1.0 - loi_cl * 0.25) if is_done else 0.0
        sp_td = float(so_luong) * max(0.0, 1.0 - loi_td * 0.25) if is_done else 0.0

        tong_so_luong += so_luong
        tong_sp_quy_doi += sp_qd
        tong_sp_cl += sp_cl
        tong_sp_td += sp_td

        cong_viec_items.append({
            "ten_cong_viec": kk.ten_cong_viec or "",
            "cap_do": "",  # Không áp dụng cho HĐ 111
            "so_luong": so_luong,
            "sp_quy_doi": sp_qd,
            "tu_dg_tien_do": 0,
            "tu_dg_chat_luong": 0,
            "so_loi_tien_do": loi_td,
            "so_loi_chat_luong": loi_cl,
            "sp_chat_luong": sp_cl,
            "sp_tien_do": sp_td,
        })

    so_ngay_lv = float(chi_tiet_xep_loai.so_ngay_lam_viec or 0) if chi_tiet_xep_loai else 0
    so_ngay_nghi = float(chi_tiet_xep_loai.so_ngay_nghi or 0) if chi_tiet_xep_loai else 0
    diem_tcc = float(chi_tiet_xep_loai.diem_tieu_chi_chung) if chi_tiet_xep_loai else 0
    diem_kpi = float(chi_tiet_xep_loai.diem_kpi) if chi_tiet_xep_loai else 0
    diem_tong = float(chi_tiet_xep_loai.diem_tong) if chi_tiet_xep_loai else 0

    return {
        "ho_ten": user.ho_ten,
        "chuc_vu": user.chuc_vu or "",
        "don_vi": user.don_vi.ten_don_vi if user.don_vi else "",
        "thang": thang,
        "nam": nam,
        "so_ngay_lam_viec": so_ngay_lv,
        "so_ngay_nghi": so_ngay_nghi,
        "cong_viec_items": cong_viec_items,
        "tong_sp_quy_doi": tong_sp_quy_doi,
        "tong_sp_chat_luong": tong_sp_cl,
        "tong_sp_tien_do": tong_sp_td,
        # HĐ 111: target = tổng CV được giao (không có ngày × 96)
        "target_sp": float(tong_so_luong),
        "diem_tieu_chi_chung": diem_tcc,
        "diem_kpi": diem_kpi,
        "diem_tong": diem_tong,
        "xep_loai": chi_tiet_xep_loai.xep_loai_cuoi_cung if chi_tiet_xep_loai else "E",
    }


def _build_mau03_data(
    bao_caos: list,
    thang: int, nam: int,
    don_vi_name: str = "",
    is_toan_chi_cuc: bool = False,
) -> dict:
    """
    Build data dict cho Mẫu 03 - Bảng tổng hợp xếp loại.

    FIX Issue #2 (27/02/2026): Sort theo chức vụ thay vì mã CC.
    """
    rows = []

    for bc in bao_caos:
        don_vi_ten = bc.don_vi.ten_don_vi if bc.don_vi else ""

        if bc.chi_tiets:
            for ct in bc.chi_tiets:
                xep_loai_cuoi = ct.xep_loai_quyet_dinh or ct.xep_loai_de_xuat or ct.xep_loai_he_thong

                # Thêm thông tin sort key
                cap_bac = ct.cong_chuc.vai_tro.cap_bac.value if ct.cong_chuc and ct.cong_chuc.vai_tro else "CONG_CHUC"
                sort_order = SORT_ORDER_CAP_BAC.get(cap_bac, 99)

                rows.append({
                    "ho_ten": ct.cong_chuc.ho_ten if ct.cong_chuc else "",
                    "don_vi": don_vi_ten,
                    "chuc_vu": ct.cong_chuc.chuc_vu if ct.cong_chuc else "",
                    "so_ngay_lv": float(ct.so_ngay_lam_viec or 0),
                    "so_ngay_nghi": float(ct.so_ngay_nghi or 0),
                    "diem_tcc": float(ct.diem_tieu_chi_chung or 0),
                    "diem_kpi": float(ct.diem_kpi or 0),
                    "diem_tong": float(ct.diem_tong or 0),
                    "xep_loai_he_thong": ct.xep_loai_he_thong or "",
                    "xep_loai_de_xuat": ct.xep_loai_de_xuat or "",
                    "xep_loai_quyet_dinh": ct.xep_loai_quyet_dinh or "",
                    "xep_loai_cuoi": xep_loai_cuoi,
                    "ghi_chu": ct.ghi_chu or "",
                    # Sort keys (không xuất ra DOCX)
                    "_sort_order": sort_order,
                })

    # FIX Issue #2: Sort theo chức vụ (cấp bậc) → họ tên
    rows.sort(key=lambda r: (r["_sort_order"], r["ho_ten"]))

    # Gán STT sau khi sort
    for i, row in enumerate(rows, 1):
        row["stt"] = i
        del row["_sort_order"]  # Xóa key tạm
    
    title = "BẢNG TỔNG HỢP KẾT QUẢ XẾP LOẠI CHẤT LƯỢNG CÔNG CHỨC"
    if is_toan_chi_cuc:
        subtitle = f"Toàn Chi cục - Tháng {thang}/{nam}"
    else:
        subtitle = f"{don_vi_name} - Tháng {thang}/{nam}"
    
    # Thống kê
    tong = len(rows)
    so_a = sum(1 for r in rows if r["xep_loai_cuoi"] == "A")
    so_b = sum(1 for r in rows if r["xep_loai_cuoi"] == "B")
    so_c = sum(1 for r in rows if r["xep_loai_cuoi"] == "C")
    so_d = sum(1 for r in rows if r["xep_loai_cuoi"] == "D")
    so_e = sum(1 for r in rows if r["xep_loai_cuoi"] == "E")
    
    return {
        "title": title,
        "subtitle": subtitle,
        "thang": thang,
        "nam": nam,
        "don_vi_name": don_vi_name,
        "is_toan_chi_cuc": is_toan_chi_cuc,
        "rows": rows,
        "thong_ke": {
            "tong": tong,
            "A": so_a, "B": so_b, "C": so_c, "D": so_d, "E": so_e,
        },
    }


def _build_mau03_data_quy(
    bao_caos: list,
    quy: int, nam: int,
    don_vi_name: str = "",
    is_toan_chi_cuc: bool = False,
) -> dict:
    """
    Build data dict cho Mẫu 03 QUÝ - Bảng tổng hợp xếp loại theo quý.

    Khác với Mẫu 03 tháng:
    - Chu kỳ theo quý (1-4) thay vì tháng (1-12).
    - ChiTietXepLoaiQuy không có so_ngay_lam_viec / so_ngay_nghi (template cũ cũng không dùng).
    - Subtitle: "Quý X/năm Y".
    """
    rows = []

    for bc in bao_caos:
        don_vi_ten = bc.don_vi.ten_don_vi if bc.don_vi else ""

        if not bc.chi_tiets:
            continue

        for ct in bc.chi_tiets:
            if not ct.cong_chuc:
                continue
            # Loại trừ CC inactive / deleted
            if hasattr(ct.cong_chuc, "is_active") and ct.cong_chuc.is_active is False:
                continue
            if hasattr(ct.cong_chuc, "deleted_at") and ct.cong_chuc.deleted_at is not None:
                continue

            xep_loai_cuoi = (
                ct.xep_loai_quyet_dinh
                or ct.xep_loai_de_xuat
                or ct.xep_loai_he_thong
            )

            cap_bac = (
                ct.cong_chuc.vai_tro.cap_bac.value
                if ct.cong_chuc and ct.cong_chuc.vai_tro
                else "CONG_CHUC"
            )
            sort_order = SORT_ORDER_CAP_BAC.get(cap_bac, 99)

            rows.append({
                "ho_ten": ct.cong_chuc.ho_ten if ct.cong_chuc else "",
                "don_vi": don_vi_ten,
                "chuc_vu": ct.cong_chuc.chuc_vu if ct.cong_chuc else "",
                # Quý không theo dõi ngày LV/ngày nghỉ — để 0 cho nhất quán schema
                "so_ngay_lv": 0.0,
                "so_ngay_nghi": 0.0,
                "diem_tcc": float(ct.diem_tieu_chi_chung or 0),
                "diem_kpi": float(ct.diem_kpi or 0),
                "diem_tong": float(ct.diem_tong or 0),
                "xep_loai_he_thong": ct.xep_loai_he_thong or "",
                "xep_loai_de_xuat": ct.xep_loai_de_xuat or "",
                "xep_loai_quyet_dinh": ct.xep_loai_quyet_dinh or "",
                "xep_loai_cuoi": xep_loai_cuoi,
                "ghi_chu": ct.ghi_chu or "",
                "_sort_order": sort_order,
            })

    rows.sort(key=lambda r: (r["_sort_order"], r["ho_ten"]))
    for i, row in enumerate(rows, 1):
        row["stt"] = i
        del row["_sort_order"]

    title = "BẢNG TỔNG HỢP KẾT QUẢ XẾP LOẠI CHẤT LƯỢNG CÔNG CHỨC"
    if is_toan_chi_cuc:
        subtitle = f"Toàn Chi cục - Quý {quy}/{nam}"
    else:
        subtitle = f"{don_vi_name} - Quý {quy}/{nam}"

    tong = len(rows)
    so_a = sum(1 for r in rows if r["xep_loai_cuoi"] == "A")
    so_b = sum(1 for r in rows if r["xep_loai_cuoi"] == "B")
    so_c = sum(1 for r in rows if r["xep_loai_cuoi"] == "C")
    so_d = sum(1 for r in rows if r["xep_loai_cuoi"] == "D")
    so_e = sum(1 for r in rows if r["xep_loai_cuoi"] == "E")

    return {
        "title": title,
        "subtitle": subtitle,
        # Báo cho generate.js biết đây là báo cáo quý → render "Quý X năm Y"
        "quy": quy,
        "nam": nam,
        "don_vi_name": don_vi_name,
        "is_toan_chi_cuc": is_toan_chi_cuc,
        "rows": rows,
        "thong_ke": {
            "tong": tong,
            "A": so_a, "B": so_b, "C": so_c, "D": so_d, "E": so_e,
        },
    }


def _build_mau04_data(
    bao_caos: list,
    thang: int, nam: int,
) -> dict:
    """
    Build data dict cho Mẫu 04 - DANH SÁCH PHÊ DUYỆT KẾT QUẢ XẾP LOẠI CHẤT LƯỢNG CÔNG CHỨC.

    Mẫu 04 bao gồm:
    - Bảng danh sách toàn bộ công chức với: STT, Mã CC, Họ tên, Năm sinh, Chức vụ, Đơn vị, Điểm, Xếp loại, Ghi chú
    - Bảng tổng hợp: Mức xếp loại | Số lượng | Tỷ lệ %
    - Bảng vinh danh: Top 5 công chức có điểm cao nhất (loại A)

    FIX Issue #2 (27/02/2026): Sort theo logic đặc biệt:
    - CCT → PCCT (toàn chi cục)
    - Mỗi đơn vị: TDV → QLDV → PDV → CC (sort theo tên đơn vị)
    """
    rows = []

    for bc in bao_caos:
        don_vi_ten = bc.don_vi.ten_don_vi if bc.don_vi else ""
        don_vi_ma = bc.don_vi.ma_don_vi if bc.don_vi else "ZZZ"  # Sort key

        if bc.chi_tiets:
            for ct in bc.chi_tiets:
                # Filter: bỏ qua user đã bị vô hiệu hóa hoặc đã xóa
                if not ct.cong_chuc:
                    continue
                if hasattr(ct.cong_chuc, 'is_active') and ct.cong_chuc.is_active == False:
                    continue
                if hasattr(ct.cong_chuc, 'deleted_at') and ct.cong_chuc.deleted_at is not None:
                    continue
                # Loại trừ ADMIN và QLDV khỏi báo cáo
                if ct.cong_chuc.vai_tro and ct.cong_chuc.vai_tro.cap_bac in (
                    CapBacVaiTro.SUPER_ADMIN, CapBacVaiTro.QUAN_LY_DON_VI,
                ):
                    continue

                xep_loai_cuoi = ct.xep_loai_quyet_dinh or ct.xep_loai_de_xuat or ct.xep_loai_he_thong

                # Lấy năm sinh từ ngày sinh
                nam_sinh = ""
                if ct.cong_chuc and ct.cong_chuc.ngay_sinh:
                    nam_sinh = ct.cong_chuc.ngay_sinh.year if hasattr(ct.cong_chuc.ngay_sinh, 'year') else str(ct.cong_chuc.ngay_sinh)[:4]

                # Sort keys
                cap_bac = ct.cong_chuc.vai_tro.cap_bac.value if ct.cong_chuc and ct.cong_chuc.vai_tro else "CONG_CHUC"
                sort_order = SORT_ORDER_CAP_BAC.get(cap_bac, 99)

                # CCT/PCCT thuộc "đơn vị" đặc biệt (hiển thị trước tất cả đơn vị)
                if cap_bac in ["CHI_CUC_TRUONG", "PHO_CHI_CUC_TRUONG"]:
                    don_vi_sort_key = "000_CHI_CUC"  # Đặt lên đầu
                else:
                    don_vi_sort_key = don_vi_ma

                rows.append({
                    "ma_cc": ct.cong_chuc.ma_cc if ct.cong_chuc else "",
                    "ho_ten": ct.cong_chuc.ho_ten if ct.cong_chuc else "",
                    "nam_sinh": nam_sinh,
                    "chuc_vu": ct.cong_chuc.chuc_vu if ct.cong_chuc else "",
                    "don_vi": don_vi_ten,
                    "diem_tong": float(ct.diem_tong or 0),
                    "xep_loai": xep_loai_cuoi,
                    "ghi_chu": ct.ghi_chu or "",
                    # Sort keys
                    "_don_vi_sort": don_vi_sort_key,
                    "_cap_bac_sort": sort_order,
                })

    # FIX Issue #2: Sort theo đơn vị (CCT/PCCT trước) → cấp bậc → họ tên
    rows.sort(key=lambda r: (r["_don_vi_sort"], r["_cap_bac_sort"], r["ho_ten"]))

    # Gán STT sau khi sort và xóa sort keys
    for i, row in enumerate(rows, 1):
        row["stt"] = i
        del row["_don_vi_sort"]
        del row["_cap_bac_sort"]

    # Thống kê
    tong = len(rows)
    so_a = sum(1 for r in rows if r["xep_loai"] == "A")
    so_b = sum(1 for r in rows if r["xep_loai"] == "B")
    so_c = sum(1 for r in rows if r["xep_loai"] == "C")
    so_d = sum(1 for r in rows if r["xep_loai"] == "D")
    so_e = sum(1 for r in rows if r["xep_loai"] == "E")
    
    # Tính tỷ lệ phần trăm
    def calc_pct(count):
        return round(count * 100 / tong, 1) if tong > 0 else 0
    
    thong_ke = [
        {"muc": "A", "so_luong": so_a, "ty_le": calc_pct(so_a)},
        {"muc": "B", "so_luong": so_b, "ty_le": calc_pct(so_b)},
        {"muc": "C", "so_luong": so_c, "ty_le": calc_pct(so_c)},
        {"muc": "D", "so_luong": so_d, "ty_le": calc_pct(so_d)},
        {"muc": "E", "so_luong": so_e, "ty_le": calc_pct(so_e)},
    ]
    
    # Top công chức xuất sắc (loại A, sắp theo điểm giảm dần)
    top_xuat_sac = sorted(
        [r for r in rows if r["xep_loai"] == "A"],
        key=lambda x: x["diem_tong"],
        reverse=True
    )[:5]  # Top 5
    
    vinh_danh = []
    for i, r in enumerate(top_xuat_sac, 1):
        vinh_danh.append({
            "stt": i,
            "ho_ten": r["ho_ten"],
            "don_vi": r["don_vi"],
            "diem": r["diem_tong"],
        })
    
    return {
        "title": "DANH SÁCH PHÊ DUYỆT KẾT QUẢ XẾP LOẠI CHẤT LƯỢNG CÔNG CHỨC",
        "subtitle": f"Tháng {thang} năm {nam}",
        "thang": thang,
        "nam": nam,
        "rows": rows,
        "thong_ke": thong_ke,
        "vinh_danh": vinh_danh,
        "tong_cong_chuc": tong,
    }

async def _build_mau05_data(
    db: AsyncSession,
    thang: int, 
    nam: int,
) -> dict:
    """
    Build data cho Mẫu 05 - Báo cáo CC có thành tích đổi mới sáng tạo.
    Chỉ lấy CC có tích ít nhất 1 tiêu chí nhóm III (nhom_tieu_chi = 3).
    """
    # 1. Lấy tất cả tiêu chí chung (master data)
    stmt_tc = select(TieuChiChung).where(TieuChiChung.is_active == True).order_by(TieuChiChung.ma_tieu_chi)
    result_tc = await db.execute(stmt_tc)
    all_tieu_chi = result_tc.scalars().all()
    
    # 2. Lấy tất cả đánh giá tháng
    stmt = (
        select(DanhGiaThang)
        .options(
            selectinload(DanhGiaThang.cong_chuc).selectinload(CongChuc.don_vi),
            selectinload(DanhGiaThang.cong_chuc).selectinload(CongChuc.vai_tro),
            selectinload(DanhGiaThang.tieu_chi_chungs),
        )
        .where(
            DanhGiaThang.thang == thang,
            DanhGiaThang.nam == nam,
        )
    )
    result = await db.execute(stmt)
    danh_gias = result.scalars().all()
    
    cong_chuc_list = []
    
    for dg in danh_gias:
        # Filter: bỏ user inactive/deleted
        if not dg.cong_chuc:
            continue
        if hasattr(dg.cong_chuc, 'is_active') and dg.cong_chuc.is_active == False:
            continue
        if hasattr(dg.cong_chuc, 'deleted_at') and dg.cong_chuc.deleted_at is not None:
            continue
        # Loại trừ ADMIN và QLDV khỏi báo cáo
        if dg.cong_chuc.vai_tro and dg.cong_chuc.vai_tro.cap_bac in (
            CapBacVaiTro.SUPER_ADMIN, CapBacVaiTro.QUAN_LY_DON_VI,
        ):
            continue

        # Kiểm tra có tích nhóm III không
        has_nhom3 = False
        tieu_chi_data = []
        tong_diem_cc = 0
        tong_diem_ld = 0
        diem_nhom3_cc = 0
        diem_nhom3_ld = 0
        
        # Map tiêu chí đánh giá theo tieu_chi_id
        tc_danh_gia_map = {str(tcdg.tieu_chi_id): tcdg for tcdg in (dg.tieu_chi_chungs or [])}
        
        for tc in all_tieu_chi:
            tcdg = tc_danh_gia_map.get(str(tc.id))
            
            is_achieved_cc = tcdg.is_achieved_cc if tcdg else False
            is_achieved_ld = tcdg.is_achieved_ld if tcdg else None
            diem_cc = float(tcdg.diem_tu_cham) if tcdg and tcdg.diem_tu_cham else 0
            diem_ld = float(tcdg.diem_phe_duyet) if tcdg and tcdg.diem_phe_duyet is not None else None
            ghi_chu = tcdg.ghi_chu_cc if tcdg else ""
            
            # Tính điểm
            tong_diem_cc += diem_cc
            if diem_ld is not None:
                tong_diem_ld += diem_ld
            
            # Kiểm tra nhóm III
            if tc.nhom_tieu_chi == 3:
                diem_nhom3_cc += diem_cc
                if diem_ld is not None:
                    diem_nhom3_ld += diem_ld
                final_achieved = is_achieved_ld if is_achieved_ld is not None else is_achieved_cc
                if final_achieved:
                    has_nhom3 = True
            
            tieu_chi_data.append({
                "ma": tc.ma_tieu_chi,
                "ten": tc.ten_tieu_chi,
                "nhom": tc.nhom_tieu_chi,
                "diem_toi_da": float(tc.diem_toi_da),
                "is_achieved_cc": is_achieved_cc,
                "is_achieved_ld": is_achieved_ld,
                "diem_cc": diem_cc,
                "diem_ld": diem_ld,
                "ghi_chu": ghi_chu or "",
            })
        
        # Chỉ lấy CC có tích nhóm III
        if has_nhom3:
            cong_chuc_list.append({
                "ho_ten": dg.cong_chuc.ho_ten,
                "ma_cc": dg.cong_chuc.ma_cc,
                "don_vi": dg.cong_chuc.don_vi.ten_don_vi if dg.cong_chuc.don_vi else "",
                "chuc_vu": dg.cong_chuc.chuc_vu or "",
                "tieu_chi": tieu_chi_data,
                "tong_diem_cc": tong_diem_cc,
                "tong_diem_ld": tong_diem_ld,
                "diem_nhom3_cc": diem_nhom3_cc,
                "diem_nhom3_ld": diem_nhom3_ld,
            })
    
    # Sắp xếp theo điểm nhóm III giảm dần
    cong_chuc_list.sort(key=lambda x: x["diem_nhom3_ld"] or x["diem_nhom3_cc"], reverse=True)
    
    return {
        "title": "PHIẾU THEO DÕI TIÊU CHÍ CHUNG - CÔNG CHỨC CÓ THÀNH TÍCH ĐỔI MỚI",
        "thang": thang,
        "nam": nam,
        "cong_chucs": cong_chuc_list,
        "tong_so_cc": len(cong_chuc_list),
    }

# =============================================================================
# HELPER: TÌM CHI TIẾT XẾP LOẠI CỦA 1 CC
# =============================================================================

async def _find_chi_tiet_xep_loai(
    db: AsyncSession,
    cong_chuc_id: UUID,
    thang: int,
    nam: int,
) -> Optional[ChiTietXepLoai]:
    """Tìm chi tiết xếp loại của 1 CC trong tháng."""
    stmt = (
        select(ChiTietXepLoai)
        .join(BaoCaoXepLoai)
        .where(
            ChiTietXepLoai.cong_chuc_id == cong_chuc_id,
            BaoCaoXepLoai.thang == thang,
            BaoCaoXepLoai.nam == nam,
            BaoCaoXepLoai.is_deleted == False,
        )
    )
    result = await db.execute(stmt)
    return result.scalar_one_or_none()


# =============================================================================
# 1. EXPORT CÁ NHÂN (Mẫu 01 + 02)
# =============================================================================

@router.get("/ca-nhan/thang/{thang}/nam/{nam}")
async def export_ca_nhan(
    db: DatabaseDep,
    current_user: ActiveUserDep,
    thang: int,
    nam: int,
    format: str = Query("docx", regex="^(docx|pdf)$"),
    cong_chuc_id: Optional[UUID] = Query(None, description="ID CC cần xuất (LĐ xuất cho nhân viên)"),
):
    """
    Xuất báo cáo cá nhân: Mẫu 01 + Mẫu 02 trong 1 file DOCX.
    
    - CC thường: Xuất cho chính mình
    - Lãnh đạo (ĐT/PCCT/CCT): Có thể xuất cho CC thuộc đơn vị (truyền cong_chuc_id)
    
    Query params:
    - format: docx | pdf (default: docx)
    - cong_chuc_id: UUID (optional, LĐ xuất cho CC khác)
    """
    # Validate
    if thang < 1 or thang > 12:
        raise HTTPException(400, detail=error_response("VAL_001", "Tháng phải từ 1-12"))
    if nam < 2025:
        raise HTTPException(400, detail=error_response("VAL_002", "Năm phải >= 2025"))
    
    # Xác định CC cần xuất
    target_user = current_user
    if cong_chuc_id and cong_chuc_id != current_user.id:
        # Kiểm tra quyền: phải là LĐ đơn vị hoặc LĐ chi cục
        if not (_is_lanh_dao_don_vi(current_user) or _is_lanh_dao_chi_cuc(current_user)):
            raise HTTPException(403, detail=error_response(
                "PERM_003", "Bạn không có quyền xuất báo cáo cho người khác"
            ))
        
        # Load target user
        stmt = (
            select(CongChuc)
            .options(selectinload(CongChuc.don_vi), selectinload(CongChuc.vai_tro))
            .where(CongChuc.id == cong_chuc_id, CongChuc.is_deleted == False)
        )
        result = await db.execute(stmt)
        target_user = result.scalar_one_or_none()
        if not target_user:
            raise HTTPException(404, detail=error_response("BIZ_001", "Không tìm thấy công chức"))
        
        # ĐT/Phó ĐT chỉ được xuất cho CC cùng đơn vị
        if _is_lanh_dao_don_vi(current_user) and target_user.don_vi_id != current_user.don_vi_id:
            raise HTTPException(403, detail=error_response(
                "PERM_004", "Bạn chỉ được xuất báo cáo cho CC trong đơn vị mình"
            ))
    else:
        # Load đầy đủ relationships cho current_user
        stmt = (
            select(CongChuc)
            .options(selectinload(CongChuc.don_vi), selectinload(CongChuc.vai_tro))
            .where(CongChuc.id == current_user.id)
        )
        result = await db.execute(stmt)
        target_user = result.scalar_one_or_none() or current_user
    
    # Lấy dữ liệu
    danh_gia = await _get_danh_gia_thang(db, target_user.id, thang, nam)
    chi_tiet_xl = await _find_chi_tiet_xep_loai(db, target_user.id, thang, nam)

    # Build data
    mau01_data = _build_mau01_data(target_user, thang, nam, danh_gia, chi_tiet_xl)

    # Lãnh đạo + HĐ 111 kê khai theo form ke_khai_lanh_dao (không có data
    # trong ke_khai_cong_viec). Đọc đúng bảng để Mẫu 02 không trống.
    # Phase 3 (29/04/2026): mở rộng cho HĐ 111.
    # Phase 3 KPI LĐ V2 (05/05/2026): từ tháng 5/2026, LĐ THẬT đã chuyển
    # sang kê khai trên ke_khai_cong_viec (form V2). HĐ 111 vẫn giữ form cũ.
    from app.core.kpi_lanh_dao_v2 import is_kpi_lanh_dao_v2_active
    use_old_form = (
        target_user.kekhai_dung_form_lanh_dao
        and not (
            target_user.is_lanh_dao
            and not target_user.is_hd_111
            and is_kpi_lanh_dao_v2_active(thang, nam)
        )
    )
    if use_old_form:
        ke_khai_ld = await _get_ke_khai_lanh_dao_list(db, target_user.id, thang, nam)
        mau02_data = _build_mau02_data_lanh_dao_style(target_user, thang, nam, ke_khai_ld, chi_tiet_xl)
    else:
        ke_khais = await _get_ke_khai_list(db, target_user.id, thang, nam)
        mau02_data = _build_mau02_data(target_user, thang, nam, ke_khais, chi_tiet_xl)
    
    # Generate DOCX via Node.js script
    import json
    combined_data = {
        "mau01": mau01_data,
        "mau02": mau02_data,
    }
    
    # Gọi Node.js generator
    docx_bytes = await _generate_docx("ca-nhan", combined_data)
    
    # Convert nếu cần PDF
    if format == "pdf":
        docx_bytes = convert_docx_to_pdf(docx_bytes)
    
    filename = f"BaoCao_CaNhan_{target_user.ma_cc}_{thang:02d}_{nam}.{format}"
    return make_file_response(docx_bytes, filename, format)


# =============================================================================
# 2. EXPORT ĐƠN VỊ (Mẫu 03)
# =============================================================================

@router.get("/don-vi/thang/{thang}/nam/{nam}")
async def export_don_vi(
    db: DatabaseDep,
    current_user: ActiveUserDep,
    thang: int,
    nam: int,
    format: str = Query("docx", regex="^(docx|pdf)$"),
    don_vi_id: Optional[UUID] = Query(None, description="ID đơn vị (CCT/PCCT chọn đơn vị)"),
):
    """
    Xuất Mẫu 03 - Bảng tổng hợp xếp loại đơn vị.
    
    - ĐT/Phó ĐT: Chỉ xuất đơn vị mình
    - PCCT/CCT: Chọn đơn vị bất kỳ
    """
    if thang < 1 or thang > 12:
        raise HTTPException(400, detail=error_response("VAL_001", "Tháng phải từ 1-12"))
    if nam < 2025:
        raise HTTPException(400, detail=error_response("VAL_002", "Năm phải >= 2025"))
    
    # Xác định đơn vị
    target_don_vi_id = don_vi_id or current_user.don_vi_id
    
    cap_bac = _get_cap_bac(current_user)
    has_view_all = getattr(current_user, 'can_view_all_units', False)
    if not has_view_all and cap_bac not in [
        CapBacVaiTro.TRUONG_DON_VI, CapBacVaiTro.PHO_DON_VI,
        CapBacVaiTro.PHO_CHI_CUC_TRUONG, CapBacVaiTro.CHI_CUC_TRUONG,
    ]:
        raise HTTPException(403, detail=error_response(
            "PERM_003", "Bạn không có quyền xuất báo cáo đơn vị"
        ))

    # ĐT/Phó ĐT: chỉ được xuất đơn vị mình (trừ user có flag can_view_all_units)
    if not has_view_all and _is_lanh_dao_don_vi(current_user) and target_don_vi_id != current_user.don_vi_id:
        raise HTTPException(403, detail=error_response(
            "PERM_004", "Bạn chỉ được xuất báo cáo đơn vị mình"
        ))
    
    # Lấy báo cáo
    bao_cao = await _get_bao_cao_don_vi(db, target_don_vi_id, thang, nam)
    if not bao_cao:
        raise HTTPException(404, detail=error_response(
            "BIZ_002", f"Chưa có báo cáo xếp loại tháng {thang}/{nam} cho đơn vị này"
        ))
    
    try:
        don_vi_name = bao_cao.don_vi.ten_don_vi if bao_cao.don_vi else ""
        mau03_data = _build_mau03_data([bao_cao], thang, nam, don_vi_name, is_toan_chi_cuc=False)
        
        # Generate DOCX
        docx_bytes = await _generate_docx("don-vi", mau03_data)
        
        if format == "pdf":
            docx_bytes = convert_docx_to_pdf(docx_bytes)
        
        filename = f"BaoCao_DonVi_{thang:02d}_{nam}.{format}"
        return make_file_response(docx_bytes, filename, format)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[EXPORT] export_don_vi error: {traceback.format_exc()}")
        raise HTTPException(500, detail=error_response(
            "SYS_099", f"Lỗi xuất báo cáo: {str(e)[:200]}"
        ))


# =============================================================================
# 3. EXPORT TOÀN CHI CỤC (Mẫu 03 tổng hợp)
# =============================================================================

@router.get("/tong-hop/thang/{thang}/nam/{nam}")
async def export_tong_hop(
    db: DatabaseDep,
    current_user: ActiveUserDep,
    thang: int,
    nam: int,
    format: str = Query("docx", regex="^(docx|pdf)$"),
):
    """
    Xuất Mẫu 04 - Danh sách phê duyệt kết quả xếp loại chất lượng công chức toàn Chi cục.
    
    Quyền: Chỉ CCT và PCCT.
    """
    if thang < 1 or thang > 12:
        raise HTTPException(400, detail=error_response("VAL_001", "Tháng phải từ 1-12"))
    if nam < 2025:
        raise HTTPException(400, detail=error_response("VAL_002", "Năm phải >= 2025"))
    
    if not _is_lanh_dao_chi_cuc(current_user):
        raise HTTPException(403, detail=error_response(
            "PERM_003", "Chỉ CCT và Phó CCT mới được xuất báo cáo tổng hợp toàn Chi cục"
        ))
    
    bao_caos = await _get_all_bao_cao(db, thang, nam)
    if not bao_caos:
        raise HTTPException(404, detail=error_response(
            "BIZ_003", f"Chưa có báo cáo xếp loại tháng {thang}/{nam}"
        ))
    
    # Sử dụng Mẫu 04 thay vì Mẫu 03
    mau04_data = _build_mau04_data(bao_caos, thang, nam)
    
    docx_bytes = await _generate_docx("tong-hop", mau04_data)
    
    if format == "pdf":
        docx_bytes = convert_docx_to_pdf(docx_bytes)
    
    filename = f"Mau04_DanhSach_XepLoai_ChiCuc_{thang:02d}_{nam}.{format}"
    return make_file_response(docx_bytes, filename, format)


@router.get("/don-vi-tong-hop/thang/{thang}/nam/{nam}")
async def export_don_vi_tong_hop(
    db: DatabaseDep,
    current_user: ActiveUserDep,
    thang: int,
    nam: int,
    format: str = Query("docx", regex="^(docx|pdf)$"),
):
    """
    Xuất Mẫu 03 tổng hợp TẤT CẢ đơn vị (mỗi đơn vị 1 trang).
    
    Quyền: Chỉ CCT và PCCT.
    """
    if thang < 1 or thang > 12:
        raise HTTPException(400, detail=error_response("VAL_001", "Tháng phải từ 1-12"))
    if nam < 2025:
        raise HTTPException(400, detail=error_response("VAL_002", "Năm phải >= 2025"))
    
    if not _is_lanh_dao_chi_cuc(current_user):
        raise HTTPException(403, detail=error_response(
            "PERM_003", "Chỉ CCT và Phó CCT mới được xuất báo cáo tổng hợp"
        ))
    
    bao_caos = await _get_all_bao_cao(db, thang, nam)
    if not bao_caos:
        raise HTTPException(404, detail=error_response(
            "BIZ_003", f"Chưa có báo cáo xếp loại tháng {thang}/{nam}"
        ))
    
    # Build data cho từng đơn vị
    don_vi_list = []
    for bc in bao_caos:
        don_vi_data = _build_mau03_data(
            [bc], thang, nam,
            don_vi_name=bc.don_vi.ten_don_vi if bc.don_vi else "",
            is_toan_chi_cuc=False,
        )
        don_vi_list.append(don_vi_data)
    
    combined_data = {
        "don_vi_list": don_vi_list,
        "thang": thang,
        "nam": nam,
    }
    
    docx_bytes = await _generate_docx("don-vi-tong-hop", combined_data)
    
    if format == "pdf":
        docx_bytes = convert_docx_to_pdf(docx_bytes)
    
    filename = f"Mau03_TatCaDonVi_{thang:02d}_{nam}.{format}"
    return make_file_response(docx_bytes, filename, format)


# =============================================================================
# 2B. EXPORT ĐƠN VỊ THEO QUÝ (Mẫu 03 quý)
# =============================================================================

@router.get("/don-vi/quy/{quy}/nam/{nam}")
async def export_don_vi_quy(
    db: DatabaseDep,
    current_user: ActiveUserDep,
    quy: int,
    nam: int,
    format: str = Query("docx", regex="^(docx|pdf)$"),
    don_vi_id: Optional[UUID] = Query(
        None, description="ID đơn vị (CCT/PCCT/TCCB chọn đơn vị bất kỳ)"
    ),
):
    """
    Xuất Mẫu 03 QUÝ - Bảng tổng hợp xếp loại đơn vị theo quý.

    - TDV/PDV: chỉ đơn vị mình.
    - CCT/PCCT/TCCB (can_view_all_units): chọn đơn vị bất kỳ.
    """
    if quy < 1 or quy > 4:
        raise HTTPException(400, detail=error_response("VAL_001", "Quý phải từ 1-4"))
    if nam < 2025:
        raise HTTPException(400, detail=error_response("VAL_002", "Năm phải >= 2025"))

    target_don_vi_id = don_vi_id or current_user.don_vi_id

    has_view_all = getattr(current_user, "can_view_all_units", False)
    cap_bac = _get_cap_bac(current_user)
    if not has_view_all and cap_bac not in [
        CapBacVaiTro.TRUONG_DON_VI, CapBacVaiTro.PHO_DON_VI,
        CapBacVaiTro.PHO_CHI_CUC_TRUONG, CapBacVaiTro.CHI_CUC_TRUONG,
    ]:
        raise HTTPException(403, detail=error_response(
            "PERM_003", "Bạn không có quyền xuất báo cáo đơn vị"
        ))
    if (
        not has_view_all
        and _is_lanh_dao_don_vi(current_user)
        and target_don_vi_id != current_user.don_vi_id
    ):
        raise HTTPException(403, detail=error_response(
            "PERM_004", "Bạn chỉ được xuất báo cáo đơn vị mình"
        ))

    bao_cao = await _get_bao_cao_don_vi_quy(db, target_don_vi_id, quy, nam)
    if not bao_cao:
        raise HTTPException(404, detail=error_response(
            "BIZ_002",
            f"Chưa có báo cáo xếp loại quý {quy}/{nam} cho đơn vị này",
        ))

    try:
        don_vi_name = bao_cao.don_vi.ten_don_vi if bao_cao.don_vi else ""
        mau03_data = _build_mau03_data_quy(
            [bao_cao], quy, nam, don_vi_name, is_toan_chi_cuc=False
        )
        docx_bytes = await _generate_docx("don-vi", mau03_data)

        if format == "pdf":
            docx_bytes = convert_docx_to_pdf(docx_bytes)

        filename = f"Mau03_DonVi_Q{quy}_{nam}.{format}"
        return make_file_response(docx_bytes, filename, format)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[EXPORT] export_don_vi_quy error: {traceback.format_exc()}")
        raise HTTPException(500, detail=error_response(
            "SYS_099", f"Lỗi xuất báo cáo quý: {str(e)[:200]}"
        ))


@router.get("/don-vi-tong-hop/quy/{quy}/nam/{nam}")
async def export_don_vi_tong_hop_quy(
    db: DatabaseDep,
    current_user: ActiveUserDep,
    quy: int,
    nam: int,
    format: str = Query("docx", regex="^(docx|pdf)$"),
):
    """
    Xuất Mẫu 03 QUÝ tổng hợp TẤT CẢ đơn vị (mỗi đơn vị 1 trang).

    Quyền: CCT, PCCT hoặc user có can_view_all_units (TCCB).
    """
    if quy < 1 or quy > 4:
        raise HTTPException(400, detail=error_response("VAL_001", "Quý phải từ 1-4"))
    if nam < 2025:
        raise HTTPException(400, detail=error_response("VAL_002", "Năm phải >= 2025"))

    if not _is_lanh_dao_chi_cuc(current_user):
        raise HTTPException(403, detail=error_response(
            "PERM_003", "Chỉ CCT/PCCT/TCCB mới được xuất báo cáo tổng hợp quý"
        ))

    bao_caos = await _get_all_bao_cao_quy(db, quy, nam)
    if not bao_caos:
        raise HTTPException(404, detail=error_response(
            "BIZ_003", f"Chưa có báo cáo xếp loại quý {quy}/{nam}"
        ))

    don_vi_list = []
    for bc in bao_caos:
        don_vi_data = _build_mau03_data_quy(
            [bc], quy, nam,
            don_vi_name=bc.don_vi.ten_don_vi if bc.don_vi else "",
            is_toan_chi_cuc=False,
        )
        don_vi_list.append(don_vi_data)

    combined_data = {
        "don_vi_list": don_vi_list,
        "quy": quy,
        "nam": nam,
    }

    docx_bytes = await _generate_docx("don-vi-tong-hop", combined_data)

    if format == "pdf":
        docx_bytes = convert_docx_to_pdf(docx_bytes)

    filename = f"Mau03_TatCaDonVi_Q{quy}_{nam}.{format}"
    return make_file_response(docx_bytes, filename, format)


@router.get("/mau05-doi-moi/thang/{thang}/nam/{nam}")
async def export_mau05_doi_moi(
    db: DatabaseDep,
    current_user: ActiveUserDep,
    thang: int,
    nam: int,
    format: str = Query("docx", regex="^(docx|pdf)$"),
):
    """
    Xuất Mẫu 05 - Báo cáo công chức có thành tích đổi mới sáng tạo.
    Chỉ lấy CC có tích ít nhất 1 tiêu chí nhóm III.
    
    Quyền: Chỉ CCT và PCCT.
    """
    if thang < 1 or thang > 12:
        raise HTTPException(400, detail=error_response("VAL_001", "Tháng phải từ 1-12"))
    if nam < 2025:
        raise HTTPException(400, detail=error_response("VAL_002", "Năm phải >= 2025"))
    
    if not _is_lanh_dao_chi_cuc(current_user):
        raise HTTPException(403, detail=error_response(
            "PERM_003", "Chỉ CCT và Phó CCT mới được xuất báo cáo này"
        ))
    
    mau05_data = await _build_mau05_data(db, thang, nam)
    
    if not mau05_data["cong_chucs"]:
        raise HTTPException(404, detail=error_response(
            "BIZ_004", f"Không có công chức nào có tích tiêu chí nhóm III trong tháng {thang}/{nam}"
        ))
    
    docx_bytes = await _generate_docx("mau05-doi-moi", mau05_data)
    
    if format == "pdf":
        docx_bytes = convert_docx_to_pdf(docx_bytes)
    
    filename = f"Mau05_DoiMoiSangTao_{thang:02d}_{nam}.{format}"
    return make_file_response(docx_bytes, filename, format)

# =============================================================================
# HELPER: GỌI NODE.JS SCRIPT TẠO DOCX
# =============================================================================

async def _generate_docx(report_type: str, data: dict) -> bytes:
    """
    Gọi Node.js script để generate DOCX file.
    
    Args:
        report_type: "ca-nhan" | "don-vi" | "tong-hop"
        data: Dict dữ liệu cho template
    
    Returns:
        bytes: Nội dung file DOCX
    """
    import json
    
    # Đường dẫn tới script generator
    script_dir = Path(__file__).parent.parent.parent.parent / "scripts" / "docx_generator"
    script_path = script_dir / "generate.js"
    
    logger.error(f"[EXPORT] _generate_docx: script_path={script_path}, exists={script_path.exists()}, resolved={script_path.resolve()}")
    
    if not script_path.exists():
        raise HTTPException(500, detail=error_response(
            "SYS_020", f"Script tạo DOCX chưa được cài đặt trên server. Path: {script_path}"
        ))
    
    with tempfile.TemporaryDirectory() as tmpdir:
        # Ghi data JSON
        data_path = Path(tmpdir) / "data.json"
        with open(data_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, default=str)
        
        output_path = Path(tmpdir) / "output.docx"
        
        try:
            cmd = ["node", str(script_path), report_type, str(data_path), str(output_path)]
            logger.error(f"[EXPORT] Running: {' '.join(cmd)}")
            logger.error(f"[EXPORT] CWD: {script_dir}")
            result = subprocess.run(
                cmd,
                capture_output=True, text=True, timeout=30,
                cwd=str(script_dir),
            )
            logger.error(f"[EXPORT] returncode={result.returncode}, stdout={result.stdout[:200]}, stderr={result.stderr[:200]}")
        except subprocess.TimeoutExpired:
            raise HTTPException(500, detail=error_response(
                "SYS_021", "Tạo DOCX quá thời gian cho phép"
            ))
        
        # Node.js có thể exit với code != 0 (ví dụ SIGABRT) nhưng vẫn tạo file OK
        # Ưu tiên kiểm tra file output tồn tại
        if output_path.exists() and output_path.stat().st_size > 0:
            logger.error(f"[EXPORT] DOCX created OK: {output_path.stat().st_size} bytes (returncode={result.returncode})")
            return output_path.read_bytes()
        
        # Nếu file không tồn tại VÀ returncode != 0 → lỗi thật
        if result.returncode != 0:
            raise HTTPException(500, detail=error_response(
                "SYS_022",
                f"Lỗi tạo DOCX: {result.stderr[:200] if result.stderr else 'Unknown error'}"
            ))
        
        if not output_path.exists():
            raise HTTPException(500, detail=error_response(
                "SYS_023", "File DOCX không được tạo thành công"
            ))

        return output_path.read_bytes()


# =============================================================================
# 4. EXPORT ALL 5 STATISTICAL REPORTS AS ZIP
# =============================================================================

@router.get("/bao-cao-tong-hop/thang/{thang}/nam/{nam}")
async def export_bao_cao_tong_hop(
    db: DatabaseDep,
    current_user: ActiveUserDep,
    thang: int,
    nam: int,
):
    """
    Xuất tất cả 5 báo cáo thống kê dưới dạng ZIP chứa 5 file Excel.

    Quyền: Chỉ CCT, PCCT, hoặc users có can_view_all_units.

    Các báo cáo:
    1. 01_TieuChiChung_MM_YYYY.xlsx - Thống kê tiêu chí chung (30 điểm)
    2. 02_DiemKPI_MM_YYYY.xlsx - Thống kê điểm KPI (70 điểm)
    3. 03_LanhDaoDDE_MM_YYYY.xlsx - Lãnh đạo bị trừ điểm d, đ, e
    4. 04_KhoiLuongCongViec_MM_YYYY.xlsx - Khối lượng công việc
    5. 05_DanhMucCongViec_MM_YYYY.xlsx - Danh mục công việc chi tiết
    6. 06_ChiSoABCDDE_MM_YYYY.xlsx - Chỉ số a/b/c (+ d/đ/e cho LĐ) bị trừ
    7. 07_SPTheoLinhVuc_MM_YYYY.xlsx - Tổng SP theo lĩnh vực + nội dung nổi bật

    Returns:
        ZIP file chứa 7 Excel files
    """
    # Validate
    if thang < 1 or thang > 12:
        raise HTTPException(400, detail=error_response("VAL_001", "Tháng phải từ 1-12"))
    if nam < 2025:
        raise HTTPException(400, detail=error_response("VAL_002", "Năm phải >= 2025"))

    # Check permission: CCT, PCCT, or can_view_all_units
    if not _is_lanh_dao_chi_cuc(current_user):
        raise HTTPException(403, detail=error_response(
            "PERM_003", "Chỉ CCT và Phó CCT mới được xuất báo cáo tổng hợp"
        ))

    try:
        # Generate all 7 Excel files
        logger.info(f"[EXPORT_ZIP] Generating 7 reports for {thang}/{nam}")

        excel_01 = await _generate_report_01_tieu_chi_chung(db, thang, nam)
        excel_02 = await _generate_report_02_diem_kpi(db, thang, nam)
        excel_03 = await _generate_report_03_lanh_dao_dde(db, thang, nam)
        excel_04 = await _generate_report_04_khoi_luong_cv(db, thang, nam)
        excel_05 = await _generate_report_05_danh_muc_cv(db, thang, nam)
        excel_06 = await _generate_report_06_chi_so_abc_dde(db, thang, nam)
        excel_07 = await _generate_report_07_sp_theo_linh_vuc(db, thang, nam)

        # Create ZIP file in memory
        import zipfile
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            zip_file.writestr(f"01_TieuChiChung_{thang:02d}_{nam}.xlsx", excel_01.getvalue())
            zip_file.writestr(f"02_DiemKPI_{thang:02d}_{nam}.xlsx", excel_02.getvalue())
            zip_file.writestr(f"03_LanhDaoDDE_{thang:02d}_{nam}.xlsx", excel_03.getvalue())
            zip_file.writestr(f"04_KhoiLuongCongViec_{thang:02d}_{nam}.xlsx", excel_04.getvalue())
            zip_file.writestr(f"05_DanhMucCongViec_{thang:02d}_{nam}.xlsx", excel_05.getvalue())
            zip_file.writestr(f"06_ChiSoABCDDE_{thang:02d}_{nam}.xlsx", excel_06.getvalue())
            zip_file.writestr(f"07_SPTheoLinhVuc_{thang:02d}_{nam}.xlsx", excel_07.getvalue())

        zip_buffer.seek(0)

        filename = f"BaoCaoThongKe_{thang:02d}_{nam}.zip"

        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Access-Control-Expose-Headers": "Content-Disposition",
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[EXPORT_ZIP] Error: {traceback.format_exc()}")
        raise HTTPException(500, detail=error_response(
            "SYS_099", f"Lỗi xuất báo cáo tổng hợp: {str(e)[:200]}"
        ))


# =============================================================================
# HELPER: GENERATE REPORT 01 - TIÊU CHÍ CHUNG
# =============================================================================

async def _generate_report_01_tieu_chi_chung(db: AsyncSession, thang: int, nam: int) -> io.BytesIO:
    """Generate Excel report 01 - Tiêu chí chung statistics."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Get data
    data = await _get_data_01_tieu_chi_chung(db, thang, nam)

    # Create Excel
    wb = Workbook()

    # Styles
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    percent_font = Font(bold=True, color="0070C0")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='center')

    total = len(data)
    duoi_20 = [cc for cc in data if cc["tong_diem"] < 20]
    tron_20 = [cc for cc in data if cc["tong_diem"] == 20]
    tren_20 = [cc for cc in data if cc["tong_diem"] > 20]

    tren_20_co_nhom3 = [cc for cc in tren_20 if cc["has_nhom3"] and cc["has_ghi_chu_nhom3"]]
    tren_20_khong_nhom3 = [cc for cc in tren_20 if cc["has_nhom3"] and not cc["has_ghi_chu_nhom3"]]
    tren_20_30diem = [cc for cc in tren_20 if cc["tong_diem"] == 30]

    def pct(count):
        return f"{count/total*100:.1f}%" if total > 0 else "0%"

    # SHEET 1: TỔNG HỢP
    ws1 = wb.active
    ws1.title = "Tổng hợp"

    ws1['A1'] = f"1. THỐNG KÊ TIÊU CHÍ CHUNG - THÁNG {thang}/{nam}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:E1')

    ws1['A2'] = f"Tổng số công chức: {total}"
    ws1['A2'].font = Font(bold=True)

    row = 4
    headers = ["Nhóm", "Số lượng", "Tỷ lệ", "Ghi chú"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    rows_data = [
        ("1. Dưới 20 điểm", len(duoi_20), pct(len(duoi_20)), "Xem Sheet 'Dưới 20 điểm'"),
        ("2. Tròn 20 điểm", len(tron_20), pct(len(tron_20)), "Hoàn thành tốt, không sai sót"),
        ("3. Trên 20 điểm", len(tren_20), pct(len(tren_20)), "Chi tiết bên dưới"),
        ("   3a. CÓ sản phẩm đổi mới", len(tren_20_co_nhom3), pct(len(tren_20_co_nhom3)), "Có minh chứng cụ thể"),
        ("   3b. KHÔNG CÓ sản phẩm đổi mới", len(tren_20_khong_nhom3), pct(len(tren_20_khong_nhom3)), "Chưa hiểu rõ quy định"),
        ("   3c. Đạt 30 điểm tối đa", len(tren_20_30diem), pct(len(tren_20_30diem)), "Xuất sắc"),
    ]

    for i, (nhom, sl, tl, gc) in enumerate(rows_data):
        r = row + 1 + i
        ws1.cell(row=r, column=1, value=nhom).border = border
        ws1.cell(row=r, column=2, value=sl).border = border
        ws1.cell(row=r, column=2).alignment = center_alignment
        ws1.cell(row=r, column=3, value=tl).border = border
        ws1.cell(row=r, column=3).alignment = center_alignment
        ws1.cell(row=r, column=3).font = percent_font
        ws1.cell(row=r, column=4, value=gc).border = border
        if i < 3:
            ws1.cell(row=r, column=1).font = Font(bold=True)

    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 40

    # SHEET 2: DƯỚI 20 ĐIỂM
    ws2 = wb.create_sheet("Dưới 20 điểm")

    ws2['A1'] = f"DANH SÁCH CÔNG CHỨC DƯỚI 20 ĐIỂM - THÁNG {thang}/{nam}"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:H1')

    ws2['A2'] = "Bao gồm cả CC ở trạng thái CHỜ_DUYỆT — xem cột 'Trạng thái' để biết bản nào LĐ đã chấm chính thức."
    ws2['A2'].font = Font(italic=True, color="666666")

    headers2 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "Tổng điểm", "Điểm trừ", "Trạng thái", "Lý do trừ điểm"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    for i, cc in enumerate(duoi_20, 1):
        r = 3 + i
        ws2.cell(row=r, column=1, value=i).border = border
        ws2.cell(row=r, column=2, value=cc["ho_ten"]).border = border
        ws2.cell(row=r, column=3, value=cc["ma_cc"]).border = border
        ws2.cell(row=r, column=4, value=cc["don_vi"]).border = border
        ws2.cell(row=r, column=5, value=cc["tong_diem"]).border = border

        diem_tru = sum(ly["diem_tru"] for ly in cc["ly_do_tru_diem"])
        ws2.cell(row=r, column=6, value=diem_tru).border = border

        # Trạng thái — TrangThaiDanhGia (cấp phiếu chung)
        tt_raw = cc.get("dgt_trang_thai", "")
        tt_display = {
            "HOAN_THANH": "Đã duyệt",
            "CHO_PHE_DUYET": "Chờ CCT duyệt",
            "DA_TONG_HOP": "Đã tổng hợp",
            "CHO_TONG_HOP": "Chờ LĐ tổng hợp",
            "DANG_DANH_GIA": "Đang đánh giá",
            "CO_KIEN_NGHI": "Có kiến nghị",
        }.get(tt_raw, tt_raw or "—")
        tt_cell = ws2.cell(row=r, column=7, value=tt_display)
        tt_cell.border = border
        tt_cell.alignment = center_alignment
        if tt_raw == "HOAN_THANH":
            tt_cell.fill = PatternFill("solid", fgColor="C6EFCE")
        elif tt_raw in ("CHO_PHE_DUYET", "DA_TONG_HOP", "CHO_TONG_HOP"):
            tt_cell.fill = PatternFill("solid", fgColor="FFEB9C")
            tt_cell.font = Font(bold=True, color="9C5700")
        elif tt_raw == "CO_KIEN_NGHI":
            tt_cell.fill = PatternFill("solid", fgColor="FFC7CE")
            tt_cell.font = Font(bold=True, color="9C0006")

        ly_do_text = "; ".join([f"{ly['ma']}: {ly['ly_do']}" for ly in cc["ly_do_tru_diem"]])
        ws2.cell(row=r, column=8, value=ly_do_text).border = border
        ws2.cell(row=r, column=8).alignment = wrap_alignment
        ws2.row_dimensions[r].height = 40

    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 30
    ws2.column_dimensions['E'].width = 10
    ws2.column_dimensions['F'].width = 10
    ws2.column_dimensions['G'].width = 12
    ws2.column_dimensions['H'].width = 70

    # SHEET 3: TRÊN 20 ĐIỂM
    ws3 = wb.create_sheet("Trên 20 điểm")

    ws3['A1'] = f"DANH SÁCH CÔNG CHỨC TRÊN 20 ĐIỂM - THÁNG {thang}/{nam}"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:H1')

    headers3 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "Tổng điểm", "Điểm nhóm III", "Có MC", "Minh chứng đổi mới sáng tạo"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    tren_20_sorted = sorted(tren_20, key=lambda x: x["tong_diem"], reverse=True)

    for i, cc in enumerate(tren_20_sorted, 1):
        r = 3 + i
        ws3.cell(row=r, column=1, value=i).border = border
        ws3.cell(row=r, column=2, value=cc["ho_ten"]).border = border
        ws3.cell(row=r, column=3, value=cc["ma_cc"]).border = border
        ws3.cell(row=r, column=4, value=cc["don_vi"]).border = border
        ws3.cell(row=r, column=5, value=cc["tong_diem"]).border = border
        ws3.cell(row=r, column=6, value=cc["diem_nhom3"]).border = border

        if cc["has_ghi_chu_nhom3"]:
            co_mc = "Có"
            ws3.cell(row=r, column=7).fill = PatternFill("solid", fgColor="C6EFCE")
        else:
            co_mc = "Không"
            ws3.cell(row=r, column=7).fill = PatternFill("solid", fgColor="FFC7CE")
        ws3.cell(row=r, column=7, value=co_mc).border = border
        ws3.cell(row=r, column=7).alignment = center_alignment

        minh_chung = cc.get("minh_chung_nhom3", "")
        if not minh_chung:
            if cc["tong_diem"] == 30:
                minh_chung = "(Đạt điểm tối đa - chưa điền minh chứng)"
            else:
                minh_chung = "(Chưa điền minh chứng cụ thể)"
        ws3.cell(row=r, column=8, value=minh_chung).border = border
        ws3.cell(row=r, column=8).alignment = wrap_alignment
        ws3.row_dimensions[r].height = 60

    ws3.column_dimensions['A'].width = 5
    ws3.column_dimensions['B'].width = 25
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 25
    ws3.column_dimensions['E'].width = 10
    ws3.column_dimensions['F'].width = 12
    ws3.column_dimensions['G'].width = 8
    ws3.column_dimensions['H'].width = 100

    # SHEET 4: TRÒN 20 ĐIỂM
    ws4 = wb.create_sheet("Tròn 20 điểm")

    ws4['A1'] = f"DANH SÁCH CÔNG CHỨC TRÒN 20 ĐIỂM - THÁNG {thang}/{nam}"
    ws4['A1'].font = title_font
    ws4.merge_cells('A1:E1')

    ws4['A2'] = "Ghi chú: Hoàn thành tốt nhiệm vụ, không có sai sót."
    ws4['A2'].font = Font(italic=True, color="666666")

    headers4 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "Tổng điểm"]
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    for i, cc in enumerate(tron_20, 1):
        r = 4 + i
        ws4.cell(row=r, column=1, value=i).border = border
        ws4.cell(row=r, column=2, value=cc["ho_ten"]).border = border
        ws4.cell(row=r, column=3, value=cc["ma_cc"]).border = border
        ws4.cell(row=r, column=4, value=cc["don_vi"]).border = border
        ws4.cell(row=r, column=5, value=cc["tong_diem"]).border = border

    ws4.column_dimensions['A'].width = 5
    ws4.column_dimensions['B'].width = 25
    ws4.column_dimensions['C'].width = 12
    ws4.column_dimensions['D'].width = 35
    ws4.column_dimensions['E'].width = 12

    # SHEET 5: TẤT CẢ CC BỊ TRỪ TIÊU CHÍ (mọi mức điểm)
    # Khác Sheet 2 ("Dưới 20 điểm") — sheet này gồm cả CC ≥20đ vẫn bị trừ 1-2 TC.
    ws5 = wb.create_sheet("TC bị trừ - tất cả CC")

    ws5['A1'] = f"DANH SÁCH CÔNG CHỨC BỊ TRỪ TIÊU CHÍ CHUNG - THÁNG {thang}/{nam}"
    ws5['A1'].font = title_font
    ws5.merge_cells('A1:J1')

    ws5['A2'] = ("Bao gồm mọi CC có ≥1 tiêu chí bị trừ (kể cả CC tổng điểm ≥20, kể cả bản CHỜ_DUYỆT). "
                 "Cột 'Trạng thái' cho biết bản LĐ đã chấm chính thức hay CC mới tự chấm chờ duyệt.")
    ws5['A2'].font = Font(italic=True, color="666666")

    bi_tru_all = [cc for cc in data if cc["ly_do_tru_diem"]]
    bi_tru_all_sorted = sorted(
        bi_tru_all,
        key=lambda x: (sum(ly["diem_tru"] for ly in x["ly_do_tru_diem"]), len(x["ly_do_tru_diem"])),
        reverse=True,
    )

    headers5 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "Tổng điểm", "Số TC bị trừ", "Tổng điểm trừ", "Mã TC bị trừ", "Trạng thái", "Lý do / Ghi chú"]
    for col, h in enumerate(headers5, 1):
        cell = ws5.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    for i, cc in enumerate(bi_tru_all_sorted, 1):
        r = 4 + i
        diem_tru = sum(ly["diem_tru"] for ly in cc["ly_do_tru_diem"])
        ma_list = ", ".join(ly["ma"] for ly in cc["ly_do_tru_diem"])
        ly_do_text = "; ".join(f"[{ly['ma']}] {ly['ly_do']}" for ly in cc["ly_do_tru_diem"])

        ws5.cell(row=r, column=1, value=i).border = border
        ws5.cell(row=r, column=2, value=cc["ho_ten"]).border = border
        ws5.cell(row=r, column=3, value=cc["ma_cc"]).border = border
        ws5.cell(row=r, column=4, value=cc["don_vi"]).border = border

        tong_cell = ws5.cell(row=r, column=5, value=cc["tong_diem"])
        tong_cell.border = border
        tong_cell.alignment = center_alignment
        if cc["tong_diem"] < 20:
            tong_cell.fill = PatternFill("solid", fgColor="FFC7CE")
            tong_cell.font = Font(bold=True, color="9C0006")

        ws5.cell(row=r, column=6, value=len(cc["ly_do_tru_diem"])).border = border
        ws5.cell(row=r, column=6).alignment = center_alignment

        ws5.cell(row=r, column=7, value=diem_tru).border = border
        ws5.cell(row=r, column=7).alignment = center_alignment
        ws5.cell(row=r, column=7).font = Font(bold=True)

        ws5.cell(row=r, column=8, value=ma_list).border = border
        ws5.cell(row=r, column=8).alignment = center_alignment

        # Cột 9: Trạng thái — TrangThaiDanhGia
        tt_raw = cc.get("dgt_trang_thai", "")
        tt_display = {
            "HOAN_THANH": "Đã duyệt",
            "CHO_PHE_DUYET": "Chờ CCT duyệt",
            "DA_TONG_HOP": "Đã tổng hợp",
            "CHO_TONG_HOP": "Chờ LĐ tổng hợp",
            "DANG_DANH_GIA": "Đang đánh giá",
            "CO_KIEN_NGHI": "Có kiến nghị",
        }.get(tt_raw, tt_raw or "—")
        tt_cell5 = ws5.cell(row=r, column=9, value=tt_display)
        tt_cell5.border = border
        tt_cell5.alignment = center_alignment
        if tt_raw in ("CHO_PHE_DUYET", "DA_TONG_HOP", "CHO_TONG_HOP"):
            tt_cell5.fill = PatternFill("solid", fgColor="FFEB9C")
            tt_cell5.font = Font(bold=True, color="9C5700")
        elif tt_raw == "HOAN_THANH":
            tt_cell5.fill = PatternFill("solid", fgColor="C6EFCE")
        elif tt_raw == "CO_KIEN_NGHI":
            tt_cell5.fill = PatternFill("solid", fgColor="FFC7CE")
            tt_cell5.font = Font(bold=True, color="9C0006")

        ws5.cell(row=r, column=10, value=ly_do_text).border = border
        ws5.cell(row=r, column=10).alignment = wrap_alignment
        ws5.row_dimensions[r].height = 40

    for c, w in [('A', 5), ('B', 25), ('C', 12), ('D', 30), ('E', 10), ('F', 12), ('G', 14), ('H', 18), ('I', 12), ('J', 60)]:
        ws5.column_dimensions[c].width = w

    # Bật autofilter để CCT/PCCT lọc nhanh
    if bi_tru_all_sorted:
        ws5.auto_filter.ref = f"A4:J{4 + len(bi_tru_all_sorted)}"

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def _get_data_01_tieu_chi_chung(db: AsyncSession, thang: int, nam: int) -> list:
    """Get data for report 01 - Tiêu chí chung."""
    from app.models.user_org import CongChuc, DonVi
    from app.models.kpi_assessment import DanhGiaThang, TieuChiChung

    # Lấy tất cả tiêu chí chung
    stmt_tc = select(TieuChiChung).where(TieuChiChung.is_active == True).order_by(TieuChiChung.ma_tieu_chi)
    result_tc = await db.execute(stmt_tc)
    all_tieu_chi = result_tc.scalars().all()

    # Lấy đánh giá tháng — TrangThaiDanhGia (cấp phiếu chung, KHÁC TrangThaiTieuChi).
    # 2026-05-14 (rev): bỏ filter trang_thai để không bị thiếu bản — production
    # data có nhiều trạng thái khác nhau (DANG_DANH_GIA / CO_KIEN_NGHI cũng có
    # tiêu_chi_chungs nếu CC đã chấm dở). Cờ "Trạng thái" trong sheet hiển thị
    # đầy đủ để CCT/PCCT phân biệt.
    stmt = (
        select(DanhGiaThang)
        .options(
            selectinload(DanhGiaThang.cong_chuc).selectinload(CongChuc.don_vi),
            selectinload(DanhGiaThang.tieu_chi_chungs),
        )
        .where(
            DanhGiaThang.thang == thang,
            DanhGiaThang.nam == nam,
            DanhGiaThang.is_deleted == False,
        )
    )
    result = await db.execute(stmt)
    danh_gias = result.scalars().all()

    cong_chuc_list = []

    for dg in danh_gias:
        if not dg.cong_chuc:
            continue
        if hasattr(dg.cong_chuc, 'is_active') and dg.cong_chuc.is_active == False:
            continue
        if hasattr(dg.cong_chuc, 'deleted_at') and dg.cong_chuc.deleted_at is not None:
            continue

        tc_danh_gia_map = {str(tcdg.tieu_chi_id): tcdg for tcdg in (dg.tieu_chi_chungs or [])}

        tong_diem = 0
        diem_nhom1 = 0
        diem_nhom2 = 0
        diem_nhom3 = 0
        has_nhom3 = False
        has_ghi_chu_nhom3 = False
        ly_do_tru_diem = []
        ghi_chu_nhom3_list = []

        for tc in all_tieu_chi:
            tcdg = tc_danh_gia_map.get(str(tc.id))

            is_achieved_ld = tcdg.is_achieved_ld if tcdg else None
            is_achieved_cc = tcdg.is_achieved_cc if tcdg else False
            final_achieved = is_achieved_ld if is_achieved_ld is not None else is_achieved_cc

            diem_ld = float(tcdg.diem_phe_duyet) if tcdg and tcdg.diem_phe_duyet is not None else None
            diem_cc = float(tcdg.diem_tu_cham) if tcdg and tcdg.diem_tu_cham else 0
            diem = diem_ld if diem_ld is not None else diem_cc

            ghi_chu = tcdg.ghi_chu_cc if tcdg else ""
            ghi_chu_ld = tcdg.ghi_chu_ld if tcdg and hasattr(tcdg, 'ghi_chu_ld') else ""
            ly_do_dieu_chinh = tcdg.ly_do_dieu_chinh if tcdg and hasattr(tcdg, 'ly_do_dieu_chinh') else ""

            tong_diem += diem

            if tc.nhom_tieu_chi == 1:
                diem_nhom1 += diem
            elif tc.nhom_tieu_chi == 2:
                diem_nhom2 += diem
            elif tc.nhom_tieu_chi == 3:
                diem_nhom3 += diem
                if final_achieved:
                    has_nhom3 = True
                    if ghi_chu:
                        has_ghi_chu_nhom3 = True
                        ghi_chu_nhom3_list.append(f"[{tc.ma_tieu_chi}] {ghi_chu}")

            if not final_achieved and float(tc.diem_toi_da) > 0:
                ly_do = ly_do_dieu_chinh or ghi_chu_ld or f"Không đạt tiêu chí {tc.ma_tieu_chi}"
                ly_do_tru_diem.append({
                    "ma": tc.ma_tieu_chi,
                    "ten": tc.ten_tieu_chi,
                    "diem_tru": float(tc.diem_toi_da),
                    "ly_do": ly_do,
                    "nhom": tc.nhom_tieu_chi,
                })

        # Trạng thái phiếu đánh giá (enum value hoặc str — chuẩn hoá str)
        tt_raw = dg.trang_thai
        if hasattr(tt_raw, "value"):
            tt_raw = tt_raw.value

        cong_chuc_list.append({
            "cong_chuc_id": str(dg.cong_chuc.id),
            "ho_ten": dg.cong_chuc.ho_ten,
            "ma_cc": dg.cong_chuc.ma_cc,
            "don_vi": dg.cong_chuc.don_vi.ten_don_vi if dg.cong_chuc.don_vi else "",
            "tong_diem": tong_diem,
            "diem_nhom1": diem_nhom1,
            "diem_nhom2": diem_nhom2,
            "diem_nhom3": diem_nhom3,
            "has_nhom3": has_nhom3,
            "has_ghi_chu_nhom3": has_ghi_chu_nhom3,
            "minh_chung_nhom3": "\n".join(ghi_chu_nhom3_list) if ghi_chu_nhom3_list else "",
            "ly_do_tru_diem": ly_do_tru_diem,
            "dgt_trang_thai": tt_raw or "",
        })

    return cong_chuc_list


# =============================================================================
# HELPER: GENERATE REPORT 02 - ĐIỂM KPI
# =============================================================================

async def _generate_report_02_diem_kpi(db: AsyncSession, thang: int, nam: int) -> io.BytesIO:
    """Generate Excel report 02 - Điểm KPI statistics."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from sqlalchemy import text

    # Get data
    data, so_ngay = await _get_data_02_diem_kpi(db, thang, nam)

    # Create Excel
    wb = Workbook()

    # Styles
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    percent_font = Font(bold=True, color="0070C0")
    alert_fill = PatternFill("solid", fgColor="FFC7CE")
    good_fill = PatternFill("solid", fgColor="C6EFCE")
    warn_fill = PatternFill("solid", fgColor="FFEB9C")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Phân loại
    total = len(data)
    total_co_kpi = len([cc for cc in data if cc["diem_kpi_70"] is not None])

    # Chỉ tính CC không phải lãnh đạo
    cc_list = [cc for cc in data if not cc.get("is_lanh_dao")]

    dat_kpi_70 = [cc for cc in cc_list if cc["diem_kpi_70"] is not None and cc["diem_kpi_70"] >= 70]
    chua_dat_kpi_70 = [cc for cc in cc_list if cc["diem_kpi_70"] is not None and cc["diem_kpi_70"] < 70]

    # Phân loại lý do chưa đạt — 2026-05-14: bỏ "Do SP chưa đạt" + "Vượt KPI
    # bất thường" vì V2_PL3 không còn target SP cố định (ngày × 96).
    chua_dat_do_chat_luong = [cc for cc in chua_dat_kpi_70 if cc["loi_cl"] > 0]
    chua_dat_do_tien_do = [cc for cc in chua_dat_kpi_70 if cc["loi_td"] > 0]

    def pct(count, base=total_co_kpi):
        return f"{count/base*100:.1f}%" if base > 0 else "0%"

    # SHEET 1: TỔNG HỢP
    ws1 = wb.active
    ws1.title = "Tổng hợp"

    ws1['A1'] = f"2. THỐNG KÊ ĐIỂM KPI - THÁNG {thang}/{nam}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:E1')

    ws1['A2'] = f"Tổng CC có dữ liệu KPI: {total_co_kpi} | Số ngày trong tháng: {so_ngay}"
    ws1['A2'].font = Font(bold=True)

    row = 4
    headers = ["Nhóm", "Số lượng", "Tỷ lệ", "Ghi chú"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    rows_data = [
        ("I. Đạt KPI 70 điểm", len(dat_kpi_70), pct(len(dat_kpi_70)), ""),
        ("II. Chưa đạt KPI 70 điểm", len(chua_dat_kpi_70), pct(len(chua_dat_kpi_70)), ""),
        ("   - Do CL bị trừ", len(chua_dat_do_chat_luong), pct(len(chua_dat_do_chat_luong)), "Có lỗi chất lượng"),
        ("   - Do TĐ bị trừ", len(chua_dat_do_tien_do), pct(len(chua_dat_do_tien_do)), "Có lỗi tiến độ"),
    ]

    for i, (nhom, sl, tl, gc) in enumerate(rows_data):
        r = row + 1 + i
        ws1.cell(row=r, column=1, value=nhom).border = border
        ws1.cell(row=r, column=2, value=sl).border = border
        ws1.cell(row=r, column=2).alignment = center_alignment
        ws1.cell(row=r, column=3, value=tl).border = border
        ws1.cell(row=r, column=3).alignment = center_alignment
        ws1.cell(row=r, column=3).font = percent_font
        ws1.cell(row=r, column=4, value=gc).border = border
        if i in [0, 1]:
            ws1.cell(row=r, column=1).font = Font(bold=True)

    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 40

    # SHEET 2: CHƯA ĐẠT KPI (đổi sheet name từ "Chưa đạt KPI" giữ nguyên).
    # Đã bỏ sheet "Vượt KPI bất thường" — V2_PL3 không có concept SP được giao
    # cố định nên "vượt" không có ý nghĩa.
    ws3 = wb.create_sheet("Chưa đạt KPI")

    ws3['A1'] = f"DANH SÁCH CHƯA ĐẠT KPI 70 ĐIỂM - THÁNG {thang}/{nam}"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:J1')

    ws3['A2'] = "Tổng điểm = Σ(so_sp_goc_quy_doi) bản DA_PHE_DUYET (đơn vị: điểm SP quy đổi). KPI < 70 do lỗi CL/TĐ."
    ws3['A2'].font = Font(italic=True, color="666666")

    headers3 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "Tổng điểm", "Lỗi CL", "Lỗi TĐ", "Điểm KPI", "Lý do chính"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    chua_dat_sorted = sorted(chua_dat_kpi_70, key=lambda x: x["diem_kpi_70"] or 0)

    for i, cc in enumerate(chua_dat_sorted, 1):
        r = 3 + i
        ws3.cell(row=r, column=1, value=i).border = border
        ws3.cell(row=r, column=2, value=cc["ho_ten"]).border = border
        ws3.cell(row=r, column=3, value=cc["ma_cc"]).border = border
        ws3.cell(row=r, column=4, value=cc["don_vi"]).border = border
        # Cột 5: Tổng SP kê khai = SUM so_sp_goc_quy_doi (sp_hoan_thanh)
        ws3.cell(row=r, column=5, value=f"{cc['sp_hoan_thanh']:,.0f}").border = border

        loi_cl_cell = ws3.cell(row=r, column=6, value=cc["loi_cl"])
        loi_cl_cell.border = border
        if cc["loi_cl"] > 0:
            loi_cl_cell.fill = alert_fill

        loi_td_cell = ws3.cell(row=r, column=7, value=cc["loi_td"])
        loi_td_cell.border = border
        if cc["loi_td"] > 0:
            loi_td_cell.fill = warn_fill

        ws3.cell(row=r, column=8, value=cc["diem_kpi_70"]).border = border

        # Lý do chính — bỏ "SP chưa đạt" vì V2_PL3 không có target SP cố định
        ly_do = []
        if cc["loi_cl"] > 0:
            ly_do.append(f"CL -{cc['loi_cl']} lỗi")
        if cc["loi_td"] > 0:
            ly_do.append(f"TĐ -{cc['loi_td']} lỗi")
        ws3.cell(row=r, column=9, value=", ".join(ly_do) if ly_do else "Khác").border = border

    for c, w in [('A', 5), ('B', 25), ('C', 12), ('D', 25), ('E', 16), ('F', 8), ('G', 8), ('H', 10), ('I', 25)]:
        ws3.column_dimensions[c].width = w

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def _get_data_02_diem_kpi(db: AsyncSession, thang: int, nam: int) -> tuple:
    """Get data for report 02 - Điểm KPI."""
    from sqlalchemy import text

    so_ngay_trong_thang = calendar.monthrange(nam, thang)[1]

    # Load KPI từ chi_tiet_xep_loai
    kpi_result = await db.execute(text("""
        SELECT ct.cong_chuc_id::text, ct.diem_kpi, ct.diem_tong,
               ct.is_lanh_dao, ct.xep_loai_he_thong,
               ct.so_ngay_lam_viec, ct.so_ngay_nghi,
               cc.ho_ten, cc.ma_cc, dv.ten_don_vi
        FROM chi_tiet_xep_loai ct
        JOIN bao_cao_xep_loai bc ON bc.id = ct.bao_cao_id
        JOIN cong_chuc cc ON cc.id = ct.cong_chuc_id
        LEFT JOIN don_vi dv ON dv.id = cc.don_vi_id
        WHERE bc.thang = :thang AND bc.nam = :nam
              AND cc.is_active = true
    """), {"thang": thang, "nam": nam})

    kpi_list = []
    for row in kpi_result:
        kpi_list.append({
            "cong_chuc_id": row[0],
            "diem_kpi_70": float(row[1]) if row[1] is not None else None,
            "diem_tong_100": float(row[2]) if row[2] is not None else None,
            "is_lanh_dao": row[3],
            "xep_loai": row[4],
            "so_ngay_lv": float(row[5]) if row[5] is not None else None,
            "so_ngay_nghi": float(row[6]) if row[6] is not None else None,
            "ho_ten": row[7],
            "ma_cc": row[8],
            "don_vi": row[9] or "",
        })

    # Load nghỉ phép
    nghi_result = await db.execute(text("""
        SELECT cong_chuc_id::text, COALESCE(SUM(so_ngay), 0) as tong_nghi
        FROM dang_ky_nghi
        WHERE thang_ap_dung = :thang AND nam_ap_dung = :nam
              AND trang_thai = 'DA_PHE_DUYET' AND is_deleted = false
        GROUP BY cong_chuc_id
    """), {"thang": thang, "nam": nam})
    nghi_by_cc = {row[0]: float(row[1]) for row in nghi_result}

    # Load SP quy đổi CC
    sp_result = await db.execute(text("""
        SELECT cong_chuc_id::text,
               COALESCE(SUM(so_sp_goc_quy_doi), 0) as tong_sp,
               COALESCE(SUM(so_sp_goc_quy_doi * GREATEST(0, 1 - COALESCE(so_loi_chat_luong, 0) * 0.25)), 0) as sp_cl,
               COALESCE(SUM(so_sp_goc_quy_doi * GREATEST(0, 1 - COALESCE(so_loi_tien_do, 0) * 0.25)), 0) as sp_td,
               SUM(so_loi_chat_luong) as loi_cl,
               SUM(so_loi_tien_do) as loi_td
        FROM ke_khai_cong_viec
        WHERE thang = :thang AND nam = :nam
              AND trang_thai = 'DA_PHE_DUYET' AND is_deleted = false
        GROUP BY cong_chuc_id
    """), {"thang": thang, "nam": nam})
    sp_by_cc = {}
    for row in sp_result:
        sp_by_cc[row[0]] = {
            "tong_sp": float(row[1]),
            "sp_cl": float(row[2]),
            "sp_td": float(row[3]),
            "loi_cl": int(row[4] or 0),
            "loi_td": int(row[5] or 0),
        }

    # Tính toán cho từng CC
    for cc in kpi_list:
        cc_id = cc["cong_chuc_id"]
        tong_nghi = nghi_by_cc.get(cc_id, 0)
        sp_data = sp_by_cc.get(cc_id, {"tong_sp": 0, "sp_cl": 0, "sp_td": 0, "loi_cl": 0, "loi_td": 0})

        # SP được giao = (ngày trong tháng - nghỉ) × 96
        sp_duoc_giao = (so_ngay_trong_thang - tong_nghi) * 96
        cc["sp_duoc_giao"] = sp_duoc_giao
        cc["sp_hoan_thanh"] = sp_data["tong_sp"]
        cc["sp_cl"] = sp_data["sp_cl"]
        cc["sp_td"] = sp_data["sp_td"]
        cc["loi_cl"] = sp_data["loi_cl"]
        cc["loi_td"] = sp_data["loi_td"]

        # Tỷ lệ vượt KPI
        if sp_duoc_giao > 0:
            cc["ty_le_vuot"] = (sp_data["tong_sp"] - sp_duoc_giao) / sp_duoc_giao * 100
        else:
            cc["ty_le_vuot"] = 0

    return kpi_list, so_ngay_trong_thang

# =============================================================================
# HELPER: GENERATE REPORT 03 - LÃNH ĐẠO BỊ TRỪ d,đ,e
# =============================================================================

async def _generate_report_03_lanh_dao_dde(db: AsyncSession, thang: int, nam: int) -> io.BytesIO:
    """Generate Excel report 03 - Lãnh đạo bị trừ điểm d,đ,e."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from sqlalchemy import text

    # Get data
    data = await _get_data_03_lanh_dao_dde(db, thang, nam)

    # Create Excel
    wb = Workbook()

    # Styles
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    percent_font = Font(bold=True, color="0070C0")
    alert_fill = PatternFill("solid", fgColor="FFC7CE")
    good_fill = PatternFill("solid", fgColor="C6EFCE")
    warn_fill = PatternFill("solid", fgColor="FFEB9C")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Phân loại
    total_ld = len(data)
    bi_tru_d = [ld for ld in data if ld["bi_tru_d"]]
    bi_tru_dd = [ld for ld in data if ld["bi_tru_dd"]]
    bi_tru_e = [ld for ld in data if ld["bi_tru_e"]]
    bi_tru_any = [ld for ld in data if ld["tong_bi_tru"] > 0]
    khong_bi_tru = [ld for ld in data if ld["tong_bi_tru"] == 0]

    def pct(count):
        return f"{count/total_ld*100:.1f}%" if total_ld > 0 else "0%"

    # SHEET 1: TỔNG HỢP
    ws1 = wb.active
    ws1.title = "Tổng hợp"

    ws1['A1'] = f"3. THỐNG KÊ LÃNH ĐẠO BỊ TRỪ ĐIỂM d, đ, e - THÁNG {thang}/{nam}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:E1')

    ws1['A2'] = f"Tổng số lãnh đạo: {total_ld}"
    ws1['A2'].font = Font(bold=True)

    row = 4
    headers = ["Tiêu chí", "Số lượng bị trừ", "Tỷ lệ", "Ghi chú"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    rows_data = [
        ("d - Kết quả đơn vị", len(bi_tru_d), pct(len(bi_tru_d)), "Đơn vị không hoàn thành nhiệm vụ"),
        ("đ - Tổ chức triển khai", len(bi_tru_dd), pct(len(bi_tru_dd)), "Triển khai không đạt yêu cầu"),
        ("e - Đoàn kết nội bộ", len(bi_tru_e), pct(len(bi_tru_e)), "Có vấn đề đoàn kết"),
        ("TỔNG BỊ TRỪ (ít nhất 1 tiêu chí)", len(bi_tru_any), pct(len(bi_tru_any)), ""),
        ("Không bị trừ", len(khong_bi_tru), pct(len(khong_bi_tru)), ""),
    ]

    for i, (nhom, sl, tl, gc) in enumerate(rows_data):
        r = row + 1 + i
        ws1.cell(row=r, column=1, value=nhom).border = border
        ws1.cell(row=r, column=2, value=sl).border = border
        ws1.cell(row=r, column=2).alignment = center_alignment
        ws1.cell(row=r, column=3, value=tl).border = border
        ws1.cell(row=r, column=3).alignment = center_alignment
        ws1.cell(row=r, column=3).font = percent_font
        ws1.cell(row=r, column=4, value=gc).border = border

        if i == 3:  # Tổng bị trừ
            ws1.cell(row=r, column=1).font = Font(bold=True)
            if sl > 0:
                ws1.cell(row=r, column=2).fill = alert_fill

    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 40

    # SHEET 2: DANH SÁCH BỊ TRỪ
    ws2 = wb.create_sheet("Danh sách bị trừ")

    ws2['A1'] = f"DANH SÁCH LÃNH ĐẠO BỊ TRỪ ĐIỂM d, đ, e - THÁNG {thang}/{nam}"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:M1')

    ws2['A2'] = ("Cột 'Điểm KPI' = engine chính thức (chỉ DA_PHE_DUYET). "
                 "Cột 'Điểm KPI (tạm tính)' = nếu bản CHỜ_DUYỆT cũng vào điểm.")
    ws2['A2'].font = Font(italic=True, color="666666")

    headers2 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "Chức vụ", "d", "đ", "e",
                "Tổng trừ", "Lý do", "Trạng thái", "Điểm KPI", "Điểm KPI (tạm tính)"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    # Sắp xếp theo số tiêu chí bị trừ giảm dần
    bi_tru_sorted = sorted(bi_tru_any, key=lambda x: x["tong_bi_tru"], reverse=True)

    for i, ld in enumerate(bi_tru_sorted, 1):
        r = 3 + i
        ws2.cell(row=r, column=1, value=i).border = border
        ws2.cell(row=r, column=2, value=ld["ho_ten"]).border = border
        ws2.cell(row=r, column=3, value=ld["ma_cc"]).border = border
        ws2.cell(row=r, column=4, value=ld["don_vi"]).border = border
        ws2.cell(row=r, column=5, value=ld["vai_tro"]).border = border

        # d
        d_cell = ws2.cell(row=r, column=6, value="✗" if ld["bi_tru_d"] else "✓")
        d_cell.border = border
        d_cell.alignment = center_alignment
        if ld["bi_tru_d"]:
            d_cell.fill = alert_fill
            d_cell.font = Font(bold=True, color="9C0006")
        else:
            d_cell.fill = good_fill

        # đ
        dd_cell = ws2.cell(row=r, column=7, value="✗" if ld["bi_tru_dd"] else "✓")
        dd_cell.border = border
        dd_cell.alignment = center_alignment
        if ld["bi_tru_dd"]:
            dd_cell.fill = alert_fill
            dd_cell.font = Font(bold=True, color="9C0006")
        else:
            dd_cell.fill = good_fill

        # e
        e_cell = ws2.cell(row=r, column=8, value="✗" if ld["bi_tru_e"] else "✓")
        e_cell.border = border
        e_cell.alignment = center_alignment
        if ld["bi_tru_e"]:
            e_cell.fill = alert_fill
            e_cell.font = Font(bold=True, color="9C0006")
        else:
            e_cell.fill = good_fill

        # Tổng trừ
        ws2.cell(row=r, column=9, value=ld["tong_bi_tru"]).border = border
        ws2.cell(row=r, column=9).alignment = center_alignment
        ws2.cell(row=r, column=9).font = Font(bold=True)

        # Lý do
        ly_do = []
        if ld["bi_tru_d"]:
            ly_do.append(f"d: {ld['d_ghi_chu']}" if ld['d_ghi_chu'] else "d")
        if ld["bi_tru_dd"]:
            ly_do.append(f"đ: {ld['dd_ghi_chu']}" if ld['dd_ghi_chu'] else "đ")
        if ld["bi_tru_e"]:
            ly_do.append(f"e: {ld['e_ghi_chu']}" if ld['e_ghi_chu'] else "e")
        ws2.cell(row=r, column=10, value="; ".join(ly_do)).border = border
        ws2.cell(row=r, column=10).alignment = wrap_alignment

        # Trạng thái d/đ/e: DA_PHE_DUYET (xanh) | CHO_PHE_DUYET (vàng)
        tt_raw = ld.get("dde_trang_thai", "")
        tt_display = {
            "DA_PHE_DUYET": "Đã duyệt",
            "CHO_PHE_DUYET": "Chờ duyệt",
        }.get(tt_raw, tt_raw or "—")
        tt_cell = ws2.cell(row=r, column=11, value=tt_display)
        tt_cell.border = border
        tt_cell.alignment = center_alignment
        if tt_raw == "CHO_PHE_DUYET":
            tt_cell.fill = PatternFill("solid", fgColor="FFEB9C")
            tt_cell.font = Font(bold=True, color="9C5700")
        elif tt_raw == "DA_PHE_DUYET":
            tt_cell.fill = good_fill

        ws2.cell(row=r, column=12, value=ld.get("diem_kpi")).border = border

        # Điểm KPI (tạm tính) — broadened CHO+DA
        tt_val = ld.get("diem_kpi_tam_tinh")
        tt_cell_dt = ws2.cell(row=r, column=13,
                              value=round(tt_val, 2) if tt_val is not None else "—")
        tt_cell_dt.border = border
        tt_cell_dt.alignment = center_alignment
        if tt_val is not None and ld.get("diem_kpi") is not None:
            if tt_val < float(ld["diem_kpi"]) - 0.01:
                # Tạm tính thấp hơn chính thức (có CHO_PHE_DUYET làm giảm điểm)
                tt_cell_dt.fill = warn_fill
                tt_cell_dt.font = Font(bold=True, color="9C5700")

        ws2.row_dimensions[r].height = 30

    for c, w in [('A', 5), ('B', 25), ('C', 12), ('D', 25), ('E', 15),
                 ('F', 6), ('G', 6), ('H', 6), ('I', 10), ('J', 50),
                 ('K', 12), ('L', 10), ('M', 16)]:
        ws2.column_dimensions[c].width = w

    # SHEET 3: TẤT CẢ LÃNH ĐẠO
    ws3 = wb.create_sheet("Tất cả lãnh đạo")

    ws3['A1'] = f"DANH SÁCH TẤT CẢ LÃNH ĐẠO - THÁNG {thang}/{nam}"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:M1')

    ws3['A2'] = ("Điểm chính thức = engine production (DA_PHE_DUYET). "
                 "Điểm tạm tính = nếu bản CHỜ_DUYỆT cũng vào điểm (mức 2: a/b/c + d/đ/e CHO+DA).")
    ws3['A2'].font = Font(italic=True, color="666666")

    headers3 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "Chức vụ", "d", "đ", "e",
                "Trạng thái", "Điểm KPI", "Điểm tổng", "Điểm KPI (tạm tính)", "Điểm tổng (tạm tính)"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    for i, ld in enumerate(data, 1):
        r = 3 + i
        ws3.cell(row=r, column=1, value=i).border = border
        ws3.cell(row=r, column=2, value=ld["ho_ten"]).border = border
        ws3.cell(row=r, column=3, value=ld["ma_cc"]).border = border
        ws3.cell(row=r, column=4, value=ld["don_vi"]).border = border
        ws3.cell(row=r, column=5, value=ld["vai_tro"]).border = border

        for col_idx, key in [(6, "bi_tru_d"), (7, "bi_tru_dd"), (8, "bi_tru_e")]:
            cell = ws3.cell(row=r, column=col_idx, value="✗" if ld[key] else "✓")
            cell.border = border
            cell.alignment = center_alignment
            if ld[key]:
                cell.fill = alert_fill
            else:
                cell.fill = good_fill

        tt_raw = ld.get("dde_trang_thai", "")
        tt_display = {
            "DA_PHE_DUYET": "Đã duyệt",
            "CHO_PHE_DUYET": "Chờ duyệt",
        }.get(tt_raw, tt_raw or "—")
        tt_cell = ws3.cell(row=r, column=9, value=tt_display)
        tt_cell.border = border
        tt_cell.alignment = center_alignment
        if tt_raw == "CHO_PHE_DUYET":
            tt_cell.fill = PatternFill("solid", fgColor="FFEB9C")
            tt_cell.font = Font(bold=True, color="9C5700")
        elif tt_raw == "DA_PHE_DUYET":
            tt_cell.fill = good_fill

        ws3.cell(row=r, column=10, value=ld["diem_kpi"]).border = border
        ws3.cell(row=r, column=11, value=ld["diem_tong"]).border = border

        # Cột 12-13: Điểm KPI (tạm tính) + Điểm tổng (tạm tính)
        kpi_tt = ld.get("diem_kpi_tam_tinh")
        tong_tt = ld.get("diem_tong_tam_tinh")
        kpi_cell = ws3.cell(row=r, column=12,
                            value=round(kpi_tt, 2) if kpi_tt is not None else "—")
        kpi_cell.border = border
        kpi_cell.alignment = center_alignment
        tong_cell_tt = ws3.cell(row=r, column=13,
                                value=round(tong_tt, 2) if tong_tt is not None else "—")
        tong_cell_tt.border = border
        tong_cell_tt.alignment = center_alignment
        # Highlight nếu tạm tính < chính thức (có bản CHO_PHE_DUYET làm giảm điểm)
        if kpi_tt is not None and ld.get("diem_kpi") is not None and kpi_tt < float(ld["diem_kpi"]) - 0.01:
            kpi_cell.fill = warn_fill
            kpi_cell.font = Font(bold=True, color="9C5700")
            tong_cell_tt.fill = warn_fill
            tong_cell_tt.font = Font(bold=True, color="9C5700")

    for c, w in [('A', 5), ('B', 25), ('C', 12), ('D', 25), ('E', 15),
                 ('F', 6), ('G', 6), ('H', 6), ('I', 12), ('J', 10), ('K', 10),
                 ('L', 14), ('M', 16)]:
        ws3.column_dimensions[c].width = w

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def _calc_diem_kpi_ld_tam_tinh(
    db: AsyncSession,
    cong_chuc_id: UUID,
    thang: int,
    nam: int,
    dde_map: dict,
) -> Optional[float]:
    """Tính Điểm KPI 70đ "tạm tính" cho LĐ — broadened CHO_PHE_DUYET + DA_PHE_DUYET.

    CHỈ DÙNG TRONG BÁO CÁO. Không ghi DB. Engine production chính thức vẫn
    là DA_PHE_DUYET only.

    Cách tính (theo lựa chọn user 2026-05-14 — Mức 2):
    - a/b/c: gọi engine sẵn có với tam_tinh=True (LĐ V2 dùng calc_kpi_lanh_dao_v2
      với scope mở rộng; LĐ V1/HĐ 111 dùng KeKhaiLanhDao filter CHO+DA).
      Lưu ý: engine tam_tinh=True của V2 cũng gồm NHAP — chấp nhận tradeoff
      vì NHAP rất hiếm trên production (CC kê khai xong thường submit ngay).
    - d/đ/e: lấy từ dde_map (đã broadened CHO+DA, do _load_dde_finals trả về).
    - KPI = (a+b+c+d+đ+e) / 6 × 70, cap 70.

    Returns:
        float diem_kpi_70 hoặc None nếu lỗi tính toán.
    """
    from app.core.kpi_lanh_dao_v2 import calc_kpi_lanh_dao_v2, is_kpi_lanh_dao_v2_active

    a = b = c = 0.0
    try:
        if is_kpi_lanh_dao_v2_active(thang, nam):
            v2 = await calc_kpi_lanh_dao_v2(db, cong_chuc_id, thang, nam, tam_tinh=True)
            a, b, c = float(v2["a"]), float(v2["b"]), float(v2["c"])
        else:
            # V1 LĐ / HĐ 111 — tính từ KeKhaiLanhDao với CHO+DA (đúng Mức 2)
            from sqlalchemy import select as _select
            stmt = (
                _select(
                    KeKhaiLanhDao.trang_thai_hoan_thanh,
                    KeKhaiLanhDao.so_loi_chat_luong,
                    KeKhaiLanhDao.so_loi_tien_do,
                )
                .where(KeKhaiLanhDao.cong_chuc_id == cong_chuc_id)
                .where(KeKhaiLanhDao.thang == thang)
                .where(KeKhaiLanhDao.nam == nam)
                .where(KeKhaiLanhDao.is_deleted == False)
                .where(KeKhaiLanhDao.trang_thai.in_((
                    TrangThaiKeKhaiLD.CHO_PHE_DUYET.value,
                    TrangThaiKeKhaiLD.DA_PHE_DUYET.value,
                )))
            )
            rows = (await db.execute(stmt)).all()
            tong_cv = len(rows)
            if tong_cv > 0:
                tong_ht = sum(
                    1 for r in rows if r[0] == TrangThaiHoanThanh.DA_HOAN_THANH
                )
                tong_diem_cl = sum(max(0.0, 1.0 - (r[1] or 0) * 0.25) for r in rows)
                tong_diem_td = sum(max(0.0, 1.0 - (r[2] or 0) * 0.25) for r in rows)
                a = min(tong_ht / tong_cv, 1.0)
                b = min(tong_diem_td / tong_cv, 1.0)
                c = min(tong_diem_cl / tong_cv, 1.0)
    except Exception as e:
        logger.warning(
            f"_calc_diem_kpi_ld_tam_tinh failed cho cc={cong_chuc_id} {thang}/{nam}: {e}"
        )
        return None

    # d/đ/e — broadened từ dde_map (mặc định 100 nếu LĐ chưa có bản nào)
    dde = dde_map.get(str(cong_chuc_id), {})
    d = dde.get("d_final", 100) / 100.0
    dd = dde.get("dd_final", 100) / 100.0
    e = dde.get("e_final", 100) / 100.0

    kpi_ratio = (a + b + c + d + dd + e) / 6
    return min(70.0, kpi_ratio * 70)


async def _load_dde_finals(db: AsyncSession, thang: int, nam: int) -> dict:
    """Helper dùng chung — load d/đ/e final cho mọi LĐ trong tháng.

    Schema thực tế (xác minh 2026-05-13):
      d_ket_qua_don_vi / dd_to_chuc_trien_khai / e_doan_ket_noi_bo: INT 50|100 (tự đánh giá)
      d_phe_duyet / dd_phe_duyet / e_phe_duyet: INT 50|100|NULL (sau phê duyệt)
    Giá trị final = COALESCE(*_phe_duyet, *_ket_qua_don_vi/_to_chuc/_doan_ket).
    Bị trừ ⟺ final == 50.

    Filter trang_thai IN ('CHO_PHE_DUYET', 'DA_PHE_DUYET') — bao gồm cả bản
    đang chờ phê duyệt (theo yêu cầu CCT 2026-05-13). Bản CHO_PHE_DUYET dùng
    giá trị tự đánh giá (d_phe_duyet còn NULL). Báo cáo gắn cờ trang_thai để
    user phân biệt bản chính thức vs tạm tính.
    LĐ không có bản ghi → mặc định 100 (không bị trừ).

    Returns:
        dict {cong_chuc_id_str → {d_final, dd_final, e_final,
                                  d_ghi_chu, dd_ghi_chu, e_ghi_chu,
                                  trang_thai}}
    """
    from sqlalchemy import text

    result = await db.execute(text("""
        SELECT cong_chuc_id::text,
               COALESCE(d_phe_duyet, d_ket_qua_don_vi)        AS d_final,
               COALESCE(dd_phe_duyet, dd_to_chuc_trien_khai)  AS dd_final,
               COALESCE(e_phe_duyet, e_doan_ket_noi_bo)       AS e_final,
               d_ghi_chu, dd_ghi_chu, e_ghi_chu,
               trang_thai
        FROM danh_gia_dde
        WHERE thang = :thang AND nam = :nam
              AND trang_thai IN ('CHO_PHE_DUYET', 'DA_PHE_DUYET')
    """), {"thang": thang, "nam": nam})

    dde_by_cc = {}
    for row in result:
        dde_by_cc[row[0]] = {
            "d_final": int(row[1]),    # 50 hoặc 100
            "dd_final": int(row[2]),
            "e_final": int(row[3]),
            "d_ghi_chu": row[4] or "",
            "dd_ghi_chu": row[5] or "",
            "e_ghi_chu": row[6] or "",
            "trang_thai": row[7],
        }
    return dde_by_cc


async def _get_data_03_lanh_dao_dde(db: AsyncSession, thang: int, nam: int) -> list:
    """Get data for report 03 - Lãnh đạo bị trừ d,đ,e.

    Fix 2026-05-13: dùng helper _load_dde_finals. Trước đó so sánh
    int 50/100 với False → không bao giờ phát hiện bị trừ.
    """
    from sqlalchemy import text

    # Lấy danh sách lãnh đạo từ chi_tiet_xep_loai (có thêm diem_tieu_chi_chung
    # để tính Điểm tổng tạm tính).
    ld_result = await db.execute(text("""
        SELECT ct.cong_chuc_id::text, cc.ho_ten, cc.ma_cc, dv.ten_don_vi,
               vt.ten_vai_tro, ct.diem_kpi, ct.diem_tong, ct.xep_loai_he_thong,
               ct.diem_tieu_chi_chung
        FROM chi_tiet_xep_loai ct
        JOIN bao_cao_xep_loai bc ON bc.id = ct.bao_cao_id
        JOIN cong_chuc cc ON cc.id = ct.cong_chuc_id
        LEFT JOIN don_vi dv ON dv.id = cc.don_vi_id
        LEFT JOIN vai_tro vt ON vt.id = cc.vai_tro_id
        WHERE bc.thang = :thang AND bc.nam = :nam
              AND ct.is_lanh_dao = true
              AND cc.is_active = true
    """), {"thang": thang, "nam": nam})

    lanh_dao_list = []
    for row in ld_result:
        lanh_dao_list.append({
            "cong_chuc_id": row[0],
            "ho_ten": row[1],
            "ma_cc": row[2],
            "don_vi": row[3] or "",
            "vai_tro": row[4] or "",
            "diem_kpi": float(row[5]) if row[5] is not None else None,
            "diem_tong": float(row[6]) if row[6] is not None else None,
            "xep_loai": row[7],
            "diem_tcc": float(row[8]) if row[8] is not None else 0.0,
        })

    dde_by_cc = await _load_dde_finals(db, thang, nam)

    # Gộp dữ liệu — LĐ không có bản ghi (CHO_PHE_DUYET hoặc DA_PHE_DUYET) →
    # mặc định 100 (không bị trừ), trạng thái rỗng.
    for ld in lanh_dao_list:
        cc_id = ld["cong_chuc_id"]
        dde = dde_by_cc.get(cc_id, {})
        ld["d_final"] = dde.get("d_final", 100)
        ld["d_ghi_chu"] = dde.get("d_ghi_chu", "")
        ld["dd_final"] = dde.get("dd_final", 100)
        ld["dd_ghi_chu"] = dde.get("dd_ghi_chu", "")
        ld["e_final"] = dde.get("e_final", 100)
        ld["e_ghi_chu"] = dde.get("e_ghi_chu", "")
        ld["dde_trang_thai"] = dde.get("trang_thai", "")

        # Bị trừ ⟺ final == 50 (giá trị int trong DanhGiaDDE, không phải boolean)
        ld["bi_tru_d"] = (ld["d_final"] == 50)
        ld["bi_tru_dd"] = (ld["dd_final"] == 50)
        ld["bi_tru_e"] = (ld["e_final"] == 50)
        ld["tong_bi_tru"] = sum([ld["bi_tru_d"], ld["bi_tru_dd"], ld["bi_tru_e"]])

        # Điểm KPI / Điểm tổng "tạm tính" — broadened CHO+DA (CHỈ trong báo cáo).
        # Đặt cùng UUID type vì _calc_diem_kpi_ld_tam_tinh expect UUID.
        from uuid import UUID as _UUID
        try:
            _cc_uuid = _UUID(cc_id)
        except Exception:
            _cc_uuid = cc_id
        diem_kpi_tt = await _calc_diem_kpi_ld_tam_tinh(
            db, _cc_uuid, thang, nam, dde_by_cc
        )
        ld["diem_kpi_tam_tinh"] = diem_kpi_tt
        ld["diem_tong_tam_tinh"] = (
            (ld["diem_tcc"] + diem_kpi_tt) if diem_kpi_tt is not None else None
        )

    return lanh_dao_list


# =============================================================================
# HELPER: GENERATE REPORT 04 - KHỐI LƯỢNG CÔNG VIỆC
# =============================================================================

async def _generate_report_04_khoi_luong_cv(db: AsyncSession, thang: int, nam: int) -> io.BytesIO:
    """Generate Excel report 04 - Khối lượng công việc (V2 native, 2026-05-14).

    4 sheet:
    1. Tổng hợp toàn Chi cục (chỉ tiêu chính)
    2. Phân bố theo Nhóm PL3 (1-5)
    3. Ranking đơn vị (sort theo tổng điểm)
    4. Pivot đơn vị × Nhóm PL3
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = await _get_data_04_khoi_luong_cv(db, thang, nam)
    tong_hop = data["tong_hop"]
    nhom_pl3 = data["nhom_pl3"]
    ranking_dv = data["ranking_dv"]
    pivot_dv_nhom = data["pivot_dv_nhom"]
    so_ngay = data["so_ngay"]

    wb = Workbook()

    # Styles
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=13, color="2F5496")
    percent_font = Font(bold=True, color="0070C0")
    alert_fill = PatternFill("solid", fgColor="FFC7CE")
    warn_fill = PatternFill("solid", fgColor="FFEB9C")
    good_fill = PatternFill("solid", fgColor="C6EFCE")
    gold_fill = PatternFill("solid", fgColor="FFD966")  # top 3
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')
    wrap_top = Alignment(wrap_text=True, vertical='top')

    NHOM_FILLS = {
        1: PatternFill("solid", fgColor="C6EFCE"),  # đơn giản — xanh
        2: PatternFill("solid", fgColor="D9EAD3"),
        3: PatternFill("solid", fgColor="FFEB9C"),
        4: PatternFill("solid", fgColor="FFC7CE"),
        5: PatternFill("solid", fgColor="E6B8AF"),  # rất khó — đậm
    }

    # =========================================================================
    # SHEET 1: TỔNG HỢP TOÀN CHI CỤC
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Tổng hợp"

    ws1['A1'] = f"4. KHỐI LƯỢNG CÔNG VIỆC - THÁNG {thang}/{nam}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:C1')

    ws1['A2'] = (f"Chi cục Hải quan Khu vực VIII | Số ngày trong tháng: {so_ngay}. "
                 "Dữ liệu V2_PL3 (kê khai chính thức từ 4/2026).")
    ws1['A2'].font = Font(italic=True, color="666666")

    rows_data = [
        ("Tổng số CC kê khai", f"{tong_hop['so_cc_ke_khai']:,}",
         f"{tong_hop['ty_le_ke_khai']:.1f}% trên {tong_hop['tong_bien_che']} biên chế"),
        ("Tổng số lượt kê khai", f"{tong_hop['so_luot_khai']:,}", ""),
        ("Tổng điểm SP toàn Chi cục", f"{tong_hop['tong_diem']:,.0f}", "đơn vị: điểm SP quy đổi"),
        ("Trung bình điểm / CC", f"{tong_hop['tb_diem_per_cc']:,.1f}", "tổng điểm ÷ số CC kê khai"),
        ("Số CC có lỗi chất lượng",
         f"{tong_hop['so_cc_co_loi_cl']:,}",
         f"{tong_hop['so_cc_co_loi_cl']/tong_hop['so_cc_ke_khai']*100:.1f}%" if tong_hop['so_cc_ke_khai'] > 0 else "0%"),
        ("Số CC có lỗi tiến độ",
         f"{tong_hop['so_cc_co_loi_td']:,}",
         f"{tong_hop['so_cc_co_loi_td']/tong_hop['so_cc_ke_khai']*100:.1f}%" if tong_hop['so_cc_ke_khai'] > 0 else "0%"),
    ]
    if tong_hop["linh_vuc_top"]:
        lv = tong_hop["linh_vuc_top"]
        rows_data.append((
            "Lĩnh vực nhiều SP nhất",
            f"LV {lv['ma_lv']} — {lv['ten_lv'][:40]}",
            f"{lv['tong_diem']:,.0f} điểm"
        ))

    row = 4
    headers = ["Chỉ tiêu", "Giá trị", "Ghi chú"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    for i, (label, val, note) in enumerate(rows_data):
        r = row + 1 + i
        ws1.cell(row=r, column=1, value=label).border = border
        ws1.cell(row=r, column=1).font = Font(bold=True)
        val_cell = ws1.cell(row=r, column=2, value=val)
        val_cell.border = border
        val_cell.alignment = center
        val_cell.font = Font(bold=True, color="0070C0")
        ws1.cell(row=r, column=3, value=note).border = border
        ws1.cell(row=r, column=3).alignment = wrap_top

    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 45

    # =========================================================================
    # SHEET 2: PHÂN BỐ THEO NHÓM PL3
    # =========================================================================
    ws2 = wb.create_sheet("Theo Nhóm PL3")

    ws2['A1'] = f"PHÂN BỐ KHỐI LƯỢNG THEO NHÓM PL3 - THÁNG {thang}/{nam}"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:E1')

    ws2['A2'] = "Nhóm PL3 phản ánh độ phức tạp công việc: 1 (đơn giản) → 5 (rất khó). Mỗi nhóm có khung điểm tối đa khác nhau."
    ws2['A2'].font = Font(italic=True, color="666666")

    headers2 = ["Nhóm", "Mô tả", "Khung điểm tối đa", "Số lượt khai", "Tổng điểm", "Tỷ lệ"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    tong_diem_nhom = sum(n["tong_diem"] for n in nhom_pl3)
    tong_luot_nhom = sum(n["so_luot"] for n in nhom_pl3)

    for i, n in enumerate(nhom_pl3):
        r = 5 + i
        nhom_cell = ws2.cell(row=r, column=1, value=n["nhom"])
        nhom_cell.border = border
        nhom_cell.alignment = center
        nhom_cell.font = Font(bold=True)
        nhom_cell.fill = NHOM_FILLS[n["nhom"]]

        ws2.cell(row=r, column=2, value=n["ten"]).border = border
        ws2.cell(row=r, column=3, value=n["khung"]).border = border
        ws2.cell(row=r, column=3).alignment = center
        ws2.cell(row=r, column=4, value=n["so_luot"]).border = border
        ws2.cell(row=r, column=4).alignment = center
        ws2.cell(row=r, column=5, value=f"{n['tong_diem']:,.0f}").border = border
        ws2.cell(row=r, column=5).alignment = center
        ws2.cell(row=r, column=5).font = Font(bold=True)

        pct_val = (n["tong_diem"] / tong_diem_nhom * 100) if tong_diem_nhom > 0 else 0
        pct_cell = ws2.cell(row=r, column=6, value=f"{pct_val:.1f}%")
        pct_cell.border = border
        pct_cell.alignment = center
        pct_cell.font = percent_font

    # Dòng tổng
    r = 5 + len(nhom_pl3)
    ws2.cell(row=r, column=1, value="TỔNG").border = border
    ws2.cell(row=r, column=1).font = Font(bold=True)
    ws2.cell(row=r, column=2).border = border
    ws2.cell(row=r, column=3).border = border
    ws2.cell(row=r, column=4, value=tong_luot_nhom).border = border
    ws2.cell(row=r, column=4).alignment = center
    ws2.cell(row=r, column=4).font = Font(bold=True)
    ws2.cell(row=r, column=5, value=f"{tong_diem_nhom:,.0f}").border = border
    ws2.cell(row=r, column=5).alignment = center
    ws2.cell(row=r, column=5).font = Font(bold=True)
    ws2.cell(row=r, column=6, value="100%").border = border
    ws2.cell(row=r, column=6).alignment = center
    ws2.cell(row=r, column=6).font = Font(bold=True, color="0070C0")

    for c, w in [('A', 8), ('B', 18), ('C', 18), ('D', 14), ('E', 16), ('F', 10)]:
        ws2.column_dimensions[c].width = w

    # =========================================================================
    # SHEET 3: RANKING ĐƠN VỊ
    # =========================================================================
    ws3 = wb.create_sheet("Ranking đơn vị")

    ws3['A1'] = f"XẾP HẠNG ĐƠN VỊ THEO TỔNG ĐIỂM - THÁNG {thang}/{nam}"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:H1')

    ws3['A2'] = "Top 3 nền vàng (cao nhất), bottom 3 nền hồng (thấp nhất). TB điểm/CC = Tổng điểm ÷ số CC."
    ws3['A2'].font = Font(italic=True, color="666666")

    headers3 = ["Hạng", "Đơn vị", "Số CC", "Số lượt khai", "Tổng điểm", "TB điểm/CC", "Lỗi CL", "Lỗi TĐ"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    n_dv = len(ranking_dv)
    for i, dv in enumerate(ranking_dv):
        r = 5 + i
        hang_cell = ws3.cell(row=r, column=1, value=i + 1)
        hang_cell.border = border
        hang_cell.alignment = center
        hang_cell.font = Font(bold=True)

        is_top3 = i < 3
        is_bottom3 = (n_dv > 6 and i >= n_dv - 3)
        row_fill = gold_fill if is_top3 else (alert_fill if is_bottom3 else None)
        if row_fill:
            hang_cell.fill = row_fill

        ws3.cell(row=r, column=2, value=dv["don_vi"]).border = border
        if row_fill:
            ws3.cell(row=r, column=2).fill = row_fill

        ws3.cell(row=r, column=3, value=dv["so_cc"]).border = border
        ws3.cell(row=r, column=3).alignment = center
        ws3.cell(row=r, column=4, value=dv["so_luot"]).border = border
        ws3.cell(row=r, column=4).alignment = center
        ws3.cell(row=r, column=5, value=f"{dv['tong_diem']:,.0f}").border = border
        ws3.cell(row=r, column=5).alignment = center
        ws3.cell(row=r, column=5).font = Font(bold=True)
        ws3.cell(row=r, column=6, value=f"{dv['tb_per_cc']:,.1f}").border = border
        ws3.cell(row=r, column=6).alignment = center

        loi_cl = dv["loi_cl"]
        cl_cell = ws3.cell(row=r, column=7, value=loi_cl)
        cl_cell.border = border
        cl_cell.alignment = center
        if loi_cl > 0:
            cl_cell.fill = alert_fill
            cl_cell.font = Font(bold=True, color="9C0006")

        loi_td = dv["loi_td"]
        td_cell = ws3.cell(row=r, column=8, value=loi_td)
        td_cell.border = border
        td_cell.alignment = center
        if loi_td > 0:
            td_cell.fill = warn_fill
            td_cell.font = Font(bold=True, color="9C5700")

    for c, w in [('A', 6), ('B', 35), ('C', 8), ('D', 14), ('E', 14), ('F', 12), ('G', 9), ('H', 9)]:
        ws3.column_dimensions[c].width = w

    if ranking_dv:
        ws3.auto_filter.ref = f"A4:H{4 + n_dv}"

    # =========================================================================
    # SHEET 4: PIVOT ĐƠN VỊ × NHÓM PL3
    # =========================================================================
    ws4 = wb.create_sheet("Đơn vị × Nhóm PL3")

    ws4['A1'] = f"PIVOT TỔNG ĐIỂM THEO ĐƠN VỊ × NHÓM PL3 - THÁNG {thang}/{nam}"
    ws4['A1'].font = title_font
    ws4.merge_cells('A1:I1')

    ws4['A2'] = "Cho thấy mỗi đơn vị làm công việc ở mức độ phức tạp nào nhiều nhất."
    ws4['A2'].font = Font(italic=True, color="666666")

    headers4 = ["STT", "Đơn vị", "Nhóm 1\n(Đơn giản)", "Nhóm 2\n(TB)", "Nhóm 3\n(Khá)",
                "Nhóm 4\n(Khó)", "Nhóm 5\n(Rất khó)", "Tổng điểm"]
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=4, column=col, value=h.replace("\n", "\n"))
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if col >= 3 and col <= 7:
            cell.fill = NHOM_FILLS[col - 2]

    dv_sorted = sorted(pivot_dv_nhom.keys())
    for i, dv in enumerate(dv_sorted):
        r = 5 + i
        ws4.cell(row=r, column=1, value=i + 1).border = border
        ws4.cell(row=r, column=1).alignment = center
        ws4.cell(row=r, column=2, value=dv).border = border

        tong_dv = 0.0
        for col_idx, n in enumerate((1, 2, 3, 4, 5), 3):
            val = pivot_dv_nhom[dv].get(n, 0)
            tong_dv += val
            cell = ws4.cell(row=r, column=col_idx, value=f"{val:,.0f}" if val > 0 else "")
            cell.border = border
            cell.alignment = center
            if val > 0:
                cell.fill = NHOM_FILLS[n]

        tong_cell = ws4.cell(row=r, column=8, value=f"{tong_dv:,.0f}")
        tong_cell.border = border
        tong_cell.alignment = center
        tong_cell.font = Font(bold=True)

    # Dòng tổng
    r = 5 + len(dv_sorted)
    ws4.cell(row=r, column=1).border = border
    ws4.cell(row=r, column=2, value="TỔNG").border = border
    ws4.cell(row=r, column=2).font = Font(bold=True)
    total_all = 0.0
    for col_idx, n in enumerate((1, 2, 3, 4, 5), 3):
        tong = sum(dv_map.get(n, 0) for dv_map in pivot_dv_nhom.values())
        total_all += tong
        cell = ws4.cell(row=r, column=col_idx, value=f"{tong:,.0f}")
        cell.border = border
        cell.font = Font(bold=True)
        cell.alignment = center
    ws4.cell(row=r, column=8, value=f"{total_all:,.0f}").border = border
    ws4.cell(row=r, column=8).font = Font(bold=True)
    ws4.cell(row=r, column=8).alignment = center

    ws4.row_dimensions[4].height = 32
    for c, w in [('A', 5), ('B', 30), ('C', 12), ('D', 11), ('E', 11), ('F', 11), ('G', 13), ('H', 13)]:
        ws4.column_dimensions[c].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output



async def _get_data_04_khoi_luong_cv(db: AsyncSession, thang: int, nam: int) -> dict:
    """Get data for report 04 - Khối lượng công việc (V2 native).

    2026-05-14 (redesign theo CCT): bỏ thống kê SP1-4 + C1-5 (V1 only).
    Thay bằng 4 phần V2 native:
      - tong_hop: chỉ tiêu toàn Chi cục (số CC, lượt, điểm, lỗi, top LV).
      - nhom_pl3: phân bố 5 nhóm độ phức tạp.
      - ranking_dv: xếp hạng đơn vị theo tổng điểm.
      - pivot_dv_nhom: đơn vị × nhóm PL3.

    Returns dict với 4 key trên + so_ngay (giữ cho callers cũ).
    """
    from sqlalchemy import text
    from collections import defaultdict

    so_ngay_trong_thang = calendar.monthrange(nam, thang)[1]

    base_where = """
        WHERE kk.thang = :thang AND kk.nam = :nam
              AND kk.trang_thai = 'DA_PHE_DUYET' AND kk.is_deleted = false
    """
    params = {"thang": thang, "nam": nam}

    # 1. Tổng hợp toàn Chi cục
    tong_hop_result = await db.execute(text(f"""
        SELECT
            COUNT(DISTINCT kk.cong_chuc_id) as so_cc_ke_khai,
            COUNT(*) as so_luot_khai,
            COALESCE(SUM(kk.so_sp_goc_quy_doi), 0) as tong_diem,
            COUNT(DISTINCT kk.cong_chuc_id) FILTER (WHERE kk.so_loi_chat_luong > 0) as so_cc_co_loi_cl,
            COUNT(DISTINCT kk.cong_chuc_id) FILTER (WHERE kk.so_loi_tien_do > 0) as so_cc_co_loi_td
        FROM ke_khai_cong_viec kk
        {base_where}
    """), params)
    th_row = tong_hop_result.first()
    so_cc_ke_khai = int(th_row[0] or 0)
    so_luot_khai = int(th_row[1] or 0)
    tong_diem_all = float(th_row[2] or 0)
    so_cc_co_loi_cl = int(th_row[3] or 0)
    so_cc_co_loi_td = int(th_row[4] or 0)

    # Tổng biên chế active để tính %
    bien_che_result = await db.execute(text("""
        SELECT COUNT(*) FROM cong_chuc WHERE is_active = true
    """))
    tong_bien_che = int(bien_che_result.scalar() or 0)

    # Top lĩnh vực
    top_lv_result = await db.execute(text(f"""
        SELECT kk.linh_vuc_snapshot,
               MAX(dm.ten_linh_vuc) as ten_lv,
               COALESCE(SUM(kk.so_sp_goc_quy_doi), 0) as tong_diem
        FROM ke_khai_cong_viec kk
        JOIN danh_muc_sp_cong_viec dm ON dm.id = kk.danh_muc_sp_id
        {base_where} AND kk.linh_vuc_snapshot IS NOT NULL
        GROUP BY kk.linh_vuc_snapshot
        ORDER BY tong_diem DESC
        LIMIT 1
    """), params)
    top_lv_row = top_lv_result.first()
    linh_vuc_top = None
    if top_lv_row:
        linh_vuc_top = {
            "ma_lv": top_lv_row[0],
            "ten_lv": top_lv_row[1] or "",
            "tong_diem": float(top_lv_row[2] or 0),
        }

    tong_hop = {
        "so_cc_ke_khai": so_cc_ke_khai,
        "so_luot_khai": so_luot_khai,
        "tong_diem": tong_diem_all,
        "tb_diem_per_cc": (tong_diem_all / so_cc_ke_khai) if so_cc_ke_khai > 0 else 0.0,
        "ty_le_ke_khai": (so_cc_ke_khai / tong_bien_che * 100) if tong_bien_che > 0 else 0.0,
        "tong_bien_che": tong_bien_che,
        "so_cc_co_loi_cl": so_cc_co_loi_cl,
        "so_cc_co_loi_td": so_cc_co_loi_td,
        "linh_vuc_top": linh_vuc_top,
    }

    # 2. Phân bố theo Nhóm PL3 (1-5)
    nhom_result = await db.execute(text(f"""
        SELECT kk.nhom_pl3_snapshot AS nhom,
               COUNT(*) as so_luot,
               COALESCE(SUM(kk.so_sp_goc_quy_doi), 0) as tong_diem
        FROM ke_khai_cong_viec kk
        {base_where} AND kk.nhom_pl3_snapshot IS NOT NULL
        GROUP BY kk.nhom_pl3_snapshot
        ORDER BY kk.nhom_pl3_snapshot
    """), params)
    nhom_map = {int(r[0]): {"so_luot": int(r[1]), "tong_diem": float(r[2])} for r in nhom_result}
    NHOM_MO_TA = {
        1: ("Đơn giản", 100),
        2: ("Trung bình", 200),
        3: ("Khá", 300),
        4: ("Khó", 400),
        5: ("Rất khó", 500),
    }
    nhom_pl3_data = []
    for n in (1, 2, 3, 4, 5):
        info = nhom_map.get(n, {"so_luot": 0, "tong_diem": 0.0})
        ten, khung = NHOM_MO_TA[n]
        nhom_pl3_data.append({
            "nhom": n,
            "ten": ten,
            "khung": khung,
            "so_luot": info["so_luot"],
            "tong_diem": info["tong_diem"],
        })

    # 3. Ranking đơn vị
    rank_result = await db.execute(text(f"""
        SELECT dv.ten_don_vi,
               COUNT(DISTINCT kk.cong_chuc_id) as so_cc,
               COUNT(*) as so_luot,
               COALESCE(SUM(kk.so_sp_goc_quy_doi), 0) as tong_diem,
               COALESCE(SUM(kk.so_loi_chat_luong), 0) as loi_cl,
               COALESCE(SUM(kk.so_loi_tien_do), 0) as loi_td
        FROM ke_khai_cong_viec kk
        JOIN cong_chuc cc ON cc.id = kk.cong_chuc_id
        LEFT JOIN don_vi dv ON dv.id = cc.don_vi_id
        {base_where}
        GROUP BY dv.ten_don_vi
        ORDER BY tong_diem DESC
    """), params)
    ranking_dv = []
    for r in rank_result:
        so_cc = int(r[1] or 0)
        tong_d = float(r[3] or 0)
        ranking_dv.append({
            "don_vi": r[0] or "(không xác định)",
            "so_cc": so_cc,
            "so_luot": int(r[2] or 0),
            "tong_diem": tong_d,
            "tb_per_cc": (tong_d / so_cc) if so_cc > 0 else 0.0,
            "loi_cl": int(r[4] or 0),
            "loi_td": int(r[5] or 0),
        })

    # 4. Pivot Đơn vị × Nhóm PL3 (tổng điểm)
    pivot_result = await db.execute(text(f"""
        SELECT dv.ten_don_vi,
               kk.nhom_pl3_snapshot AS nhom,
               COALESCE(SUM(kk.so_sp_goc_quy_doi), 0) as tong_diem
        FROM ke_khai_cong_viec kk
        JOIN cong_chuc cc ON cc.id = kk.cong_chuc_id
        LEFT JOIN don_vi dv ON dv.id = cc.don_vi_id
        {base_where} AND kk.nhom_pl3_snapshot IS NOT NULL
        GROUP BY dv.ten_don_vi, kk.nhom_pl3_snapshot
    """), params)
    pivot_dv_nhom = defaultdict(lambda: defaultdict(float))
    for r in pivot_result:
        pivot_dv_nhom[r[0] or "(không xác định)"][int(r[1])] = float(r[2] or 0)

    return {
        "tong_hop": tong_hop,
        "nhom_pl3": nhom_pl3_data,
        "ranking_dv": ranking_dv,
        "pivot_dv_nhom": dict(pivot_dv_nhom),
        "so_ngay": so_ngay_trong_thang,
    }


# =============================================================================
# HELPER: GENERATE REPORT 05 - DANH MỤC CÔNG VIỆC
# =============================================================================

async def _generate_report_05_danh_muc_cv(db: AsyncSession, thang: int, nam: int) -> io.BytesIO:
    """Generate Excel report 05 - Danh mục công việc chi tiết."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Get data
    data = await _get_data_05_danh_muc_cv(db, thang, nam)

    # Create Excel
    wb = Workbook()

    # Styles
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    dm_header_fill = PatternFill("solid", fgColor="2F5496")
    sub_header_fill = PatternFill("solid", fgColor="D6DCE4")
    sp1_fill = PatternFill("solid", fgColor="DAEEF3")
    sp2_fill = PatternFill("solid", fgColor="E2EFDA")
    sp3_fill = PatternFill("solid", fgColor="FDE9D9")
    sp4_fill = PatternFill("solid", fgColor="E4DFEC")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_alignment = Alignment(wrap_text=True, vertical='top')

    # SHEET 1: TỔNG HỢP DANH MỤC — V2 only (production đã chuyển hoàn toàn từ 4/2026)
    ws1 = wb.active
    ws1.title = "Tổng hợp"

    ws1['A1'] = f"5. THỐNG KÊ DANH MỤC CÔNG VIỆC - THÁNG {thang}/{nam}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:H1')

    ws1['A2'] = f"Tổng số đầu mục công việc: {len(data)} (V2_PL3 — theo Lĩnh vực + Nhóm PL3 + Điểm chấm)."
    ws1['A2'].font = Font(bold=True)

    headers = ["STT", "Lĩnh vực", "Tên công việc",
               "Số user kê khai", "Số lần kê khai", "Điểm",
               "Điểm chấm", "Nhóm PL3", "Khung điểm"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    row = 4
    for i, dm in enumerate(data, 1):
        row += 1
        ws1.cell(row=row, column=1, value=i).border = border
        ws1.cell(row=row, column=1).alignment = center_alignment

        # Lĩnh vực
        lv_text = f"{dm['linh_vuc'] or ''}{' — ' + dm['ten_linh_vuc'] if dm['ten_linh_vuc'] else ''}".strip(" —")
        lv_cell = ws1.cell(row=row, column=2, value=lv_text or "—")
        lv_cell.border = border
        lv_cell.alignment = center_alignment
        lv_cell.font = Font(bold=True)

        ws1.cell(row=row, column=3, value=dm["ten_cong_viec"]).border = border

        ws1.cell(row=row, column=4, value=dm["so_user"]).border = border
        ws1.cell(row=row, column=4).alignment = center_alignment

        ws1.cell(row=row, column=5, value=dm["tong_lan_khai"]).border = border
        ws1.cell(row=row, column=5).alignment = center_alignment

        ws1.cell(row=row, column=6, value=f"{dm['tong_sp']:,.0f}").border = border
        ws1.cell(row=row, column=6).alignment = center_alignment

        # Điểm chấm (1-500, cố định mỗi danh mục)
        dc_cell = ws1.cell(row=row, column=7, value=dm.get("diem_cham") or "—")
        dc_cell.border = border
        dc_cell.alignment = center_alignment

        # Nhóm PL3 (1-5)
        nhom_cell = ws1.cell(row=row, column=8, value=dm.get("nhom_pl3") or "—")
        nhom_cell.border = border
        nhom_cell.alignment = center_alignment
        # Highlight nhóm cao (4-5 = phức tạp)
        nhom_val = dm.get("nhom_pl3")
        if isinstance(nhom_val, int):
            if nhom_val >= 4:
                nhom_cell.fill = PatternFill("solid", fgColor="FFC7CE")
                nhom_cell.font = Font(bold=True, color="9C0006")
            elif nhom_val == 3:
                nhom_cell.fill = PatternFill("solid", fgColor="FFEB9C")

        # Khung điểm
        khung_val = dm.get("khung_diem_toi_da")
        kh_cell = ws1.cell(row=row, column=9, value=khung_val or "—")
        kh_cell.border = border
        kh_cell.alignment = center_alignment

    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 50
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 14
    ws1.column_dimensions['F'].width = 14
    ws1.column_dimensions['G'].width = 12
    ws1.column_dimensions['H'].width = 11
    ws1.column_dimensions['I'].width = 12

    # Autofilter để CCT/PCCT lọc theo Lĩnh vực / Nhóm PL3
    if data:
        ws1.auto_filter.ref = f"A4:I{4 + len(data)}"

    # SHEET 2: CHI TIẾT TỪNG DANH MỤC VÀ USER
    ws2 = wb.create_sheet("Chi tiết theo công việc")

    ws2['A1'] = f"CHI TIẾT DANH MỤC CÔNG VIỆC VÀ USER KÊ KHAI - THÁNG {thang}/{nam}"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:H1')

    row = 3
    for dm_idx, dm in enumerate(data, 1):
        # Header cho mỗi danh mục — gắn nhãn lĩnh vực
        label = f"[LV {dm.get('linh_vuc') or '—'}]"
        ws2.merge_cells(f'A{row}:H{row}')
        header_text = f"{dm_idx}. {label} {dm['ten_cong_viec']} ({dm['so_user']} user | {dm['tong_lan_khai']} lần | {dm['tong_sp']:,.0f} điểm)"
        cell = ws2.cell(row=row, column=1, value=header_text)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = dm_header_fill
        for col in range(2, 9):
            ws2.cell(row=row, column=col).fill = dm_header_fill
        row += 1

        # Sub-header (V2: không còn "Cấp độ sử dụng")
        sub_headers = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "Số lần khai", "Điểm"]
        for col, h in enumerate(sub_headers, 1):
            cell = ws2.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True, size=10)
            cell.fill = sub_header_fill
            cell.border = border
            cell.alignment = center_alignment
        row += 1

        # Users
        for u_idx, user in enumerate(dm["users"], 1):
            ws2.cell(row=row, column=1, value=u_idx).border = border
            ws2.cell(row=row, column=1).alignment = center_alignment

            ws2.cell(row=row, column=2, value=user["ho_ten"]).border = border
            ws2.cell(row=row, column=3, value=user["ma_cc"]).border = border
            ws2.cell(row=row, column=4, value=user["don_vi"]).border = border

            ws2.cell(row=row, column=5, value=user["so_lan_khai"]).border = border
            ws2.cell(row=row, column=5).alignment = center_alignment

            ws2.cell(row=row, column=6, value=f"{user['tong_sp']:,.0f}").border = border
            ws2.cell(row=row, column=6).alignment = center_alignment

            row += 1

        row += 1  # Khoảng cách giữa các danh mục

    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 30
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 12
    ws2.column_dimensions['G'].width = 18

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def _get_data_05_danh_muc_cv(db: AsyncSession, thang: int, nam: int) -> list:
    """Get data for report 05 - Danh mục công việc.

    2026-05-14: hỗ trợ V2_PL3 bằng LEFT JOIN sp_chuan + cap_do. Thêm trường
    nguon_du_lieu (V1/PL3), linh_vuc, nhom_pl3, diem_cham, khung_diem_toi_da
    để render "Version" + "Cấp độ phổ biến / Điểm chấm" trong sheet.
    """
    from sqlalchemy import text
    from collections import defaultdict

    result = await db.execute(text("""
        SELECT
            dm.id as danh_muc_id,
            dm.ten_cong_viec,
            COALESCE(sp.ma_sp, '') AS ma_sp,
            COALESCE(sp.ten_sp, '') AS ten_sp,
            cc.id as cong_chuc_id,
            cc.ho_ten,
            cc.ma_cc,
            dv.ten_don_vi,
            COALESCE(cd.ma_cap_do, '') AS ma_cap_do,
            COALESCE(cd.ten_cap_do, '') AS ten_cap_do,
            COUNT(*) as so_lan_khai,
            COALESCE(SUM(kk.so_sp_goc_quy_doi), 0) as tong_sp_quy_doi,
            COALESCE(SUM(kk.so_luong), 0) as tong_so_luong,
            COALESCE(dm.nguon_du_lieu, 'V1') AS nguon_du_lieu,
            dm.linh_vuc,
            dm.ten_linh_vuc,
            dm.nhom_pl3,
            dm.diem_cham,
            dm.khung_diem_toi_da
        FROM ke_khai_cong_viec kk
        JOIN danh_muc_sp_cong_viec dm ON dm.id = kk.danh_muc_sp_id
        LEFT JOIN sp_cong_viec_chuan sp ON sp.id = dm.sp_chuan_id
        JOIN cong_chuc cc ON cc.id = kk.cong_chuc_id
        LEFT JOIN don_vi dv ON dv.id = cc.don_vi_id
        LEFT JOIN cap_do_phuc_tap cd ON cd.id = kk.cap_do_id
        WHERE kk.thang = :thang AND kk.nam = :nam
              AND kk.trang_thai = 'DA_PHE_DUYET' AND kk.is_deleted = false
        GROUP BY dm.id, dm.ten_cong_viec,
                 sp.ma_sp, sp.ten_sp,
                 cc.id, cc.ho_ten, cc.ma_cc, dv.ten_don_vi,
                 cd.ma_cap_do, cd.ten_cap_do,
                 dm.nguon_du_lieu, dm.linh_vuc, dm.ten_linh_vuc,
                 dm.nhom_pl3, dm.diem_cham, dm.khung_diem_toi_da
        ORDER BY dm.nguon_du_lieu, sp.ma_sp NULLS LAST, dm.linh_vuc, dm.ten_cong_viec
    """), {"thang": thang, "nam": nam})

    raw_data = []
    for row in result:
        raw_data.append({
            "danh_muc_id": row[0],
            "ten_cong_viec": row[1],
            "ma_sp": row[2],
            "ten_sp": row[3],
            "cong_chuc_id": row[4],
            "ho_ten": row[5],
            "ma_cc": row[6],
            "don_vi": row[7] or "",
            "ma_cap_do": row[8],
            "ten_cap_do": row[9],
            "so_lan_khai": int(row[10]),
            "tong_sp_quy_doi": float(row[11]),
            "tong_so_luong": float(row[12]),
            "nguon_du_lieu": row[13] or "V1",
            "linh_vuc": row[14],
            "ten_linh_vuc": row[15],
            "nhom_pl3": row[16],
            "diem_cham": row[17],
            "khung_diem_toi_da": row[18],
        })

    # Tổng hợp theo danh mục công việc
    danh_muc_map = defaultdict(lambda: {
        "ten_cong_viec": "",
        "ma_sp": "",
        "ten_sp": "",
        "users": [],
        "tong_sp": 0,
        "tong_lan_khai": 0,
        "cap_do_stats": defaultdict(int),
        # V2 fields
        "nguon_du_lieu": "V1",
        "linh_vuc": None,
        "ten_linh_vuc": None,
        "nhom_pl3": None,
        "diem_cham": None,
        "khung_diem_toi_da": None,
    })

    for item in raw_data:
        dm_id = item["danh_muc_id"]
        dm = danh_muc_map[dm_id]
        dm["ten_cong_viec"] = item["ten_cong_viec"]
        dm["ma_sp"] = item["ma_sp"]
        dm["ten_sp"] = item["ten_sp"]
        dm["tong_sp"] += item["tong_sp_quy_doi"]
        dm["tong_lan_khai"] += item["so_lan_khai"]
        # V2 không có cap_do → bỏ qua "" khỏi stats
        if item["ma_cap_do"]:
            dm["cap_do_stats"][item["ma_cap_do"]] += item["so_lan_khai"]
        dm["nguon_du_lieu"] = item["nguon_du_lieu"]
        dm["linh_vuc"] = item["linh_vuc"]
        dm["ten_linh_vuc"] = item["ten_linh_vuc"]
        dm["nhom_pl3"] = item["nhom_pl3"]
        dm["diem_cham"] = item["diem_cham"]
        dm["khung_diem_toi_da"] = item["khung_diem_toi_da"]

        # Tìm user đã có chưa
        user_found = False
        for u in dm["users"]:
            if u["cong_chuc_id"] == item["cong_chuc_id"]:
                u["so_lan_khai"] += item["so_lan_khai"]
                u["tong_sp"] += item["tong_sp_quy_doi"]
                if item["ma_cap_do"]:
                    u["cap_do_list"].add(item["ma_cap_do"])
                user_found = True
                break

        if not user_found:
            dm["users"].append({
                "cong_chuc_id": item["cong_chuc_id"],
                "ho_ten": item["ho_ten"],
                "ma_cc": item["ma_cc"],
                "don_vi": item["don_vi"],
                "so_lan_khai": item["so_lan_khai"],
                "tong_sp": item["tong_sp_quy_doi"],
                "cap_do_list": {item["ma_cap_do"]} if item["ma_cap_do"] else set(),
            })

    # Convert to list và sort: V1 trước (ma_sp), V2 sau (linh_vuc)
    danh_muc_list = []
    for dm_id, dm in danh_muc_map.items():
        dm["danh_muc_id"] = dm_id
        dm["so_user"] = len(dm["users"])
        dm["users"] = sorted(dm["users"], key=lambda x: (x["don_vi"], x["ho_ten"]))
        danh_muc_list.append(dm)

    danh_muc_list = sorted(
        danh_muc_list,
        key=lambda x: (x["nguon_du_lieu"], x["ma_sp"] or "", x["linh_vuc"] or "", x["ten_cong_viec"]),
    )

    return danh_muc_list


# =============================================================================
# HELPER: GENERATE REPORT 06 - CHỈ SỐ a, b, c (+ d, đ, e CHO LÃNH ĐẠO)
# =============================================================================
# Bổ sung 2026-05-13 theo yêu cầu CCT:
#   - Thống kê CC bị trừ ở 3 chỉ số a (số lượng), b (chất lượng), c (tiến độ).
#   - Lãnh đạo: gộp thêm d (kết quả đv), đ (tổ chức triển khai), e (đoàn kết).
#   - d/đ/e đọc từ danh_gia_dde (INT 50|100, DA_PHE_DUYET) → hiển thị 50% / 100%.
# Tiêu chí "bị trừ" — đơn giản và đồng nhất với cách CCT/PCCT thường nhìn:
#   - a bị trừ ⟺ sp_hoan_thanh < sp_duoc_giao  (làm chưa đủ target)
#   - b bị trừ ⟺ so_loi_chat_luong > 0
#   - c bị trừ ⟺ so_loi_tien_do > 0
#   - d/đ/e bị trừ ⟺ COALESCE(*_phe_duyet, *_ket_qua/_to_chuc/_doan_ket) == 50
# =============================================================================

async def _generate_report_06_chi_so_abc_dde(db: AsyncSession, thang: int, nam: int) -> io.BytesIO:
    """Generate Excel report 06 — Chỉ số a, b, c, d, đ, e."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = await _get_data_06_chi_so_abc_dde(db, thang, nam)

    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    percent_font = Font(bold=True, color="0070C0")
    alert_fill = PatternFill("solid", fgColor="FFC7CE")
    warn_fill = PatternFill("solid", fgColor="FFEB9C")
    good_fill = PatternFill("solid", fgColor="C6EFCE")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='center')

    cc_thuong = [cc for cc in data if not cc["is_lanh_dao"]]
    lanh_dao = [cc for cc in data if cc["is_lanh_dao"]]

    cc_bi_tru_a = [cc for cc in cc_thuong if cc["bi_tru_a"]]
    cc_bi_tru_b = [cc for cc in cc_thuong if cc["bi_tru_b"]]
    cc_bi_tru_c = [cc for cc in cc_thuong if cc["bi_tru_c"]]
    cc_bi_tru_any = [cc for cc in cc_thuong if cc["bi_tru_a"] or cc["bi_tru_b"] or cc["bi_tru_c"]]

    ld_bi_tru_a = [ld for ld in lanh_dao if ld["bi_tru_a"]]
    ld_bi_tru_b = [ld for ld in lanh_dao if ld["bi_tru_b"]]
    ld_bi_tru_c = [ld for ld in lanh_dao if ld["bi_tru_c"]]
    ld_bi_tru_d = [ld for ld in lanh_dao if ld["bi_tru_d"]]
    ld_bi_tru_dd = [ld for ld in lanh_dao if ld["bi_tru_dd"]]
    ld_bi_tru_e = [ld for ld in lanh_dao if ld["bi_tru_e"]]

    total_cc = len(cc_thuong)
    total_ld = len(lanh_dao)

    def pct(count, base):
        return f"{count/base*100:.1f}%" if base > 0 else "0%"

    # SHEET 1: TỔNG HỢP
    ws1 = wb.active
    ws1.title = "Tổng hợp"

    ws1['A1'] = f"6. THỐNG KÊ CHỈ SỐ a, b, c, d, đ, e - THÁNG {thang}/{nam}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:E1')

    ws1['A2'] = f"Công chức thường: {total_cc} | Lãnh đạo: {total_ld}"
    ws1['A2'].font = Font(bold=True)

    row = 4
    headers = ["Chỉ số", "CC thường bị trừ", "% / tổng CC", "LĐ bị trừ", "% / tổng LĐ"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    rows_data = [
        ("a — Số lượng (sp_hoan_thanh < sp_duoc_giao)", len(cc_bi_tru_a), pct(len(cc_bi_tru_a), total_cc), len(ld_bi_tru_a), pct(len(ld_bi_tru_a), total_ld)),
        ("b — Chất lượng (có lỗi CL)", len(cc_bi_tru_b), pct(len(cc_bi_tru_b), total_cc), len(ld_bi_tru_b), pct(len(ld_bi_tru_b), total_ld)),
        ("c — Tiến độ (có lỗi TĐ)", len(cc_bi_tru_c), pct(len(cc_bi_tru_c), total_cc), len(ld_bi_tru_c), pct(len(ld_bi_tru_c), total_ld)),
        ("d — Kết quả đơn vị (chỉ LĐ)", "-", "-", len(ld_bi_tru_d), pct(len(ld_bi_tru_d), total_ld)),
        ("đ — Tổ chức triển khai (chỉ LĐ)", "-", "-", len(ld_bi_tru_dd), pct(len(ld_bi_tru_dd), total_ld)),
        ("e — Đoàn kết nội bộ (chỉ LĐ)", "-", "-", len(ld_bi_tru_e), pct(len(ld_bi_tru_e), total_ld)),
    ]
    for i, r_data in enumerate(rows_data):
        r = row + 1 + i
        for ci, val in enumerate(r_data, 1):
            cell = ws1.cell(row=r, column=ci, value=val)
            cell.border = border
            if ci != 1:
                cell.alignment = center_alignment
            if ci in (3, 5):
                cell.font = percent_font
        ws1.cell(row=r, column=1).font = Font(bold=True)

    for c, w in [('A', 45), ('B', 18), ('C', 14), ('D', 14), ('E', 14)]:
        ws1.column_dimensions[c].width = w

    # SHEET 2: CC THƯỜNG BỊ TRỪ a/b/c
    ws2 = wb.create_sheet("CC thường bị trừ")

    ws2['A1'] = f"CÔNG CHỨC THƯỜNG BỊ TRỪ CHỈ SỐ a / b / c - THÁNG {thang}/{nam}"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:L1')

    ws2['A2'] = "Chỉ liệt kê CC có ≥1 chỉ số bị trừ. a (%): tỷ lệ SP hoàn thành / SP được giao."
    ws2['A2'].font = Font(italic=True, color="666666")

    headers2 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "SP giao", "SP HT", "a (%)", "Lỗi CL", "b", "Lỗi TĐ", "c", "Lý do chính"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    cc_sorted = sorted(
        cc_bi_tru_any,
        key=lambda x: (x["bi_tru_a"] + x["bi_tru_b"] + x["bi_tru_c"], -x["a_pct"]),
        reverse=True,
    )

    for i, cc in enumerate(cc_sorted, 1):
        r = 4 + i
        ws2.cell(row=r, column=1, value=i).border = border
        ws2.cell(row=r, column=2, value=cc["ho_ten"]).border = border
        ws2.cell(row=r, column=3, value=cc["ma_cc"]).border = border
        ws2.cell(row=r, column=4, value=cc["don_vi"]).border = border

        ws2.cell(row=r, column=5, value=f"{cc['sp_duoc_giao']:,.0f}").border = border
        ws2.cell(row=r, column=5).alignment = center_alignment
        ws2.cell(row=r, column=6, value=f"{cc['sp_hoan_thanh']:,.0f}").border = border
        ws2.cell(row=r, column=6).alignment = center_alignment

        a_cell = ws2.cell(row=r, column=7, value=f"{cc['a_pct']:.1f}%")
        a_cell.border = border
        a_cell.alignment = center_alignment
        if cc["bi_tru_a"]:
            a_cell.fill = alert_fill
            a_cell.font = Font(bold=True, color="9C0006")
        else:
            a_cell.fill = good_fill

        loi_cl_cell = ws2.cell(row=r, column=8, value=cc["loi_cl"])
        loi_cl_cell.border = border
        loi_cl_cell.alignment = center_alignment
        if cc["loi_cl"] > 0:
            loi_cl_cell.fill = alert_fill

        b_cell = ws2.cell(row=r, column=9, value="✗" if cc["bi_tru_b"] else "✓")
        b_cell.border = border
        b_cell.alignment = center_alignment
        b_cell.fill = alert_fill if cc["bi_tru_b"] else good_fill
        if cc["bi_tru_b"]:
            b_cell.font = Font(bold=True, color="9C0006")

        loi_td_cell = ws2.cell(row=r, column=10, value=cc["loi_td"])
        loi_td_cell.border = border
        loi_td_cell.alignment = center_alignment
        if cc["loi_td"] > 0:
            loi_td_cell.fill = warn_fill

        c_cell = ws2.cell(row=r, column=11, value="✗" if cc["bi_tru_c"] else "✓")
        c_cell.border = border
        c_cell.alignment = center_alignment
        c_cell.fill = alert_fill if cc["bi_tru_c"] else good_fill
        if cc["bi_tru_c"]:
            c_cell.font = Font(bold=True, color="9C0006")

        ly_do = []
        if cc["bi_tru_a"]:
            ly_do.append(f"a: thiếu {cc['sp_duoc_giao'] - cc['sp_hoan_thanh']:,.0f} SP")
        if cc["bi_tru_b"]:
            ly_do.append(f"b: {cc['loi_cl']} lỗi CL")
        if cc["bi_tru_c"]:
            ly_do.append(f"c: {cc['loi_td']} lỗi TĐ")
        ws2.cell(row=r, column=12, value="; ".join(ly_do)).border = border
        ws2.cell(row=r, column=12).alignment = wrap_alignment

    for c, w in [('A', 5), ('B', 25), ('C', 12), ('D', 25), ('E', 10), ('F', 10), ('G', 9), ('H', 8), ('I', 5), ('J', 8), ('K', 5), ('L', 30)]:
        ws2.column_dimensions[c].width = w

    if cc_sorted:
        ws2.auto_filter.ref = f"A4:L{4 + len(cc_sorted)}"

    # SHEET 3: LÃNH ĐẠO — gộp a, b, c + d, đ, e
    ws3 = wb.create_sheet("LĐ gộp a-b-c + d-đ-e")

    ws3['A1'] = f"LÃNH ĐẠO — CHỈ SỐ a, b, c + d, đ, e - THÁNG {thang}/{nam}"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:Q1')

    ws3['A2'] = ("Điểm KPI chính thức = engine production (DA_PHE_DUYET). "
                 "Điểm KPI tạm tính = nếu bản CHỜ_DUYỆT cũng vào điểm (mức 2: a/b/c + d/đ/e CHO+DA).")
    ws3['A2'].font = Font(italic=True, color="666666")

    headers3 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "Chức vụ",
                "a (%)", "Lỗi CL", "Lỗi TĐ", "d", "đ", "e", "Hệ số LĐ",
                "Trạng thái d/đ/e", "Số CS bị trừ", "Ghi chú d/đ/e",
                "Điểm KPI (chính thức)", "Điểm KPI (tạm tính)"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    ld_sorted = sorted(
        lanh_dao,
        key=lambda x: (
            x["bi_tru_a"] + x["bi_tru_b"] + x["bi_tru_c"]
            + x["bi_tru_d"] + x["bi_tru_dd"] + x["bi_tru_e"],
            -x["a_pct"],
        ),
        reverse=True,
    )

    for i, ld in enumerate(ld_sorted, 1):
        r = 4 + i
        ws3.cell(row=r, column=1, value=i).border = border
        ws3.cell(row=r, column=2, value=ld["ho_ten"]).border = border
        ws3.cell(row=r, column=3, value=ld["ma_cc"]).border = border
        ws3.cell(row=r, column=4, value=ld["don_vi"]).border = border
        ws3.cell(row=r, column=5, value=ld["chuc_vu"]).border = border

        a_cell = ws3.cell(row=r, column=6, value=f"{ld['a_pct']:.1f}%")
        a_cell.border = border
        a_cell.alignment = center_alignment
        if ld["bi_tru_a"]:
            a_cell.fill = alert_fill
            a_cell.font = Font(bold=True, color="9C0006")
        else:
            a_cell.fill = good_fill

        cl_cell = ws3.cell(row=r, column=7, value=ld["loi_cl"])
        cl_cell.border = border
        cl_cell.alignment = center_alignment
        if ld["loi_cl"] > 0:
            cl_cell.fill = alert_fill

        td_cell = ws3.cell(row=r, column=8, value=ld["loi_td"])
        td_cell.border = border
        td_cell.alignment = center_alignment
        if ld["loi_td"] > 0:
            td_cell.fill = warn_fill

        for col_idx, (chi_so, bi_tru) in enumerate(
            [("d", ld["bi_tru_d"]), ("dd", ld["bi_tru_dd"]), ("e", ld["bi_tru_e"])],
            start=9,
        ):
            val = "50%" if bi_tru else "100%"
            cell = ws3.cell(row=r, column=col_idx, value=val)
            cell.border = border
            cell.alignment = center_alignment
            if bi_tru:
                cell.fill = alert_fill
                cell.font = Font(bold=True, color="9C0006")
            else:
                cell.fill = good_fill

        he_so = (0.5 if ld["bi_tru_d"] else 1.0) * (0.5 if ld["bi_tru_dd"] else 1.0) * (0.5 if ld["bi_tru_e"] else 1.0)
        hs_cell = ws3.cell(row=r, column=12, value=f"{he_so*100:.1f}%")
        hs_cell.border = border
        hs_cell.alignment = center_alignment
        hs_cell.font = Font(bold=True)
        if he_so < 1.0:
            hs_cell.fill = warn_fill

        # Trạng thái d/đ/e (CHO_PHE_DUYET → vàng, DA_PHE_DUYET → xanh, không có → "—")
        tt_raw = ld.get("dde_trang_thai", "")
        tt_display = {
            "DA_PHE_DUYET": "Đã duyệt",
            "CHO_PHE_DUYET": "Chờ duyệt",
        }.get(tt_raw, tt_raw or "—")
        tt_cell = ws3.cell(row=r, column=13, value=tt_display)
        tt_cell.border = border
        tt_cell.alignment = center_alignment
        if tt_raw == "CHO_PHE_DUYET":
            tt_cell.fill = PatternFill("solid", fgColor="FFEB9C")
            tt_cell.font = Font(bold=True, color="9C5700")
        elif tt_raw == "DA_PHE_DUYET":
            tt_cell.fill = good_fill

        tong_bi_tru = sum([ld["bi_tru_a"], ld["bi_tru_b"], ld["bi_tru_c"], ld["bi_tru_d"], ld["bi_tru_dd"], ld["bi_tru_e"]])
        tong_cell = ws3.cell(row=r, column=14, value=tong_bi_tru)
        tong_cell.border = border
        tong_cell.alignment = center_alignment
        tong_cell.font = Font(bold=True)

        ghi_chu = []
        if ld["bi_tru_d"] and ld.get("d_ghi_chu"):
            ghi_chu.append(f"d: {ld['d_ghi_chu']}")
        if ld["bi_tru_dd"] and ld.get("dd_ghi_chu"):
            ghi_chu.append(f"đ: {ld['dd_ghi_chu']}")
        if ld["bi_tru_e"] and ld.get("e_ghi_chu"):
            ghi_chu.append(f"e: {ld['e_ghi_chu']}")
        ws3.cell(row=r, column=15, value="; ".join(ghi_chu)).border = border
        ws3.cell(row=r, column=15).alignment = wrap_alignment

        # Cột 16-17: Điểm KPI (chính thức) + Điểm KPI (tạm tính)
        kpi_ct = ld.get("diem_kpi_chinh_thuc")
        kpi_tt = ld.get("diem_kpi_tam_tinh")
        ct_cell = ws3.cell(row=r, column=16,
                           value=round(kpi_ct, 2) if kpi_ct is not None else "—")
        ct_cell.border = border
        ct_cell.alignment = center_alignment
        tt_cell_kpi = ws3.cell(row=r, column=17,
                               value=round(kpi_tt, 2) if kpi_tt is not None else "—")
        tt_cell_kpi.border = border
        tt_cell_kpi.alignment = center_alignment
        if (kpi_tt is not None and kpi_ct is not None
                and kpi_tt < kpi_ct - 0.01):
            tt_cell_kpi.fill = warn_fill
            tt_cell_kpi.font = Font(bold=True, color="9C5700")

        ws3.row_dimensions[r].height = 28

    for c, w in [('A', 5), ('B', 24), ('C', 12), ('D', 22), ('E', 16),
                 ('F', 9), ('G', 8), ('H', 8), ('I', 6), ('J', 6), ('K', 6),
                 ('L', 11), ('M', 13), ('N', 13), ('O', 45),
                 ('P', 14), ('Q', 14)]:
        ws3.column_dimensions[c].width = w

    if ld_sorted:
        ws3.auto_filter.ref = f"A4:Q{4 + len(ld_sorted)}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def _get_data_06_chi_so_abc_dde(db: AsyncSession, thang: int, nam: int) -> list:
    """Get data for report 06 — gộp a/b/c (từ ke_khai) + d/đ/e (từ danh_gia_dde)."""
    from sqlalchemy import text

    so_ngay_trong_thang = calendar.monthrange(nam, thang)[1]

    # 1. Tất cả CC (cả thường + LĐ) có row trong chi_tiet_xep_loai
    cc_result = await db.execute(text("""
        SELECT ct.cong_chuc_id::text, ct.is_lanh_dao,
               cc.ho_ten, cc.ma_cc, cc.chuc_vu,
               dv.ten_don_vi
        FROM chi_tiet_xep_loai ct
        JOIN bao_cao_xep_loai bc ON bc.id = ct.bao_cao_id
        JOIN cong_chuc cc ON cc.id = ct.cong_chuc_id
        LEFT JOIN don_vi dv ON dv.id = cc.don_vi_id
        WHERE bc.thang = :thang AND bc.nam = :nam
              AND cc.is_active = true
    """), {"thang": thang, "nam": nam})

    cc_map = {}
    for row in cc_result:
        cc_id = row[0]
        cc_map[cc_id] = {
            "cong_chuc_id": cc_id,
            "is_lanh_dao": bool(row[1]),
            "ho_ten": row[2],
            "ma_cc": row[3],
            "chuc_vu": row[4] or "",
            "don_vi": row[5] or "",
            "sp_hoan_thanh": 0.0,
            "loi_cl": 0,
            "loi_td": 0,
            # Mặc định 100 (không bị trừ) nếu LĐ chưa có bản CHO_PHE_DUYET / DA_PHE_DUYET
            "d_final": 100, "dd_final": 100, "e_final": 100,
            "d_ghi_chu": "", "dd_ghi_chu": "", "e_ghi_chu": "",
            "dde_trang_thai": "",
        }

    # 2. Số ngày nghỉ (để tính sp_duoc_giao = (ngày tháng - nghỉ) × 96)
    nghi_result = await db.execute(text("""
        SELECT cong_chuc_id::text, COALESCE(SUM(so_ngay), 0)
        FROM dang_ky_nghi
        WHERE thang_ap_dung = :thang AND nam_ap_dung = :nam
              AND trang_thai = 'DA_PHE_DUYET' AND is_deleted = false
        GROUP BY cong_chuc_id
    """), {"thang": thang, "nam": nam})
    nghi_by_cc = {row[0]: float(row[1]) for row in nghi_result}

    # 3. SP + lỗi từ kê khai công chức (CC thường + LĐ V2)
    sp_result = await db.execute(text("""
        SELECT cong_chuc_id::text,
               COALESCE(SUM(so_sp_goc_quy_doi), 0) as tong_sp,
               COALESCE(SUM(so_loi_chat_luong), 0) as loi_cl,
               COALESCE(SUM(so_loi_tien_do), 0) as loi_td
        FROM ke_khai_cong_viec
        WHERE thang = :thang AND nam = :nam
              AND trang_thai = 'DA_PHE_DUYET' AND is_deleted = false
        GROUP BY cong_chuc_id
    """), {"thang": thang, "nam": nam})
    for row in sp_result:
        cc_id = row[0]
        if cc_id in cc_map:
            cc_map[cc_id]["sp_hoan_thanh"] = float(row[1])
            cc_map[cc_id]["loi_cl"] = int(row[2] or 0)
            cc_map[cc_id]["loi_td"] = int(row[3] or 0)

    # 4. SP từ kê khai lãnh đạo (LĐ thật + HĐ 111 dùng form ke_khai_lanh_dao)
    #    Quy ước cộng: mỗi CV "hoàn thành" = 1 đơn vị SP để có a (%) cho LĐ.
    ld_kk_result = await db.execute(text("""
        SELECT cong_chuc_id::text,
               COUNT(*) FILTER (WHERE trang_thai_hoan_thanh = 'DA_HOAN_THANH') as so_cv_ht,
               COALESCE(SUM(so_loi_chat_luong), 0) as loi_cl,
               COALESCE(SUM(so_loi_tien_do), 0) as loi_td
        FROM ke_khai_lanh_dao
        WHERE thang = :thang AND nam = :nam
              AND trang_thai = 'DA_PHE_DUYET' AND is_deleted = false
        GROUP BY cong_chuc_id
    """), {"thang": thang, "nam": nam})
    for row in ld_kk_result:
        cc_id = row[0]
        if cc_id in cc_map:
            # Nếu CC đã có data từ ke_khai_cong_viec thì giữ; nếu rỗng (LĐ V1) thì
            # dùng số lượng công việc hoàn thành làm proxy cho sp_hoan_thanh.
            if cc_map[cc_id]["sp_hoan_thanh"] == 0:
                cc_map[cc_id]["sp_hoan_thanh"] = float(row[1] or 0)
                cc_map[cc_id]["loi_cl"] = int(row[2] or 0)
                cc_map[cc_id]["loi_td"] = int(row[3] or 0)

    # 5. d/đ/e final từ danh_gia_dde — dùng helper chung (filter
    #    IN ('CHO_PHE_DUYET','DA_PHE_DUYET'), COALESCE phe_duyet → ket_qua tự đánh giá).
    dde_by_cc = await _load_dde_finals(db, thang, nam)
    for cc_id, dde in dde_by_cc.items():
        if cc_id in cc_map:
            cc_map[cc_id]["d_final"] = dde["d_final"]
            cc_map[cc_id]["d_ghi_chu"] = dde["d_ghi_chu"]
            cc_map[cc_id]["dd_final"] = dde["dd_final"]
            cc_map[cc_id]["dd_ghi_chu"] = dde["dd_ghi_chu"]
            cc_map[cc_id]["e_final"] = dde["e_final"]
            cc_map[cc_id]["e_ghi_chu"] = dde["e_ghi_chu"]
            cc_map[cc_id]["dde_trang_thai"] = dde["trang_thai"]

    # 5b. Điểm KPI chính thức từ chi_tiet_xep_loai (engine production)
    kpi_result = await db.execute(text("""
        SELECT ct.cong_chuc_id::text, ct.diem_kpi
        FROM chi_tiet_xep_loai ct
        JOIN bao_cao_xep_loai bc ON bc.id = ct.bao_cao_id
        WHERE bc.thang = :thang AND bc.nam = :nam
              AND bc.is_deleted = false
    """), {"thang": thang, "nam": nam})
    for row in kpi_result:
        if row[0] in cc_map:
            cc_map[row[0]]["diem_kpi_chinh_thuc"] = (
                float(row[1]) if row[1] is not None else None
            )

    # 6. Tính derived fields
    for cc in cc_map.values():
        nghi = nghi_by_cc.get(cc["cong_chuc_id"], 0)
        sp_duoc_giao = (so_ngay_trong_thang - nghi) * 96 if not cc["is_lanh_dao"] else 0
        cc["sp_duoc_giao"] = sp_duoc_giao

        if sp_duoc_giao > 0:
            cc["a_pct"] = min(100.0, cc["sp_hoan_thanh"] / sp_duoc_giao * 100)
        else:
            # LĐ hoặc CC ngoại lệ: lấy 100% nếu có làm việc, 0% nếu không kê khai
            cc["a_pct"] = 100.0 if cc["sp_hoan_thanh"] > 0 else 0.0

        cc["bi_tru_a"] = (sp_duoc_giao > 0 and cc["sp_hoan_thanh"] < sp_duoc_giao)
        cc["bi_tru_b"] = cc["loi_cl"] > 0
        cc["bi_tru_c"] = cc["loi_td"] > 0
        # Bị trừ ⟺ final == 50 (DanhGiaDDE.*_final là int 50|100)
        cc["bi_tru_d"] = (cc["d_final"] == 50)
        cc["bi_tru_dd"] = (cc["dd_final"] == 50)
        cc["bi_tru_e"] = (cc["e_final"] == 50)

    # 7. Điểm KPI tạm tính cho LĐ (broadened CHO+DA). CC thường không cần
    #    vì không có d/đ/e ảnh hưởng.
    from uuid import UUID as _UUID
    for cc in cc_map.values():
        cc["diem_kpi_chinh_thuc"] = cc.get("diem_kpi_chinh_thuc")
        cc["diem_kpi_tam_tinh"] = None
        if cc["is_lanh_dao"]:
            try:
                _cc_uuid = _UUID(cc["cong_chuc_id"])
            except Exception:
                _cc_uuid = cc["cong_chuc_id"]
            cc["diem_kpi_tam_tinh"] = await _calc_diem_kpi_ld_tam_tinh(
                db, _cc_uuid, thang, nam, dde_by_cc
            )

    return list(cc_map.values())


# =============================================================================
# HELPER: GENERATE REPORT 07 - SẢN PHẨM THEO LĨNH VỰC
# =============================================================================
# Bổ sung 2026-05-13 theo yêu cầu CCT:
#   - Tổng SP (= SL × he_so_quy_doi_snapshot, đơn vị "điểm SP1") theo từng
#     lĩnh vực La Mã I–XV.
#   - Nội dung (danh mục) có nhiều lượt kê khai nhất per lĩnh vực.
#   - Pivot đơn vị × lĩnh vực để CCT/PCCT thấy phân bổ.
#   - Cảnh báo các bản kê khai V1 không có lĩnh vực (sẽ không vào báo cáo này).
# =============================================================================

LINH_VUC_ORDER = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII",
                  "IX", "X", "XI", "XII", "XIII", "XIV", "XV"]


async def _generate_report_07_sp_theo_linh_vuc(db: AsyncSession, thang: int, nam: int) -> io.BytesIO:
    """Generate Excel report 07 — Tổng SP theo lĩnh vực + top nội dung."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from collections import defaultdict

    tong_hop, top_noi_dung, pivot_dv, canh_bao_v1 = await _get_data_07_sp_theo_linh_vuc(db, thang, nam)

    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    percent_font = Font(bold=True, color="0070C0")
    lv_header_fill = PatternFill("solid", fgColor="2F5496")
    warn_fill = PatternFill("solid", fgColor="FFEB9C")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_alignment = Alignment(wrap_text=True, vertical='top')

    tong_diem_all = sum(lv["tong_diem_sp"] for lv in tong_hop)

    def pct_lv(val):
        return f"{val/tong_diem_all*100:.1f}%" if tong_diem_all > 0 else "0%"

    # SHEET 1: TỔNG HỢP 15 LĨNH VỰC
    ws1 = wb.active
    ws1.title = "Tổng hợp 15 lĩnh vực"

    ws1['A1'] = f"7. TỔNG SẢN PHẨM THEO LĨNH VỰC - THÁNG {thang}/{nam}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:G1')

    ws1['A2'] = "Tổng điểm SP = Σ(số_lượng × hệ_số_quy_đổi_snapshot). Dữ liệu V2_PL3."
    ws1['A2'].font = Font(italic=True, color="666666")

    headers = ["LV", "Tên lĩnh vực", "Số CC kê khai", "Số lượt kê khai", "Tổng SL gốc", "Tổng điểm SP", "% so toàn CC"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    for i, lv in enumerate(tong_hop, 1):
        r = 4 + i
        ws1.cell(row=r, column=1, value=lv["ma_lv"]).border = border
        ws1.cell(row=r, column=1).font = Font(bold=True)
        ws1.cell(row=r, column=1).alignment = center_alignment

        ws1.cell(row=r, column=2, value=lv["ten_lv"]).border = border
        ws1.cell(row=r, column=3, value=lv["so_cc"]).border = border
        ws1.cell(row=r, column=3).alignment = center_alignment
        ws1.cell(row=r, column=4, value=lv["so_lan_khai"]).border = border
        ws1.cell(row=r, column=4).alignment = center_alignment
        ws1.cell(row=r, column=5, value=f"{lv['tong_so_luong']:,.0f}").border = border
        ws1.cell(row=r, column=5).alignment = center_alignment
        ws1.cell(row=r, column=6, value=f"{lv['tong_diem_sp']:,.2f}").border = border
        ws1.cell(row=r, column=6).alignment = center_alignment
        ws1.cell(row=r, column=6).font = Font(bold=True)

        pct_cell = ws1.cell(row=r, column=7, value=pct_lv(lv["tong_diem_sp"]))
        pct_cell.border = border
        pct_cell.alignment = center_alignment
        pct_cell.font = percent_font

    # Dòng tổng
    r = 4 + len(tong_hop) + 1
    ws1.cell(row=r, column=1, value="TỔNG").border = border
    ws1.cell(row=r, column=1).font = Font(bold=True)
    ws1.cell(row=r, column=2).border = border
    ws1.cell(row=r, column=3, value=sum(lv["so_cc"] for lv in tong_hop)).border = border
    ws1.cell(row=r, column=3).alignment = center_alignment
    ws1.cell(row=r, column=4, value=sum(lv["so_lan_khai"] for lv in tong_hop)).border = border
    ws1.cell(row=r, column=4).alignment = center_alignment
    ws1.cell(row=r, column=5, value=f"{sum(lv['tong_so_luong'] for lv in tong_hop):,.0f}").border = border
    ws1.cell(row=r, column=5).alignment = center_alignment
    ws1.cell(row=r, column=6, value=f"{tong_diem_all:,.2f}").border = border
    ws1.cell(row=r, column=6).font = Font(bold=True)
    ws1.cell(row=r, column=6).alignment = center_alignment
    ws1.cell(row=r, column=7, value="100%").border = border
    ws1.cell(row=r, column=7).font = Font(bold=True, color="0070C0")
    ws1.cell(row=r, column=7).alignment = center_alignment

    for c, w in [('A', 6), ('B', 50), ('C', 15), ('D', 16), ('E', 14), ('F', 16), ('G', 14)]:
        ws1.column_dimensions[c].width = w

    # SHEET 2: TOP NỘI DUNG MỖI LĨNH VỰC (sort theo số lượt kê khai)
    ws2 = wb.create_sheet("Nội dung chiếm nhiều nhất")

    ws2['A1'] = f"NỘI DUNG (DANH MỤC) CHIẾM NHIỀU NHẤT THEO LĨNH VỰC - THÁNG {thang}/{nam}"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:F1')

    ws2['A2'] = "Top 10 danh mục theo SỐ LƯỢT KÊ KHAI (mỗi lĩnh vực)."
    ws2['A2'].font = Font(italic=True, color="666666")

    row = 4
    for lv_ma in LINH_VUC_ORDER:
        items = top_noi_dung.get(lv_ma, [])
        if not items:
            continue

        ten_lv = items[0]["ten_lv"] if items else ""

        # Block header
        ws2.merge_cells(f"A{row}:F{row}")
        cell = ws2.cell(row=row, column=1, value=f"Lĩnh vực {lv_ma} — {ten_lv}")
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = lv_header_fill
        for col in range(2, 7):
            ws2.cell(row=row, column=col).fill = lv_header_fill
        row += 1

        # Sub-header
        sub_headers = ["STT", "Tên công việc / Nội dung", "Sản phẩm đầu ra (catalog)", "Số lượt khai", "Tổng SL", "Tổng điểm SP"]
        for col, h in enumerate(sub_headers, 1):
            cell = ws2.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill("solid", fgColor="D6DCE4")
            cell.border = border
            cell.alignment = center_alignment
        row += 1

        for i, it in enumerate(items, 1):
            ws2.cell(row=row, column=1, value=i).border = border
            ws2.cell(row=row, column=1).alignment = center_alignment
            ws2.cell(row=row, column=2, value=it["ten_cong_viec"]).border = border
            ws2.cell(row=row, column=2).alignment = wrap_alignment

            ws2.cell(row=row, column=3, value=it.get("san_pham_dau_ra") or "").border = border
            ws2.cell(row=row, column=3).alignment = wrap_alignment

            ws2.cell(row=row, column=4, value=it["so_lan_khai"]).border = border
            ws2.cell(row=row, column=4).alignment = center_alignment
            ws2.cell(row=row, column=4).font = Font(bold=True)

            ws2.cell(row=row, column=5, value=f"{it['tong_so_luong']:,.0f}").border = border
            ws2.cell(row=row, column=5).alignment = center_alignment

            ws2.cell(row=row, column=6, value=f"{it['tong_diem_sp']:,.2f}").border = border
            ws2.cell(row=row, column=6).alignment = center_alignment

            ws2.row_dimensions[row].height = 28
            row += 1

        row += 1  # khoảng trắng giữa các lĩnh vực

    for c, w in [('A', 5), ('B', 45), ('C', 45), ('D', 13), ('E', 12), ('F', 14)]:
        ws2.column_dimensions[c].width = w

    # SHEET 3: PIVOT ĐƠN VỊ × LĨNH VỰC (giá trị = tổng điểm SP)
    ws3 = wb.create_sheet("Đơn vị × Lĩnh vực")

    ws3['A1'] = f"PIVOT TỔNG ĐIỂM SP THEO ĐƠN VỊ × LĨNH VỰC - THÁNG {thang}/{nam}"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:Q1')

    headers3 = ["STT", "Đơn vị"] + LINH_VUC_ORDER + ["Tổng"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    dv_sorted = sorted(pivot_dv.keys())
    for i, dv in enumerate(dv_sorted, 1):
        r = 3 + i
        ws3.cell(row=r, column=1, value=i).border = border
        ws3.cell(row=r, column=1).alignment = center_alignment
        ws3.cell(row=r, column=2, value=dv).border = border

        tong_dv = 0.0
        for lv_idx, lv_ma in enumerate(LINH_VUC_ORDER, 3):
            val = pivot_dv[dv].get(lv_ma, 0)
            tong_dv += val
            cell = ws3.cell(row=r, column=lv_idx, value=f"{val:,.1f}" if val > 0 else "")
            cell.border = border
            cell.alignment = center_alignment

        tong_cell = ws3.cell(row=r, column=2 + len(LINH_VUC_ORDER) + 1, value=f"{tong_dv:,.1f}")
        tong_cell.border = border
        tong_cell.font = Font(bold=True)
        tong_cell.alignment = center_alignment

    ws3.column_dimensions['A'].width = 5
    ws3.column_dimensions['B'].width = 30
    for ci in range(3, 3 + len(LINH_VUC_ORDER)):
        ws3.column_dimensions[chr(ord('A') + ci - 1)].width = 9
    # Cột Tổng cuối (Q)
    ws3.column_dimensions['R'].width = 12

    # SHEET 4: CẢNH BÁO V1 (KÊ KHAI KHÔNG CÓ LĨNH VỰC)
    ws4 = wb.create_sheet("Cảnh báo V1")

    ws4['A1'] = f"CÁC KÊ KHAI KHÔNG CÓ LĨNH VỰC (V1 hoặc snapshot NULL) - THÁNG {thang}/{nam}"
    ws4['A1'].font = title_font
    ws4.merge_cells('A1:F1')

    if not canh_bao_v1:
        ws4['A3'] = "✓ Tất cả kê khai trong tháng đã có lĩnh vực — báo cáo 07 phủ 100%."
        ws4['A3'].font = Font(bold=True, color="006100")
    else:
        ws4['A2'] = "Các đơn vị/CC sau có kê khai V1 (không có lĩnh vực). Báo cáo 07 KHÔNG bao gồm các bản này."
        ws4['A2'].font = Font(italic=True, color="9C0006")

        headers4 = ["STT", "Đơn vị", "Họ và tên", "Mã CC", "Số bản V1", "Tổng SL"]
        for col, h in enumerate(headers4, 1):
            cell = ws4.cell(row=4, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment

        for i, w_row in enumerate(canh_bao_v1, 1):
            r = 4 + i
            ws4.cell(row=r, column=1, value=i).border = border
            ws4.cell(row=r, column=2, value=w_row["don_vi"]).border = border
            ws4.cell(row=r, column=3, value=w_row["ho_ten"]).border = border
            ws4.cell(row=r, column=4, value=w_row["ma_cc"]).border = border
            ws4.cell(row=r, column=5, value=w_row["so_ban"]).border = border
            ws4.cell(row=r, column=5).alignment = center_alignment
            ws4.cell(row=r, column=5).fill = warn_fill
            ws4.cell(row=r, column=6, value=f"{w_row['tong_so_luong']:,.0f}").border = border
            ws4.cell(row=r, column=6).alignment = center_alignment

    for c, w in [('A', 5), ('B', 25), ('C', 25), ('D', 12), ('E', 12), ('F', 12)]:
        ws4.column_dimensions[c].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def _get_data_07_sp_theo_linh_vuc(db: AsyncSession, thang: int, nam: int) -> tuple:
    """Get data for report 07 — SP theo lĩnh vực.

    Returns:
        (tong_hop, top_noi_dung, pivot_dv, canh_bao_v1)
        - tong_hop: list[dict] 15 lĩnh vực với số CC / số lần khai / tổng SL / tổng điểm
        - top_noi_dung: dict{ma_lv → list[dict]} top 10 danh mục per lĩnh vực
        - pivot_dv: dict{don_vi → dict{ma_lv → tổng_điểm_sp}}
        - canh_bao_v1: list[dict] CC kê khai V1 (không có lĩnh vực)
    """
    from sqlalchemy import text
    from collections import defaultdict

    # Sub-query chung — chỉ lấy bản V2 (linh_vuc_snapshot không NULL).
    # JOIN phải đứng TRƯỚC WHERE: gộp tất cả JOIN vào base_sql.
    base_sql = """
        FROM ke_khai_cong_viec kk
        JOIN cong_chuc cc ON cc.id = kk.cong_chuc_id
        LEFT JOIN don_vi dv ON dv.id = cc.don_vi_id
        LEFT JOIN danh_muc_sp_cong_viec dm ON dm.id = kk.danh_muc_sp_id
        WHERE kk.thang = :thang AND kk.nam = :nam
              AND kk.trang_thai = 'DA_PHE_DUYET' AND kk.is_deleted = false
              AND kk.linh_vuc_snapshot IS NOT NULL
    """

    # 1. Tổng hợp 15 lĩnh vực
    tong_result = await db.execute(text(f"""
        SELECT kk.linh_vuc_snapshot as ma_lv,
               MAX(dm.ten_linh_vuc) as ten_lv,
               COUNT(DISTINCT kk.cong_chuc_id) as so_cc,
               COUNT(*) as so_lan_khai,
               COALESCE(SUM(kk.so_luong), 0) as tong_so_luong,
               COALESCE(SUM(kk.so_luong * COALESCE(kk.he_so_quy_doi_snapshot, 0)), 0) as tong_diem_sp
        {base_sql}
        GROUP BY kk.linh_vuc_snapshot
        ORDER BY kk.linh_vuc_snapshot
    """), {"thang": thang, "nam": nam})

    tong_hop_map = {}
    for row in tong_result:
        tong_hop_map[row[0]] = {
            "ma_lv": row[0],
            "ten_lv": row[1] or "",
            "so_cc": int(row[2]),
            "so_lan_khai": int(row[3]),
            "tong_so_luong": float(row[4]),
            "tong_diem_sp": float(row[5]),
        }

    # Đảm bảo đủ 15 lĩnh vực (kể cả lĩnh vực không có data)
    tong_hop = []
    for lv_ma in LINH_VUC_ORDER:
        tong_hop.append(tong_hop_map.get(lv_ma, {
            "ma_lv": lv_ma,
            "ten_lv": "",
            "so_cc": 0,
            "so_lan_khai": 0,
            "tong_so_luong": 0.0,
            "tong_diem_sp": 0.0,
        }))

    # 2. Top 10 nội dung (danh mục) per lĩnh vực — sort theo SỐ LƯỢT KHAI
    nd_result = await db.execute(text(f"""
        SELECT kk.linh_vuc_snapshot as ma_lv,
               MAX(dm.ten_linh_vuc) as ten_lv,
               dm.id as dm_id,
               MAX(dm.ten_cong_viec) as ten_cong_viec,
               MAX(dm.san_pham_dau_ra) as san_pham_dau_ra,
               COUNT(*) as so_lan_khai,
               COALESCE(SUM(kk.so_luong), 0) as tong_so_luong,
               COALESCE(SUM(kk.so_luong * COALESCE(kk.he_so_quy_doi_snapshot, 0)), 0) as tong_diem_sp
        {base_sql}
        GROUP BY kk.linh_vuc_snapshot, dm.id
        ORDER BY kk.linh_vuc_snapshot, so_lan_khai DESC
    """), {"thang": thang, "nam": nam})

    top_noi_dung = defaultdict(list)
    for row in nd_result:
        lv_ma = row[0]
        if len(top_noi_dung[lv_ma]) >= 10:
            continue
        top_noi_dung[lv_ma].append({
            "ma_lv": lv_ma,
            "ten_lv": row[1] or "",
            "danh_muc_id": str(row[2]),
            "ten_cong_viec": row[3] or "",
            "san_pham_dau_ra": row[4] or "",
            "so_lan_khai": int(row[5]),
            "tong_so_luong": float(row[6]),
            "tong_diem_sp": float(row[7]),
        })

    # 3. Pivot đơn vị × lĩnh vực
    pv_result = await db.execute(text(f"""
        SELECT COALESCE(dv.ten_don_vi, '(không xác định)') as don_vi,
               kk.linh_vuc_snapshot as ma_lv,
               COALESCE(SUM(kk.so_luong * COALESCE(kk.he_so_quy_doi_snapshot, 0)), 0) as tong_diem_sp
        {base_sql}
        GROUP BY dv.ten_don_vi, kk.linh_vuc_snapshot
    """), {"thang": thang, "nam": nam})

    pivot_dv = defaultdict(lambda: defaultdict(float))
    for row in pv_result:
        pivot_dv[row[0]][row[1]] = float(row[2])

    # 4. Cảnh báo V1 — kê khai không có lĩnh vực
    v1_result = await db.execute(text("""
        SELECT COALESCE(dv.ten_don_vi, '(không xác định)') as don_vi,
               cc.ho_ten, cc.ma_cc,
               COUNT(*) as so_ban,
               COALESCE(SUM(kk.so_luong), 0) as tong_so_luong
        FROM ke_khai_cong_viec kk
        JOIN cong_chuc cc ON cc.id = kk.cong_chuc_id
        LEFT JOIN don_vi dv ON dv.id = cc.don_vi_id
        WHERE kk.thang = :thang AND kk.nam = :nam
              AND kk.trang_thai = 'DA_PHE_DUYET' AND kk.is_deleted = false
              AND kk.linh_vuc_snapshot IS NULL
        GROUP BY dv.ten_don_vi, cc.ho_ten, cc.ma_cc
        ORDER BY so_ban DESC
    """), {"thang": thang, "nam": nam})

    canh_bao_v1 = []
    for row in v1_result:
        canh_bao_v1.append({
            "don_vi": row[0],
            "ho_ten": row[1],
            "ma_cc": row[2],
            "so_ban": int(row[3]),
            "tong_so_luong": float(row[4]),
        })

    return tong_hop, dict(top_noi_dung), dict(pivot_dv), canh_bao_v1


# =============================================================================
# ═══════════════════════════════════════════════════════════════════════════
# BÁO CÁO TỔNG HỢP THEO QUÝ (ZIP 5 FILES)
# ═══════════════════════════════════════════════════════════════════════════
#
# Nguyên tắc xử lý (đã chốt với CCT 24/04/2026):
# - File 01 (TC chung) + 02 (KPI): HYBRID
#     + Nếu có snapshot chi_tiet_xep_loai_quy → dùng final_xep_loai
#       (quyết định > đề xuất > hệ thống) và điểm snapshot.
#     + Nếu chưa có snapshot → gọi tinh_diem_quy() on-the-fly
#       (đúng công thức: lũy kế a/b/c, MIN d/đ/e, TB các tháng thực tế).
# - File 03 (LĐ d,đ,e): UNION 3 tháng, dedup theo cong_chuc_id,
#   thêm cột "Tháng bị trừ" (T1/T2/T3).
# - File 04 (Khối lượng CV): CỘNG DỒN 3 tháng SQL.
# - File 05 (Danh mục CV): UNION 3 tháng SQL.
# =============================================================================


# =============================================================================
# HELPER: QUERY SNAPSHOT QUÝ THEO BATCH
# =============================================================================

async def _bulk_get_snapshot_quy(
    db: AsyncSession, quy: int, nam: int
) -> dict:
    """
    Lấy snapshot điểm quý của tất cả CC trong quý/năm.

    Returns:
        dict[cong_chuc_id_str] = {
            "diem_tieu_chi_chung": float,  # TB 3 tháng (snapshot)
            "diem_kpi": float,              # 0-70, TB 3 tháng (snapshot)
            "diem_tong": float,             # 0-100, TB 3 tháng (snapshot)
            "final_xep_loai": str,          # quyết định > đề xuất > hệ thống
            "is_lanh_dao": bool,
        }
        dict rỗng nếu chưa có báo cáo quý nào.
    """
    from sqlalchemy import text as sa_text

    result = await db.execute(sa_text("""
        SELECT ct.cong_chuc_id::text,
               ct.diem_tieu_chi_chung, ct.diem_kpi, ct.diem_tong,
               ct.xep_loai_he_thong, ct.xep_loai_de_xuat, ct.xep_loai_quyet_dinh,
               ct.is_lanh_dao
        FROM chi_tiet_xep_loai_quy ct
        JOIN bao_cao_xep_loai_quy bc ON bc.id = ct.bao_cao_quy_id
        WHERE bc.quy = :quy AND bc.nam = :nam
    """), {"quy": quy, "nam": nam})

    snap_by_cc = {}
    for row in result:
        snap_by_cc[row[0]] = {
            "diem_tieu_chi_chung": float(row[1]) if row[1] is not None else 0.0,
            "diem_kpi": float(row[2]) if row[2] is not None else 0.0,
            "diem_tong": float(row[3]) if row[3] is not None else 0.0,
            "final_xep_loai": row[6] or row[5] or row[4],
            "is_lanh_dao": bool(row[7]),
        }
    return snap_by_cc


async def _compute_diem_quy_hybrid(
    db: AsyncSession, cong_chuc_id: str, quy: int, nam: int,
    snap_by_cc: dict
) -> dict:
    """
    Lấy điểm quý (tc_chung/kpi/tong/xep_loai) theo hybrid:
    - Có snapshot → dùng snapshot + final_xep_loai
    - Không có → gọi tinh_diem_quy() on-the-fly, dùng xep_loai_he_thong

    Returns dict có các key:
        diem_tc_chung, diem_kpi, diem_tong, xep_loai, nguon
        (nguon = "snapshot" hoặc "on_the_fly")
    """
    from app.api.v1.endpoints.xep_loai_quy_helpers import tinh_diem_quy

    if cong_chuc_id in snap_by_cc:
        snap = snap_by_cc[cong_chuc_id]
        return {
            "diem_tc_chung": snap["diem_tieu_chi_chung"],
            "diem_kpi": snap["diem_kpi"],
            "diem_tong": snap["diem_tong"],
            "xep_loai": snap["final_xep_loai"],
            "nguon": "snapshot",
        }

    # On-the-fly (snapshot chưa có)
    dq = await tinh_diem_quy(db, UUID(cong_chuc_id), quy, nam, tam_tinh=True)
    return {
        "diem_tc_chung": float(dq.get("diem_tc_quy") or 0.0),
        "diem_kpi": float(dq.get("diem_kpi_quy") or 0.0),
        "diem_tong": float(dq.get("diem_tong_quy") or 0.0),
        "xep_loai": dq.get("xep_loai_quy") or "E",
        "nguon": "on_the_fly",
    }


# =============================================================================
# GET DATA 01 QUÝ — TIÊU CHÍ CHUNG
# =============================================================================

async def _get_data_01_quy(db: AsyncSession, quy: int, nam: int) -> list:
    """Lấy data File 01 cho quý — điểm TC chung TB quý + chi tiết gộp 3 tháng."""
    from app.api.v1.endpoints.xep_loai_quy_helpers import QUY_TO_THANG

    thang_list = QUY_TO_THANG[quy]

    # 1. Lấy data từng tháng (tái dùng hàm có sẵn)
    data_by_thang = {}
    for t in thang_list:
        data_by_thang[t] = await _get_data_01_tieu_chi_chung(db, t, nam)

    # 2. Snapshot quý
    snap_by_cc = await _bulk_get_snapshot_quy(db, quy, nam)

    # 3. Gộp theo cong_chuc_id
    cc_map = {}
    for thang, items in data_by_thang.items():
        for cc in items:
            cc_id = cc["cong_chuc_id"]
            if cc_id not in cc_map:
                cc_map[cc_id] = {
                    "cong_chuc_id": cc_id,
                    "ma_cc": cc["ma_cc"],
                    "ho_ten": cc["ho_ten"],
                    "don_vi": cc["don_vi"],
                    "by_thang": {},
                }
            cc_map[cc_id]["by_thang"][thang] = cc

    # 4. Tính toán theo quý
    result = []
    for cc_id, info in cc_map.items():
        # Điểm TC chung quý (hybrid)
        dq = await _compute_diem_quy_hybrid(db, cc_id, quy, nam, snap_by_cc)
        diem_tc_quy = dq["diem_tc_chung"]

        # Gộp dữ liệu chi tiết 3 tháng
        tong_diem_nhom3 = 0.0
        has_nhom3 = False
        has_ghi_chu_nhom3 = False
        minh_chung_parts = []
        ly_do_tru_diem = []  # mỗi item thêm "thang"

        for thang, cc_thang in info["by_thang"].items():
            tong_diem_nhom3 += cc_thang["diem_nhom3"]
            if cc_thang["has_nhom3"]:
                has_nhom3 = True
            if cc_thang["has_ghi_chu_nhom3"]:
                has_ghi_chu_nhom3 = True
            if cc_thang.get("minh_chung_nhom3"):
                minh_chung_parts.append(f"[T{thang}] {cc_thang['minh_chung_nhom3']}")
            for ly_do in cc_thang["ly_do_tru_diem"]:
                ly_do_tru_diem.append({**ly_do, "thang": thang})

        so_thang_co_data = len(info["by_thang"])
        diem_nhom3_tb = (
            tong_diem_nhom3 / so_thang_co_data if so_thang_co_data > 0 else 0.0
        )

        result.append({
            "cong_chuc_id": cc_id,
            "ho_ten": info["ho_ten"],
            "ma_cc": info["ma_cc"],
            "don_vi": info["don_vi"],
            "tong_diem": round(diem_tc_quy, 2),
            "diem_nhom3": round(diem_nhom3_tb, 2),
            "has_nhom3": has_nhom3,
            "has_ghi_chu_nhom3": has_ghi_chu_nhom3,
            "minh_chung_nhom3": "\n".join(minh_chung_parts) if minh_chung_parts else "",
            "ly_do_tru_diem": ly_do_tru_diem,
            "nguon": dq["nguon"],
        })

    return result


# =============================================================================
# GET DATA 02 QUÝ — ĐIỂM KPI
# =============================================================================

async def _get_data_02_quy(db: AsyncSession, quy: int, nam: int) -> tuple:
    """Lấy data File 02 cho quý — KPI TB quý + SP cộng dồn 3 tháng."""
    from app.api.v1.endpoints.xep_loai_quy_helpers import QUY_TO_THANG

    thang_list = QUY_TO_THANG[quy]
    so_ngay_quy = sum(calendar.monthrange(nam, t)[1] for t in thang_list)

    # 1. Snapshot quý
    snap_by_cc = await _bulk_get_snapshot_quy(db, quy, nam)

    # 2. Aggregate SP/errors từ 3 tháng (tái dùng _get_data_02)
    monthly = {}
    for t in thang_list:
        kpi_list, _ = await _get_data_02_diem_kpi(db, t, nam)
        for cc in kpi_list:
            cc_id = cc["cong_chuc_id"]
            if cc_id not in monthly:
                monthly[cc_id] = {
                    "cong_chuc_id": cc_id,
                    "ho_ten": cc["ho_ten"],
                    "ma_cc": cc["ma_cc"],
                    "don_vi": cc["don_vi"],
                    "is_lanh_dao": cc.get("is_lanh_dao", False),
                    "sp_duoc_giao": 0.0,
                    "sp_hoan_thanh": 0.0,
                    "sp_cl": 0.0,
                    "sp_td": 0.0,
                    "loi_cl": 0,
                    "loi_td": 0,
                    "so_ngay_lv": 0.0,
                    "so_ngay_nghi": 0.0,
                }
            m = monthly[cc_id]
            m["sp_duoc_giao"] += cc["sp_duoc_giao"]
            m["sp_hoan_thanh"] += cc["sp_hoan_thanh"]
            m["sp_cl"] += cc["sp_cl"]
            m["sp_td"] += cc["sp_td"]
            m["loi_cl"] += cc["loi_cl"]
            m["loi_td"] += cc["loi_td"]
            if cc.get("so_ngay_lv") is not None:
                m["so_ngay_lv"] += cc["so_ngay_lv"]
            if cc.get("so_ngay_nghi") is not None:
                m["so_ngay_nghi"] += cc["so_ngay_nghi"]
            # LĐ flag: đánh dấu nếu bất kỳ tháng nào là LĐ
            if cc.get("is_lanh_dao"):
                m["is_lanh_dao"] = True

    # 3. Build output list
    result = []
    for cc_id, m in monthly.items():
        dq = await _compute_diem_quy_hybrid(db, cc_id, quy, nam, snap_by_cc)

        ty_le_vuot = (
            (m["sp_hoan_thanh"] - m["sp_duoc_giao"]) / m["sp_duoc_giao"] * 100
            if m["sp_duoc_giao"] > 0 else 0.0
        )

        result.append({
            "cong_chuc_id": cc_id,
            "ho_ten": m["ho_ten"],
            "ma_cc": m["ma_cc"],
            "don_vi": m["don_vi"],
            "is_lanh_dao": m["is_lanh_dao"],
            "diem_kpi_70": round(dq["diem_kpi"], 2),
            "diem_tong_100": round(dq["diem_tong"], 2),
            "xep_loai": dq["xep_loai"],
            "sp_duoc_giao": m["sp_duoc_giao"],
            "sp_hoan_thanh": m["sp_hoan_thanh"],
            "sp_cl": m["sp_cl"],
            "sp_td": m["sp_td"],
            "loi_cl": m["loi_cl"],
            "loi_td": m["loi_td"],
            "ty_le_vuot": ty_le_vuot,
            "so_ngay_lv": m["so_ngay_lv"],
            "so_ngay_nghi": m["so_ngay_nghi"],
            "nguon": dq["nguon"],
        })

    return result, so_ngay_quy


# =============================================================================
# GET DATA 03 QUÝ — LÃNH ĐẠO BỊ TRỪ d,đ,e
# =============================================================================

async def _get_data_03_quy(db: AsyncSession, quy: int, nam: int) -> list:
    """Lấy data File 03 cho quý — LĐ d/đ/e union 3 tháng + tháng bị trừ."""
    from app.api.v1.endpoints.xep_loai_quy_helpers import QUY_TO_THANG
    from sqlalchemy import text as sa_text, bindparam
    from collections import defaultdict

    thang_list = QUY_TO_THANG[quy]

    # 1. Lấy danh sách LĐ (distinct, từ chi_tiet_xep_loai 3 tháng)
    ld_stmt = sa_text("""
        SELECT DISTINCT ON (ct.cong_chuc_id)
               ct.cong_chuc_id::text, cc.ho_ten, cc.ma_cc,
               dv.ten_don_vi, vt.ten_vai_tro
        FROM chi_tiet_xep_loai ct
        JOIN bao_cao_xep_loai bc ON bc.id = ct.bao_cao_id
        JOIN cong_chuc cc ON cc.id = ct.cong_chuc_id
        LEFT JOIN don_vi dv ON dv.id = cc.don_vi_id
        LEFT JOIN vai_tro vt ON vt.id = cc.vai_tro_id
        WHERE bc.thang IN :thang_list AND bc.nam = :nam
              AND ct.is_lanh_dao = true
              AND cc.is_active = true
        ORDER BY ct.cong_chuc_id, bc.thang DESC
    """).bindparams(bindparam('thang_list', expanding=True))

    ld_result = await db.execute(ld_stmt, {"thang_list": thang_list, "nam": nam})
    lanh_dao_list = []
    for row in ld_result:
        lanh_dao_list.append({
            "cong_chuc_id": row[0],
            "ho_ten": row[1],
            "ma_cc": row[2],
            "don_vi": row[3] or "",
            "vai_tro": row[4] or "",
        })

    # 2. Lấy đánh giá d/đ/e của 3 tháng — final = COALESCE(phe_duyet, ket_qua_don_vi).
    # Bị trừ tháng X ⟺ final_X == 50.
    # Filter mở rộng (2026-05-13): bao gồm cả CHO_PHE_DUYET để báo cáo phản ánh
    # cả các bản LĐ tự đánh giá nhưng cấp trên chưa duyệt.
    dde_stmt = sa_text("""
        SELECT cong_chuc_id::text, thang,
               COALESCE(d_phe_duyet, d_ket_qua_don_vi)        AS d_final, d_ghi_chu,
               COALESCE(dd_phe_duyet, dd_to_chuc_trien_khai)  AS dd_final, dd_ghi_chu,
               COALESCE(e_phe_duyet, e_doan_ket_noi_bo)       AS e_final, e_ghi_chu
        FROM danh_gia_dde
        WHERE thang IN :thang_list AND nam = :nam
              AND trang_thai IN ('CHO_PHE_DUYET', 'DA_PHE_DUYET')
    """).bindparams(bindparam('thang_list', expanding=True))

    dde_result = await db.execute(dde_stmt, {"thang_list": thang_list, "nam": nam})

    dde_by_cc = defaultdict(lambda: {
        "d_thang_tru": [], "dd_thang_tru": [], "e_thang_tru": [],
        "d_ghi_chu": [], "dd_ghi_chu": [], "e_ghi_chu": [],
    })
    for row in dde_result:
        cc_id, thang_dde = row[0], int(row[1])
        d_final, dd_final, e_final = row[2], row[4], row[6]
        if d_final == 50:
            dde_by_cc[cc_id]["d_thang_tru"].append(thang_dde)
            if row[3]:
                dde_by_cc[cc_id]["d_ghi_chu"].append(f"T{thang_dde}: {row[3]}")
        if dd_final == 50:
            dde_by_cc[cc_id]["dd_thang_tru"].append(thang_dde)
            if row[5]:
                dde_by_cc[cc_id]["dd_ghi_chu"].append(f"T{thang_dde}: {row[5]}")
        if e_final == 50:
            dde_by_cc[cc_id]["e_thang_tru"].append(thang_dde)
            if row[7]:
                dde_by_cc[cc_id]["e_ghi_chu"].append(f"T{thang_dde}: {row[7]}")

    # 3. Snapshot quý cho KPI
    snap_by_cc = await _bulk_get_snapshot_quy(db, quy, nam)

    # 4. Gộp
    for ld in lanh_dao_list:
        cc_id = ld["cong_chuc_id"]
        dde = dde_by_cc.get(cc_id, {})

        d_tru = sorted(dde.get("d_thang_tru", []))
        dd_tru = sorted(dde.get("dd_thang_tru", []))
        e_tru = sorted(dde.get("e_thang_tru", []))

        ld["bi_tru_d"] = len(d_tru) > 0
        ld["bi_tru_dd"] = len(dd_tru) > 0
        ld["bi_tru_e"] = len(e_tru) > 0
        ld["tong_bi_tru"] = sum([ld["bi_tru_d"], ld["bi_tru_dd"], ld["bi_tru_e"]])

        ld["d_thang_tru"] = d_tru
        ld["dd_thang_tru"] = dd_tru
        ld["e_thang_tru"] = e_tru
        ld["d_ghi_chu"] = "; ".join(dde.get("d_ghi_chu", []))
        ld["dd_ghi_chu"] = "; ".join(dde.get("dd_ghi_chu", []))
        ld["e_ghi_chu"] = "; ".join(dde.get("e_ghi_chu", []))

        # KPI quý (hybrid)
        dq = await _compute_diem_quy_hybrid(db, cc_id, quy, nam, snap_by_cc)
        ld["diem_kpi"] = round(dq["diem_kpi"], 2)
        ld["diem_tong"] = round(dq["diem_tong"], 2)
        ld["xep_loai"] = dq["xep_loai"]
        ld["nguon"] = dq["nguon"]

    return lanh_dao_list


# =============================================================================
# GET DATA 04 QUÝ — KHỐI LƯỢNG CÔNG VIỆC (cộng dồn 3 tháng)
# =============================================================================

async def _get_data_04_quy(db: AsyncSession, quy: int, nam: int) -> dict:
    """Get data File 04 cho QUÝ — V2 native, cộng dồn 3 tháng (2026-05-14).

    Returns dict same shape as _get_data_04_khoi_luong_cv:
      tong_hop, nhom_pl3, ranking_dv, pivot_dv_nhom, so_ngay.
    """
    from app.api.v1.endpoints.xep_loai_quy_helpers import QUY_TO_THANG
    from sqlalchemy import text as sa_text, bindparam
    from collections import defaultdict

    thang_list = QUY_TO_THANG[quy]
    so_ngay_quy = sum(calendar.monthrange(nam, t)[1] for t in thang_list)
    params = {"thang_list": thang_list, "nam": nam}

    base_where = (
        "WHERE kk.thang IN :thang_list AND kk.nam = :nam "
        "AND kk.trang_thai = 'DA_PHE_DUYET' AND kk.is_deleted = false"
    )

    # 1. Tổng hợp
    th_stmt = sa_text(f"""
        SELECT
            COUNT(DISTINCT kk.cong_chuc_id) as so_cc_ke_khai,
            COUNT(*) as so_luot_khai,
            COALESCE(SUM(kk.so_sp_goc_quy_doi), 0) as tong_diem,
            COUNT(DISTINCT kk.cong_chuc_id) FILTER (WHERE kk.so_loi_chat_luong > 0) as so_cc_co_loi_cl,
            COUNT(DISTINCT kk.cong_chuc_id) FILTER (WHERE kk.so_loi_tien_do > 0) as so_cc_co_loi_td
        FROM ke_khai_cong_viec kk
        {base_where}
    """).bindparams(bindparam('thang_list', expanding=True))
    th_row = (await db.execute(th_stmt, params)).first()
    so_cc_ke_khai = int(th_row[0] or 0)
    so_luot_khai = int(th_row[1] or 0)
    tong_diem_all = float(th_row[2] or 0)
    so_cc_co_loi_cl = int(th_row[3] or 0)
    so_cc_co_loi_td = int(th_row[4] or 0)

    bien_che_result = await db.execute(sa_text(
        "SELECT COUNT(*) FROM cong_chuc WHERE is_active = true"
    ))
    tong_bien_che = int(bien_che_result.scalar() or 0)

    top_lv_stmt = sa_text(f"""
        SELECT kk.linh_vuc_snapshot,
               MAX(dm.ten_linh_vuc) as ten_lv,
               COALESCE(SUM(kk.so_sp_goc_quy_doi), 0) as tong_diem
        FROM ke_khai_cong_viec kk
        JOIN danh_muc_sp_cong_viec dm ON dm.id = kk.danh_muc_sp_id
        {base_where} AND kk.linh_vuc_snapshot IS NOT NULL
        GROUP BY kk.linh_vuc_snapshot
        ORDER BY tong_diem DESC
        LIMIT 1
    """).bindparams(bindparam('thang_list', expanding=True))
    top_lv_row = (await db.execute(top_lv_stmt, params)).first()
    linh_vuc_top = None
    if top_lv_row:
        linh_vuc_top = {
            "ma_lv": top_lv_row[0],
            "ten_lv": top_lv_row[1] or "",
            "tong_diem": float(top_lv_row[2] or 0),
        }

    tong_hop = {
        "so_cc_ke_khai": so_cc_ke_khai,
        "so_luot_khai": so_luot_khai,
        "tong_diem": tong_diem_all,
        "tb_diem_per_cc": (tong_diem_all / so_cc_ke_khai) if so_cc_ke_khai > 0 else 0.0,
        "ty_le_ke_khai": (so_cc_ke_khai / tong_bien_che * 100) if tong_bien_che > 0 else 0.0,
        "tong_bien_che": tong_bien_che,
        "so_cc_co_loi_cl": so_cc_co_loi_cl,
        "so_cc_co_loi_td": so_cc_co_loi_td,
        "linh_vuc_top": linh_vuc_top,
    }

    # 2. Nhóm PL3
    nhom_stmt = sa_text(f"""
        SELECT kk.nhom_pl3_snapshot, COUNT(*),
               COALESCE(SUM(kk.so_sp_goc_quy_doi), 0)
        FROM ke_khai_cong_viec kk
        {base_where} AND kk.nhom_pl3_snapshot IS NOT NULL
        GROUP BY kk.nhom_pl3_snapshot
        ORDER BY kk.nhom_pl3_snapshot
    """).bindparams(bindparam('thang_list', expanding=True))
    nhom_result = await db.execute(nhom_stmt, params)
    nhom_map = {int(r[0]): {"so_luot": int(r[1]), "tong_diem": float(r[2])} for r in nhom_result}
    NHOM_MO_TA = {1: ("Đơn giản", 100), 2: ("Trung bình", 200),
                  3: ("Khá", 300), 4: ("Khó", 400), 5: ("Rất khó", 500)}
    nhom_pl3_data = []
    for n in (1, 2, 3, 4, 5):
        info = nhom_map.get(n, {"so_luot": 0, "tong_diem": 0.0})
        ten, khung = NHOM_MO_TA[n]
        nhom_pl3_data.append({"nhom": n, "ten": ten, "khung": khung,
                              "so_luot": info["so_luot"], "tong_diem": info["tong_diem"]})

    # 3. Ranking đơn vị
    rank_stmt = sa_text(f"""
        SELECT dv.ten_don_vi,
               COUNT(DISTINCT kk.cong_chuc_id) as so_cc,
               COUNT(*) as so_luot,
               COALESCE(SUM(kk.so_sp_goc_quy_doi), 0) as tong_diem,
               COALESCE(SUM(kk.so_loi_chat_luong), 0) as loi_cl,
               COALESCE(SUM(kk.so_loi_tien_do), 0) as loi_td
        FROM ke_khai_cong_viec kk
        JOIN cong_chuc cc ON cc.id = kk.cong_chuc_id
        LEFT JOIN don_vi dv ON dv.id = cc.don_vi_id
        {base_where}
        GROUP BY dv.ten_don_vi
        ORDER BY tong_diem DESC
    """).bindparams(bindparam('thang_list', expanding=True))
    rank_result = await db.execute(rank_stmt, params)
    ranking_dv = []
    for r in rank_result:
        so_cc = int(r[1] or 0)
        tong_d = float(r[3] or 0)
        ranking_dv.append({
            "don_vi": r[0] or "(không xác định)",
            "so_cc": so_cc, "so_luot": int(r[2] or 0),
            "tong_diem": tong_d,
            "tb_per_cc": (tong_d / so_cc) if so_cc > 0 else 0.0,
            "loi_cl": int(r[4] or 0), "loi_td": int(r[5] or 0),
        })

    # 4. Pivot đơn vị × nhóm
    pivot_stmt = sa_text(f"""
        SELECT dv.ten_don_vi, kk.nhom_pl3_snapshot,
               COALESCE(SUM(kk.so_sp_goc_quy_doi), 0) as tong_diem
        FROM ke_khai_cong_viec kk
        JOIN cong_chuc cc ON cc.id = kk.cong_chuc_id
        LEFT JOIN don_vi dv ON dv.id = cc.don_vi_id
        {base_where} AND kk.nhom_pl3_snapshot IS NOT NULL
        GROUP BY dv.ten_don_vi, kk.nhom_pl3_snapshot
    """).bindparams(bindparam('thang_list', expanding=True))
    pivot_result = await db.execute(pivot_stmt, params)
    pivot_dv_nhom = defaultdict(lambda: defaultdict(float))
    for r in pivot_result:
        pivot_dv_nhom[r[0] or "(không xác định)"][int(r[1])] = float(r[2] or 0)

    return {
        "tong_hop": tong_hop,
        "nhom_pl3": nhom_pl3_data,
        "ranking_dv": ranking_dv,
        "pivot_dv_nhom": dict(pivot_dv_nhom),
        "so_ngay": so_ngay_quy,
    }


# =============================================================================
# GET DATA 05 QUÝ — DANH MỤC CÔNG VIỆC (union 3 tháng)
# =============================================================================

async def _get_data_05_quy(db: AsyncSession, quy: int, nam: int) -> list:
    """Lấy data File 05 cho quý — union 3 tháng SQL."""
    from app.api.v1.endpoints.xep_loai_quy_helpers import QUY_TO_THANG
    from sqlalchemy import text as sa_text, bindparam
    from collections import defaultdict

    thang_list = QUY_TO_THANG[quy]

    # 2026-05-14: LEFT JOIN sp_chuan + cap_do để V2 không bị lọc; thêm các
    # field V2 (nguon_du_lieu, linh_vuc, nhom_pl3, diem_cham, khung_diem).
    stmt = sa_text("""
        SELECT
            dm.id as danh_muc_id,
            dm.ten_cong_viec,
            COALESCE(sp.ma_sp, '') AS ma_sp,
            COALESCE(sp.ten_sp, '') AS ten_sp,
            cc.id as cong_chuc_id,
            cc.ho_ten,
            cc.ma_cc,
            dv.ten_don_vi,
            COALESCE(cd.ma_cap_do, '') AS ma_cap_do,
            COALESCE(cd.ten_cap_do, '') AS ten_cap_do,
            COUNT(*) as so_lan_khai,
            COALESCE(SUM(kk.so_sp_goc_quy_doi), 0) as tong_sp_quy_doi,
            COALESCE(SUM(kk.so_luong), 0) as tong_so_luong,
            COALESCE(dm.nguon_du_lieu, 'V1') AS nguon_du_lieu,
            dm.linh_vuc, dm.ten_linh_vuc, dm.nhom_pl3,
            dm.diem_cham, dm.khung_diem_toi_da
        FROM ke_khai_cong_viec kk
        JOIN danh_muc_sp_cong_viec dm ON dm.id = kk.danh_muc_sp_id
        LEFT JOIN sp_cong_viec_chuan sp ON sp.id = dm.sp_chuan_id
        JOIN cong_chuc cc ON cc.id = kk.cong_chuc_id
        LEFT JOIN don_vi dv ON dv.id = cc.don_vi_id
        LEFT JOIN cap_do_phuc_tap cd ON cd.id = kk.cap_do_id
        WHERE kk.thang IN :thang_list AND kk.nam = :nam
              AND kk.trang_thai = 'DA_PHE_DUYET' AND kk.is_deleted = false
        GROUP BY dm.id, dm.ten_cong_viec,
                 sp.ma_sp, sp.ten_sp,
                 cc.id, cc.ho_ten, cc.ma_cc, dv.ten_don_vi,
                 cd.ma_cap_do, cd.ten_cap_do,
                 dm.nguon_du_lieu, dm.linh_vuc, dm.ten_linh_vuc,
                 dm.nhom_pl3, dm.diem_cham, dm.khung_diem_toi_da
        ORDER BY dm.nguon_du_lieu, sp.ma_sp NULLS LAST, dm.linh_vuc, dm.ten_cong_viec
    """).bindparams(bindparam('thang_list', expanding=True))

    result = await db.execute(stmt, {"thang_list": thang_list, "nam": nam})
    raw_data = [{
        "danh_muc_id": row[0], "ten_cong_viec": row[1],
        "ma_sp": row[2], "ten_sp": row[3],
        "cong_chuc_id": row[4], "ho_ten": row[5], "ma_cc": row[6],
        "don_vi": row[7] or "",
        "ma_cap_do": row[8], "ten_cap_do": row[9],
        "so_lan_khai": int(row[10]),
        "tong_sp_quy_doi": float(row[11]),
        "tong_so_luong": float(row[12]),
        "nguon_du_lieu": row[13] or "V1",
        "linh_vuc": row[14], "ten_linh_vuc": row[15],
        "nhom_pl3": row[16], "diem_cham": row[17],
        "khung_diem_toi_da": row[18],
    } for row in result]

    danh_muc_map = defaultdict(lambda: {
        "ten_cong_viec": "", "ma_sp": "", "ten_sp": "",
        "users": [], "tong_sp": 0, "tong_lan_khai": 0,
        "cap_do_stats": defaultdict(int),
        "nguon_du_lieu": "V1", "linh_vuc": None, "ten_linh_vuc": None,
        "nhom_pl3": None, "diem_cham": None, "khung_diem_toi_da": None,
    })

    for item in raw_data:
        dm_id = item["danh_muc_id"]
        dm = danh_muc_map[dm_id]
        dm["ten_cong_viec"] = item["ten_cong_viec"]
        dm["ma_sp"] = item["ma_sp"]
        dm["ten_sp"] = item["ten_sp"]
        dm["tong_sp"] += item["tong_sp_quy_doi"]
        dm["tong_lan_khai"] += item["so_lan_khai"]
        if item["ma_cap_do"]:
            dm["cap_do_stats"][item["ma_cap_do"]] += item["so_lan_khai"]
        dm["nguon_du_lieu"] = item["nguon_du_lieu"]
        dm["linh_vuc"] = item["linh_vuc"]
        dm["ten_linh_vuc"] = item["ten_linh_vuc"]
        dm["nhom_pl3"] = item["nhom_pl3"]
        dm["diem_cham"] = item["diem_cham"]
        dm["khung_diem_toi_da"] = item["khung_diem_toi_da"]

        user_found = False
        for u in dm["users"]:
            if u["cong_chuc_id"] == item["cong_chuc_id"]:
                u["so_lan_khai"] += item["so_lan_khai"]
                u["tong_sp"] += item["tong_sp_quy_doi"]
                if item["ma_cap_do"]:
                    u["cap_do_list"].add(item["ma_cap_do"])
                user_found = True
                break
        if not user_found:
            dm["users"].append({
                "cong_chuc_id": item["cong_chuc_id"],
                "ho_ten": item["ho_ten"],
                "ma_cc": item["ma_cc"],
                "don_vi": item["don_vi"],
                "so_lan_khai": item["so_lan_khai"],
                "tong_sp": item["tong_sp_quy_doi"],
                "cap_do_list": {item["ma_cap_do"]} if item["ma_cap_do"] else set(),
            })

    danh_muc_list = []
    for dm_id, dm in danh_muc_map.items():
        dm["danh_muc_id"] = dm_id
        dm["so_user"] = len(dm["users"])
        dm["users"] = sorted(dm["users"], key=lambda x: (x["don_vi"], x["ho_ten"]))
        danh_muc_list.append(dm)

    danh_muc_list = sorted(
        danh_muc_list,
        key=lambda x: (x["nguon_du_lieu"], x["ma_sp"] or "", x["linh_vuc"] or "", x["ten_cong_viec"]),
    )
    return danh_muc_list


# =============================================================================
# GENERATE REPORT 01 QUÝ — TIÊU CHÍ CHUNG
# =============================================================================

async def _generate_report_01_quy(db: AsyncSession, quy: int, nam: int) -> io.BytesIO:
    """Generate Excel File 01 quý — Tiêu chí chung (điểm TB quý)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = await _get_data_01_quy(db, quy, nam)
    co_snapshot = any(cc["nguon"] == "snapshot" for cc in data)
    nguon_note = (
        "Dữ liệu từ báo cáo quý đã chốt (snapshot)"
        if co_snapshot else
        "Báo cáo quý chưa tổng hợp — số liệu tính từ dữ liệu tháng tại thời điểm xuất"
    )

    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    note_font = Font(italic=True, size=10, color="7F7F7F")
    percent_font = Font(bold=True, color="0070C0")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='center')

    total = len(data)
    # Ngưỡng điểm quý: dùng TB 3 tháng → giữ <20 / =20 / >20 / =30
    # (trong đó =20 / =30 kiểm tra trong biên độ ±0.05 để xử lý số thập phân)
    def near(v, target):
        return abs(v - target) < 0.05

    duoi_20 = [cc for cc in data if cc["tong_diem"] < 20 and not near(cc["tong_diem"], 20)]
    tron_20 = [cc for cc in data if near(cc["tong_diem"], 20)]
    tren_20 = [cc for cc in data if cc["tong_diem"] > 20 and not near(cc["tong_diem"], 20)]

    tren_20_co_nhom3 = [cc for cc in tren_20 if cc["has_nhom3"] and cc["has_ghi_chu_nhom3"]]
    tren_20_khong_nhom3 = [cc for cc in tren_20 if cc["has_nhom3"] and not cc["has_ghi_chu_nhom3"]]
    tren_20_30diem = [cc for cc in tren_20 if near(cc["tong_diem"], 30)]

    def pct(count):
        return f"{count/total*100:.1f}%" if total > 0 else "0%"

    # SHEET 1: TỔNG HỢP
    ws1 = wb.active
    ws1.title = "Tổng hợp"

    ws1['A1'] = f"1. THỐNG KÊ TIÊU CHÍ CHUNG - QUÝ {quy}/{nam}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:E1')

    ws1['A2'] = f"Tổng số công chức: {total}"
    ws1['A2'].font = Font(bold=True)

    ws1['A3'] = nguon_note
    ws1['A3'].font = note_font
    ws1.merge_cells('A3:E3')

    row = 5
    headers = ["Nhóm", "Số lượng", "Tỷ lệ", "Ghi chú"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    rows_data = [
        ("1. Dưới 20 điểm (TB quý)", len(duoi_20), pct(len(duoi_20)), "Xem Sheet 'Dưới 20 điểm'"),
        ("2. Bằng 20 điểm (TB quý)", len(tron_20), pct(len(tron_20)), "Hoàn thành tốt, không sai sót"),
        ("3. Trên 20 điểm (TB quý)", len(tren_20), pct(len(tren_20)), "Chi tiết bên dưới"),
        ("   3a. CÓ sản phẩm đổi mới", len(tren_20_co_nhom3), pct(len(tren_20_co_nhom3)), "Có minh chứng cụ thể"),
        ("   3b. KHÔNG CÓ sản phẩm đổi mới", len(tren_20_khong_nhom3), pct(len(tren_20_khong_nhom3)), "Chưa hiểu rõ quy định"),
        ("   3c. Đạt TB 30 điểm", len(tren_20_30diem), pct(len(tren_20_30diem)), "Xuất sắc"),
    ]

    for i, (nhom, sl, tl, gc) in enumerate(rows_data):
        r = row + 1 + i
        ws1.cell(row=r, column=1, value=nhom).border = border
        ws1.cell(row=r, column=2, value=sl).border = border
        ws1.cell(row=r, column=2).alignment = center_alignment
        ws1.cell(row=r, column=3, value=tl).border = border
        ws1.cell(row=r, column=3).alignment = center_alignment
        ws1.cell(row=r, column=3).font = percent_font
        ws1.cell(row=r, column=4, value=gc).border = border
        if i < 3:
            ws1.cell(row=r, column=1).font = Font(bold=True)

    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 40

    # SHEET 2: DƯỚI 20 ĐIỂM
    ws2 = wb.create_sheet("Dưới 20 điểm")
    ws2['A1'] = f"DANH SÁCH CÔNG CHỨC TB QUÝ DƯỚI 20 ĐIỂM - QUÝ {quy}/{nam}"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:G1')

    headers2 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "TB quý", "Điểm trừ", "Lý do trừ điểm (3 tháng)"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    for i, cc in enumerate(duoi_20, 1):
        r = 3 + i
        ws2.cell(row=r, column=1, value=i).border = border
        ws2.cell(row=r, column=2, value=cc["ho_ten"]).border = border
        ws2.cell(row=r, column=3, value=cc["ma_cc"]).border = border
        ws2.cell(row=r, column=4, value=cc["don_vi"]).border = border
        ws2.cell(row=r, column=5, value=cc["tong_diem"]).border = border

        diem_tru = sum(ly["diem_tru"] for ly in cc["ly_do_tru_diem"])
        ws2.cell(row=r, column=6, value=diem_tru).border = border

        ly_do_text = "; ".join(
            [f"[T{ly['thang']}-{ly['ma']}] {ly['ly_do']}" for ly in cc["ly_do_tru_diem"]]
        )
        ws2.cell(row=r, column=7, value=ly_do_text).border = border
        ws2.cell(row=r, column=7).alignment = wrap_alignment
        ws2.row_dimensions[r].height = 40

    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 30
    ws2.column_dimensions['E'].width = 10
    ws2.column_dimensions['F'].width = 10
    ws2.column_dimensions['G'].width = 90

    # SHEET 3: TRÊN 20 ĐIỂM
    ws3 = wb.create_sheet("Trên 20 điểm")
    ws3['A1'] = f"DANH SÁCH CÔNG CHỨC TB QUÝ TRÊN 20 ĐIỂM - QUÝ {quy}/{nam}"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:H1')

    headers3 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "TB quý", "TB nhóm III", "Có MC", "Minh chứng đổi mới (3 tháng)"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    tren_20_sorted = sorted(tren_20, key=lambda x: x["tong_diem"], reverse=True)

    for i, cc in enumerate(tren_20_sorted, 1):
        r = 3 + i
        ws3.cell(row=r, column=1, value=i).border = border
        ws3.cell(row=r, column=2, value=cc["ho_ten"]).border = border
        ws3.cell(row=r, column=3, value=cc["ma_cc"]).border = border
        ws3.cell(row=r, column=4, value=cc["don_vi"]).border = border
        ws3.cell(row=r, column=5, value=cc["tong_diem"]).border = border
        ws3.cell(row=r, column=6, value=cc["diem_nhom3"]).border = border

        if cc["has_ghi_chu_nhom3"]:
            co_mc = "Có"
            ws3.cell(row=r, column=7).fill = PatternFill("solid", fgColor="C6EFCE")
        else:
            co_mc = "Không"
            ws3.cell(row=r, column=7).fill = PatternFill("solid", fgColor="FFC7CE")
        ws3.cell(row=r, column=7, value=co_mc).border = border
        ws3.cell(row=r, column=7).alignment = center_alignment

        minh_chung = cc.get("minh_chung_nhom3", "")
        if not minh_chung:
            if near(cc["tong_diem"], 30):
                minh_chung = "(TB đạt 30 - chưa điền minh chứng)"
            else:
                minh_chung = "(Chưa điền minh chứng cụ thể)"
        ws3.cell(row=r, column=8, value=minh_chung).border = border
        ws3.cell(row=r, column=8).alignment = wrap_alignment
        ws3.row_dimensions[r].height = 60

    ws3.column_dimensions['A'].width = 5
    ws3.column_dimensions['B'].width = 25
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 25
    ws3.column_dimensions['E'].width = 10
    ws3.column_dimensions['F'].width = 12
    ws3.column_dimensions['G'].width = 8
    ws3.column_dimensions['H'].width = 100

    # SHEET 4: TRÒN 20 ĐIỂM
    ws4 = wb.create_sheet("Bằng 20 điểm")
    ws4['A1'] = f"DANH SÁCH CÔNG CHỨC TB QUÝ = 20 ĐIỂM - QUÝ {quy}/{nam}"
    ws4['A1'].font = title_font
    ws4.merge_cells('A1:E1')

    ws4['A2'] = "Ghi chú: Hoàn thành tốt nhiệm vụ, không có sai sót."
    ws4['A2'].font = Font(italic=True, color="666666")

    headers4 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "TB quý"]
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    for i, cc in enumerate(tron_20, 1):
        r = 4 + i
        ws4.cell(row=r, column=1, value=i).border = border
        ws4.cell(row=r, column=2, value=cc["ho_ten"]).border = border
        ws4.cell(row=r, column=3, value=cc["ma_cc"]).border = border
        ws4.cell(row=r, column=4, value=cc["don_vi"]).border = border
        ws4.cell(row=r, column=5, value=cc["tong_diem"]).border = border

    ws4.column_dimensions['A'].width = 5
    ws4.column_dimensions['B'].width = 25
    ws4.column_dimensions['C'].width = 12
    ws4.column_dimensions['D'].width = 35
    ws4.column_dimensions['E'].width = 12

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =============================================================================
# GENERATE REPORT 02 QUÝ — ĐIỂM KPI
# =============================================================================

async def _generate_report_02_quy(db: AsyncSession, quy: int, nam: int) -> io.BytesIO:
    """Generate Excel File 02 quý — Điểm KPI (TB quý + SP cộng dồn)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data, so_ngay_quy = await _get_data_02_quy(db, quy, nam)
    co_snapshot = any(cc.get("nguon") == "snapshot" for cc in data)
    nguon_note = (
        "Điểm KPI từ báo cáo quý đã chốt (snapshot)"
        if co_snapshot else
        "Báo cáo quý chưa tổng hợp — điểm KPI tính on-the-fly (lũy kế a/b/c, MIN d/đ/e)"
    )

    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    note_font = Font(italic=True, size=10, color="7F7F7F")
    percent_font = Font(bold=True, color="0070C0")
    alert_fill = PatternFill("solid", fgColor="FFC7CE")
    warn_fill = PatternFill("solid", fgColor="FFEB9C")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    total = len(data)
    total_co_kpi = len([cc for cc in data if cc["diem_kpi_70"] is not None])

    # Phân loại: loại LĐ khỏi SP-based logic
    cc_list = [cc for cc in data if not cc.get("is_lanh_dao")]

    dat_kpi_70 = [cc for cc in cc_list if cc["diem_kpi_70"] is not None and cc["diem_kpi_70"] >= 70]
    chua_dat_kpi_70 = [cc for cc in cc_list if cc["diem_kpi_70"] is not None and cc["diem_kpi_70"] < 70]
    vuot_kpi_bat_thuong = [cc for cc in dat_kpi_70 if cc["ty_le_vuot"] > 50]

    chua_dat_do_so_luong = [cc for cc in chua_dat_kpi_70 if cc["sp_hoan_thanh"] < cc["sp_duoc_giao"]]
    chua_dat_do_chat_luong = [cc for cc in chua_dat_kpi_70 if cc["loi_cl"] > 0]
    chua_dat_do_tien_do = [cc for cc in chua_dat_kpi_70 if cc["loi_td"] > 0]

    def pct(count, base=total_co_kpi):
        return f"{count/base*100:.1f}%" if base > 0 else "0%"

    # SHEET 1: TỔNG HỢP
    ws1 = wb.active
    ws1.title = "Tổng hợp"

    ws1['A1'] = f"2. THỐNG KÊ ĐIỂM KPI - QUÝ {quy}/{nam}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:E1')

    ws1['A2'] = f"Tổng CC có dữ liệu KPI: {total_co_kpi} | Tổng số ngày trong quý: {so_ngay_quy}"
    ws1['A2'].font = Font(bold=True)

    ws1['A3'] = nguon_note
    ws1['A3'].font = note_font
    ws1.merge_cells('A3:E3')

    row = 5
    headers = ["Nhóm", "Số lượng", "Tỷ lệ", "Ghi chú"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    rows_data = [
        ("I. Đạt KPI 70 điểm (TB quý)", len(dat_kpi_70), pct(len(dat_kpi_70)), ""),
        ("   - Vượt KPI bất thường (>50%)", len(vuot_kpi_bat_thuong), pct(len(vuot_kpi_bat_thuong)), "Cần xem xét cấp độ phức tạp"),
        ("II. Chưa đạt KPI 70 điểm", len(chua_dat_kpi_70), pct(len(chua_dat_kpi_70)), ""),
        ("   - Do SP chưa đạt", len(chua_dat_do_so_luong), pct(len(chua_dat_do_so_luong)), "SP hoàn thành < SP được giao"),
        ("   - Do CL bị trừ", len(chua_dat_do_chat_luong), pct(len(chua_dat_do_chat_luong)), "Có lỗi chất lượng"),
        ("   - Do TĐ bị trừ", len(chua_dat_do_tien_do), pct(len(chua_dat_do_tien_do)), "Có lỗi tiến độ"),
    ]

    for i, (nhom, sl, tl, gc) in enumerate(rows_data):
        r = row + 1 + i
        ws1.cell(row=r, column=1, value=nhom).border = border
        ws1.cell(row=r, column=2, value=sl).border = border
        ws1.cell(row=r, column=2).alignment = center_alignment
        ws1.cell(row=r, column=3, value=tl).border = border
        ws1.cell(row=r, column=3).alignment = center_alignment
        ws1.cell(row=r, column=3).font = percent_font
        ws1.cell(row=r, column=4, value=gc).border = border
        if i in [0, 2]:
            ws1.cell(row=r, column=1).font = Font(bold=True)

    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 40

    # SHEET 2: VƯỢT KPI BẤT THƯỜNG
    ws2 = wb.create_sheet("Vượt KPI bất thường")
    ws2['A1'] = f"DANH SÁCH VƯỢT KPI BẤT THƯỜNG (>50%) - QUÝ {quy}/{nam}"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:I1')
    ws2['A2'] = "Ghi chú: Cộng dồn SP 3 tháng. Cần xem xét việc kê khai cấp độ phức tạp."
    ws2['A2'].font = Font(italic=True, color="C00000")

    headers2 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "SP giao (quý)", "SP HT (quý)", "Tỷ lệ vượt", "KPI quý", "Xếp loại"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    vuot_sorted = sorted(vuot_kpi_bat_thuong, key=lambda x: x["ty_le_vuot"], reverse=True)
    for i, cc in enumerate(vuot_sorted, 1):
        r = 4 + i
        ws2.cell(row=r, column=1, value=i).border = border
        ws2.cell(row=r, column=2, value=cc["ho_ten"]).border = border
        ws2.cell(row=r, column=3, value=cc["ma_cc"]).border = border
        ws2.cell(row=r, column=4, value=cc["don_vi"]).border = border
        ws2.cell(row=r, column=5, value=f"{cc['sp_duoc_giao']:,.0f}").border = border
        ws2.cell(row=r, column=6, value=f"{cc['sp_hoan_thanh']:,.0f}").border = border

        ty_le_cell = ws2.cell(row=r, column=7, value=f"{cc['ty_le_vuot']:.1f}%")
        ty_le_cell.border = border
        ty_le_cell.alignment = center_alignment
        if cc["ty_le_vuot"] > 100:
            ty_le_cell.fill = alert_fill
            ty_le_cell.font = Font(bold=True, color="9C0006")
        elif cc["ty_le_vuot"] > 50:
            ty_le_cell.fill = warn_fill

        ws2.cell(row=r, column=8, value=cc["diem_kpi_70"]).border = border
        ws2.cell(row=r, column=9, value=cc.get("xep_loai", "")).border = border
        ws2.cell(row=r, column=9).alignment = center_alignment

    for c, w in [('A', 5), ('B', 25), ('C', 12), ('D', 25), ('E', 14), ('F', 14), ('G', 12), ('H', 10), ('I', 10)]:
        ws2.column_dimensions[c].width = w

    # SHEET 3: CHƯA ĐẠT KPI
    ws3 = wb.create_sheet("Chưa đạt KPI")
    ws3['A1'] = f"DANH SÁCH CHƯA ĐẠT KPI 70 ĐIỂM (TB QUÝ) - QUÝ {quy}/{nam}"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:K1')

    headers3 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "SP giao", "SP HT", "Lỗi CL", "Lỗi TĐ", "KPI quý", "Xếp loại", "Lý do chính"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    chua_dat_sorted = sorted(chua_dat_kpi_70, key=lambda x: x["diem_kpi_70"] or 0)
    for i, cc in enumerate(chua_dat_sorted, 1):
        r = 3 + i
        ws3.cell(row=r, column=1, value=i).border = border
        ws3.cell(row=r, column=2, value=cc["ho_ten"]).border = border
        ws3.cell(row=r, column=3, value=cc["ma_cc"]).border = border
        ws3.cell(row=r, column=4, value=cc["don_vi"]).border = border
        ws3.cell(row=r, column=5, value=f"{cc['sp_duoc_giao']:,.0f}").border = border
        ws3.cell(row=r, column=6, value=f"{cc['sp_hoan_thanh']:,.0f}").border = border

        loi_cl_cell = ws3.cell(row=r, column=7, value=cc["loi_cl"])
        loi_cl_cell.border = border
        if cc["loi_cl"] > 0:
            loi_cl_cell.fill = alert_fill

        loi_td_cell = ws3.cell(row=r, column=8, value=cc["loi_td"])
        loi_td_cell.border = border
        if cc["loi_td"] > 0:
            loi_td_cell.fill = warn_fill

        ws3.cell(row=r, column=9, value=cc["diem_kpi_70"]).border = border
        ws3.cell(row=r, column=10, value=cc.get("xep_loai", "")).border = border
        ws3.cell(row=r, column=10).alignment = center_alignment

        ly_do = []
        if cc["sp_hoan_thanh"] < cc["sp_duoc_giao"]:
            ly_do.append("SP chưa đạt")
        if cc["loi_cl"] > 0:
            ly_do.append(f"CL -{cc['loi_cl']} lỗi")
        if cc["loi_td"] > 0:
            ly_do.append(f"TĐ -{cc['loi_td']} lỗi")
        ws3.cell(row=r, column=11, value=", ".join(ly_do) if ly_do else "Khác").border = border

    for c, w in [('A', 5), ('B', 25), ('C', 12), ('D', 25), ('E', 10), ('F', 10), ('G', 8), ('H', 8), ('I', 10), ('J', 10), ('K', 25)]:
        ws3.column_dimensions[c].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =============================================================================
# GENERATE REPORT 03 QUÝ — LÃNH ĐẠO BỊ TRỪ d,đ,e
# =============================================================================

async def _generate_report_03_quy(db: AsyncSession, quy: int, nam: int) -> io.BytesIO:
    """Generate Excel File 03 quý — LĐ d/đ/e union 3 tháng."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = await _get_data_03_quy(db, quy, nam)

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    note_font = Font(italic=True, size=10, color="7F7F7F")
    percent_font = Font(bold=True, color="0070C0")
    alert_fill = PatternFill("solid", fgColor="FFC7CE")
    good_fill = PatternFill("solid", fgColor="C6EFCE")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='center')

    total_ld = len(data)
    bi_tru_d = [ld for ld in data if ld["bi_tru_d"]]
    bi_tru_dd = [ld for ld in data if ld["bi_tru_dd"]]
    bi_tru_e = [ld for ld in data if ld["bi_tru_e"]]
    bi_tru_any = [ld for ld in data if ld["tong_bi_tru"] > 0]
    khong_bi_tru = [ld for ld in data if ld["tong_bi_tru"] == 0]

    def pct(count):
        return f"{count/total_ld*100:.1f}%" if total_ld > 0 else "0%"

    def thang_str(lst):
        return ", ".join(f"T{t}" for t in lst) if lst else ""

    # SHEET 1: TỔNG HỢP
    ws1 = wb.active
    ws1.title = "Tổng hợp"

    ws1['A1'] = f"3. THỐNG KÊ LÃNH ĐẠO BỊ TRỪ d, đ, e (GỘP 3 THÁNG) - QUÝ {quy}/{nam}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:E1')

    ws1['A2'] = f"Tổng số lãnh đạo: {total_ld}"
    ws1['A2'].font = Font(bold=True)

    ws1['A3'] = "Lưu ý: LĐ bị trừ nếu bất kỳ tháng nào trong quý bị đánh giá không đạt (union 3 tháng)."
    ws1['A3'].font = note_font
    ws1.merge_cells('A3:E3')

    row = 5
    headers = ["Tiêu chí", "Số lượng bị trừ", "Tỷ lệ", "Ghi chú"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    rows_data = [
        ("d - Kết quả đơn vị", len(bi_tru_d), pct(len(bi_tru_d)), "Bị trừ ít nhất 1 tháng trong quý"),
        ("đ - Tổ chức triển khai", len(bi_tru_dd), pct(len(bi_tru_dd)), "Bị trừ ít nhất 1 tháng trong quý"),
        ("e - Đoàn kết nội bộ", len(bi_tru_e), pct(len(bi_tru_e)), "Bị trừ ít nhất 1 tháng trong quý"),
        ("TỔNG BỊ TRỪ (≥1 tiêu chí, ≥1 tháng)", len(bi_tru_any), pct(len(bi_tru_any)), ""),
        ("Không bị trừ cả quý", len(khong_bi_tru), pct(len(khong_bi_tru)), ""),
    ]

    for i, (nhom, sl, tl, gc) in enumerate(rows_data):
        r = row + 1 + i
        ws1.cell(row=r, column=1, value=nhom).border = border
        ws1.cell(row=r, column=2, value=sl).border = border
        ws1.cell(row=r, column=2).alignment = center_alignment
        ws1.cell(row=r, column=3, value=tl).border = border
        ws1.cell(row=r, column=3).alignment = center_alignment
        ws1.cell(row=r, column=3).font = percent_font
        ws1.cell(row=r, column=4, value=gc).border = border
        if i == 3:
            ws1.cell(row=r, column=1).font = Font(bold=True)
            if sl > 0:
                ws1.cell(row=r, column=2).fill = alert_fill

    ws1.column_dimensions['A'].width = 38
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 45

    # SHEET 2: DANH SÁCH BỊ TRỪ
    ws2 = wb.create_sheet("Danh sách bị trừ")
    ws2['A1'] = f"DANH SÁCH LÃNH ĐẠO BỊ TRỪ d, đ, e (GỘP 3 THÁNG) - QUÝ {quy}/{nam}"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:L1')

    headers2 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "Chức vụ",
                "d (tháng tr)", "đ (tháng tr)", "e (tháng tr)",
                "Tổng tiêu chí bị trừ", "Lý do", "KPI quý", "Xếp loại"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    bi_tru_sorted = sorted(bi_tru_any, key=lambda x: x["tong_bi_tru"], reverse=True)
    for i, ld in enumerate(bi_tru_sorted, 1):
        r = 3 + i
        ws2.cell(row=r, column=1, value=i).border = border
        ws2.cell(row=r, column=2, value=ld["ho_ten"]).border = border
        ws2.cell(row=r, column=3, value=ld["ma_cc"]).border = border
        ws2.cell(row=r, column=4, value=ld["don_vi"]).border = border
        ws2.cell(row=r, column=5, value=ld["vai_tro"]).border = border

        # d
        d_val = thang_str(ld["d_thang_tru"]) if ld["bi_tru_d"] else "✓"
        d_cell = ws2.cell(row=r, column=6, value=d_val)
        d_cell.border = border
        d_cell.alignment = center_alignment
        if ld["bi_tru_d"]:
            d_cell.fill = alert_fill
            d_cell.font = Font(bold=True, color="9C0006")
        else:
            d_cell.fill = good_fill

        # đ
        dd_val = thang_str(ld["dd_thang_tru"]) if ld["bi_tru_dd"] else "✓"
        dd_cell = ws2.cell(row=r, column=7, value=dd_val)
        dd_cell.border = border
        dd_cell.alignment = center_alignment
        if ld["bi_tru_dd"]:
            dd_cell.fill = alert_fill
            dd_cell.font = Font(bold=True, color="9C0006")
        else:
            dd_cell.fill = good_fill

        # e
        e_val = thang_str(ld["e_thang_tru"]) if ld["bi_tru_e"] else "✓"
        e_cell = ws2.cell(row=r, column=8, value=e_val)
        e_cell.border = border
        e_cell.alignment = center_alignment
        if ld["bi_tru_e"]:
            e_cell.fill = alert_fill
            e_cell.font = Font(bold=True, color="9C0006")
        else:
            e_cell.fill = good_fill

        ws2.cell(row=r, column=9, value=ld["tong_bi_tru"]).border = border
        ws2.cell(row=r, column=9).alignment = center_alignment
        ws2.cell(row=r, column=9).font = Font(bold=True)

        ly_do_parts = []
        if ld["bi_tru_d"] and ld["d_ghi_chu"]:
            ly_do_parts.append(f"d: {ld['d_ghi_chu']}")
        if ld["bi_tru_dd"] and ld["dd_ghi_chu"]:
            ly_do_parts.append(f"đ: {ld['dd_ghi_chu']}")
        if ld["bi_tru_e"] and ld["e_ghi_chu"]:
            ly_do_parts.append(f"e: {ld['e_ghi_chu']}")
        ws2.cell(row=r, column=10, value=" | ".join(ly_do_parts)).border = border
        ws2.cell(row=r, column=10).alignment = wrap_alignment

        ws2.cell(row=r, column=11, value=ld["diem_kpi"]).border = border
        ws2.cell(row=r, column=12, value=ld.get("xep_loai", "")).border = border
        ws2.cell(row=r, column=12).alignment = center_alignment

        ws2.row_dimensions[r].height = 35

    for c, w in [('A', 5), ('B', 25), ('C', 12), ('D', 25), ('E', 15),
                 ('F', 12), ('G', 12), ('H', 12), ('I', 10), ('J', 60), ('K', 10), ('L', 10)]:
        ws2.column_dimensions[c].width = w

    # SHEET 3: TẤT CẢ LÃNH ĐẠO
    ws3 = wb.create_sheet("Tất cả lãnh đạo")
    ws3['A1'] = f"DANH SÁCH TẤT CẢ LÃNH ĐẠO - QUÝ {quy}/{nam}"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:J1')

    headers3 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "Chức vụ",
                "d", "đ", "e", "KPI quý", "Xếp loại"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    for i, ld in enumerate(data, 1):
        r = 3 + i
        ws3.cell(row=r, column=1, value=i).border = border
        ws3.cell(row=r, column=2, value=ld["ho_ten"]).border = border
        ws3.cell(row=r, column=3, value=ld["ma_cc"]).border = border
        ws3.cell(row=r, column=4, value=ld["don_vi"]).border = border
        ws3.cell(row=r, column=5, value=ld["vai_tro"]).border = border

        for col_idx, key in [(6, "bi_tru_d"), (7, "bi_tru_dd"), (8, "bi_tru_e")]:
            cell = ws3.cell(row=r, column=col_idx, value="✗" if ld[key] else "✓")
            cell.border = border
            cell.alignment = center_alignment
            if ld[key]:
                cell.fill = alert_fill
            else:
                cell.fill = good_fill

        ws3.cell(row=r, column=9, value=ld["diem_kpi"]).border = border
        ws3.cell(row=r, column=10, value=ld.get("xep_loai", "")).border = border
        ws3.cell(row=r, column=10).alignment = center_alignment

    for c, w in [('A', 5), ('B', 25), ('C', 12), ('D', 25), ('E', 15),
                 ('F', 6), ('G', 6), ('H', 6), ('I', 10), ('J', 10)]:
        ws3.column_dimensions[c].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =============================================================================
# GENERATE REPORT 04 QUÝ — KHỐI LƯỢNG CÔNG VIỆC
# =============================================================================

async def _generate_report_04_quy(db: AsyncSession, quy: int, nam: int) -> io.BytesIO:
    """Generate Excel File 04 QUÝ — V2 native, cộng dồn 3 tháng (2026-05-14).

    Tái sử dụng cùng layout 4 sheet như tháng (chỉ đổi title).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = await _get_data_04_quy(db, quy, nam)
    tong_hop = data["tong_hop"]
    nhom_pl3 = data["nhom_pl3"]
    ranking_dv = data["ranking_dv"]
    pivot_dv_nhom = data["pivot_dv_nhom"]
    so_ngay_quy = data["so_ngay"]

    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    percent_font = Font(bold=True, color="0070C0")
    alert_fill = PatternFill("solid", fgColor="FFC7CE")
    warn_fill = PatternFill("solid", fgColor="FFEB9C")
    gold_fill = PatternFill("solid", fgColor="FFD966")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')
    wrap_top = Alignment(wrap_text=True, vertical='top')

    NHOM_FILLS = {
        1: PatternFill("solid", fgColor="C6EFCE"),
        2: PatternFill("solid", fgColor="D9EAD3"),
        3: PatternFill("solid", fgColor="FFEB9C"),
        4: PatternFill("solid", fgColor="FFC7CE"),
        5: PatternFill("solid", fgColor="E6B8AF"),
    }

    # SHEET 1: TỔNG HỢP
    ws1 = wb.active
    ws1.title = "Tổng hợp"
    ws1['A1'] = f"4. KHỐI LƯỢNG CÔNG VIỆC (CỘNG DỒN 3 THÁNG) - QUÝ {quy}/{nam}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:C1')
    ws1['A2'] = f"Chi cục Hải quan Khu vực VIII | Tổng ngày làm việc quý: {so_ngay_quy}. Dữ liệu V2_PL3."
    ws1['A2'].font = Font(italic=True, color="666666")

    rows_data = [
        ("Tổng số CC kê khai (quý)", f"{tong_hop['so_cc_ke_khai']:,}",
         f"{tong_hop['ty_le_ke_khai']:.1f}% trên {tong_hop['tong_bien_che']} biên chế"),
        ("Tổng số lượt kê khai (quý)", f"{tong_hop['so_luot_khai']:,}", ""),
        ("Tổng điểm SP toàn Chi cục (quý)", f"{tong_hop['tong_diem']:,.0f}", "đơn vị: điểm SP quy đổi"),
        ("Trung bình điểm / CC (quý)", f"{tong_hop['tb_diem_per_cc']:,.1f}", "tổng điểm ÷ số CC kê khai"),
        ("Số CC có lỗi CL (≥1 lần trong quý)",
         f"{tong_hop['so_cc_co_loi_cl']:,}",
         f"{tong_hop['so_cc_co_loi_cl']/tong_hop['so_cc_ke_khai']*100:.1f}%" if tong_hop['so_cc_ke_khai'] > 0 else "0%"),
        ("Số CC có lỗi TĐ (≥1 lần trong quý)",
         f"{tong_hop['so_cc_co_loi_td']:,}",
         f"{tong_hop['so_cc_co_loi_td']/tong_hop['so_cc_ke_khai']*100:.1f}%" if tong_hop['so_cc_ke_khai'] > 0 else "0%"),
    ]
    if tong_hop["linh_vuc_top"]:
        lv = tong_hop["linh_vuc_top"]
        rows_data.append((
            "Lĩnh vực nhiều SP nhất (quý)",
            f"LV {lv['ma_lv']} — {lv['ten_lv'][:40]}",
            f"{lv['tong_diem']:,.0f} điểm"
        ))

    row = 4
    for col, h in enumerate(["Chỉ tiêu", "Giá trị", "Ghi chú"], 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    for i, (label, val, note) in enumerate(rows_data):
        r = row + 1 + i
        ws1.cell(row=r, column=1, value=label).border = border
        ws1.cell(row=r, column=1).font = Font(bold=True)
        val_cell = ws1.cell(row=r, column=2, value=val)
        val_cell.border = border
        val_cell.alignment = center
        val_cell.font = Font(bold=True, color="0070C0")
        ws1.cell(row=r, column=3, value=note).border = border
        ws1.cell(row=r, column=3).alignment = wrap_top

    ws1.column_dimensions['A'].width = 38
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 45

    # SHEET 2: NHÓM PL3
    ws2 = wb.create_sheet("Theo Nhóm PL3")
    ws2['A1'] = f"PHÂN BỐ KHỐI LƯỢNG THEO NHÓM PL3 (QUÝ) - QUÝ {quy}/{nam}"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:F1')
    ws2['A2'] = "Cộng dồn 3 tháng. Nhóm 1 (đơn giản) → 5 (rất khó)."
    ws2['A2'].font = Font(italic=True, color="666666")

    for col, h in enumerate(["Nhóm", "Mô tả", "Khung điểm tối đa", "Số lượt khai (quý)", "Tổng điểm (quý)", "Tỷ lệ"], 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    tong_diem_nhom = sum(n["tong_diem"] for n in nhom_pl3)
    tong_luot_nhom = sum(n["so_luot"] for n in nhom_pl3)
    for i, n in enumerate(nhom_pl3):
        r = 5 + i
        nhom_cell = ws2.cell(row=r, column=1, value=n["nhom"])
        nhom_cell.border = border
        nhom_cell.alignment = center
        nhom_cell.font = Font(bold=True)
        nhom_cell.fill = NHOM_FILLS[n["nhom"]]
        ws2.cell(row=r, column=2, value=n["ten"]).border = border
        ws2.cell(row=r, column=3, value=n["khung"]).border = border
        ws2.cell(row=r, column=3).alignment = center
        ws2.cell(row=r, column=4, value=n["so_luot"]).border = border
        ws2.cell(row=r, column=4).alignment = center
        ws2.cell(row=r, column=5, value=f"{n['tong_diem']:,.0f}").border = border
        ws2.cell(row=r, column=5).alignment = center
        ws2.cell(row=r, column=5).font = Font(bold=True)
        pct_val = (n["tong_diem"] / tong_diem_nhom * 100) if tong_diem_nhom > 0 else 0
        pct_cell = ws2.cell(row=r, column=6, value=f"{pct_val:.1f}%")
        pct_cell.border = border
        pct_cell.alignment = center
        pct_cell.font = percent_font

    r = 5 + len(nhom_pl3)
    ws2.cell(row=r, column=1, value="TỔNG").border = border
    ws2.cell(row=r, column=1).font = Font(bold=True)
    ws2.cell(row=r, column=2).border = border
    ws2.cell(row=r, column=3).border = border
    ws2.cell(row=r, column=4, value=tong_luot_nhom).border = border
    ws2.cell(row=r, column=4).alignment = center
    ws2.cell(row=r, column=4).font = Font(bold=True)
    ws2.cell(row=r, column=5, value=f"{tong_diem_nhom:,.0f}").border = border
    ws2.cell(row=r, column=5).alignment = center
    ws2.cell(row=r, column=5).font = Font(bold=True)
    ws2.cell(row=r, column=6, value="100%").border = border
    ws2.cell(row=r, column=6).alignment = center
    ws2.cell(row=r, column=6).font = Font(bold=True, color="0070C0")

    for c, w in [('A', 8), ('B', 18), ('C', 18), ('D', 16), ('E', 16), ('F', 10)]:
        ws2.column_dimensions[c].width = w

    # SHEET 3: RANKING ĐƠN VỊ
    ws3 = wb.create_sheet("Ranking đơn vị")
    ws3['A1'] = f"XẾP HẠNG ĐƠN VỊ THEO TỔNG ĐIỂM (QUÝ) - QUÝ {quy}/{nam}"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:H1')
    ws3['A2'] = "Top 3 nền vàng, bottom 3 nền hồng. TB điểm/CC = Tổng điểm ÷ số CC kê khai."
    ws3['A2'].font = Font(italic=True, color="666666")

    for col, h in enumerate(["Hạng", "Đơn vị", "Số CC", "Số lượt khai (quý)", "Tổng điểm (quý)", "TB điểm/CC", "Lỗi CL", "Lỗi TĐ"], 1):
        cell = ws3.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    n_dv = len(ranking_dv)
    for i, dv in enumerate(ranking_dv):
        r = 5 + i
        hang_cell = ws3.cell(row=r, column=1, value=i + 1)
        hang_cell.border = border
        hang_cell.alignment = center
        hang_cell.font = Font(bold=True)
        is_top3 = i < 3
        is_bottom3 = (n_dv > 6 and i >= n_dv - 3)
        row_fill = gold_fill if is_top3 else (alert_fill if is_bottom3 else None)
        if row_fill:
            hang_cell.fill = row_fill
        ws3.cell(row=r, column=2, value=dv["don_vi"]).border = border
        if row_fill:
            ws3.cell(row=r, column=2).fill = row_fill
        ws3.cell(row=r, column=3, value=dv["so_cc"]).border = border
        ws3.cell(row=r, column=3).alignment = center
        ws3.cell(row=r, column=4, value=dv["so_luot"]).border = border
        ws3.cell(row=r, column=4).alignment = center
        ws3.cell(row=r, column=5, value=f"{dv['tong_diem']:,.0f}").border = border
        ws3.cell(row=r, column=5).alignment = center
        ws3.cell(row=r, column=5).font = Font(bold=True)
        ws3.cell(row=r, column=6, value=f"{dv['tb_per_cc']:,.1f}").border = border
        ws3.cell(row=r, column=6).alignment = center
        cl_cell = ws3.cell(row=r, column=7, value=dv["loi_cl"])
        cl_cell.border = border
        cl_cell.alignment = center
        if dv["loi_cl"] > 0:
            cl_cell.fill = alert_fill
            cl_cell.font = Font(bold=True, color="9C0006")
        td_cell = ws3.cell(row=r, column=8, value=dv["loi_td"])
        td_cell.border = border
        td_cell.alignment = center
        if dv["loi_td"] > 0:
            td_cell.fill = warn_fill
            td_cell.font = Font(bold=True, color="9C5700")

    for c, w in [('A', 6), ('B', 35), ('C', 8), ('D', 16), ('E', 16), ('F', 12), ('G', 9), ('H', 9)]:
        ws3.column_dimensions[c].width = w

    if ranking_dv:
        ws3.auto_filter.ref = f"A4:H{4 + n_dv}"

    # SHEET 4: PIVOT ĐƠN VỊ × NHÓM PL3
    ws4 = wb.create_sheet("Đơn vị × Nhóm PL3")
    ws4['A1'] = f"PIVOT TỔNG ĐIỂM THEO ĐƠN VỊ × NHÓM PL3 (QUÝ) - QUÝ {quy}/{nam}"
    ws4['A1'].font = title_font
    ws4.merge_cells('A1:I1')
    ws4['A2'] = "Cộng dồn 3 tháng. Cho thấy mỗi đơn vị làm công việc ở mức độ phức tạp nào nhiều nhất."
    ws4['A2'].font = Font(italic=True, color="666666")

    headers4 = ["STT", "Đơn vị", "Nhóm 1\n(Đơn giản)", "Nhóm 2\n(TB)", "Nhóm 3\n(Khá)",
                "Nhóm 4\n(Khó)", "Nhóm 5\n(Rất khó)", "Tổng điểm"]
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=4, column=col, value=h.replace("\n", "\n"))
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if col >= 3 and col <= 7:
            cell.fill = NHOM_FILLS[col - 2]

    dv_sorted = sorted(pivot_dv_nhom.keys())
    for i, dv in enumerate(dv_sorted):
        r = 5 + i
        ws4.cell(row=r, column=1, value=i + 1).border = border
        ws4.cell(row=r, column=1).alignment = center
        ws4.cell(row=r, column=2, value=dv).border = border
        tong_dv = 0.0
        for col_idx, n in enumerate((1, 2, 3, 4, 5), 3):
            val = pivot_dv_nhom[dv].get(n, 0)
            tong_dv += val
            cell = ws4.cell(row=r, column=col_idx, value=f"{val:,.0f}" if val > 0 else "")
            cell.border = border
            cell.alignment = center
            if val > 0:
                cell.fill = NHOM_FILLS[n]
        tong_cell = ws4.cell(row=r, column=8, value=f"{tong_dv:,.0f}")
        tong_cell.border = border
        tong_cell.alignment = center
        tong_cell.font = Font(bold=True)

    r = 5 + len(dv_sorted)
    ws4.cell(row=r, column=1).border = border
    ws4.cell(row=r, column=2, value="TỔNG").border = border
    ws4.cell(row=r, column=2).font = Font(bold=True)
    total_all = 0.0
    for col_idx, n in enumerate((1, 2, 3, 4, 5), 3):
        tong = sum(dv_map.get(n, 0) for dv_map in pivot_dv_nhom.values())
        total_all += tong
        cell = ws4.cell(row=r, column=col_idx, value=f"{tong:,.0f}")
        cell.border = border
        cell.font = Font(bold=True)
        cell.alignment = center
    ws4.cell(row=r, column=8, value=f"{total_all:,.0f}").border = border
    ws4.cell(row=r, column=8).font = Font(bold=True)
    ws4.cell(row=r, column=8).alignment = center

    ws4.row_dimensions[4].height = 32
    for c, w in [('A', 5), ('B', 30), ('C', 12), ('D', 11), ('E', 11), ('F', 11), ('G', 13), ('H', 13)]:
        ws4.column_dimensions[c].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =============================================================================
# GENERATE REPORT 05 QUÝ — DANH MỤC CÔNG VIỆC
# =============================================================================

async def _generate_report_05_quy(db: AsyncSession, quy: int, nam: int) -> io.BytesIO:
    """Generate Excel File 05 quý — Danh mục công việc union 3 tháng."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = await _get_data_05_quy(db, quy, nam)

    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    dm_header_fill = PatternFill("solid", fgColor="2F5496")
    sub_header_fill = PatternFill("solid", fgColor="D6DCE4")
    sp1_fill = PatternFill("solid", fgColor="DAEEF3")
    sp2_fill = PatternFill("solid", fgColor="E2EFDA")
    sp3_fill = PatternFill("solid", fgColor="FDE9D9")
    sp4_fill = PatternFill("solid", fgColor="E4DFEC")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    # SHEET 1: TỔNG HỢP — V2 only
    ws1 = wb.active
    ws1.title = "Tổng hợp"

    ws1['A1'] = f"5. DANH MỤC CÔNG VIỆC (GỘP 3 THÁNG) - QUÝ {quy}/{nam}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:I1')

    ws1['A2'] = f"Tổng số đầu mục công việc: {len(data)} (V2_PL3 — theo Lĩnh vực + Nhóm PL3 + Điểm chấm)."
    ws1['A2'].font = Font(bold=True)

    headers = ["STT", "Lĩnh vực", "Tên công việc",
               "Số user kê khai", "Số lần kê khai (quý)", "Điểm (quý)",
               "Điểm chấm", "Nhóm PL3", "Khung điểm"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    row = 4
    for i, dm in enumerate(data, 1):
        row += 1
        ws1.cell(row=row, column=1, value=i).border = border
        ws1.cell(row=row, column=1).alignment = center_alignment

        lv_text = f"{dm['linh_vuc'] or ''}{' — ' + dm['ten_linh_vuc'] if dm['ten_linh_vuc'] else ''}".strip(" —")
        lv_cell = ws1.cell(row=row, column=2, value=lv_text or "—")
        lv_cell.border = border
        lv_cell.alignment = center_alignment
        lv_cell.font = Font(bold=True)

        ws1.cell(row=row, column=3, value=dm["ten_cong_viec"]).border = border
        ws1.cell(row=row, column=4, value=dm["so_user"]).border = border
        ws1.cell(row=row, column=4).alignment = center_alignment
        ws1.cell(row=row, column=5, value=dm["tong_lan_khai"]).border = border
        ws1.cell(row=row, column=5).alignment = center_alignment
        ws1.cell(row=row, column=6, value=f"{dm['tong_sp']:,.0f}").border = border
        ws1.cell(row=row, column=6).alignment = center_alignment

        dc_cell = ws1.cell(row=row, column=7, value=dm.get("diem_cham") or "—")
        dc_cell.border = border
        dc_cell.alignment = center_alignment

        nhom_cell = ws1.cell(row=row, column=8, value=dm.get("nhom_pl3") or "—")
        nhom_cell.border = border
        nhom_cell.alignment = center_alignment
        nhom_val = dm.get("nhom_pl3")
        if isinstance(nhom_val, int):
            if nhom_val >= 4:
                nhom_cell.fill = PatternFill("solid", fgColor="FFC7CE")
                nhom_cell.font = Font(bold=True, color="9C0006")
            elif nhom_val == 3:
                nhom_cell.fill = PatternFill("solid", fgColor="FFEB9C")

        kh_cell = ws1.cell(row=row, column=9, value=dm.get("khung_diem_toi_da") or "—")
        kh_cell.border = border
        kh_cell.alignment = center_alignment

    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 50
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 16
    ws1.column_dimensions['F'].width = 14
    ws1.column_dimensions['G'].width = 12
    ws1.column_dimensions['H'].width = 11
    ws1.column_dimensions['I'].width = 12

    if data:
        ws1.auto_filter.ref = f"A4:I{4 + len(data)}"

    # SHEET 2: CHI TIẾT — chỉ nhãn lĩnh vực (đã chuyển hoàn toàn V2)
    ws2 = wb.create_sheet("Chi tiết theo công việc")
    ws2['A1'] = f"CHI TIẾT DANH MỤC VÀ USER KÊ KHAI (3 THÁNG) - QUÝ {quy}/{nam}"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:H1')

    row = 3
    for dm_idx, dm in enumerate(data, 1):
        label = f"[LV {dm.get('linh_vuc') or '—'}]"
        ws2.merge_cells(f'A{row}:H{row}')
        header_text = (
            f"{dm_idx}. {label} {dm['ten_cong_viec']} "
            f"({dm['so_user']} user | {dm['tong_lan_khai']} lần | {dm['tong_sp']:,.0f} điểm)"
        )
        cell = ws2.cell(row=row, column=1, value=header_text)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = dm_header_fill
        for col in range(2, 9):
            ws2.cell(row=row, column=col).fill = dm_header_fill
        row += 1

        sub_headers = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "Số lần khai (quý)", "Điểm (quý)"]
        for col, h in enumerate(sub_headers, 1):
            cell = ws2.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True, size=10)
            cell.fill = sub_header_fill
            cell.border = border
            cell.alignment = center_alignment
        row += 1

        for u_idx, user in enumerate(dm["users"], 1):
            ws2.cell(row=row, column=1, value=u_idx).border = border
            ws2.cell(row=row, column=1).alignment = center_alignment
            ws2.cell(row=row, column=2, value=user["ho_ten"]).border = border
            ws2.cell(row=row, column=3, value=user["ma_cc"]).border = border
            ws2.cell(row=row, column=4, value=user["don_vi"]).border = border
            ws2.cell(row=row, column=5, value=user["so_lan_khai"]).border = border
            ws2.cell(row=row, column=5).alignment = center_alignment
            ws2.cell(row=row, column=6, value=f"{user['tong_sp']:,.0f}").border = border
            ws2.cell(row=row, column=6).alignment = center_alignment
            row += 1

        row += 1

    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 30
    ws2.column_dimensions['E'].width = 16
    ws2.column_dimensions['F'].width = 14
    ws2.column_dimensions['G'].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =============================================================================
# ENDPOINT: ZIP 5 BÁO CÁO THEO QUÝ
# =============================================================================

@router.get("/bao-cao-tong-hop/quy/{quy}/nam/{nam}")
async def export_bao_cao_tong_hop_quy(
    db: DatabaseDep,
    current_user: ActiveUserDep,
    quy: int,
    nam: int,
):
    """
    Xuất ZIP chứa 5 báo cáo thống kê Excel cho 1 QUÝ.

    Quyền: CCT, PCCT, hoặc users có can_view_all_units.

    5 báo cáo:
    1. 01_TieuChiChung_Q{quy}_{nam}.xlsx     - TC chung TB quý (hybrid snapshot/on-the-fly)
    2. 02_DiemKPI_Q{quy}_{nam}.xlsx          - Điểm KPI TB quý (lũy kế a/b/c, MIN d/đ/e)
    3. 03_LanhDaoDDE_Q{quy}_{nam}.xlsx       - LĐ bị trừ d,đ,e (union 3 tháng)
    4. 04_KhoiLuongCongViec_Q{quy}_{nam}.xlsx - Khối lượng CV (cộng dồn 3 tháng)
    5. 05_DanhMucCongViec_Q{quy}_{nam}.xlsx  - Danh mục CV (union 3 tháng)
    """
    # Validate
    if quy < 1 or quy > 4:
        raise HTTPException(400, detail=error_response("VAL_001", "Quý phải từ 1-4"))
    if nam < 2025:
        raise HTTPException(400, detail=error_response("VAL_002", "Năm phải >= 2025"))

    if not _is_lanh_dao_chi_cuc(current_user):
        raise HTTPException(403, detail=error_response(
            "PERM_003", "Chỉ CCT và Phó CCT mới được xuất báo cáo tổng hợp"
        ))

    try:
        logger.info(f"[EXPORT_ZIP_QUY] Generating 5 reports for Q{quy}/{nam}")

        excel_01 = await _generate_report_01_quy(db, quy, nam)
        excel_02 = await _generate_report_02_quy(db, quy, nam)
        excel_03 = await _generate_report_03_quy(db, quy, nam)
        excel_04 = await _generate_report_04_quy(db, quy, nam)
        excel_05 = await _generate_report_05_quy(db, quy, nam)

        import zipfile
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            zip_file.writestr(f"01_TieuChiChung_Q{quy}_{nam}.xlsx", excel_01.getvalue())
            zip_file.writestr(f"02_DiemKPI_Q{quy}_{nam}.xlsx", excel_02.getvalue())
            zip_file.writestr(f"03_LanhDaoDDE_Q{quy}_{nam}.xlsx", excel_03.getvalue())
            zip_file.writestr(f"04_KhoiLuongCongViec_Q{quy}_{nam}.xlsx", excel_04.getvalue())
            zip_file.writestr(f"05_DanhMucCongViec_Q{quy}_{nam}.xlsx", excel_05.getvalue())

        zip_buffer.seek(0)
        filename = f"BaoCaoTongHop_Q{quy}_{nam}.zip"

        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Access-Control-Expose-Headers": "Content-Disposition",
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[EXPORT_ZIP_QUY] Error: {traceback.format_exc()}")
        raise HTTPException(500, detail=error_response(
            "SYS_099", f"Lỗi xuất báo cáo tổng hợp quý: {str(e)[:200]}"
        ))
