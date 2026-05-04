"""
app/core/pl3_excel_parser.py
============================
Module dùng chung để parse file Excel PL3.

Dùng bởi:
- scripts/seed_pl3_catalog.py (CLI seed lần đầu)
- API endpoint admin POST /admin/danh-muc-pl3/import/{dry-run|commit}

Logic (thống nhất với Phase A):
- Detect 15 section headers I-XV bằng regex.
- Carry-forward nhiệm vụ: dòng STT N.0 chỉ có cột B (header nhiệm vụ),
  các dòng N.M bên dưới chỉ có cột C (công việc chi tiết) thuộc nhiệm
  vụ N. Parser maintain `current_nhiem_vu` cập nhật mỗi khi cột B có
  giá trị, reset khi sang lĩnh vực mới. Bug cũ (pre-2026-05): chỉ đọc
  cột B của row hiện tại → 93% mục bị nhiem_vu=NULL.
- Skip dòng nhiệm vụ header (chỉ có cột B, không có cột C/D) SAU KHI
  đã update state.
- Parse rows có đủ cong_viec_chi_tiet + san_pham_dau_ra + diem_cham + he_so + phan_nhom.
- Fallback nhom_pl3 từ khung_diem nếu phan_nhom rỗng.
- Append -r{row} cho duplicate ma_danh_muc.
- Trust Excel: chỉ validate (a) he_so ≈ diem_cham/25 (tol 0.05), (b) khung match nhóm.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from decimal import Decimal
from typing import Optional

import openpyxl


# =============================================================================
# CONSTANTS
# =============================================================================

SHEET_NAME = "PL3"
DATA_START_ROW = 8

NHOM_KHUNG_MAP = {1: 100, 2: 200, 3: 300, 4: 400, 5: 500}
KHUNG_TO_NHOM = {v: k for k, v in NHOM_KHUNG_MAP.items()}

# Cột Excel (1-based)
COL_STT = 1
COL_NHIEM_VU = 2
COL_CONG_VIEC = 3
COL_SAN_PHAM = 4
COL_PHAN_NHOM = 5
COL_KHUNG_DIEM = 6
COL_DIEM_CHAM = 11
COL_HE_SO = 12
COL_GHI_CHU = 13

SECTION_REGEX = re.compile(r"^([IVXLCDM]+)\.\s*(.+)$")

# Tolerance khi check he_so == diem_cham/25 (Excel có thể nhập làm tròn)
HE_SO_TOLERANCE = Decimal("0.05")


# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass
class ParsedRow:
    """1 row PL3 đã parse + validate."""
    ma_danh_muc: str
    ten_cong_viec: str
    nguon_du_lieu: str  # luôn 'PL3'
    linh_vuc: str
    ten_linh_vuc: Optional[str]
    nhiem_vu: Optional[str]
    cong_viec_chi_tiet: str
    san_pham_dau_ra: str
    nhom_pl3: int
    khung_diem_toi_da: int
    diem_cham: int
    he_so_quy_doi: Decimal
    mo_ta: Optional[str]
    is_active: bool = True
    excel_row: int = 0  # row gốc trong Excel


@dataclass
class ParseError:
    """1 lỗi parse/validate."""
    row: int
    ma_danh_muc: Optional[str]
    error: str


@dataclass
class ParseResult:
    """Kết quả parse 1 file Excel."""
    sections: list[tuple[int, str, str]]  # (row, ma_lv, ten_lv)
    rows: list[ParsedRow]
    errors: list[ParseError]
    skipped_no_data: int = 0
    skipped_pre_section: int = 0
    duplicate_count: int = 0
    total_excel_rows: int = 0


# =============================================================================
# HELPERS
# =============================================================================

def _to_int(val) -> Optional[int]:
    if val is None or val == "":
        return None
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return None


def _to_decimal(val) -> Optional[Decimal]:
    if val is None or val == "":
        return None
    try:
        return Decimal(str(val))
    except (TypeError, ValueError):
        return None


def _detect_nhom_from_phan_nhom(s) -> Optional[int]:
    """'Nhóm 3' → 3."""
    if not s:
        return None
    m = re.search(r"(\d+)", str(s))
    if not m:
        return None
    n = int(m.group(1))
    return n if 1 <= n <= 5 else None


def _is_section_header(val) -> Optional[tuple[str, str]]:
    if val is None:
        return None
    s = str(val).strip()
    m = SECTION_REGEX.match(s)
    if m:
        return m.group(1).upper(), m.group(2).strip()
    return None


# =============================================================================
# PARSE
# =============================================================================

def parse_pl3_excel(source) -> ParseResult:
    """
    Parse file Excel PL3 từ nguồn source.

    Args:
        source: path str / Path / bytes / BinaryIO. openpyxl chấp nhận tất cả.

    Returns:
        ParseResult với rows + errors (không insert DB).
    """
    if isinstance(source, bytes):
        source = io.BytesIO(source)

    wb = openpyxl.load_workbook(source, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{SHEET_NAME}' không tồn tại. Có: {wb.sheetnames}"
        )
    ws = wb[SHEET_NAME]

    result = ParseResult(sections=[], rows=[], errors=[])
    result.total_excel_rows = ws.max_row

    # 1) Pre-scan section headers
    for r in range(1, ws.max_row + 1):
        info = _is_section_header(ws.cell(row=r, column=COL_STT).value)
        if info:
            result.sections.append((r, info[0], info[1]))

    if len(result.sections) != 15:
        result.errors.append(ParseError(
            row=0,
            ma_danh_muc=None,
            error=f"Kỳ vọng 15 section headers (I-XV), tìm thấy {len(result.sections)}",
        ))
        return result

    def _linh_vuc_at(row: int) -> tuple[Optional[str], Optional[str]]:
        cur_lv, cur_ten = None, None
        for sr, lv, ten in result.sections:
            if sr <= row:
                cur_lv, cur_ten = lv, ten
            else:
                break
        return cur_lv, cur_ten

    # 2) Parse data rows
    seen_ma: set[str] = set()
    duplicate_count = 0
    # State machine: nhiệm vụ "đang mở" — cập nhật khi gặp row có cột B,
    # reset khi sang lĩnh vực mới. Áp cho mọi row có công việc chi tiết
    # nhưng cột B trống (= row con thuộc nhiệm vụ ở trên).
    current_nhiem_vu: Optional[str] = None

    for r in range(DATA_START_ROW, ws.max_row + 1):
        stt_raw = ws.cell(row=r, column=COL_STT).value

        if _is_section_header(stt_raw):
            # Sang lĩnh vực mới → reset state nhiệm vụ.
            # Sub-heading nội bộ (vd "Công tác tổng hợp...") không match
            # SECTION_REGEX (không bắt đầu bằng La Mã + dấu chấm) nên KHÔNG
            # reset — giữ nhiệm vụ đang mở.
            current_nhiem_vu = None
            continue

        nhiem_vu_cell = ws.cell(row=r, column=COL_NHIEM_VU).value
        cong_viec = ws.cell(row=r, column=COL_CONG_VIEC).value
        san_pham = ws.cell(row=r, column=COL_SAN_PHAM).value
        phan_nhom_raw = ws.cell(row=r, column=COL_PHAN_NHOM).value
        khung_diem_excel = _to_int(ws.cell(row=r, column=COL_KHUNG_DIEM).value)
        diem_cham = _to_int(ws.cell(row=r, column=COL_DIEM_CHAM).value)
        he_so = _to_decimal(ws.cell(row=r, column=COL_HE_SO).value)
        ghi_chu = ws.cell(row=r, column=COL_GHI_CHU).value

        # Cột B có data → cập nhật state nhiệm vụ. Trường hợp R8 (STT=1.0)
        # và R41 (STT=7.0): vừa khai báo nhiệm vụ vừa kèm 1 công việc chi
        # tiết → vẫn pass guard bên dưới và row đó nhận chính nhiệm vụ này.
        if nhiem_vu_cell:
            current_nhiem_vu = str(nhiem_vu_cell).strip()

        if not cong_viec or not san_pham or diem_cham is None or he_so is None:
            result.skipped_no_data += 1
            continue

        linh_vuc, ten_linh_vuc = _linh_vuc_at(r)
        if not linh_vuc:
            result.skipped_pre_section += 1
            continue

        nhom = _detect_nhom_from_phan_nhom(phan_nhom_raw)
        if nhom is None and khung_diem_excel in KHUNG_TO_NHOM:
            nhom = KHUNG_TO_NHOM[khung_diem_excel]
        if nhom is None:
            result.errors.append(ParseError(
                row=r,
                ma_danh_muc=None,
                error=f"Không xác định được nhóm: phan_nhom='{phan_nhom_raw}', khung_diem='{khung_diem_excel}'",
            ))
            continue

        # Validate (a): he_so ≈ diem_cham / 25 (tolerance)
        expected_he_so = Decimal(diem_cham) / Decimal("25")
        if abs(he_so - expected_he_so) > HE_SO_TOLERANCE:
            result.errors.append(ParseError(
                row=r,
                ma_danh_muc=None,
                error=f"he_so={he_so} không khớp diem_cham/25={expected_he_so} (tolerance {HE_SO_TOLERANCE})",
            ))
            continue

        # Validate (b): khung_diem từ nhóm
        khung_from_nhom = NHOM_KHUNG_MAP[nhom]
        # Build ma_danh_muc
        if isinstance(stt_raw, (int, float)):
            stt_str = str(int(stt_raw)) if float(stt_raw).is_integer() else str(stt_raw)
        else:
            stt_str = str(stt_raw).strip() if stt_raw is not None else f"r{r}"

        ma_danh_muc = f"PL3-{linh_vuc}-{stt_str}"
        if ma_danh_muc in seen_ma:
            ma_danh_muc = f"{ma_danh_muc}-r{r}"
            duplicate_count += 1
        seen_ma.add(ma_danh_muc)

        if len(ma_danh_muc) > 30:
            result.errors.append(ParseError(
                row=r,
                ma_danh_muc=ma_danh_muc,
                error=f"ma_danh_muc='{ma_danh_muc}' > 30 chars",
            ))
            continue

        result.rows.append(ParsedRow(
            ma_danh_muc=ma_danh_muc,
            ten_cong_viec=str(cong_viec).strip()[:500],
            nguon_du_lieu="PL3",
            linh_vuc=linh_vuc,
            ten_linh_vuc=ten_linh_vuc,
            nhiem_vu=current_nhiem_vu[:500] if current_nhiem_vu else None,
            cong_viec_chi_tiet=str(cong_viec).strip(),
            san_pham_dau_ra=str(san_pham).strip(),
            nhom_pl3=nhom,
            khung_diem_toi_da=khung_from_nhom,
            diem_cham=diem_cham,
            he_so_quy_doi=he_so,
            mo_ta=str(ghi_chu).strip() if ghi_chu else None,
            is_active=True,
            excel_row=r,
        ))

    result.duplicate_count = duplicate_count
    return result
