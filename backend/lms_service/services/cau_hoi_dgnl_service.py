"""
lms_service/services/cau_hoi_dgnl_service.py
=============================================
Business logic cho ngan hang cau hoi DGNL.
CRUD + import Excel + thong ke.
"""

import io
import math
import uuid
from datetime import datetime
from decimal import Decimal
from typing import Optional

from fastapi import HTTPException, UploadFile, status
from sqlalchemy import select, func, case
from sqlalchemy.ext.asyncio import AsyncSession

from lms_service.models.base import CongChucRef
from lms_service.models.cau_hoi_dgnl import CauHoiDgnl
from lms_service.models.linh_vuc import LinhVuc
from lms_service.schemas.cau_hoi_dgnl import (
    CauHoiDgnlCreate,
    CauHoiDgnlUpdate,
    CauHoiDgnlXoaNhieu,
    ThongKeNganHang,
)
from shared.auth import TokenPayload

LOAI_CAU_HOI = ["TRAC_NGHIEM_1", "TRAC_NGHIEM_NHIEU", "DUNG_SAI", "TU_LUAN"]
DO_KHO_VALUES = ["DE", "TRUNG_BINH", "KHO"]


class CauHoiDgnlService:
    def __init__(self, db: AsyncSession):
        self.db = db

    def _is_manager(self, user: TokenPayload) -> bool:
        return user.vai_tro == "SUPER_ADMIN" or "QT_DAO_TAO" in (user.platform_roles or [])

    # ================================================================
    # CRUD
    # ================================================================

    async def danh_sach(
        self,
        linh_vuc_id: Optional[uuid.UUID] = None,
        do_kho: Optional[str] = None,
        loai: Optional[str] = None,
        search: Optional[str] = None,
        page: int = 1,
        page_size: int = 20,
    ) -> dict:
        base_where = [CauHoiDgnl.is_active == True]
        if linh_vuc_id:
            base_where.append(CauHoiDgnl.linh_vuc_id == linh_vuc_id)
        if do_kho:
            base_where.append(CauHoiDgnl.do_kho == do_kho)
        if loai:
            base_where.append(CauHoiDgnl.loai == loai)
        if search:
            base_where.append(CauHoiDgnl.noi_dung.ilike(f"%{search}%"))

        count_stmt = select(func.count(CauHoiDgnl.id)).where(*base_where)
        total = (await self.db.execute(count_stmt)).scalar() or 0

        nt = CongChucRef.__table__.alias("nt")
        stmt = (
            select(
                CauHoiDgnl,
                LinhVuc.ten_linh_vuc.label("linh_vuc_ten"),
                nt.c.ho_ten.label("nguoi_tao_ho_ten"),
            )
            .join(LinhVuc, CauHoiDgnl.linh_vuc_id == LinhVuc.id)
            .outerjoin(nt, CauHoiDgnl.nguoi_tao_id == nt.c.id)
            .where(*base_where)
            .order_by(CauHoiDgnl.created_at.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
        result = await self.db.execute(stmt)
        rows = result.all()

        items = []
        for ch, lv_ten, nt_ten in rows:
            items.append({
                **{c.key: getattr(ch, c.key) for c in ch.__table__.columns},
                "linh_vuc_ten": lv_ten,
                "nguoi_tao_ho_ten": nt_ten,
            })

        return {
            "items": items,
            "pagination": {
                "page": page, "page_size": page_size,
                "total_items": total,
                "total_pages": math.ceil(total / page_size) if total > 0 else 0,
            },
        }

    async def chi_tiet(self, id: uuid.UUID) -> CauHoiDgnl:
        stmt = select(CauHoiDgnl).where(CauHoiDgnl.id == id, CauHoiDgnl.is_active == True)
        result = await self.db.execute(stmt)
        ch = result.scalar_one_or_none()
        if not ch:
            raise HTTPException(status_code=404, detail={
                "success": False, "error": {"code": "DGNL_040", "message": "Câu hỏi ĐGNL không tồn tại"}
            })
        return ch

    async def tao_moi(self, data: CauHoiDgnlCreate, user: TokenPayload) -> CauHoiDgnl:
        # Validate linh_vuc
        lv = await self.db.execute(
            select(LinhVuc).where(LinhVuc.id == data.linh_vuc_id, LinhVuc.is_active == True)
        )
        if not lv.scalar_one_or_none():
            raise HTTPException(status_code=404, detail={
                "success": False, "error": {"code": "DGNL_001", "message": "Lĩnh vực không tồn tại"}
            })

        ch = CauHoiDgnl(
            linh_vuc_id=data.linh_vuc_id,
            noi_dung=data.noi_dung,
            giai_thich=data.giai_thich,
            loai=data.loai,
            dap_an=data.dap_an,
            diem=data.diem,
            do_kho=data.do_kho,
            nguoi_tao_id=uuid.UUID(user.sub),
        )
        self.db.add(ch)
        await self.db.commit()
        await self.db.refresh(ch)
        return ch

    async def cap_nhat(self, id: uuid.UUID, data: CauHoiDgnlUpdate, user: TokenPayload) -> CauHoiDgnl:
        ch = await self.chi_tiet(id)
        if not self._is_manager(user) and str(ch.nguoi_tao_id) != user.sub:
            raise HTTPException(status_code=403, detail={
                "success": False, "error": {"code": "DGNL_041", "message": "Bạn chỉ được sửa câu hỏi của mình"}
            })
        for field, value in data.model_dump(exclude_unset=True).items():
            setattr(ch, field, value)
        ch.updated_at = datetime.utcnow()
        await self.db.commit()
        await self.db.refresh(ch)
        return ch

    async def xoa(self, id: uuid.UUID, user: TokenPayload) -> None:
        ch = await self.chi_tiet(id)
        ch.is_active = False
        ch.updated_at = datetime.utcnow()
        await self.db.commit()

    async def xoa_nhieu(self, payload: "CauHoiDgnlXoaNhieu", user: TokenPayload) -> int:
        """Xoa mem nhieu cau hoi cung luc. Tra ve so cau da xoa.

        Chi QT_DAO_TAO (hoac SUPER_ADMIN) moi duoc xoa hang loat.
        Xoa mem (is_active=False) — giu nguyen id de cac bai thi cu
        (thi_sinh.de_thi_ids) van tham chieu duoc lich su.
        """
        if not self._is_manager(user):
            raise HTTPException(status_code=403, detail={
                "success": False,
                "error": {"code": "DGNL_045", "message": "Chỉ Quản trị đào tạo được xóa hàng loạt"},
            })

        from sqlalchemy import update

        conds = [CauHoiDgnl.is_active == True]

        if payload.tat_ca_theo_bo_loc:
            # Xoa toan bo theo bo loc — BAT BUOC co it nhat 1 dieu kien loc
            # de tranh xoa nham toan bo ngan hang.
            loc = []
            if payload.linh_vuc_id:
                loc.append(CauHoiDgnl.linh_vuc_id == payload.linh_vuc_id)
            if payload.do_kho:
                loc.append(CauHoiDgnl.do_kho == payload.do_kho)
            if payload.loai:
                loc.append(CauHoiDgnl.loai == payload.loai)
            if payload.search:
                loc.append(CauHoiDgnl.noi_dung.ilike(f"%{payload.search}%"))
            if not loc:
                raise HTTPException(status_code=400, detail={
                    "success": False,
                    "error": {"code": "DGNL_046",
                              "message": "Phải chọn ít nhất một bộ lọc (lĩnh vực/độ khó/loại/tìm kiếm) khi xóa tất cả"},
                })
            conds.extend(loc)
        else:
            if not payload.ids:
                raise HTTPException(status_code=400, detail={
                    "success": False,
                    "error": {"code": "DGNL_047", "message": "Chưa chọn câu hỏi nào để xóa"},
                })
            conds.append(CauHoiDgnl.id.in_(payload.ids))

        stmt = (
            update(CauHoiDgnl)
            .where(*conds)
            .values(is_active=False, updated_at=datetime.utcnow())
            .execution_options(synchronize_session=False)
        )
        result = await self.db.execute(stmt)
        await self.db.commit()
        return result.rowcount or 0

    # ================================================================
    # THONG KE
    # ================================================================

    async def thong_ke(self) -> list[ThongKeNganHang]:
        stmt = (
            select(
                LinhVuc.id,
                LinhVuc.ten_linh_vuc,
                func.count(CauHoiDgnl.id).filter(CauHoiDgnl.do_kho == "DE").label("so_cau_de"),
                func.count(CauHoiDgnl.id).filter(CauHoiDgnl.do_kho == "TRUNG_BINH").label("so_cau_trung_binh"),
                func.count(CauHoiDgnl.id).filter(CauHoiDgnl.do_kho == "KHO").label("so_cau_kho"),
                func.count(CauHoiDgnl.id).label("tong"),
            )
            .outerjoin(CauHoiDgnl, (LinhVuc.id == CauHoiDgnl.linh_vuc_id) & (CauHoiDgnl.is_active == True))
            .where(LinhVuc.is_active == True)
            .group_by(LinhVuc.id, LinhVuc.ten_linh_vuc)
            .order_by(LinhVuc.thu_tu)
        )
        result = await self.db.execute(stmt)
        return [
            ThongKeNganHang(
                linh_vuc_id=str(row.id),
                linh_vuc_ten=row.ten_linh_vuc,
                so_cau_de=row.so_cau_de,
                so_cau_trung_binh=row.so_cau_trung_binh,
                so_cau_kho=row.so_cau_kho,
                tong=row.tong,
            )
            for row in result.all()
        ]

    # ================================================================
    # IMPORT EXCEL
    # ================================================================

    def generate_template(self) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Câu hỏi ĐGNL"

        headers = ["noi_dung", "loai", "do_kho", "linh_vuc",
                    "dap_an_a", "dap_an_b", "dap_an_c", "dap_an_d",
                    "dap_an_dung", "diem", "giai_thich"]
        widths = [50, 20, 15, 25, 25, 25, 25, 25, 15, 8, 40]

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            ws.column_dimensions[cell.column_letter].width = w

        # Sample rows
        samples = [
            ["Hải quan là gì?", "TRAC_NGHIEM_1", "DE", "KIEN_THUC_CHUNG",
             "Cơ quan nhà nước", "Tổ chức xã hội", "Doanh nghiệp", "Đoàn thể",
             "A", 1, "Hải quan là cơ quan nhà nước"],
            ["Thuế XNK gồm?", "TRAC_NGHIEM_NHIEU", "TRUNG_BINH", "THUE_HQ",
             "Thuế NK", "Thuế XK", "Thuế GTGT", "Phí hải quan",
             "A,B", 1, ""],
            ["Buôn lậu là hành vi vi phạm pháp luật", "DUNG_SAI", "DE", "CHONG_BUON_LAU",
             "Đúng", "Sai", "", "",
             "A", 1, ""],
        ]
        for i, row in enumerate(samples, 2):
            for col, val in enumerate(row, 1):
                ws.cell(row=i, column=col, value=val)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def import_from_file(self, file: UploadFile, db: AsyncSession, user_id: str) -> dict:
        filename = file.filename or ""
        if not filename.endswith((".xlsx", ".csv")):
            raise HTTPException(status_code=400, detail={
                "success": False, "error": {"code": "DGNL_042", "message": "Chỉ hỗ trợ file .xlsx hoặc .csv"}
            })

        content = await file.read()

        if filename.endswith(".xlsx"):
            rows = self._parse_xlsx(content)
        else:
            rows = self._parse_csv(content)

        if not rows:
            raise HTTPException(status_code=400, detail={
                "success": False, "error": {"code": "DGNL_043", "message": "File không có dữ liệu"}
            })

        # Lay map linh_vuc
        lv_result = await db.execute(select(LinhVuc).where(LinhVuc.is_active == True))
        lv_map = {lv.ma_linh_vuc: lv.id for lv in lv_result.scalars().all()}

        thanh_cong = 0
        that_bai = 0
        loi_chi_tiet = []
        user_uuid = uuid.UUID(user_id)

        for i, row in enumerate(rows, 2):
            try:
                noi_dung = (row.get("noi_dung") or "").strip()
                if not noi_dung:
                    continue  # skip empty rows

                loai = (row.get("loai") or "TRAC_NGHIEM_1").strip().upper()
                do_kho = (row.get("do_kho") or "TRUNG_BINH").strip().upper()
                ma_lv = (row.get("linh_vuc") or "").strip().upper()

                if loai not in LOAI_CAU_HOI:
                    raise ValueError(f"Loại '{loai}' không hợp lệ")
                if do_kho not in DO_KHO_VALUES:
                    raise ValueError(f"Độ khó '{do_kho}' không hợp lệ")
                if ma_lv not in lv_map:
                    raise ValueError(f"Lĩnh vực '{ma_lv}' không tồn tại. Mã hợp lệ: {list(lv_map.keys())}")

                dap_an = self._parse_dap_an(row, loai)
                diem = float(row.get("diem") or 1)
                giai_thich = (row.get("giai_thich") or "").strip() or None

                ch = CauHoiDgnl(
                    linh_vuc_id=lv_map[ma_lv],
                    noi_dung=noi_dung,
                    giai_thich=giai_thich,
                    loai=loai,
                    dap_an=dap_an,
                    diem=Decimal(str(diem)),
                    do_kho=do_kho,
                    nguoi_tao_id=user_uuid,
                )
                db.add(ch)
                thanh_cong += 1

            except Exception as e:
                that_bai += 1
                loi_chi_tiet.append({"dong": i, "loi": str(e)})

        if thanh_cong > 0:
            await db.commit()

        return {
            "thanh_cong": thanh_cong,
            "that_bai": that_bai,
            "tong": thanh_cong + that_bai,
            "loi_chi_tiet": loi_chi_tiet,
        }

    def _parse_xlsx(self, content: bytes) -> list[dict]:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip().lower().replace(" ", "_") if h else f"col_{i}" for i, h in enumerate(next(rows_iter, []))]
        result = []
        for row in rows_iter:
            d = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row) if i < len(headers)}
            if any(v for v in d.values()):
                result.append(d)
        return result

    def _parse_csv(self, content: bytes) -> list[dict]:
        import csv
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [{k.strip().lower().replace(" ", "_"): (v or "").strip() for k, v in row.items()} for row in reader]

    def _parse_dap_an(self, row: dict, loai: str) -> dict:
        dap_an_dung = (row.get("dap_an_dung") or "").strip().upper()
        options = []
        for key in ["dap_an_a", "dap_an_b", "dap_an_c", "dap_an_d"]:
            val = (row.get(key) or "").strip()
            if val:
                options.append(val)

        if loai == "TRAC_NGHIEM_1":
            idx_map = {"A": 0, "B": 1, "C": 2, "D": 3}
            correct = idx_map.get(dap_an_dung, 0)
            return {
                "lua_chon": [{"key": chr(65 + i), "noi_dung": o} for i, o in enumerate(options)],
                "dap_an_dung": dap_an_dung,
            }

        elif loai == "TRAC_NGHIEM_NHIEU":
            correct_keys = [k.strip() for k in dap_an_dung.split(",")]
            return {
                "lua_chon": [{"key": chr(65 + i), "noi_dung": o} for i, o in enumerate(options)],
                "dap_an_dung": correct_keys,
            }

        elif loai == "DUNG_SAI":
            return {
                "lua_chon": [{"key": "A", "noi_dung": options[0] if options else "Đúng"},
                             {"key": "B", "noi_dung": options[1] if len(options) > 1 else "Sai"}],
                "dap_an_dung": dap_an_dung,
            }

        elif loai == "TU_LUAN":
            return {"goi_y": dap_an_dung or ""}

        return {}
