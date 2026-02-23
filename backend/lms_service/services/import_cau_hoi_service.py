"""
lms_service/services/import_cau_hoi_service.py
===============================================
Service import câu hỏi hàng loạt từ file Excel (.xlsx) hoặc CSV (.csv).

Format file mong đợi:
  noi_dung | loai | dap_an_a | dap_an_b | dap_an_c | dap_an_d | dap_an_dung | do_kho | diem | giai_thich

Quy tắc đáp án (cột dap_an_dung):
  - TRAC_NGHIEM_1:    "A" | "B" | "C" | "D"          → index 0/1/2/3
  - TRAC_NGHIEM_NHIEU: "A,C" | "A,B,D"               → [0, 2] | [0, 1, 3]
  - DUNG_SAI:          "A" = Đúng, "B" = Sai          → True / False
  - TU_LUAN:           text tự do (gợi ý đáp án)
"""

import csv
import io
import uuid
from pathlib import Path
from typing import Optional

from fastapi import HTTPException, UploadFile
from sqlalchemy.ext.asyncio import AsyncSession

from lms_service.models.cau_hoi import CauHoi


# Các cột bắt buộc phải có trong header
REQUIRED_COLUMNS = {"noi_dung", "loai", "dap_an_dung"}

# Loại câu hỏi hợp lệ
LOAI_HOP_LE = {"TRAC_NGHIEM_1", "TRAC_NGHIEM_NHIEU", "DUNG_SAI", "TU_LUAN"}

# Độ khó hợp lệ
DO_KHO_HOP_LE = {"DE", "TRUNG_BINH", "KHO"}


class ImportCauHoiService:
    """Service import câu hỏi từ file Excel/CSV."""

    async def import_from_file(
        self,
        file: UploadFile,
        khoa_hoc_id: Optional[str],
        db: AsyncSession,
        user_id: str,
        bai_kiem_tra_id: Optional[str] = None,
    ) -> dict:
        """
        Đọc file Excel/CSV → parse từng dòng → tạo câu hỏi trong DB.

        Nếu bai_kiem_tra_id được cung cấp, tự động lấy khoa_hoc_id từ BKT
        và gắn tất cả câu hỏi vào BKT đó.

        Returns:
            {
              "thanh_cong": int,
              "that_bai": int,
              "tong": int,
              "loi_chi_tiet": [{"dong": int, "loi": str}]
            }
        """
        # Xác định định dạng file
        filename = file.filename or ""
        ext = Path(filename).suffix.lower()

        if ext == ".xlsx":
            rows = self._read_xlsx(file)
        elif ext == ".csv":
            rows = self._read_csv(file)
        else:
            raise HTTPException(
                status_code=400,
                detail={
                    "success": False,
                    "error": {
                        "code": "LMS_IMPORT_001",
                        "message": "Chỉ hỗ trợ file .xlsx hoặc .csv",
                    },
                },
            )

        if not rows:
            raise HTTPException(
                status_code=400,
                detail={
                    "success": False,
                    "error": {
                        "code": "LMS_IMPORT_002",
                        "message": "File trống hoặc không đọc được",
                    },
                },
            )

        # Kiểm tra header — cột bắt buộc
        header_set = set(rows[0].keys()) if rows else set()
        missing = REQUIRED_COLUMNS - header_set
        if missing:
            raise HTTPException(
                status_code=400,
                detail={
                    "success": False,
                    "error": {
                        "code": "LMS_IMPORT_003",
                        "message": f"File thiếu cột bắt buộc: {', '.join(sorted(missing))}",
                    },
                },
            )

        # Validate và resolve bai_kiem_tra_id
        bkt_uuid: Optional[uuid.UUID] = None
        if bai_kiem_tra_id:
            try:
                bkt_uuid = uuid.UUID(bai_kiem_tra_id)
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail={
                        "success": False,
                        "error": {
                            "code": "LMS_IMPORT_004",
                            "message": "bai_kiem_tra_id không hợp lệ",
                        },
                    },
                )
            # Nếu không có khoa_hoc_id, tự lấy từ BKT
            if not khoa_hoc_id:
                from lms_service.models.bai_kiem_tra import BaiKiemTra
                from sqlalchemy import select
                bkt_r = await db.execute(select(BaiKiemTra).where(BaiKiemTra.id == bkt_uuid))
                bkt = bkt_r.scalar_one_or_none()
                if bkt is None:
                    raise HTTPException(
                        status_code=404,
                        detail={
                            "success": False,
                            "error": {
                                "code": "LMS_ERR_001",
                                "message": "Không tìm thấy bài kiểm tra",
                            },
                        },
                    )
                if bkt.khoa_hoc_id:
                    khoa_hoc_id = str(bkt.khoa_hoc_id)

        # Chuyển khoa_hoc_id thành UUID nếu có
        khoa_hoc_uuid: Optional[uuid.UUID] = None
        if khoa_hoc_id:
            try:
                khoa_hoc_uuid = uuid.UUID(khoa_hoc_id)
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail={
                        "success": False,
                        "error": {
                            "code": "LMS_IMPORT_004",
                            "message": "khoa_hoc_id không hợp lệ",
                        },
                    },
                )

        user_uuid = uuid.UUID(user_id)

        # Parse và tạo từng câu hỏi
        thanh_cong = 0
        that_bai = 0
        loi_chi_tiet = []
        created_objects: list[CauHoi] = []  # Thu thập để lấy ID sau flush

        for i, row in enumerate(rows, start=2):  # dòng 2 trở đi (dòng 1 là header)
            try:
                # Làm sạch giá trị — xử lý cả None từ ô trống Excel
                cleaned = {
                    k: (str(v).strip() if v is not None else "")
                    for k, v in row.items()
                }

                # Nội dung câu hỏi — bắt buộc
                noi_dung = cleaned.get("noi_dung", "")
                if not noi_dung:
                    loi_chi_tiet.append({"dong": i, "loi": "Thiếu nội dung câu hỏi (cột noi_dung)"})
                    that_bai += 1
                    continue

                # Loại câu hỏi — mặc định TRAC_NGHIEM_1
                loai = cleaned.get("loai", "").upper() or "TRAC_NGHIEM_1"
                if loai not in LOAI_HOP_LE:
                    loai = "TRAC_NGHIEM_1"

                # Parse đáp án
                try:
                    dap_an = self._parse_dap_an(cleaned, loai)
                except ValueError as ve:
                    loi_chi_tiet.append({"dong": i, "loi": f"Lỗi đáp án: {ve}"})
                    that_bai += 1
                    continue

                # Độ khó — mặc định TRUNG_BINH
                do_kho = cleaned.get("do_kho", "").upper() or "TRUNG_BINH"
                if do_kho not in DO_KHO_HOP_LE:
                    do_kho = "TRUNG_BINH"

                # Điểm — mặc định 1.0
                try:
                    diem_raw = cleaned.get("diem", "") or "1.0"
                    diem = float(diem_raw)
                    if diem < 0:
                        diem = 1.0
                except (ValueError, TypeError):
                    diem = 1.0

                # Giải thích — tuỳ chọn
                giai_thich = cleaned.get("giai_thich", "") or None

                # Tạo câu hỏi (gắn bai_kiem_tra_id nếu có)
                cau_hoi = CauHoi(
                    noi_dung=noi_dung,
                    giai_thich=giai_thich,
                    loai=loai,
                    do_kho=do_kho,
                    diem=diem,
                    dap_an=dap_an,
                    khoa_hoc_id=khoa_hoc_uuid,
                    bai_kiem_tra_id=bkt_uuid,
                    nguoi_tao_id=user_uuid,
                    is_active=True,
                )
                db.add(cau_hoi)
                created_objects.append(cau_hoi)
                thanh_cong += 1

            except Exception as e:
                loi_chi_tiet.append({"dong": i, "loi": str(e)})
                that_bai += 1

        # Flush để DB sinh UUID (server_default) và resolve id trên mỗi object
        if thanh_cong > 0:
            await db.flush()

        # Lấy IDs sau flush — dùng cho frontend thêm vào local state
        cau_hoi_ids = [str(ch.id) for ch in created_objects]

        return {
            "thanh_cong": thanh_cong,
            "that_bai": that_bai,
            "tong": thanh_cong + that_bai,
            "loi_chi_tiet": loi_chi_tiet,
            "cau_hoi_ids": cau_hoi_ids,
        }

    def _parse_dap_an(self, row: dict, loai: str) -> dict:
        """
        Chuyển đổi cột dap_an_a/b/c/d + dap_an_dung thành JSONB format.

        Format JSONB cho từng loại:
          TRAC_NGHIEM_1:    { "options": [...], "correct": 0 }
          TRAC_NGHIEM_NHIEU: { "options": [...], "correct": [0, 2] }
          DUNG_SAI:         { "correct": true/false }
          TU_LUAN:          { "goi_y": "..." }
        """
        if loai == "TU_LUAN":
            goi_y = row.get("dap_an_dung", "")
            return {"goi_y": goi_y}

        if loai == "DUNG_SAI":
            dap_an_dung = row.get("dap_an_dung", "A").upper()
            return {"correct": dap_an_dung == "A"}

        # Thu thập options A, B, C, D (bỏ qua ô trống)
        options = []
        for col in ["dap_an_a", "dap_an_b", "dap_an_c", "dap_an_d"]:
            val = row.get(col, "")
            if val:
                options.append(val)

        if len(options) < 2:
            raise ValueError("Cần ít nhất 2 đáp án (dap_an_a, dap_an_b)")

        dap_an_dung_raw = row.get("dap_an_dung", "").upper().strip()

        if loai == "TRAC_NGHIEM_1":
            if not dap_an_dung_raw:
                raise ValueError("Thiếu dap_an_dung cho TRAC_NGHIEM_1")
            letter = dap_an_dung_raw[0]
            idx = ord(letter) - ord("A")
            if idx < 0 or idx >= len(options):
                raise ValueError(
                    f"dap_an_dung '{letter}' không hợp lệ (chỉ A-{chr(ord('A') + len(options) - 1)})"
                )
            return {"options": options, "correct": idx}

        if loai == "TRAC_NGHIEM_NHIEU":
            if not dap_an_dung_raw:
                raise ValueError("Thiếu dap_an_dung cho TRAC_NGHIEM_NHIEU")
            parts = [p.strip() for p in dap_an_dung_raw.split(",") if p.strip()]
            indices = []
            for p in parts:
                idx = ord(p[0]) - ord("A")
                if idx < 0 or idx >= len(options):
                    raise ValueError(f"dap_an_dung '{p}' không hợp lệ")
                indices.append(idx)
            if not indices:
                raise ValueError("Không có đáp án đúng nào cho TRAC_NGHIEM_NHIEU")
            return {"options": options, "correct": sorted(set(indices))}

        return {"options": options, "correct": 0}

    def _read_xlsx(self, file: UploadFile) -> list[dict]:
        """Đọc file Excel .xlsx → list of dicts (mỗi dict là 1 dòng data)."""
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail={
                    "success": False,
                    "error": {
                        "code": "LMS_IMPORT_005",
                        "message": "Server chưa cài thư viện openpyxl. Chạy: pip install openpyxl",
                    },
                },
            )

        content = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not all_rows:
            return []

        # Dòng đầu là header — chuẩn hóa về snake_case lowercase
        headers = [
            str(h).strip().lower().replace(" ", "_") if h is not None else f"col_{i}"
            for i, h in enumerate(all_rows[0])
        ]

        result = []
        for row in all_rows[1:]:
            # Bỏ qua dòng hoàn toàn trống
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            result.append(dict(zip(headers, row)))

        return result

    def _read_csv(self, file: UploadFile) -> list[dict]:
        """Đọc file CSV → list of dicts (mỗi dict là 1 dòng data)."""
        content = file.file.read()

        # Thử UTF-8 BOM trước (Excel thường export với BOM)
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("utf-8", errors="replace")

        reader = csv.DictReader(io.StringIO(text))

        # Chuẩn hóa tên cột
        rows = []
        for raw_row in reader:
            normalized = {
                k.strip().lower().replace(" ", "_"): v
                for k, v in raw_row.items()
                if k
            }
            # Bỏ qua dòng hoàn toàn trống
            if all(not str(v or "").strip() for v in normalized.values()):
                continue
            rows.append(normalized)

        return rows

    def generate_template_xlsx(self) -> bytes:
        """
        Tạo file Excel mẫu với header + 3 dòng ví dụ.
        Trả về bytes để stream về client.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail={
                    "success": False,
                    "error": {
                        "code": "LMS_IMPORT_005",
                        "message": "Server chưa cài thư viện openpyxl",
                    },
                },
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mẫu Import Câu Hỏi"

        # Header
        headers = [
            "noi_dung", "loai", "dap_an_a", "dap_an_b", "dap_an_c", "dap_an_d",
            "dap_an_dung", "do_kho", "diem", "giai_thich",
        ]
        ws.append(headers)

        # Định dạng header — nền xanh, chữ trắng, in đậm
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 3 dòng dữ liệu mẫu
        samples = [
            [
                "Cơ quan hải quan có chức năng gì?",
                "TRAC_NGHIEM_1",
                "Quản lý thuế nội địa",
                "Quản lý xuất nhập khẩu và thu thuế",
                "Quản lý biên giới",
                "Quản lý an ninh",
                "B",
                "DE",
                "1.0",
                "Hải quan quản lý hoạt động XNK và thu thuế XNK.",
            ],
            [
                "Các loại thuế hải quan bao gồm?",
                "TRAC_NGHIEM_NHIEU",
                "Thuế xuất khẩu",
                "Thuế nhập khẩu",
                "Thuế GTGT",
                "Thuế TNCN",
                "A,B,C",
                "TRUNG_BINH",
                "1.5",
                "Thuế hải quan gồm thuế XK, NK và GTGT.",
            ],
            [
                "Khai báo hải quan điện tử là bắt buộc?",
                "DUNG_SAI",
                "Đúng",
                "Sai",
                "",
                "",
                "A",
                "DE",
                "1.0",
                "",
            ],
        ]

        for row_data in samples:
            ws.append(row_data)

        # Điều chỉnh độ rộng cột
        col_widths = [50, 20, 25, 25, 25, 25, 15, 15, 8, 40]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

        # Xuất ra bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
