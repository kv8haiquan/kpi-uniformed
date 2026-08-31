"""
lms_service/services/thi_sinh_service.py
========================================
Business logic: giao thi sinh, bat dau thi (random de), nop bai (cham diem),
xem ket qua, export Excel.
"""

import random
import uuid
from datetime import datetime, timedelta
from decimal import Decimal
from typing import Optional

from fastapi import HTTPException, status
from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession

from lms_service.models.base import CongChucRef, DonViRef
from lms_service.models.ky_thi import KyThi
from lms_service.models.cau_truc_de import CauTrucDe
from lms_service.models.thi_sinh import ThiSinh
from lms_service.models.phien_thi import PhienThi
from lms_service.models.vi_pham_thi import ViPhamThi
from lms_service.models.linh_vuc import LinhVuc
from lms_service.models.vi_tri_viec_lam import ViTriViecLam
from lms_service.models.cau_hoi import CauHoi
from lms_service.models.cau_hoi_dgnl import CauHoiDgnl
from lms_service.schemas.thi_sinh import (
    ThiSinhBatchCreate, ThiSinhResponse,
    NopBaiRequest, KetQuaResponse, DiemLinhVuc,
)
from lms_service.services.thong_bao_helper import gui_thong_bao_bulk
from shared.auth import TokenPayload
from lms_service.core.timezone import fmt_vn, now_vn

# So phut ke tu thoi_gian_nop de tu dong xac nhan ca thi (khong cho thi lai)
XAC_NHAN_TIMEOUT_PHUT = 10


class ThiSinhService:
    """Service thi sinh DGNL."""

    def __init__(self, db: AsyncSession):
        self.db = db

    def _is_manager(self, user: TokenPayload) -> bool:
        return user.vai_tro == "SUPER_ADMIN" or "QT_DAO_TAO" in (user.platform_roles or [])

    def _is_lanh_dao(self, user: TokenPayload) -> bool:
        # Chi cap CCT/PCCT moi duoc xem ket qua DGNL (yeu cau nghiep vu).
        # TDV/PDV (is_lanh_dao=True) bi chan boi day du don vi nho ko nen thay
        # bai lam cua thi sinh don vi khac.
        return user.vai_tro in ("CCT", "PCCT")

    async def _get_ky_thi(self, ky_thi_id: uuid.UUID) -> KyThi:
        stmt = select(KyThi).where(KyThi.id == ky_thi_id, KyThi.is_active == True)
        result = await self.db.execute(stmt)
        kt = result.scalar_one_or_none()
        if not kt:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail={"success": False, "error": {"code": "DGNL_010", "message": "Kỳ thi không tồn tại"}},
            )
        return kt

    # ================================================================
    # PHIEN THI — 1 phien/tai khoan (chong dung chung tai khoan)
    # ================================================================

    async def _upsert_phien(
        self,
        cc_id: uuid.UUID,
        ky_thi_id: uuid.UUID,
        thi_sinh_id: uuid.UUID,
        thiet_bi: Optional[str],
    ) -> str:
        """Sinh phien_token moi cho cong_chuc_id va ghi de phien cu (1 dong/tai khoan).

        Thiet bi nao goi bat_dau_thi gan nhat se so huu phien -> cac thiet bi
        khac (token cu) bi tu choi o luu-nhap/nop-bai. Tra ve token moi.
        """
        token = uuid.uuid4().hex
        now = now_vn()
        r = await self.db.execute(
            select(PhienThi).where(PhienThi.cong_chuc_id == cc_id)
        )
        phien = r.scalar_one_or_none()
        if phien:
            phien.phien_token = token
            phien.ky_thi_id = ky_thi_id
            phien.thi_sinh_id = thi_sinh_id
            phien.thiet_bi = (thiet_bi or "")[:255]
            phien.last_seen = now
        else:
            self.db.add(PhienThi(
                cong_chuc_id=cc_id,
                phien_token=token,
                ky_thi_id=ky_thi_id,
                thi_sinh_id=thi_sinh_id,
                thiet_bi=(thiet_bi or "")[:255],
                last_seen=now,
            ))
        return token

    async def _validate_phien(self, cc_id: uuid.UUID, phien_token: Optional[str]) -> None:
        """Kiem tra token phien con hop le va cap nhat last_seen (heartbeat).

        - phien_token rong (client cu chua gui) -> bo qua enforce (tranh khoa
          nguoi dung do cache JS cu trong luc trien khai).
        - phien_token khac token dang luu -> 409: tai khoan dang thi o thiet bi khac.
        """
        if not phien_token:
            return
        r = await self.db.execute(
            select(PhienThi).where(PhienThi.cong_chuc_id == cc_id)
        )
        phien = r.scalar_one_or_none()
        if phien is None:
            return
        if phien.phien_token != phien_token:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail={"success": False, "error": {
                    "code": "PHIEN_001",
                    "message": "Tài khoản đang được dùng để thi ở thiết bị khác. "
                               "Bài làm trên thiết bị này đã bị khóa.",
                }},
            )
        phien.last_seen = now_vn()

    # ================================================================
    # GIAO THI SINH
    # ================================================================

    async def giao_thi_sinh(
        self, ky_thi_id: uuid.UUID, data: ThiSinhBatchCreate, user: TokenPayload
    ) -> dict:
        """Giao thi sinh theo danh sach hoac theo don vi."""
        kt = await self._get_ky_thi(ky_thi_id)
        if kt.trang_thai not in ("NHAP", "CHO_DUYET", "DANG_MO"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"success": False, "error": {"code": "DGNL_021", "message": "Không thể giao thí sinh khi kỳ thi đã đóng"}},
            )

        items_to_add = []

        if data.danh_sach:
            # Giao theo danh sach cu the
            for item in data.danh_sach:
                items_to_add.append((item.cong_chuc_id, item.vi_tri_id))
        elif data.don_vi_ids and data.vi_tri_id:
            # Giao theo don vi
            cc_table = CongChucRef.__table__
            stmt = select(cc_table.c.id).where(
                cc_table.c.don_vi_id.in_(data.don_vi_ids),
                cc_table.c.is_active == True,
            )
            result = await self.db.execute(stmt)
            cc_ids = result.scalars().all()
            for cc_id in cc_ids:
                items_to_add.append((cc_id, data.vi_tri_id))
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"success": False, "error": {"code": "DGNL_022", "message": "Phải cung cấp danh_sach hoặc (don_vi_ids + vi_tri_id)"}},
            )

        thanh_cong = 0
        bo_qua = 0
        loi = []

        for cc_id, vt_id in items_to_add:
            # Check da ton tai
            existing = await self.db.execute(
                select(ThiSinh).where(
                    ThiSinh.ky_thi_id == ky_thi_id,
                    ThiSinh.cong_chuc_id == cc_id,
                )
            )
            if existing.scalar_one_or_none():
                bo_qua += 1
                continue

            # Validate vi tri ton tai
            vt_r = await self.db.execute(
                select(ViTriViecLam.id).where(ViTriViecLam.id == vt_id, ViTriViecLam.is_active == True)
            )
            if not vt_r.scalar_one_or_none():
                loi.append(f"Vị trí {vt_id} không tồn tại")
                continue

            ts = ThiSinh(
                ky_thi_id=ky_thi_id,
                cong_chuc_id=cc_id,
                vi_tri_id=vt_id,
                trang_thai="CHUA_THI",
            )
            self.db.add(ts)
            thanh_cong += 1

        await self.db.commit()

        # Gui thong bao cho cac thi sinh vua duoc giao
        if thanh_cong > 0:
            new_cc_ids = [cc_id for cc_id, _ in items_to_add
                          if any(cc_id == ts_cc for ts_cc, _ in items_to_add)]
            # Lay danh sach cc_id thuc su them thanh cong (loai bo_qua + loi)
            ts_result = await self.db.execute(
                select(ThiSinh.cong_chuc_id).where(ThiSinh.ky_thi_id == ky_thi_id)
            )
            all_cc_ids = ts_result.scalars().all()
            await gui_thong_bao_bulk(
                nguoi_nhan_ids=all_cc_ids[-thanh_cong:] if thanh_cong <= len(all_cc_ids) else all_cc_ids,
                tieu_de=f"Bạn được giao thi: {kt.ten_ky_thi}",
                noi_dung=f"Kỳ thi \"{kt.ten_ky_thi}\" ({kt.ma_ky_thi}). Vui lòng vào mục Đào tạo → Kỳ thi ĐGNL để xem chi tiết.",
                muc_do="QUAN_TRONG",
                link_url="/dao-tao/ky-thi",
                doi_tuong_type="KY_THI",
                doi_tuong_id=ky_thi_id,
            )

        return {
            "thanh_cong": thanh_cong,
            "bo_qua": bo_qua,
            "loi": loi,
            "tong": len(items_to_add),
        }

    # ================================================================
    # IMPORT THI SINH TU EXCEL (file mau chi co cot ma_cc)
    # ================================================================

    @staticmethod
    def generate_template_import_thi_sinh() -> bytes:
        """Sinh file Excel mau import thi sinh — 1 cot ma_cc + huong dan."""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Danh sách thí sinh"

        # Huong dan
        ws.merge_cells("A1:C1")
        hd = ws.cell(row=1, column=1, value="HƯỚNG DẪN: Nhập mã công chức vào cột ma_cc (mỗi dòng 1 người). "
                                            "Vị trí việc làm được chọn chung trên form khi upload. "
                                            "Không đổi tên cột. Xóa các dòng ví dụ trước khi import.")
        hd.font = Font(italic=True, size=10, color="808080")
        hd.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[1].height = 45

        # Header
        cell = ws.cell(row=2, column=1, value="ma_cc")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions["A"].width = 20

        # Vi du
        ws.cell(row=3, column=1, value="20ZZ-0224")
        ws.cell(row=4, column=1, value="20ZZ-0225")

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def import_thi_sinh_excel(
        self, ky_thi_id: uuid.UUID, vi_tri_id: uuid.UUID,
        file_content: bytes, user: TokenPayload,
    ) -> dict:
        """Import thi sinh tu file Excel (cot ma_cc). Vi tri chon chung tren form.

        Loi tung dong (ma khong ton tai / ngung hoat dong / trung) KHONG chan
        cac dong hop le — tra ve loi_chi_tiet de FE hien bang ket qua.
        """
        import io
        from openpyxl import load_workbook

        kt = await self._get_ky_thi(ky_thi_id)
        if kt.trang_thai not in ("NHAP", "CHO_DUYET", "DANG_MO"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"success": False, "error": {"code": "DGNL_021", "message": "Không thể giao thí sinh khi kỳ thi đã đóng"}},
            )

        vt_r = await self.db.execute(
            select(ViTriViecLam.id).where(ViTriViecLam.id == vi_tri_id, ViTriViecLam.is_active == True)
        )
        if not vt_r.scalar_one_or_none():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"success": False, "error": {"code": "DGNL_022", "message": "Vị trí việc làm không tồn tại hoặc đã ngừng hoạt động"}},
            )

        try:
            wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        except Exception:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"success": False, "error": {"code": "DGNL_060", "message": "File không đúng định dạng Excel (.xlsx)"}},
            )
        ws = wb.active

        # Tim cot ma_cc trong 2 dong dau (file mau co dong huong dan o dong 1)
        header_row = None
        ma_cc_idx = None
        for row_idx in (1, 2):
            values = [str(c.value or "").strip().lower() for c in (ws[row_idx] if ws.max_row >= row_idx else [])]
            if "ma_cc" in values:
                header_row = row_idx
                ma_cc_idx = values.index("ma_cc")
                break
        if ma_cc_idx is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"success": False, "error": {"code": "DGNL_061", "message": "File thiếu cột 'ma_cc' — hãy dùng file mẫu"}},
            )

        # Doc danh sach (dong, ma_cc)
        ma_cc_rows: list[tuple[int, str]] = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            ma = str(row[ma_cc_idx] or "").strip() if len(row) > ma_cc_idx else ""
            if not ma:
                continue
            ma_cc_rows.append((row_idx, ma))

        if not ma_cc_rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"success": False, "error": {"code": "DGNL_062", "message": "File không có dữ liệu mã công chức"}},
            )

        # Resolve ma_cc -> cong_chuc (READONLY public.cong_chuc), match khong phan biet hoa/thuong
        cc_table = CongChucRef.__table__
        ma_upper_list = list({ma.upper() for _, ma in ma_cc_rows})
        r = await self.db.execute(
            select(cc_table.c.id, cc_table.c.ma_cc, cc_table.c.is_active).where(
                func.upper(cc_table.c.ma_cc).in_(ma_upper_list)
            )
        )
        cc_map = {row.ma_cc.upper(): (row.id, row.is_active) for row in r.all()}

        # Danh sach da duoc giao san
        existing_r = await self.db.execute(
            select(ThiSinh.cong_chuc_id).where(ThiSinh.ky_thi_id == ky_thi_id)
        )
        existing_cc_ids = set(existing_r.scalars().all())

        thanh_cong = 0
        loi_chi_tiet: list[dict] = []
        da_gap_trong_file: set[str] = set()
        new_cc_ids: list[uuid.UUID] = []

        for row_idx, ma in ma_cc_rows:
            key = ma.upper()
            if key in da_gap_trong_file:
                loi_chi_tiet.append({"dong": row_idx, "ma_cc": ma, "loi": "Trùng lặp trong file"})
                continue
            da_gap_trong_file.add(key)

            if key not in cc_map:
                loi_chi_tiet.append({"dong": row_idx, "ma_cc": ma, "loi": "Mã công chức không tồn tại"})
                continue
            cc_id, is_active = cc_map[key]
            if not is_active:
                loi_chi_tiet.append({"dong": row_idx, "ma_cc": ma, "loi": "Công chức đã ngừng hoạt động"})
                continue
            if cc_id in existing_cc_ids:
                loi_chi_tiet.append({"dong": row_idx, "ma_cc": ma, "loi": "Đã được giao thi kỳ thi này"})
                continue

            self.db.add(ThiSinh(
                ky_thi_id=ky_thi_id,
                cong_chuc_id=cc_id,
                vi_tri_id=vi_tri_id,
                trang_thai="CHUA_THI",
            ))
            existing_cc_ids.add(cc_id)
            new_cc_ids.append(cc_id)
            thanh_cong += 1

        await self.db.commit()

        # Thong bao cho cac thi sinh vua duoc giao
        if new_cc_ids:
            await gui_thong_bao_bulk(
                nguoi_nhan_ids=new_cc_ids,
                tieu_de=f"Bạn được giao thi: {kt.ten_ky_thi}",
                noi_dung=f"Kỳ thi \"{kt.ten_ky_thi}\" ({kt.ma_ky_thi}). Vui lòng vào mục Đào tạo → Kỳ thi ĐGNL để xem chi tiết.",
                muc_do="QUAN_TRONG",
                link_url="/dao-tao/ky-thi",
                doi_tuong_type="KY_THI",
                doi_tuong_id=ky_thi_id,
            )

        return {
            "tong": len(ma_cc_rows),
            "thanh_cong": thanh_cong,
            "that_bai": len(loi_chi_tiet),
            "loi_chi_tiet": loi_chi_tiet,
        }

    async def danh_sach_thi_sinh(
        self, ky_thi_id: uuid.UUID, user: TokenPayload,
        trang_thai: Optional[str] = None,
        page: int = 1, page_size: int = 50,
    ) -> dict:
        """Danh sach thi sinh + ket qua."""
        await self._get_ky_thi(ky_thi_id)

        cc = CongChucRef.__table__.alias("cc")
        dv = DonViRef.__table__.alias("dv")

        base_where = [ThiSinh.ky_thi_id == ky_thi_id]
        if trang_thai:
            base_where.append(ThiSinh.trang_thai == trang_thai)

        # Module DGNL chi admin (QT_DAO_TAO/SUPER_ADMIN) duoc xem -> khong scope don vi.

        count_stmt = select(func.count()).select_from(ThiSinh).where(*base_where)
        total = (await self.db.execute(count_stmt)).scalar() or 0

        stmt = (
            select(
                ThiSinh,
                cc.c.ho_ten,
                cc.c.ma_cc,
                dv.c.ten_don_vi.label("don_vi_ten"),
                ViTriViecLam.ten_vi_tri.label("vi_tri_ten"),
            )
            .outerjoin(cc, ThiSinh.cong_chuc_id == cc.c.id)
            .outerjoin(dv, cc.c.don_vi_id == dv.c.id)
            .outerjoin(ViTriViecLam, ThiSinh.vi_tri_id == ViTriViecLam.id)
            .where(*base_where)
            .order_by(cc.c.ho_ten.asc())
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
        result = await self.db.execute(stmt)
        rows = result.all()

        items = []
        for ts, ho_ten, ma_cc, dv_ten, vt_ten in rows:
            items.append({
                **{c.key: getattr(ts, c.key) for c in ts.__table__.columns},
                "ho_ten": ho_ten,
                "ma_cc": ma_cc,
                "don_vi_ten": dv_ten,
                "vi_tri_ten": vt_ten,
                # Override raw lich_su_thi voi summary projection (an chi_tiet_tra_loi)
                "lich_su_thi": self._project_lich_su_summary(ts.lich_su_thi),
            })

        return {
            "items": items,
            "pagination": {
                "page": page, "page_size": page_size,
                "total_items": total,
                "total_pages": (total + page_size - 1) // page_size if page_size > 0 else 0,
            },
        }

    async def giam_sat(self, ky_thi_id: uuid.UUID, user: TokenPayload) -> dict:
        """Giam sat truc tiep: tien do + vi pham + online cua tung thi sinh.

        Dung cho man hinh admin theo doi realtime (FE poll ~7s). Chi tinh
        thi sinh DANG_THI/DA_NOP/VANG (CHUA_THI khong co tien do).
        """
        kt = await self._get_ky_thi(ky_thi_id)

        cc = CongChucRef.__table__.alias("cc")
        dv = DonViRef.__table__.alias("dv")

        stmt = (
            select(
                ThiSinh,
                cc.c.ho_ten,
                cc.c.ma_cc,
                dv.c.ten_don_vi.label("don_vi_ten"),
                ViTriViecLam.ten_vi_tri.label("vi_tri_ten"),
            )
            .outerjoin(cc, ThiSinh.cong_chuc_id == cc.c.id)
            .outerjoin(dv, cc.c.don_vi_id == dv.c.id)
            .outerjoin(ViTriViecLam, ThiSinh.vi_tri_id == ViTriViecLam.id)
            .where(ThiSinh.ky_thi_id == ky_thi_id)
            .order_by(cc.c.ho_ten.asc())
        )
        result = await self.db.execute(stmt)
        rows = result.all()

        # Map last_seen / thiet_bi theo cong_chuc_id (chi phien dang gan ky thi nay)
        cc_ids = [ts.cong_chuc_id for ts, *_ in rows]
        phien_map: dict[uuid.UUID, PhienThi] = {}
        if cc_ids:
            ph_r = await self.db.execute(
                select(PhienThi).where(
                    PhienThi.cong_chuc_id.in_(cc_ids),
                    PhienThi.ky_thi_id == ky_thi_id,
                )
            )
            phien_map = {p.cong_chuc_id: p for p in ph_r.scalars().all()}

        now = now_vn()
        thi_sinh = []
        dang_thi = chua_thi = da_nop = vang = online_count = co_vi_pham = 0

        for ts, ho_ten, ma_cc, dv_ten, vt_ten in rows:
            if ts.trang_thai == "DANG_THI":
                dang_thi += 1
            elif ts.trang_thai == "DA_NOP":
                da_nop += 1
            elif ts.trang_thai == "VANG":
                vang += 1
            else:
                chua_thi += 1

            # So cau da lam: dang thi -> tu chi_tiet_nhap; da nop -> tong da tra loi
            if ts.trang_thai == "DANG_THI":
                so_cau_da_lam = len(ts.chi_tiet_nhap or [])
            elif ts.trang_thai == "DA_NOP":
                so_cau_da_lam = len(ts.chi_tiet_tra_loi or [])
            else:
                so_cau_da_lam = 0

            tg_con = self._tinh_thoi_gian_con(ts, kt) if ts.trang_thai == "DANG_THI" else None

            phien = phien_map.get(ts.cong_chuc_id)
            last_seen = phien.last_seen if phien else None
            online = bool(
                ts.trang_thai == "DANG_THI"
                and last_seen
                and (now - last_seen).total_seconds() < 60
            )
            if online:
                online_count += 1
            if (ts.so_lan_vi_pham or 0) > 0:
                co_vi_pham += 1

            thi_sinh.append({
                "cong_chuc_id": ts.cong_chuc_id,
                "ho_ten": ho_ten,
                "ma_cc": ma_cc,
                "don_vi_ten": dv_ten,
                "vi_tri_ten": vt_ten,
                "trang_thai": ts.trang_thai,
                "lan_thi_hien_tai": ts.lan_thi_hien_tai or 0,
                "so_cau_da_lam": so_cau_da_lam,
                "tong_so_cau": ts.tong_so_cau or len(ts.de_thi_ids or []),
                "so_lan_vi_pham": ts.so_lan_vi_pham or 0,
                "thoi_gian_bat_dau": ts.thoi_gian_bat_dau,
                "thoi_gian_con_lai_giay": tg_con,
                "diem_tong": ts.diem_tong,
                "xep_loai": ts.xep_loai,
                "last_seen": last_seen,
                "online": online,
                "thiet_bi": phien.thiet_bi if phien else None,
            })

        return {
            "tong_quan": {
                "tong_thi_sinh": len(rows),
                "dang_thi": dang_thi,
                "da_nop": da_nop,
                "chua_thi": chua_thi,
                "vang": vang,
                "online": online_count,
                "co_vi_pham": co_vi_pham,
            },
            "thi_sinh": thi_sinh,
        }

    async def xoa_thi_sinh(
        self, ky_thi_id: uuid.UUID, cong_chuc_id: uuid.UUID, user: TokenPayload
    ) -> None:
        """Xoa thi sinh — chi khi CHUA_THI."""
        stmt = select(ThiSinh).where(
            ThiSinh.ky_thi_id == ky_thi_id,
            ThiSinh.cong_chuc_id == cong_chuc_id,
        )
        result = await self.db.execute(stmt)
        ts = result.scalar_one_or_none()
        if not ts:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail={"success": False, "error": {"code": "DGNL_023", "message": "Thí sinh không tồn tại trong kỳ thi này"}},
            )
        if ts.trang_thai != "CHUA_THI":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"success": False, "error": {"code": "DGNL_024", "message": "Chỉ được xóa thí sinh chưa thi"}},
            )

        await self.db.delete(ts)
        await self.db.commit()

    # ================================================================
    # BAT DAU THI — RANDOM DE
    # ================================================================

    async def bat_dau_thi(
        self, ky_thi_id: uuid.UUID, user: TokenPayload, thiet_bi: Optional[str] = None
    ) -> dict:
        """Bat dau thi: random de thi cho thi sinh."""
        kt = await self._get_ky_thi(ky_thi_id)

        # Validate trang thai ky thi
        if kt.trang_thai != "DANG_MO":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"success": False, "error": {"code": "DGNL_025", "message": "Kỳ thi chưa mở hoặc đã đóng"}},
            )

        # Validate thoi gian
        now = now_vn()
        if now < kt.ngay_bat_dau:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"success": False, "error": {"code": "DGNL_026", "message": "Chưa đến thời gian thi"}},
            )
        if now > kt.ngay_ket_thuc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"success": False, "error": {"code": "DGNL_027", "message": "Đã hết thời gian thi"}},
            )

        # Lay thi sinh
        cc_id = uuid.UUID(user.sub)
        stmt = select(ThiSinh).where(
            ThiSinh.ky_thi_id == ky_thi_id,
            ThiSinh.cong_chuc_id == cc_id,
        )
        result = await self.db.execute(stmt)
        ts = result.scalar_one_or_none()
        if not ts:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail={"success": False, "error": {"code": "DGNL_028", "message": "Bạn không được giao thi kỳ thi này"}},
            )

        # Check thi lai
        if ts.trang_thai == "DANG_THI":
            # Dang thi — tra lai de cu kem bai lam nhap (resume autosave).
            # Thiet bi nay gianh quyen so huu phien (token moi) -> thiet bi cu bi 409.
            token = await self._upsert_phien(cc_id, ky_thi_id, ts.id, thiet_bi)
            await self.db.commit()
            cau_hoi_list = await self._lay_de_thi_hien_tai(ts)
            return {
                "thi_sinh_id": ts.id,
                "thoi_gian_con_lai_giay": self._tinh_thoi_gian_con(ts, kt),
                "tong_so_cau": ts.tong_so_cau or len(ts.de_thi_ids or []),
                "cau_hoi": cau_hoi_list,
                "dang_tiep_tuc": True,
                "chi_tiet_nhap": ts.chi_tiet_nhap or [],
                "so_lan_vi_pham": ts.so_lan_vi_pham or 0,
                "yeu_cau_toan_man_hinh": bool(kt.yeu_cau_toan_man_hinh),
                "phien_token": token,
            }

        if ts.trang_thai == "DA_NOP":
            # Da xac nhan ca thi -> chot ket qua, khong duoc thi lai du con luot
            if ts.da_xac_nhan:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail={"success": False, "error": {"code": "DGNL_048", "message": "Bạn đã xác nhận ca thi, không thể thi lại"}},
                )
            # Qua 10 phut ke tu khi nop ma khong xac nhan -> tu dong xac nhan
            # (lazy: enforce tai thoi diem xin thi lai, khong can background job)
            if ts.thoi_gian_nop and now > ts.thoi_gian_nop + timedelta(minutes=XAC_NHAN_TIMEOUT_PHUT):
                ts.da_xac_nhan = True
                ts.thoi_gian_xac_nhan = ts.thoi_gian_nop + timedelta(minutes=XAC_NHAN_TIMEOUT_PHUT)
                ts.updated_at = now
                await self.db.commit()
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail={"success": False, "error": {"code": "DGNL_049", "message": f"Đã quá {XAC_NHAN_TIMEOUT_PHUT} phút kể từ khi nộp bài, ca thi đã tự động được xác nhận — không thể thi lại"}},
                )
            # Check con luot thi lai khong
            if ts.lan_thi_hien_tai >= (kt.so_lan_thi_toi_da or 1):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail={"success": False, "error": {"code": "DGNL_029", "message": f"Đã hết lượt thi (tối đa {kt.so_lan_thi_toi_da} lần)"}},
                )
            # Snapshot lan vua nop vao lich_su (upsert idempotent theo lan).
            # 30/05/2026: mo rong shape -> luu day du chi_tiet_tra_loi, de_thi_ids,
            # diem_theo_linh_vuc... de QT_DAO_TAO drill-down lai bai lam lan cu.
            self._upsert_lan_thi(ts, self._snapshot_lan_thi(ts))

        if ts.trang_thai == "VANG":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"success": False, "error": {"code": "DGNL_030", "message": "Bạn đã bị đánh dấu vắng, không thể thi"}},
            )

        # Random de thi
        de_thi_ids = await self._tao_de_thi(kt, ts)

        # Cap nhat thi sinh — reset toan bo state cua lan thi truoc (Bug A fix
        # 27/05/2026): khong reset cac field diem cu -> kep DANG_THI van hien
        # diem cu, gay nham lan cho QT_DAO_TAO.
        ts.de_thi_ids = [str(cid) for cid in de_thi_ids]
        ts.chi_tiet_tra_loi = []
        ts.chi_tiet_nhap = None
        ts.so_lan_vi_pham = 0
        ts.trang_thai = "DANG_THI"
        ts.lan_thi_hien_tai = (ts.lan_thi_hien_tai or 0) + 1
        ts.thoi_gian_bat_dau = now_vn()
        ts.thoi_gian_nop = None
        ts.thoi_gian_lam_giay = None
        ts.tong_so_cau = len(de_thi_ids)
        ts.diem_tong = None
        ts.xep_loai = None
        ts.so_cau_dung = None
        ts.so_cau_sai = None
        ts.diem_theo_linh_vuc = {}
        ts.updated_at = now_vn()

        # Sinh phien moi cho thiet bi nay (chong dung chung tai khoan)
        token = await self._upsert_phien(cc_id, ky_thi_id, ts.id, thiet_bi)

        await self.db.commit()
        await self.db.refresh(ts)

        # Lay cau hoi (khong co dap an)
        cau_hoi_list = await self._lay_de_thi_hien_tai(ts)

        return {
            "thi_sinh_id": ts.id,
            "thoi_gian_con_lai_giay": kt.thoi_gian_lam_bai_phut * 60,
            "tong_so_cau": len(de_thi_ids),
            "cau_hoi": cau_hoi_list,
            "dang_tiep_tuc": False,
            "chi_tiet_nhap": [],
            "so_lan_vi_pham": 0,
            "yeu_cau_toan_man_hinh": bool(kt.yeu_cau_toan_man_hinh),
            "phien_token": token,
        }

    async def _tao_de_thi(self, kt: KyThi, ts: ThiSinh) -> list[uuid.UUID]:
        """Random de thi theo cau truc de cua vi tri."""
        # Lay cau truc de cho vi tri cua thi sinh
        stmt = select(CauTrucDe).where(
            CauTrucDe.ky_thi_id == kt.id,
            CauTrucDe.vi_tri_id == ts.vi_tri_id,
        )
        result = await self.db.execute(stmt)
        cau_trucs = result.scalars().all()

        if not cau_trucs:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"success": False, "error": {"code": "DGNL_031", "message": "Chưa cấu hình cấu trúc đề cho vị trí này"}},
            )

        de_thi: list[uuid.UUID] = []

        for ct in cau_trucs:
            # Cau hoi cua TUNG linh vuc gom thanh 1 block lien tuc — de FE danh
            # so cau 1..N tuan tu theo nhom linh vuc (fix 30/07/2026: shuffle
            # toan cuc lam so cau trong sidebar nhay lung tung).
            block: list[uuid.UUID] = []
            for do_kho, so_cau in [
                ("DE", ct.so_cau_de or 0),
                ("TRUNG_BINH", ct.so_cau_trung_binh or 0),
                ("KHO", ct.so_cau_kho or 0),
            ]:
                if so_cau <= 0:
                    continue

                # Random tu ngan hang DGNL
                stmt = (
                    select(CauHoiDgnl.id)
                    .where(
                        CauHoiDgnl.linh_vuc_id == ct.linh_vuc_id,
                        CauHoiDgnl.do_kho == do_kho,
                        CauHoiDgnl.is_active == True,
                    )
                    .order_by(func.random())
                    .limit(so_cau)
                )
                r = await self.db.execute(stmt)
                ids = r.scalars().all()

                if len(ids) < so_cau:
                    # Lay ten linh vuc
                    lv_r = await self.db.execute(select(LinhVuc.ten_linh_vuc).where(LinhVuc.id == ct.linh_vuc_id))
                    lv_ten = lv_r.scalar() or str(ct.linh_vuc_id)
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail={
                            "success": False,
                            "error": {
                                "code": "DGNL_032",
                                "message": f"Không đủ câu hỏi {do_kho} cho lĩnh vực {lv_ten} "
                                           f"(cần {so_cau}, có {len(ids)})",
                            },
                        },
                    )
                block.extend(ids)

            # Tron cau hoi TRONG pham vi linh vuc (khong tron xuyen linh vuc)
            if kt.tron_cau_hoi:
                random.shuffle(block)
            de_thi.extend(block)

        return de_thi

    async def _lay_de_thi_hien_tai(self, ts: ThiSinh) -> list[dict]:
        """Lay danh sach cau hoi cua de thi (KHONG co dap an dung)."""
        if not ts.de_thi_ids:
            return []

        cau_hoi_ids = [uuid.UUID(cid) for cid in ts.de_thi_ids]
        stmt = select(CauHoiDgnl).where(CauHoiDgnl.id.in_(cau_hoi_ids))
        result = await self.db.execute(stmt)
        ch_map = {ch.id: ch for ch in result.scalars().all()}

        # Lay ten linh vuc
        lv_ids = set(ch.linh_vuc_id for ch in ch_map.values() if ch.linh_vuc_id)
        lv_map = {}
        if lv_ids:
            lv_r = await self.db.execute(select(LinhVuc).where(LinhVuc.id.in_(lv_ids)))
            lv_map = {lv.id: lv.ten_linh_vuc for lv in lv_r.scalars().all()}

        cau_hoi_list = []
        for i, cid in enumerate(cau_hoi_ids):
            ch = ch_map.get(cid)
            if not ch:
                continue

            # Xay dung lua chon (khong co dap an dung)
            lua_chon = None
            if ch.dap_an and ch.loai in ("TRAC_NGHIEM_1", "TRAC_NGHIEM_NHIEU", "DUNG_SAI"):
                dap_an = ch.dap_an
                if isinstance(dap_an, dict) and "lua_chon" in dap_an:
                    lua_chon = dap_an["lua_chon"]
                elif isinstance(dap_an, list):
                    lua_chon = [{"key": d.get("key", ""), "noi_dung": d.get("noi_dung", "")} for d in dap_an if isinstance(d, dict)]

            cau_hoi_list.append({
                "id": str(ch.id),
                "thu_tu": i + 1,
                "noi_dung": ch.noi_dung,
                "loai": ch.loai,
                "diem": float(ch.diem) if ch.diem else 1.0,
                "lua_chon": lua_chon,
                "linh_vuc": lv_map.get(ch.linh_vuc_id),
            })

        return cau_hoi_list

    def _tinh_thoi_gian_con(self, ts: ThiSinh, kt: KyThi) -> int:
        """Tinh so giay con lai."""
        if not ts.thoi_gian_bat_dau:
            return kt.thoi_gian_lam_bai_phut * 60
        now = now_vn()
        bd = ts.thoi_gian_bat_dau
        elapsed = (now - bd).total_seconds()
        remaining = (kt.thoi_gian_lam_bai_phut * 60) - elapsed
        return max(0, int(remaining))

    # ================================================================
    # NOP BAI — CHAM DIEM
    # ================================================================

    async def nop_bai(
        self, ky_thi_id: uuid.UUID, data: NopBaiRequest, user: TokenPayload,
        phien_token: Optional[str] = None,
    ) -> KetQuaResponse:
        """Nop bai va cham diem tu dong."""
        kt = await self._get_ky_thi(ky_thi_id)
        cc_id = uuid.UUID(user.sub)

        stmt = select(ThiSinh).where(
            ThiSinh.ky_thi_id == ky_thi_id,
            ThiSinh.cong_chuc_id == cc_id,
        )
        result = await self.db.execute(stmt)
        ts = result.scalar_one_or_none()

        if not ts:
            raise HTTPException(status_code=404, detail={"success": False, "error": {"code": "DGNL_028", "message": "Bạn không có trong kỳ thi này"}})
        # Bug B fix (27/05/2026): nop bai idempotent. Neu da DA_NOP, tra lai ket
        # qua da co thay vi loi 400 — tranh truong hop user retry sau khi network
        # timeout o lan submit dau (commit DB thanh cong, response mat).
        if ts.trang_thai == "DA_NOP":
            return await self._build_ket_qua(ts, kt)
        if ts.trang_thai != "DANG_THI":
            raise HTTPException(status_code=400, detail={"success": False, "error": {"code": "DGNL_033", "message": "Bạn chưa bắt đầu thi hoặc đã nộp bài"}})

        # Chi thiet bi dang so huu phien moi duoc nop (chong dung chung tai khoan)
        await self._validate_phien(cc_id, phien_token)

        # Cham diem
        cau_hoi_ids = [uuid.UUID(cid) for cid in (ts.de_thi_ids or [])]
        ch_stmt = select(CauHoiDgnl).where(CauHoiDgnl.id.in_(cau_hoi_ids))
        ch_result = await self.db.execute(ch_stmt)
        ch_map = {ch.id: ch for ch in ch_result.scalars().all()}

        # Map tra loi
        tra_loi_map = {tl.cau_hoi_id: tl.tra_loi for tl in data.cau_tra_loi}

        chi_tiet = []
        tong_dung = 0
        tong_sai = 0
        diem_lv: dict[uuid.UUID, dict] = {}  # linh_vuc_id -> {dung, tong}

        for cid in cau_hoi_ids:
            ch = ch_map.get(cid)
            if not ch:
                continue

            tra_loi = tra_loi_map.get(cid)
            dung = self._cham_1_cau(ch, tra_loi)

            if dung:
                tong_dung += 1
            else:
                tong_sai += 1

            # Diem theo linh vuc
            lv_id = ch.linh_vuc_id
            if lv_id:
                if lv_id not in diem_lv:
                    diem_lv[lv_id] = {"dung": 0, "tong": 0}
                diem_lv[lv_id]["tong"] += 1
                if dung:
                    diem_lv[lv_id]["dung"] += 1

            chi_tiet.append({
                "cau_hoi_id": str(cid),
                "tra_loi": tra_loi,
                "dap_an_dung": self._normalize_dap_an_dung(ch.dap_an),
                "dung": dung,
            })

        tong_cau = len(cau_hoi_ids)
        diem_tong = round(Decimal(tong_dung) / Decimal(tong_cau) * 100, 2) if tong_cau > 0 else Decimal("0")
        xep_loai = "DAT" if diem_tong >= (kt.diem_dat or Decimal("50")) else "KHONG_DAT"

        # Tinh diem theo linh vuc
        diem_theo_lv = {}
        lv_ids = list(diem_lv.keys())
        if lv_ids:
            lv_r = await self.db.execute(select(LinhVuc).where(LinhVuc.id.in_(lv_ids)))
            lv_name_map = {lv.id: lv.ma_linh_vuc for lv in lv_r.scalars().all()}
            for lv_id, stats in diem_lv.items():
                ma = lv_name_map.get(lv_id, str(lv_id))
                phan_tram = round(stats["dung"] / stats["tong"] * 100, 1) if stats["tong"] > 0 else 0
                diem_theo_lv[ma] = {"dung": stats["dung"], "tong": stats["tong"], "phan_tram": phan_tram}

        # Thoi gian lam bai
        now = now_vn()
        thoi_gian_lam = 0
        if ts.thoi_gian_bat_dau:
            bd = ts.thoi_gian_bat_dau
            thoi_gian_lam = int((now - bd).total_seconds())

        # Ghi ket qua lan vua thi truoc, snapshot vao lich su, roi ap dung
        # "diem chinh thuc = diem CAO NHAT giua cac lan" (quyet dinh 30/07/2026,
        # thay latest-score cu).
        ts.diem_tong = diem_tong
        ts.xep_loai = xep_loai
        ts.so_cau_dung = tong_dung
        ts.so_cau_sai = tong_sai
        ts.tong_so_cau = tong_cau
        ts.diem_theo_linh_vuc = diem_theo_lv

        ts.chi_tiet_tra_loi = chi_tiet
        ts.chi_tiet_nhap = None  # clear autosave draft sau khi nop
        ts.trang_thai = "DA_NOP"
        ts.thoi_gian_nop = now
        ts.thoi_gian_lam_giay = thoi_gian_lam
        ts.updated_at = now

        # Snapshot lan vua nop vao lich_su_thi (upsert theo lan).
        # Ly do: voi ky thi 1 lan (so_lan_thi_toi_da=1), bat_dau_thi khong duoc
        # goi lai -> neu chi snapshot tai bat_dau_thi thi lich_su_thi mai mai
        # rong va trang thong-ke khong drill-down duoc lan duy nhat.
        self._upsert_lan_thi(ts, self._snapshot_lan_thi(ts))

        # Cot chinh = ket qua lan tot nhat (lich_su_thi da bao gom lan vua nop).
        # thoi_gian_bat_dau/nop/lam_giay + chi_tiet_tra_loi van la lan cuoi
        # (dung cho han xac nhan + drill-down tung lan qua lich_su_thi).
        self._ap_dung_diem_cao_nhat(ts)

        await self.db.commit()
        await self.db.refresh(ts)

        return await self._build_ket_qua(ts, kt)

    @staticmethod
    def _ap_dung_diem_cao_nhat(ts: ThiSinh) -> None:
        """Ghi cot ket qua chinh cua thi_sinh = ket qua lan thi co diem cao nhat.

        Neu nhieu lan bang diem -> lay lan som nhat dat diem do.
        """
        entries = [e for e in (ts.lich_su_thi or []) if e and e.get("lan") is not None]
        if not entries:
            return
        best = max(entries, key=lambda e: ((e.get("diem") or 0), -(e.get("lan") or 0)))
        ts.diem_tong = Decimal(str(best.get("diem") or 0))
        ts.xep_loai = best.get("xep_loai")
        ts.so_cau_dung = best.get("so_cau_dung")
        ts.so_cau_sai = best.get("so_cau_sai")
        ts.tong_so_cau = best.get("tong_so_cau")
        ts.diem_theo_linh_vuc = dict(best.get("diem_theo_linh_vuc") or {})

    # ================================================================
    # XAC NHAN CA THI
    # ================================================================

    async def xac_nhan_ca_thi(self, ky_thi_id: uuid.UUID, user: TokenPayload) -> dict:
        """Thi sinh xac nhan ca thi — chot ket qua, khong duoc thi lai du con luot.

        Idempotent: da xac nhan roi thi tra lai trang thai hien tai.
        """
        cc_id = uuid.UUID(user.sub)
        stmt = select(ThiSinh).where(
            ThiSinh.ky_thi_id == ky_thi_id,
            ThiSinh.cong_chuc_id == cc_id,
        )
        result = await self.db.execute(stmt)
        ts = result.scalar_one_or_none()

        if not ts:
            raise HTTPException(
                status_code=404,
                detail={"success": False, "error": {"code": "DGNL_028", "message": "Bạn không có trong kỳ thi này"}},
            )
        if ts.trang_thai != "DA_NOP":
            raise HTTPException(
                status_code=400,
                detail={"success": False, "error": {"code": "DGNL_047", "message": "Chưa nộp bài, không thể xác nhận ca thi"}},
            )

        if not ts.da_xac_nhan:
            ts.da_xac_nhan = True
            ts.thoi_gian_xac_nhan = now_vn()
            ts.updated_at = now_vn()
            await self.db.commit()
            await self.db.refresh(ts)

        return {
            "da_xac_nhan": True,
            "thoi_gian_xac_nhan": ts.thoi_gian_xac_nhan.isoformat() if ts.thoi_gian_xac_nhan else None,
        }

    # ================================================================
    # LUU NHAP — AUTO-SAVE 30s/lan
    # ================================================================

    async def luu_nhap(
        self,
        ky_thi_id: uuid.UUID,
        cau_tra_loi: list,
        so_lan_vi_pham: int,
        user: TokenPayload,
        phien_token: Optional[str] = None,
    ) -> dict:
        """Luu bai lam nhap (auto-save moi 30s).

        Khac nop_bai: KHONG cham diem, KHONG doi trang_thai. Chi luu chi_tiet_nhap
        + so_lan_vi_pham de FE restore khi mat ket noi / dong tab / treo may.
        """
        cc_id = uuid.UUID(user.sub)
        stmt = select(ThiSinh).where(
            ThiSinh.ky_thi_id == ky_thi_id,
            ThiSinh.cong_chuc_id == cc_id,
        )
        result = await self.db.execute(stmt)
        ts = result.scalar_one_or_none()

        if not ts:
            raise HTTPException(
                status_code=404,
                detail={"success": False, "error": {"code": "DGNL_028", "message": "Bạn không có trong kỳ thi này"}},
            )
        if ts.trang_thai != "DANG_THI":
            # Da nop hoac chua bat dau -> bo qua autosave (idempotent, khong raise
            # de FE khong spam loi voi user).
            return {"saved": False, "trang_thai": ts.trang_thai}

        # Chi thiet bi dang so huu phien moi duoc luu (chong dung chung tai khoan).
        # Cap nhat last_seen luon -> feed man hinh giam sat truc tiep.
        await self._validate_phien(cc_id, phien_token)

        # Chuan hoa payload
        nhap_data = [
            {"cau_hoi_id": str(tl.cau_hoi_id), "tra_loi": tl.tra_loi}
            for tl in cau_tra_loi
        ]
        ts.chi_tiet_nhap = nhap_data
        # Fallback: counter chinh do server tang khi ghi_vi_pham (30/07/2026).
        # Chi nang len theo gia tri FE gui (phong client offline luc POST vi-pham
        # that bai), KHONG bao gio ha xuong.
        ts.so_lan_vi_pham = max(ts.so_lan_vi_pham or 0, so_lan_vi_pham or 0)
        ts.updated_at = now_vn()
        await self.db.commit()
        return {"saved": True, "so_cau_da_luu": len(nhap_data)}

    # ================================================================
    # VI PHAM — LOG CHI TIET (gio + ly do giai trinh)
    # ================================================================

    VI_PHAM_LOAI_HOP_LE = {"EXIT_FULLSCREEN", "SWITCH_TAB"}

    async def ghi_vi_pham(
        self, ky_thi_id: uuid.UUID, loai_vi_pham: str, user: TokenPayload,
        phien_token: Optional[str] = None,
    ) -> dict:
        """Ghi nhan 1 lan vi pham NGAY khi xay ra (khong doi auto-save 30s).

        Server tu tang counter so_lan_vi_pham -> counter va log luon khop nhau.
        """
        if loai_vi_pham not in self.VI_PHAM_LOAI_HOP_LE:
            raise HTTPException(
                status_code=400,
                detail={"success": False, "error": {"code": "DGNL_050", "message": "Loại vi phạm không hợp lệ"}},
            )

        cc_id = uuid.UUID(user.sub)
        stmt = select(ThiSinh).where(
            ThiSinh.ky_thi_id == ky_thi_id,
            ThiSinh.cong_chuc_id == cc_id,
        )
        result = await self.db.execute(stmt)
        ts = result.scalar_one_or_none()
        if not ts:
            raise HTTPException(
                status_code=404,
                detail={"success": False, "error": {"code": "DGNL_028", "message": "Bạn không có trong kỳ thi này"}},
            )
        if ts.trang_thai != "DANG_THI":
            raise HTTPException(
                status_code=400,
                detail={"success": False, "error": {"code": "DGNL_051", "message": "Không ở trạng thái đang thi"}},
            )

        # Chi thiet bi dang so huu phien moi duoc ghi (thiet bi cu bi kick -> 409)
        await self._validate_phien(cc_id, phien_token)

        vp = ViPhamThi(
            thi_sinh_id=ts.id,
            ky_thi_id=ky_thi_id,
            lan_thi=ts.lan_thi_hien_tai or 0,
            loai_vi_pham=loai_vi_pham,
            thoi_gian=now_vn(),
        )
        self.db.add(vp)
        ts.so_lan_vi_pham = (ts.so_lan_vi_pham or 0) + 1
        ts.updated_at = now_vn()
        await self.db.commit()
        await self.db.refresh(vp)

        return {
            "id": str(vp.id),
            "thoi_gian": vp.thoi_gian.isoformat() if vp.thoi_gian else None,
            "so_lan_vi_pham": ts.so_lan_vi_pham,
        }

    async def cap_nhat_ly_do_vi_pham(
        self, ky_thi_id: uuid.UUID, vp_id: uuid.UUID, ly_do: str, user: TokenPayload
    ) -> None:
        """Thi sinh nhap ly do giai trinh cho 1 lan vi pham cua chinh minh."""
        cc_id = uuid.UUID(user.sub)
        stmt = (
            select(ViPhamThi)
            .join(ThiSinh, ViPhamThi.thi_sinh_id == ThiSinh.id)
            .where(
                ViPhamThi.id == vp_id,
                ViPhamThi.ky_thi_id == ky_thi_id,
                ThiSinh.cong_chuc_id == cc_id,
            )
        )
        result = await self.db.execute(stmt)
        vp = result.scalar_one_or_none()
        if not vp:
            raise HTTPException(
                status_code=404,
                detail={"success": False, "error": {"code": "DGNL_052", "message": "Không tìm thấy vi phạm"}},
            )
        vp.ly_do = (ly_do or "").strip() or None
        await self.db.commit()

    async def danh_sach_vi_pham(
        self, ky_thi_id: uuid.UUID, cong_chuc_id: uuid.UUID
    ) -> list[dict]:
        """Danh sach vi pham chi tiet cua 1 thi sinh (admin/TCCB xem)."""
        stmt = (
            select(ViPhamThi)
            .join(ThiSinh, ViPhamThi.thi_sinh_id == ThiSinh.id)
            .where(
                ViPhamThi.ky_thi_id == ky_thi_id,
                ThiSinh.cong_chuc_id == cong_chuc_id,
            )
            .order_by(ViPhamThi.thoi_gian.asc())
        )
        result = await self.db.execute(stmt)
        return [
            {
                "id": str(vp.id),
                "lan_thi": vp.lan_thi,
                "loai_vi_pham": vp.loai_vi_pham,
                "thoi_gian": vp.thoi_gian.isoformat() if vp.thoi_gian else None,
                "ly_do": vp.ly_do,
            }
            for vp in result.scalars().all()
        ]

    @staticmethod
    def _project_lich_su_summary(lich_su: Optional[list]) -> Optional[list[dict]]:
        """Project raw lich_su_thi (co chi_tiet_tra_loi) -> summary array.

        Dung cho ket_qua endpoints + danh_sach_thi_sinh — han che payload va
        an chi tiet cau tra loi (FE lay rieng qua /ket-qua/{lan}).
        """
        if not lich_su:
            return None
        return [
            {
                "lan": e.get("lan", 0),
                "diem": e.get("diem", 0) or 0,
                "xep_loai": e.get("xep_loai"),
                "so_cau_dung": e.get("so_cau_dung"),
                "so_cau_sai": e.get("so_cau_sai"),
                "tong_so_cau": e.get("tong_so_cau"),
                "thoi_gian_bat_dau": e.get("thoi_gian_bat_dau"),
                "thoi_gian_nop": e.get("thoi_gian_nop"),
                "thoi_gian_lam_giay": e.get("thoi_gian_lam_giay"),
                "has_chi_tiet": bool(e.get("chi_tiet_tra_loi")),
            }
            for e in lich_su if e
        ]

    @staticmethod
    def _snapshot_lan_thi(ts: ThiSinh) -> dict:
        """Build snapshot 1 entry lich_su_thi tu state hien tai cua thi sinh.

        Luu day du chi_tiet_tra_loi + de_thi_ids + diem_theo_linh_vuc de
        QT_DAO_TAO co the drill-down lai bai lam cua lan thi cu.
        """
        return {
            "lan": ts.lan_thi_hien_tai or 0,
            "diem": float(ts.diem_tong) if ts.diem_tong is not None else 0,
            "xep_loai": ts.xep_loai,
            "so_cau_dung": ts.so_cau_dung,
            "so_cau_sai": ts.so_cau_sai,
            "tong_so_cau": ts.tong_so_cau,
            "thoi_gian_bat_dau": ts.thoi_gian_bat_dau.isoformat() if ts.thoi_gian_bat_dau else None,
            "thoi_gian_nop": ts.thoi_gian_nop.isoformat() if ts.thoi_gian_nop else None,
            "thoi_gian_lam_giay": ts.thoi_gian_lam_giay,
            "diem_theo_linh_vuc": dict(ts.diem_theo_linh_vuc) if ts.diem_theo_linh_vuc else {},
            "de_thi_ids": list(ts.de_thi_ids or []),
            "chi_tiet_tra_loi": list(ts.chi_tiet_tra_loi or []),
        }

    @staticmethod
    def _upsert_lan_thi(ts: ThiSinh, snapshot: dict) -> None:
        """Upsert snapshot vao ts.lich_su_thi (theo lan). Idempotent.

        - Neu da co entry voi cung `lan` -> replace.
        - Neu chua co -> append.
        Reassign list moi de SQLAlchemy detect change tren JSONB.
        """
        lich_su = list(ts.lich_su_thi or [])
        lan = snapshot.get("lan")
        idx = next((i for i, e in enumerate(lich_su) if (e or {}).get("lan") == lan), None)
        if idx is not None:
            lich_su[idx] = snapshot
        else:
            lich_su.append(snapshot)
        ts.lich_su_thi = lich_su

    @staticmethod
    def _normalize_dap_an_dung(da):
        """Chuan hoa dap_an_dung: tu full dict ch.dap_an → scalar/list/None.

        Backend lich su luu nguyen ch.dap_an ({lua_chon, dap_an_dung, ...} hoac {goi_y})
        vao chi_tiet_tra_loi. Frontend khi do gap dict se hien '[object Object]'.
        Helper nay extract gia tri thuc su can hien thi.
        """
        if isinstance(da, dict):
            if "dap_an_dung" in da:
                return da["dap_an_dung"]
            if "goi_y" in da:
                return da["goi_y"]
            return None
        return da

    def _cham_1_cau(self, ch: CauHoiDgnl, tra_loi: Optional[dict]) -> bool:
        """Cham 1 cau hoi. Tra ve True/False."""
        if tra_loi is None:
            return False

        dap_an = ch.dap_an
        if not dap_an:
            return False

        loai = ch.loai

        if loai == "TRAC_NGHIEM_1":
            # tra_loi: {"dap_an": "A"}, dap_an: {"dap_an_dung": "A", ...} hoac {"lua_chon": [...], "dap_an_dung": "A"}
            correct = dap_an.get("dap_an_dung", "")
            answer = tra_loi.get("dap_an", "")
            return str(answer).strip().upper() == str(correct).strip().upper()

        elif loai == "TRAC_NGHIEM_NHIEU":
            # tra_loi: {"dap_an": ["A", "C"]}, dap_an: {"dap_an_dung": ["A", "C"]}
            correct = set(str(x).upper() for x in (dap_an.get("dap_an_dung") or []))
            answer = set(str(x).upper() for x in (tra_loi.get("dap_an") or []))
            return correct == answer

        elif loai == "DUNG_SAI":
            correct = str(dap_an.get("dap_an_dung", "")).upper()
            answer = str(tra_loi.get("dap_an", "")).upper()
            return correct == answer

        elif loai == "GHEP_DOI":
            correct = dap_an.get("dap_an_dung", {})
            answer = tra_loi.get("dap_an", {})
            return correct == answer

        # TU_LUAN: khong cham tu dong
        return False

    # ================================================================
    # KET QUA
    # ================================================================

    async def ket_qua_ca_nhan(self, ky_thi_id: uuid.UUID, user: TokenPayload) -> KetQuaResponse:
        """Xem ket qua ca nhan."""
        kt = await self._get_ky_thi(ky_thi_id)
        cc_id = uuid.UUID(user.sub)

        stmt = select(ThiSinh).where(ThiSinh.ky_thi_id == ky_thi_id, ThiSinh.cong_chuc_id == cc_id)
        result = await self.db.execute(stmt)
        ts = result.scalar_one_or_none()
        if not ts:
            raise HTTPException(status_code=404, detail={"success": False, "error": {"code": "DGNL_028", "message": "Bạn không có trong kỳ thi này"}})
        if ts.trang_thai not in ("DA_NOP",):
            raise HTTPException(status_code=400, detail={"success": False, "error": {"code": "DGNL_034", "message": "Bạn chưa nộp bài"}})

        return await self._build_ket_qua(ts, kt)

    async def ket_qua_cbcc(
        self, ky_thi_id: uuid.UUID, cong_chuc_id: uuid.UUID, user: TokenPayload
    ) -> KetQuaResponse:
        """Xem ket qua 1 CBCC cu the (QT/LD)."""
        kt = await self._get_ky_thi(ky_thi_id)

        stmt = select(ThiSinh).where(ThiSinh.ky_thi_id == ky_thi_id, ThiSinh.cong_chuc_id == cong_chuc_id)
        result = await self.db.execute(stmt)
        ts = result.scalar_one_or_none()
        if not ts:
            raise HTTPException(status_code=404, detail={"success": False, "error": {"code": "DGNL_023", "message": "Thí sinh không tồn tại trong kỳ thi này"}})

        return await self._build_ket_qua(ts, kt)

    async def ket_qua_lan_thi(
        self,
        ky_thi_id: uuid.UUID,
        cong_chuc_id: uuid.UUID,
        lan: int,
        user: TokenPayload,
    ) -> KetQuaResponse:
        """Xem ket qua 1 lan thi cu the (QT/LD).

        Uu tien doc tu lich_su_thi (tu 30/07/2026 cot chinh mang ket qua lan
        TOT NHAT, khong con dai dien cho lan hien tai). Fallback ve state hien
        tai cho data cu chua co snapshot lan hien tai trong lich_su_thi.
        - Entry cu khong co chi_tiet_tra_loi (data truoc 30/05/2026) -> tra ve
          response nhung chi_tiet=None, FE hien "Chi tiet khong kha dung".
        """
        kt = await self._get_ky_thi(ky_thi_id)

        stmt = select(ThiSinh).where(ThiSinh.ky_thi_id == ky_thi_id, ThiSinh.cong_chuc_id == cong_chuc_id)
        result = await self.db.execute(stmt)
        ts = result.scalar_one_or_none()
        if not ts:
            raise HTTPException(status_code=404, detail={"success": False, "error": {"code": "DGNL_023", "message": "Thí sinh không tồn tại trong kỳ thi này"}})

        # Tim trong lich su (nop_bai luon snapshot ca lan hien tai)
        entry = next(
            (e for e in (ts.lich_su_thi or []) if (e or {}).get("lan") == lan),
            None,
        )
        if entry:
            return await self._build_ket_qua_lan(ts, kt, entry)

        # Backward compat: data cu chua snapshot lan hien tai vao lich_su_thi
        if lan == (ts.lan_thi_hien_tai or 0):
            return await self._build_ket_qua(ts, kt)

        raise HTTPException(
            status_code=404,
            detail={"success": False, "error": {"code": "DGNL_035", "message": f"Không tìm thấy lần thi {lan}"}},
        )

    async def _build_ket_qua(self, ts: ThiSinh, kt: KyThi) -> KetQuaResponse:
        """Build response ket qua tu state hien tai cua thi sinh (lan moi nhat)."""
        cc_info = await self._lay_thi_sinh_info(ts)
        diem_lv_list = await self._build_diem_lv_list(ts.diem_theo_linh_vuc)
        chi_tiet = await self._enrich_chi_tiet(ts.chi_tiet_tra_loi) if kt.hien_dap_an else None

        # Lan dat diem cao nhat (cot chinh dang mang ket qua lan nay)
        entries = [e for e in (ts.lich_su_thi or []) if e and e.get("lan") is not None]
        lan_tot_nhat = None
        if entries:
            best = max(entries, key=lambda e: ((e.get("diem") or 0), -(e.get("lan") or 0)))
            lan_tot_nhat = best.get("lan")

        # Han tu dong xac nhan ca thi (chi khi vua nop, chua xac nhan)
        han_xac_nhan = None
        if ts.trang_thai == "DA_NOP" and not ts.da_xac_nhan and ts.thoi_gian_nop:
            han_xac_nhan = (ts.thoi_gian_nop + timedelta(minutes=XAC_NHAN_TIMEOUT_PHUT)).isoformat()

        return KetQuaResponse(
            ky_thi={"id": str(kt.id), "ten_ky_thi": kt.ten_ky_thi, "diem_dat": float(kt.diem_dat) if kt.diem_dat else 50},
            thi_sinh=cc_info,
            ket_qua={
                "diem_tong": float(ts.diem_tong) if ts.diem_tong else 0,
                "xep_loai": ts.xep_loai,
                "so_cau_dung": ts.so_cau_dung or 0,
                "so_cau_sai": ts.so_cau_sai or 0,
                "tong_so_cau": ts.tong_so_cau or 0,
                "thoi_gian_lam_giay": ts.thoi_gian_lam_giay or 0,
                "lan_thi": ts.lan_thi_hien_tai or 0,
                "lan_tot_nhat": lan_tot_nhat,
                "so_lan_thi_toi_da": kt.so_lan_thi_toi_da or 1,
                "da_xac_nhan": bool(ts.da_xac_nhan),
                "thoi_gian_nop": ts.thoi_gian_nop.isoformat() if ts.thoi_gian_nop else None,
                "han_xac_nhan": han_xac_nhan,
            },
            diem_theo_linh_vuc=diem_lv_list,
            chi_tiet=chi_tiet,
            lich_su_thi=self._project_lich_su_summary(ts.lich_su_thi),
        )

    async def _build_ket_qua_lan(
        self, ts: ThiSinh, kt: KyThi, entry: dict
    ) -> KetQuaResponse:
        """Build response ket qua tu 1 entry lich_su_thi (lan cu)."""
        cc_info = await self._lay_thi_sinh_info(ts)
        diem_lv_list = await self._build_diem_lv_list(entry.get("diem_theo_linh_vuc"))
        chi_tiet = None
        if kt.hien_dap_an and entry.get("chi_tiet_tra_loi"):
            chi_tiet = await self._enrich_chi_tiet(entry.get("chi_tiet_tra_loi"))

        return KetQuaResponse(
            ky_thi={"id": str(kt.id), "ten_ky_thi": kt.ten_ky_thi, "diem_dat": float(kt.diem_dat) if kt.diem_dat else 50},
            thi_sinh=cc_info,
            ket_qua={
                "diem_tong": float(entry.get("diem") or 0),
                "xep_loai": entry.get("xep_loai"),
                "so_cau_dung": entry.get("so_cau_dung") or 0,
                "so_cau_sai": entry.get("so_cau_sai") or 0,
                "tong_so_cau": entry.get("tong_so_cau") or 0,
                "thoi_gian_lam_giay": entry.get("thoi_gian_lam_giay") or 0,
                "lan_thi": entry.get("lan") or 0,
            },
            diem_theo_linh_vuc=diem_lv_list,
            chi_tiet=chi_tiet,
            lich_su_thi=self._project_lich_su_summary(ts.lich_su_thi),
        )

    async def _lay_thi_sinh_info(self, ts: ThiSinh) -> dict:
        """Lay ho_ten/ma_cc/don_vi/vi_tri cua thi sinh."""
        cc = CongChucRef.__table__
        dv = DonViRef.__table__
        stmt = (
            select(cc.c.ho_ten, cc.c.ma_cc, dv.c.ten_don_vi)
            .outerjoin(dv, cc.c.don_vi_id == dv.c.id)
            .where(cc.c.id == ts.cong_chuc_id)
        )
        r = await self.db.execute(stmt)
        cc_row = r.first()

        vt_r = await self.db.execute(
            select(ViTriViecLam.ten_vi_tri).where(ViTriViecLam.id == ts.vi_tri_id)
        )
        vt_ten = vt_r.scalar() or ""

        return {
            "ho_ten": cc_row.ho_ten if cc_row else None,
            "ma_cc": cc_row.ma_cc if cc_row else None,
            "don_vi": cc_row.ten_don_vi if cc_row else None,
            "vi_tri_thi": vt_ten,
        }

    async def _build_diem_lv_list(self, diem_theo_lv: Optional[dict]) -> list[DiemLinhVuc]:
        """Build list DiemLinhVuc tu dict {ma_lv: {dung, tong, phan_tram}}."""
        out: list[DiemLinhVuc] = []
        if not diem_theo_lv:
            return out
        for ma, stats in diem_theo_lv.items():
            lv_r = await self.db.execute(
                select(LinhVuc.ten_linh_vuc).where(LinhVuc.ma_linh_vuc == ma)
            )
            lv_ten = lv_r.scalar() or ma
            out.append(DiemLinhVuc(
                linh_vuc=lv_ten,
                so_cau_dung=(stats or {}).get("dung", 0),
                tong_cau=(stats or {}).get("tong", 0),
                phan_tram=(stats or {}).get("phan_tram", 0),
            ))
        return out

    async def _enrich_chi_tiet(self, chi_tiet_raw: Optional[list]) -> Optional[list]:
        """Enrich chi_tiet_tra_loi them noi_dung/loai/giai_thich tu CauHoiDgnl.

        Dung cho ca lan hien tai (ts.chi_tiet_tra_loi) va lan cu (entry["chi_tiet_tra_loi"]).
        """
        if not chi_tiet_raw:
            return None
        try:
            ch_ids = [uuid.UUID(item["cau_hoi_id"]) for item in chi_tiet_raw if item.get("cau_hoi_id")]
        except (ValueError, TypeError):
            ch_ids = []
        ch_map: dict[uuid.UUID, CauHoiDgnl] = {}
        if ch_ids:
            ch_r = await self.db.execute(select(CauHoiDgnl).where(CauHoiDgnl.id.in_(ch_ids)))
            ch_map = {ch.id: ch for ch in ch_r.scalars().all()}

        out: list = []
        for item in chi_tiet_raw:
            enriched = dict(item)
            try:
                cid = uuid.UUID(item["cau_hoi_id"]) if item.get("cau_hoi_id") else None
            except (ValueError, TypeError):
                cid = None
            ch = ch_map.get(cid) if cid else None
            if ch:
                enriched["noi_dung"] = ch.noi_dung
                enriched["loai"] = ch.loai
                enriched["giai_thich"] = ch.giai_thich
                enriched["diem_toi_da"] = float(ch.diem) if ch.diem else 1.0
                enriched["diem_dat"] = enriched["diem_toi_da"] if item.get("dung") else 0.0
            # Chuan hoa data cu — unwrap dict ch.dap_an de FE khong hien "[object Object]"
            enriched["dap_an_dung"] = self._normalize_dap_an_dung(enriched.get("dap_an_dung"))
            out.append(enriched)
        return out

    # ================================================================
    # EXPORT EXCEL
    # ================================================================

    async def export_excel(self, ky_thi_id: uuid.UUID, user: TokenPayload) -> bytes:
        """Export ket qua ky thi ra Excel: sheet Tong quan (thong ke) + sheet Danh sach thi sinh.

        Lay TAT CA thi sinh (khong gioi han so luong). Lanh dao chi export don vi minh.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        import io

        kt = await self._get_ky_thi(ky_thi_id)

        # Lay danh sach thi sinh (toan bo). Lanh dao chi xem don vi minh.
        cc = CongChucRef.__table__.alias("cc")
        dv = DonViRef.__table__.alias("dv")

        base_where = [ThiSinh.ky_thi_id == ky_thi_id]
        # Module DGNL chi admin (QT_DAO_TAO/SUPER_ADMIN) duoc export -> khong scope don vi.

        stmt = (
            select(
                ThiSinh,
                cc.c.ho_ten, cc.c.ma_cc,
                dv.c.ten_don_vi.label("don_vi_ten"),
                ViTriViecLam.ten_vi_tri.label("vi_tri_ten"),
            )
            .outerjoin(cc, ThiSinh.cong_chuc_id == cc.c.id)
            .outerjoin(dv, cc.c.don_vi_id == dv.c.id)
            .outerjoin(ViTriViecLam, ThiSinh.vi_tri_id == ViTriViecLam.id)
            .where(*base_where)
            .order_by(cc.c.ho_ten.asc())
        )
        result = await self.db.execute(stmt)
        rows = result.all()

        # Lay danh sach linh vuc de lam header
        lv_stmt = select(LinhVuc).where(LinhVuc.is_active == True).order_by(LinhVuc.thu_tu)
        lv_result = await self.db.execute(lv_stmt)
        all_lv = lv_result.scalars().all()
        lv_ma_to_ten = {lv.ma_linh_vuc: lv.ten_linh_vuc for lv in all_lv}

        # Styles dung chung
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        dat_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        kdat_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        def auto_width(ws, n_cols: int, from_row: int = 1) -> None:
            """Tinh do rong cot — duyet theo index, bo qua merged cell de tranh
            AttributeError tren MergedCell.column_letter."""
            for col_idx in range(1, n_cols + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                for row_idx in range(from_row, ws.max_row + 1):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val is not None:
                        max_length = max(max_length, len(str(val)))
                ws.column_dimensions[col_letter].width = min(max_length + 3, 30)

        wb = Workbook()

        # ============================================================
        # SHEET 1: TONG QUAN (thong ke)
        # ============================================================
        ws_tq = wb.active
        ws_tq.title = "Tổng quan"

        all_ts = [r[0] for r in rows]
        tong = len(all_ts)
        da_thi = sum(1 for t in all_ts if t.trang_thai == "DA_NOP")
        chua_thi = sum(1 for t in all_ts if t.trang_thai == "CHUA_THI")
        dang_thi = sum(1 for t in all_ts if t.trang_thai == "DANG_THI")
        vang = sum(1 for t in all_ts if t.trang_thai == "VANG")
        dat = sum(1 for t in all_ts if t.xep_loai == "DAT")
        khong_dat = sum(1 for t in all_ts if t.xep_loai == "KHONG_DAT")
        diem_list = [float(t.diem_tong) for t in all_ts if t.diem_tong is not None]
        diem_tb = round(sum(diem_list) / len(diem_list), 1) if diem_list else 0
        diem_cao = round(max(diem_list), 1) if diem_list else 0
        diem_thap = round(min(diem_list), 1) if diem_list else 0
        ti_le_dat = round(dat / da_thi * 100, 1) if da_thi > 0 else 0

        ws_tq.merge_cells("A1:C1")
        ws_tq["A1"] = f"THỐNG KÊ KỲ THI: {kt.ten_ky_thi}"
        ws_tq["A1"].font = Font(bold=True, size=14)
        ws_tq["A1"].alignment = center
        ws_tq.merge_cells("A2:C2")
        ws_tq["A2"] = f"Mã kỳ thi: {kt.ma_ky_thi} | Điểm đạt: {kt.diem_dat}%"
        ws_tq["A2"].font = Font(size=11)

        tq_rows = [
            ("Tổng thí sinh", tong),
            ("Đã thi", da_thi),
            ("Chưa thi", chua_thi),
            ("Đang thi", dang_thi),
            ("Vắng", vang),
            ("Đạt", dat),
            ("Không đạt", khong_dat),
            ("Tỷ lệ đạt (%)", ti_le_dat),
            ("Điểm trung bình", diem_tb),
            ("Điểm cao nhất", diem_cao),
            ("Điểm thấp nhất", diem_thap),
        ]
        r = 4
        for label, val in tq_rows:
            ws_tq.cell(row=r, column=1, value=label).font = Font(bold=True)
            c = ws_tq.cell(row=r, column=2, value=val)
            c.alignment = center
            ws_tq.cell(row=r, column=1).border = thin_border
            c.border = thin_border
            r += 1

        # Theo vi tri
        vt_groups: dict[str, list] = {}
        for ts, _ho, _ma, _dv, vt_ten in rows:
            vt_groups.setdefault(vt_ten or "Không xác định", []).append(ts)
        if vt_groups:
            r += 1
            ws_tq.cell(row=r, column=1, value="THEO VỊ TRÍ").font = Font(bold=True, size=12)
            r += 1
            for col_idx, h in enumerate(["Vị trí", "Tổng", "Đạt", "Không đạt", "Điểm TB"], 1):
                cell = ws_tq.cell(row=r, column=col_idx, value=h)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = center
                cell.border = thin_border
            r += 1
            for vt_ten, vt_ts in vt_groups.items():
                vt_dat = sum(1 for t in vt_ts if t.xep_loai == "DAT")
                vt_kdat = sum(1 for t in vt_ts if t.xep_loai == "KHONG_DAT")
                vt_diem = [float(t.diem_tong) for t in vt_ts if t.diem_tong is not None]
                vals = [
                    vt_ten, len(vt_ts), vt_dat, vt_kdat,
                    round(sum(vt_diem) / len(vt_diem), 1) if vt_diem else 0,
                ]
                for col_idx, v in enumerate(vals, 1):
                    cell = ws_tq.cell(row=r, column=col_idx, value=v)
                    cell.border = thin_border
                    if col_idx > 1:
                        cell.alignment = center
                r += 1

        # Theo don vi
        dv_groups: dict[str, list] = {}
        for ts, _ho, _ma, dv_ten, _vt in rows:
            dv_groups.setdefault(dv_ten or "Không xác định", []).append(ts)
        if dv_groups:
            r += 1
            ws_tq.cell(row=r, column=1, value="THEO ĐƠN VỊ").font = Font(bold=True, size=12)
            r += 1
            for col_idx, h in enumerate(["Đơn vị", "Tổng", "Đạt", "Tỷ lệ đạt (%)", "Điểm TB"], 1):
                cell = ws_tq.cell(row=r, column=col_idx, value=h)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = center
                cell.border = thin_border
            r += 1
            for dv_ten, dv_ts in dv_groups.items():
                dv_dat = sum(1 for t in dv_ts if t.xep_loai == "DAT")
                dv_diem = [float(t.diem_tong) for t in dv_ts if t.diem_tong is not None]
                vals = [
                    dv_ten, len(dv_ts), dv_dat,
                    round(dv_dat / len(dv_ts) * 100, 1) if dv_ts else 0,
                    round(sum(dv_diem) / len(dv_diem), 1) if dv_diem else 0,
                ]
                for col_idx, v in enumerate(vals, 1):
                    cell = ws_tq.cell(row=r, column=col_idx, value=v)
                    cell.border = thin_border
                    if col_idx > 1:
                        cell.alignment = center
                r += 1

        auto_width(ws_tq, 5, from_row=4)

        # ============================================================
        # SHEET 2: DANH SACH THI SINH
        # ============================================================
        ws = wb.create_sheet("Danh sách thí sinh")

        base_headers = ["STT", "Mã CC", "Họ tên", "Đơn vị", "Vị trí thi",
                        "Trạng thái", "Điểm (%)", "Xếp loại", "Số câu đúng",
                        "Tổng câu", "Thời gian (phút)", "Lần thi"]
        lv_headers = [lv_ma_to_ten.get(ma, ma) for ma in lv_ma_to_ten.keys()]
        headers = base_headers + lv_headers
        n_cols = len(headers)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        ws.cell(row=1, column=1, value=f"KẾT QUẢ KỲ THI: {kt.ten_ky_thi}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = center

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        ws.cell(row=2, column=1, value=f"Mã kỳ thi: {kt.ma_ky_thi} | Điểm đạt: {kt.diem_dat}%")
        ws.cell(row=2, column=1).font = Font(size=11)

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        for i, (ts, ho_ten, ma_cc, dv_ten, vt_ten) in enumerate(rows, 1):
            row_idx = i + 4
            thoi_gian_phut = round(ts.thoi_gian_lam_giay / 60, 1) if ts.thoi_gian_lam_giay else ""

            base_data = [
                i, ma_cc, ho_ten, dv_ten, vt_ten,
                ts.trang_thai, float(ts.diem_tong) if ts.diem_tong else "",
                ts.xep_loai or "", ts.so_cau_dung or "",
                ts.tong_so_cau or "", thoi_gian_phut,
                ts.lan_thi_hien_tai or 0,
            ]

            lv_data = []
            for ma in lv_ma_to_ten.keys():
                if ts.diem_theo_linh_vuc and ma in ts.diem_theo_linh_vuc:
                    stats = ts.diem_theo_linh_vuc[ma]
                    lv_data.append(f"{stats.get('dung', 0)}/{stats.get('tong', 0)}")
                else:
                    lv_data.append("")

            all_data = base_data + lv_data

            for col_idx, val in enumerate(all_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = center

            xep_loai_cell = ws.cell(row=row_idx, column=8)
            if ts.xep_loai == "DAT":
                xep_loai_cell.fill = dat_fill
            elif ts.xep_loai == "KHONG_DAT":
                xep_loai_cell.fill = kdat_fill

        auto_width(ws, n_cols, from_row=4)

        # ============================================================
        # SHEET 3: LICH SU LUOT THI — 1 row / (thi sinh × lan thi)
        # Bat AutoFilter de QT_DAO_TAO/LD filter theo cot "Lan" hoac "Xep loai"
        # ============================================================
        ws_ls = wb.create_sheet("Lịch sử lượt thi")

        ls_headers = [
            "STT", "Mã CC", "Họ tên", "Đơn vị", "Vị trí thi",
            "Lần", "Điểm (%)", "Xếp loại", "Số câu đúng", "Tổng câu",
            "Thời gian (phút)", "Thời gian nộp",
        ]
        ls_n_cols = len(ls_headers)

        ws_ls.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ls_n_cols)
        ws_ls.cell(row=1, column=1, value=f"LỊCH SỬ LƯỢT THI: {kt.ten_ky_thi}")
        ws_ls.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws_ls.cell(row=1, column=1).alignment = center

        ws_ls.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ls_n_cols)
        ws_ls.cell(
            row=2, column=1,
            value=f"Mã kỳ thi: {kt.ma_ky_thi} | Điểm đạt: {kt.diem_dat}% | "
                  f"Số lần tối đa: {kt.so_lan_thi_toi_da or 1} | "
                  f"Mẹo: Dùng AutoFilter ▼ trên cột \"Lần\" để xem 1 lượt cụ thể"
        )
        ws_ls.cell(row=2, column=1).font = Font(size=10, italic=True, color="666666")

        for col_idx, h in enumerate(ls_headers, 1):
            cell = ws_ls.cell(row=4, column=col_idx, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        # Build rows: 1 row / (thi_sinh × lan)
        ls_row_idx = 5
        stt = 0
        for ts, ho_ten, ma_cc, dv_ten, vt_ten in rows:
            lich_su = ts.lich_su_thi or []
            lan_hien_tai = ts.lan_thi_hien_tai or 0
            cu_co_lan_ht = any((e or {}).get("lan") == lan_hien_tai for e in lich_su)

            # Cac lan tu lich_su_thi (sap xep theo lan tang dan)
            entries = sorted(
                [e for e in lich_su if e and e.get("lan") is not None],
                key=lambda e: e.get("lan", 0),
            )

            # Backward-compat: data cu chua snapshot lan hien tai vao lich_su_thi.
            # Bo sung tu ts.* truc tiep.
            if ts.trang_thai == "DA_NOP" and lan_hien_tai > 0 and not cu_co_lan_ht:
                entries.append({
                    "lan": lan_hien_tai,
                    "diem": float(ts.diem_tong) if ts.diem_tong is not None else 0,
                    "xep_loai": ts.xep_loai,
                    "so_cau_dung": ts.so_cau_dung,
                    "tong_so_cau": ts.tong_so_cau,
                    "thoi_gian_lam_giay": ts.thoi_gian_lam_giay,
                    "thoi_gian_nop": ts.thoi_gian_nop.isoformat() if ts.thoi_gian_nop else None,
                })

            for e in entries:
                stt += 1
                tg_phut = round(e["thoi_gian_lam_giay"] / 60, 1) if e.get("thoi_gian_lam_giay") else ""
                tg_nop_str = ""
                if e.get("thoi_gian_nop"):
                    try:
                        # fmt_vn: chuoi cu khong offset coi la UTC, chuoi moi co +07:00
                        tg_nop_str = fmt_vn(e["thoi_gian_nop"])
                    except (ValueError, TypeError):
                        tg_nop_str = str(e["thoi_gian_nop"])

                row_data = [
                    stt, ma_cc, ho_ten, dv_ten, vt_ten,
                    e.get("lan"),
                    float(e["diem"]) if e.get("diem") is not None else "",
                    e.get("xep_loai") or "",
                    e.get("so_cau_dung") or "",
                    e.get("tong_so_cau") or "",
                    tg_phut,
                    tg_nop_str,
                ]
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws_ls.cell(row=ls_row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                    cell.alignment = center

                xl_cell = ws_ls.cell(row=ls_row_idx, column=8)
                if e.get("xep_loai") == "DAT":
                    xl_cell.fill = dat_fill
                elif e.get("xep_loai") == "KHONG_DAT":
                    xl_cell.fill = kdat_fill

                ls_row_idx += 1

        # Khong co data -> hien thong bao
        if stt == 0:
            ws_ls.merge_cells(start_row=5, start_column=1, end_row=5, end_column=ls_n_cols)
            ws_ls.cell(row=5, column=1, value="Chưa có lượt thi nào được nộp.").alignment = center
            ws_ls.cell(row=5, column=1).font = Font(italic=True, color="888888")
        else:
            # AutoFilter tu row header (4) den row cuoi cung
            ws_ls.auto_filter.ref = f"A4:{get_column_letter(ls_n_cols)}{ls_row_idx - 1}"
            # Freeze header de scroll giu tieu de
            ws_ls.freeze_panes = "A5"

        auto_width(ws_ls, ls_n_cols, from_row=4)

        # ============================================================
        # SHEET 4: VI PHAM CHI TIET (gio VN + ly do giai trinh)
        # ============================================================
        ws_vp = wb.create_sheet("Vi phạm chi tiết")
        vp_headers = ["STT", "Mã CC", "Họ tên", "Đơn vị", "Lần thi", "Loại vi phạm", "Thời gian", "Lý do giải trình"]
        vp_n_cols = len(vp_headers)

        ws_vp.merge_cells(start_row=1, start_column=1, end_row=1, end_column=vp_n_cols)
        tieu_de_vp = ws_vp.cell(row=1, column=1, value=f"VI PHẠM CHI TIẾT — {kt.ten_ky_thi}")
        tieu_de_vp.font = Font(bold=True, size=13)
        tieu_de_vp.alignment = center

        for col_idx, h in enumerate(vp_headers, 1):
            cell = ws_vp.cell(row=3, column=col_idx, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        vp_stmt = (
            select(
                ViPhamThi,
                cc.c.ho_ten, cc.c.ma_cc,
                dv.c.ten_don_vi.label("don_vi_ten"),
            )
            .join(ThiSinh, ViPhamThi.thi_sinh_id == ThiSinh.id)
            .outerjoin(cc, ThiSinh.cong_chuc_id == cc.c.id)
            .outerjoin(dv, cc.c.don_vi_id == dv.c.id)
            .where(ViPhamThi.ky_thi_id == ky_thi_id)
            .order_by(cc.c.ho_ten.asc(), ViPhamThi.thoi_gian.asc())
        )
        vp_result = await self.db.execute(vp_stmt)
        vp_rows = vp_result.all()

        LOAI_VP_LABEL = {"EXIT_FULLSCREEN": "Thoát toàn màn hình", "SWITCH_TAB": "Chuyển tab/cửa sổ"}
        vp_row_idx = 4
        for idx, (vp, ho_ten, ma_cc, don_vi_ten) in enumerate(vp_rows, 1):
            row_data = [
                idx, ma_cc, ho_ten, don_vi_ten,
                vp.lan_thi,
                LOAI_VP_LABEL.get(vp.loai_vi_pham, vp.loai_vi_pham),
                fmt_vn(vp.thoi_gian, "%d/%m/%Y %H:%M:%S"),
                vp.ly_do or "",
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws_vp.cell(row=vp_row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if col_idx != 8:  # ly do canh trai
                    cell.alignment = center
            vp_row_idx += 1

        if not vp_rows:
            ws_vp.merge_cells(start_row=4, start_column=1, end_row=4, end_column=vp_n_cols)
            ws_vp.cell(row=4, column=1, value="Không có vi phạm nào được ghi nhận.").alignment = center
            ws_vp.cell(row=4, column=1).font = Font(italic=True, color="888888")
        else:
            ws_vp.auto_filter.ref = f"A3:{get_column_letter(vp_n_cols)}{vp_row_idx - 1}"
            ws_vp.freeze_panes = "A4"
        auto_width(ws_vp, vp_n_cols, from_row=3)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
