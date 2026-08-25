"""
lms_service/tests/test_dgnl.py
==============================
Tests cho module Danh gia Nang luc (DGNL):
  - Linh vuc CRUD (4 tests)
  - Vi tri viec lam CRUD (4 tests)
  - Ky thi CRUD + trang thai (8 tests)
  - Cau truc de (4 tests)
  - Validate ngan hang (2 tests)
  - Thi sinh + lam thi + cham diem (8 tests)
  - Phan quyen (4 tests)
  Total: 34 tests
"""

import uuid
from datetime import datetime, timedelta

import pytest
import pytest_asyncio
from httpx import AsyncClient
from lms_service.core.timezone import now_vn

pytestmark = pytest.mark.asyncio

BASE = "/api/v1/lms"


# =========================================================================
# HELPERS
# =========================================================================

async def _create_linh_vuc(client: AsyncClient, ma: str = None, ten: str = None) -> dict:
    ma = ma or f"LV-{uuid.uuid4().hex[:6]}"
    ten = ten or f"Lĩnh vực {ma}"
    resp = await client.post(f"{BASE}/linh-vuc", json={
        "ma_linh_vuc": ma,
        "ten_linh_vuc": ten,
    })
    assert resp.status_code == 201
    return resp.json()["data"]


async def _create_vi_tri(client: AsyncClient, ma: str = None, ten: str = None) -> dict:
    ma = ma or f"VT-{uuid.uuid4().hex[:6]}"
    ten = ten or f"Vị trí {ma}"
    resp = await client.post(f"{BASE}/vi-tri-viec-lam", json={
        "ma_vi_tri": ma,
        "ten_vi_tri": ten,
    })
    assert resp.status_code == 201
    return resp.json()["data"]


async def _create_ky_thi(client: AsyncClient, ma: str = None) -> dict:
    ma = ma or f"KT-{uuid.uuid4().hex[:6]}"
    now = now_vn()
    resp = await client.post(f"{BASE}/ky-thi", json={
        "ma_ky_thi": ma,
        "ten_ky_thi": f"Kỳ thi {ma}",
        "ngay_bat_dau": (now - timedelta(hours=1)).isoformat(),
        "ngay_ket_thuc": (now + timedelta(days=7)).isoformat(),
        "thoi_gian_lam_bai_phut": 30,
        "diem_dat": 50,
        "so_lan_thi_toi_da": 2,
    })
    assert resp.status_code == 201
    return resp.json()["data"]


async def _setup_full_exam(client: AsyncClient, admin_user) -> dict:
    """Setup hoan chinh: linh_vuc + vi_tri + ky_thi + cau_truc_de + cau_hoi + thi_sinh."""
    # Linh vuc
    uid = uuid.uuid4().hex[:6]
    lv = await _create_linh_vuc(client, f"LV-FULL-{uid}", "Lĩnh vực test")

    # Vi tri
    vt = await _create_vi_tri(client, f"VT-FULL-{uid}", "Vị trí test")

    # Ky thi
    kt = await _create_ky_thi(client, f"KT-FULL-{uid}")

    # Tao cau hoi co linh_vuc_id
    # Can 1 khoa hoc + BKT de gan cau hoi (bai_kiem_tra_id required)
    kh_resp = await client.post(f"{BASE}/khoa-hoc", json={
        "ma_khoa_hoc": f"KH-DGNL-{uid}",
        "ten_khoa_hoc": "Khóa test ĐGNL",
        "loai": "TU_HOC",
    })
    kh_id = kh_resp.json()["data"]["id"]

    # Tao cau hoi truoc (se tu dong tao junction + cap nhat so_cau_hoi)
    def _ch(noi_dung, do_kho, da_dung):
        return {
            "noi_dung": noi_dung, "loai": "TRAC_NGHIEM_1", "do_kho": do_kho,
            "dap_an": {"lua_chon": [{"key": "A", "noi_dung": "Đáp A"}, {"key": "B", "noi_dung": "Đáp B"}], "dap_an_dung": da_dung},
        }

    # Tao BKT voi cau_hoi_ids (bat buoc co it nhat 1 cau)
    ch_bodies = []
    for i in range(5):
        ch_bodies.append(_ch(f"Câu DỄ {i+1}", "DE", "A"))
    for i in range(3):
        ch_bodies.append(_ch(f"Câu TB {i+1}", "TRUNG_BINH", "B"))
    for i in range(2):
        ch_bodies.append(_ch(f"Câu KHÓ {i+1}", "KHO", "A"))

    # Tao BKT voi cau_hoi_moi (inline)
    bkt_resp = await client.post(f"{BASE}/khoa-hoc/{kh_id}/bai-kiem-tra", json={
        "tieu_de": f"BKT-DGNL-{uid}",
        "diem_dat": 30,
        "cau_hoi_moi": ch_bodies,
    })
    assert bkt_resp.status_code == 201, f"BKT create failed: {bkt_resp.json()}"
    bkt_id = bkt_resp.json()["data"]["id"]

    # Cap nhat linh_vuc_id cho cac cau hoi vua tao
    ch_resp = await client.get(f"{BASE}/cau-hoi", params={"bai_kiem_tra_id": bkt_id, "page_size": 100})
    cau_hoi_ids = []
    for ch in ch_resp.json()["data"]:
        await client.put(f"{BASE}/cau-hoi/{ch['id']}", json={"linh_vuc_id": lv["id"]})
        cau_hoi_ids.append(ch["id"])

    # Seed NGAN HANG DGNL (bang cau_hoi_dgnl) — KHAC voi cau_hoi khoa hoc o tren.
    # bat_dau_thi()/validate random tu cau_hoi_dgnl, nen phai co du DE/TB/KHO cho lv nay.
    def _ch_dgnl(noi_dung, do_kho, da_dung="A"):
        return {
            "linh_vuc_id": lv["id"], "do_kho": do_kho, "loai": "TRAC_NGHIEM_1",
            "noi_dung": noi_dung,
            "dap_an": {"lua_chon": [{"key": "A", "noi_dung": "Đáp A"}, {"key": "B", "noi_dung": "Đáp B"}], "dap_an_dung": da_dung},
        }
    for i in range(5):
        r = await client.post(f"{BASE}/dgnl/ngan-hang", json=_ch_dgnl(f"DGNL DỄ {i+1}", "DE"))
        assert r.status_code == 201, f"seed dgnl DE failed: {r.json()}"
    for i in range(3):
        await client.post(f"{BASE}/dgnl/ngan-hang", json=_ch_dgnl(f"DGNL TB {i+1}", "TRUNG_BINH", "B"))
    for i in range(2):
        await client.post(f"{BASE}/dgnl/ngan-hang", json=_ch_dgnl(f"DGNL KHÓ {i+1}", "KHO"))

    # Cau truc de: 2 DE + 1 TB + 1 KHO = 4 cau
    resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de", json={
        "vi_tri_id": vt["id"],
        "cau_truc": [
            {"linh_vuc_id": lv["id"], "so_cau_de": 2, "so_cau_trung_binh": 1, "so_cau_kho": 1},
        ],
    })
    assert resp.status_code == 201

    return {
        "linh_vuc": lv,
        "vi_tri": vt,
        "ky_thi": kt,
        "cau_hoi_ids": cau_hoi_ids,
    }


# =========================================================================
# LINH VUC CRUD
# =========================================================================

class TestLinhVucCRUD:
    async def test_tao_linh_vuc(self, client, admin_user):
        lv = await _create_linh_vuc(client)
        assert lv["ma_linh_vuc"]
        assert lv["is_active"] is True

    async def test_danh_sach_linh_vuc(self, client, admin_user):
        await _create_linh_vuc(client, f"LV-LIST-{uuid.uuid4().hex[:6]}")
        resp = await client.get(f"{BASE}/linh-vuc")
        assert resp.status_code == 200
        assert resp.json()["success"] is True
        # Co it nhat seed data (4) + 1 vua tao
        assert len(resp.json()["data"]) >= 1

    async def test_cap_nhat_linh_vuc(self, client, admin_user):
        lv = await _create_linh_vuc(client)
        resp = await client.put(f"{BASE}/linh-vuc/{lv['id']}", json={
            "ten_linh_vuc": "Tên mới",
        })
        assert resp.status_code == 200
        assert resp.json()["data"]["ten_linh_vuc"] == "Tên mới"

    async def test_xoa_linh_vuc(self, client, admin_user):
        lv = await _create_linh_vuc(client)
        resp = await client.delete(f"{BASE}/linh-vuc/{lv['id']}")
        assert resp.status_code == 200

    async def test_tao_linh_vuc_trung_ma(self, client, admin_user):
        ma = f"LV-DUP-{uuid.uuid4().hex[:4]}"
        await _create_linh_vuc(client, ma)
        resp = await client.post(f"{BASE}/linh-vuc", json={
            "ma_linh_vuc": ma,
            "ten_linh_vuc": "Trùng mã",
        })
        assert resp.status_code == 400


# =========================================================================
# VI TRI VIEC LAM CRUD
# =========================================================================

class TestViTriViecLamCRUD:
    async def test_tao_vi_tri(self, client, admin_user):
        vt = await _create_vi_tri(client)
        assert vt["ma_vi_tri"]

    async def test_danh_sach_vi_tri(self, client, admin_user):
        await _create_vi_tri(client, f"VT-LIST-{uuid.uuid4().hex[:6]}")
        resp = await client.get(f"{BASE}/vi-tri-viec-lam")
        assert resp.status_code == 200
        assert len(resp.json()["data"]) >= 1

    async def test_cap_nhat_vi_tri(self, client, admin_user):
        vt = await _create_vi_tri(client)
        resp = await client.put(f"{BASE}/vi-tri-viec-lam/{vt['id']}", json={
            "ten_vi_tri": "Vị trí mới",
        })
        assert resp.status_code == 200
        assert resp.json()["data"]["ten_vi_tri"] == "Vị trí mới"

    async def test_xoa_vi_tri(self, client, admin_user):
        vt = await _create_vi_tri(client)
        resp = await client.delete(f"{BASE}/vi-tri-viec-lam/{vt['id']}")
        assert resp.status_code == 200


# =========================================================================
# KY THI CRUD + TRANG THAI
# =========================================================================

class TestKyThiCRUD:
    async def test_tao_ky_thi(self, client, admin_user):
        kt = await _create_ky_thi(client)
        assert kt["trang_thai"] == "NHAP"
        assert kt["ma_ky_thi"]

    async def test_danh_sach_ky_thi(self, client, admin_user):
        await _create_ky_thi(client)
        resp = await client.get(f"{BASE}/ky-thi")
        assert resp.status_code == 200
        assert resp.json()["success"] is True

    async def test_chi_tiet_ky_thi(self, client, admin_user):
        kt = await _create_ky_thi(client)
        resp = await client.get(f"{BASE}/ky-thi/{kt['id']}")
        assert resp.status_code == 200
        assert resp.json()["data"]["id"] == kt["id"]

    async def test_cap_nhat_ky_thi(self, client, admin_user):
        kt = await _create_ky_thi(client)
        resp = await client.put(f"{BASE}/ky-thi/{kt['id']}", json={
            "ten_ky_thi": "Tên kỳ thi mới",
        })
        assert resp.status_code == 200
        assert resp.json()["data"]["ten_ky_thi"] == "Tên kỳ thi mới"

    async def test_xoa_ky_thi_nhap(self, client, admin_user):
        kt = await _create_ky_thi(client)
        resp = await client.delete(f"{BASE}/ky-thi/{kt['id']}")
        assert resp.status_code == 200

    async def test_tao_ky_thi_trung_ma(self, client, admin_user):
        ma = f"KT-DUP-{uuid.uuid4().hex[:4]}"
        await _create_ky_thi(client, ma)
        now = now_vn()
        resp = await client.post(f"{BASE}/ky-thi", json={
            "ma_ky_thi": ma,
            "ten_ky_thi": "Trùng mã",
            "ngay_bat_dau": now.isoformat(),
            "ngay_ket_thuc": (now + timedelta(days=1)).isoformat(),
        })
        assert resp.status_code == 400

    async def test_cap_nhat_ky_thi_khi_khong_phai_nhap(self, client, admin_user):
        """Sua THONG TIN ky thi duoc o moi trang thai (khac voi sua CAU TRUC DE).

        Rang buoc "chi sua khi NHAP" da duoc go tu 17/04/2026 (bac5ea1) — FE cung
        hien nut "Sua" o moi trang thai. Test nay giu nguyen assert cu tu do nen
        do lien tuc; nay sua lai cho khop hanh vi that.
        """
        data = await _setup_full_exam(client, admin_user)
        kt = data["ky_thi"]
        # Chuyen sang CHO_DUYET
        await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "CHO_DUYET"})
        # Van sua duoc thong tin ky thi
        resp = await client.put(f"{BASE}/ky-thi/{kt['id']}", json={"ten_ky_thi": "Moi"})
        assert resp.status_code == 200, resp.json()
        assert resp.json()["data"]["ten_ky_thi"] == "Moi"


class TestKyThiTrangThai:
    async def test_nhap_to_cho_duyet(self, client, admin_user):
        data = await _setup_full_exam(client, admin_user)
        kt = data["ky_thi"]
        resp = await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "CHO_DUYET"})
        assert resp.status_code == 200
        assert resp.json()["data"]["trang_thai"] == "CHO_DUYET"

    async def test_cho_duyet_to_nhap(self, client, admin_user):
        """Tra lai tu CHO_DUYET ve NHAP."""
        data = await _setup_full_exam(client, admin_user)
        kt = data["ky_thi"]
        await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "CHO_DUYET"})
        resp = await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "NHAP"})
        assert resp.status_code == 200
        assert resp.json()["data"]["trang_thai"] == "NHAP"

    async def test_cho_duyet_to_dang_mo(self, client, admin_user):
        data = await _setup_full_exam(client, admin_user)
        kt = data["ky_thi"]
        await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "CHO_DUYET"})
        resp = await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "DANG_MO"})
        assert resp.status_code == 200
        assert resp.json()["data"]["trang_thai"] == "DANG_MO"

    async def test_nhap_to_dang_mo_invalid(self, client, admin_user):
        """Khong the nhay tu NHAP sang DANG_MO."""
        data = await _setup_full_exam(client, admin_user)
        kt = data["ky_thi"]
        resp = await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "DANG_MO"})
        assert resp.status_code == 400

    async def test_nhap_to_cho_duyet_no_cau_truc(self, client, admin_user):
        """Khong the gui duyet khi chua co cau truc de."""
        kt = await _create_ky_thi(client)
        resp = await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "CHO_DUYET"})
        assert resp.status_code == 400


# =========================================================================
# CAU TRUC DE
# =========================================================================

class TestCauTrucDe:
    async def test_upsert_cau_truc_de(self, client, admin_user):
        lv = await _create_linh_vuc(client, f"LV-CTD-{uuid.uuid4().hex[:4]}")
        vt = await _create_vi_tri(client, f"VT-CTD-{uuid.uuid4().hex[:4]}")
        kt = await _create_ky_thi(client)

        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de", json={
            "vi_tri_id": vt["id"],
            "cau_truc": [
                {"linh_vuc_id": lv["id"], "so_cau_de": 5, "so_cau_trung_binh": 3, "so_cau_kho": 2},
            ],
        })
        assert resp.status_code == 201
        data = resp.json()["data"]
        assert len(data) >= 1
        assert data[0]["tong_cau"] == 10

    async def test_lay_cau_truc_de(self, client, admin_user):
        data = await _setup_full_exam(client, admin_user)
        kt = data["ky_thi"]
        resp = await client.get(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de")
        assert resp.status_code == 200
        assert len(resp.json()["data"]) >= 1

    async def test_xoa_cau_truc_de_vi_tri(self, client, admin_user):
        data = await _setup_full_exam(client, admin_user)
        kt = data["ky_thi"]
        vt = data["vi_tri"]
        resp = await client.delete(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de/{vt['id']}")
        assert resp.status_code == 200

    async def test_upsert_overwrite(self, client, admin_user):
        """Upsert ghi de cau truc cu."""
        lv = await _create_linh_vuc(client, f"LV-OVW-{uuid.uuid4().hex[:4]}")
        vt = await _create_vi_tri(client, f"VT-OVW-{uuid.uuid4().hex[:4]}")
        kt = await _create_ky_thi(client)

        await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de", json={
            "vi_tri_id": vt["id"],
            "cau_truc": [{"linh_vuc_id": lv["id"], "so_cau_de": 5, "so_cau_trung_binh": 0, "so_cau_kho": 0}],
        })

        # Overwrite
        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de", json={
            "vi_tri_id": vt["id"],
            "cau_truc": [{"linh_vuc_id": lv["id"], "so_cau_de": 10, "so_cau_trung_binh": 5, "so_cau_kho": 0}],
        })
        assert resp.status_code == 201
        # Tim vi tri trong response
        vt_data = [d for d in resp.json()["data"] if d["vi_tri_id"] == vt["id"]]
        assert len(vt_data) == 1
        assert vt_data[0]["tong_cau"] == 15


# =========================================================================
# VALIDATE NGAN HANG
# =========================================================================

class TestValidate:
    async def test_validate_du_cau_hoi(self, client, admin_user):
        data = await _setup_full_exam(client, admin_user)
        kt = data["ky_thi"]
        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/validate")
        assert resp.status_code == 200
        assert resp.json()["data"]["tat_ca_du"] is True

    async def test_validate_thieu_cau_hoi(self, client, admin_user):
        """Cau truc de yeu cau nhieu hon cau hoi co san."""
        lv = await _create_linh_vuc(client, f"LV-VALI-{uuid.uuid4().hex[:4]}")
        vt = await _create_vi_tri(client, f"VT-VALI-{uuid.uuid4().hex[:4]}")
        kt = await _create_ky_thi(client)

        # Cau truc de: 100 cau DE — khong co san
        await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de", json={
            "vi_tri_id": vt["id"],
            "cau_truc": [{"linh_vuc_id": lv["id"], "so_cau_de": 100, "so_cau_trung_binh": 0, "so_cau_kho": 0}],
        })

        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/validate")
        assert resp.status_code == 200
        assert resp.json()["data"]["tat_ca_du"] is False


# =========================================================================
# THI SINH + LAM THI + CHAM DIEM
# =========================================================================

class TestThiSinh:
    async def test_giao_thi_sinh(self, client, admin_user):
        data = await _setup_full_exam(client, admin_user)
        kt = data["ky_thi"]
        vt = data["vi_tri"]

        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/thi-sinh", json={
            "danh_sach": [
                {"cong_chuc_id": "00327c43-c9a3-44d7-8306-7084e75cb2b5", "vi_tri_id": vt["id"]},
                {"cong_chuc_id": "01014af6-1505-495b-95ab-8c1503cfa061", "vi_tri_id": vt["id"]},
            ],
        })
        assert resp.status_code == 201
        assert resp.json()["data"]["thanh_cong"] == 2

    async def test_danh_sach_thi_sinh(self, client, admin_user):
        data = await _setup_full_exam(client, admin_user)
        kt = data["ky_thi"]
        vt = data["vi_tri"]

        await client.post(f"{BASE}/ky-thi/{kt['id']}/thi-sinh", json={
            "danh_sach": [{"cong_chuc_id": "00327c43-c9a3-44d7-8306-7084e75cb2b5", "vi_tri_id": vt["id"]}],
        })

        resp = await client.get(f"{BASE}/ky-thi/{kt['id']}/thi-sinh")
        assert resp.status_code == 200
        assert len(resp.json()["data"]) == 1

    async def test_xoa_thi_sinh_chua_thi(self, client, admin_user):
        data = await _setup_full_exam(client, admin_user)
        kt = data["ky_thi"]
        vt = data["vi_tri"]
        cc_id = "01014af6-1505-495b-95ab-8c1503cfa061"

        await client.post(f"{BASE}/ky-thi/{kt['id']}/thi-sinh", json={
            "danh_sach": [{"cong_chuc_id": cc_id, "vi_tri_id": vt["id"]}],
        })

        resp = await client.delete(f"{BASE}/ky-thi/{kt['id']}/thi-sinh/{cc_id}")
        assert resp.status_code == 200

    async def test_giao_trung_bo_qua(self, client, admin_user):
        """Giao thi sinh trung lap — bo qua khong loi."""
        data = await _setup_full_exam(client, admin_user)
        kt = data["ky_thi"]
        vt = data["vi_tri"]
        cc_id = "00327c43-c9a3-44d7-8306-7084e75cb2b5"

        await client.post(f"{BASE}/ky-thi/{kt['id']}/thi-sinh", json={
            "danh_sach": [{"cong_chuc_id": cc_id, "vi_tri_id": vt["id"]}],
        })
        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/thi-sinh", json={
            "danh_sach": [{"cong_chuc_id": cc_id, "vi_tri_id": vt["id"]}],
        })
        assert resp.status_code == 201
        assert resp.json()["data"]["bo_qua"] == 1


class TestLamThi:
    async def _setup_and_open(self, client, admin_user) -> dict:
        """Setup + mo ky thi + giao thi sinh cho admin_user."""
        data = await _setup_full_exam(client, admin_user)
        kt = data["ky_thi"]
        vt = data["vi_tri"]
        cc_id = "00327c43-c9a3-44d7-8306-7084e75cb2b5"  # admin_user idx=0

        # Giao thi sinh
        await client.post(f"{BASE}/ky-thi/{kt['id']}/thi-sinh", json={
            "danh_sach": [{"cong_chuc_id": cc_id, "vi_tri_id": vt["id"]}],
        })

        # Mo ky thi: NHAP -> CHO_DUYET -> DANG_MO
        await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "CHO_DUYET"})
        await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "DANG_MO"})

        data["cc_id"] = cc_id
        return data

    async def test_bat_dau_thi(self, client, admin_user):
        data = await self._setup_and_open(client, admin_user)
        kt = data["ky_thi"]

        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/bat-dau")
        assert resp.status_code == 200
        result = resp.json()["data"]
        assert result["tong_so_cau"] == 4  # 2 DE + 1 TB + 1 KHO
        assert len(result["cau_hoi"]) == 4
        # Khong co dap an dung
        for ch in result["cau_hoi"]:
            assert "dap_an_dung" not in str(ch)

    async def test_nop_bai_va_cham_diem(self, client, admin_user):
        data = await self._setup_and_open(client, admin_user)
        kt = data["ky_thi"]

        # Bat dau
        start_resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/bat-dau")
        cau_hoi = start_resp.json()["data"]["cau_hoi"]

        # Tra loi tat ca cau A
        cau_tra_loi = [{"cau_hoi_id": ch["id"], "tra_loi": {"dap_an": "A"}} for ch in cau_hoi]

        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/nop-bai", json={
            "cau_tra_loi": cau_tra_loi,
        })
        assert resp.status_code == 200
        result = resp.json()["data"]
        assert result["ket_qua"]["tong_so_cau"] == 4
        assert result["ket_qua"]["xep_loai"] in ("DAT", "KHONG_DAT")
        assert result["ket_qua"]["diem_tong"] >= 0

    async def test_ket_qua_ca_nhan(self, client, admin_user):
        data = await self._setup_and_open(client, admin_user)
        kt = data["ky_thi"]

        # Bat dau + nop bai
        start_resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/bat-dau")
        cau_hoi = start_resp.json()["data"]["cau_hoi"]
        cau_tra_loi = [{"cau_hoi_id": ch["id"], "tra_loi": {"dap_an": "A"}} for ch in cau_hoi]
        await client.post(f"{BASE}/ky-thi/{kt['id']}/nop-bai", json={"cau_tra_loi": cau_tra_loi})

        # Xem ket qua
        resp = await client.get(f"{BASE}/ky-thi/{kt['id']}/ket-qua")
        assert resp.status_code == 200
        assert resp.json()["data"]["thi_sinh"]["ho_ten"] is not None


    async def test_thong_ke_ky_thi(self, client, admin_user):
        data = await self._setup_and_open(client, admin_user)
        kt = data["ky_thi"]

        # Bat dau + nop
        start_resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/bat-dau")
        cau_hoi = start_resp.json()["data"]["cau_hoi"]
        cau_tra_loi = [{"cau_hoi_id": ch["id"], "tra_loi": {"dap_an": "A"}} for ch in cau_hoi]
        await client.post(f"{BASE}/ky-thi/{kt['id']}/nop-bai", json={"cau_tra_loi": cau_tra_loi})

        resp = await client.get(f"{BASE}/ky-thi/{kt['id']}/thong-ke")
        assert resp.status_code == 200
        tq = resp.json()["data"]["tong_quan"]
        assert tq["tong_thi_sinh"] == 1
        assert tq["da_thi"] == 1


# =========================================================================
# LICH SU LAN THI — snapshot + drill-down (30/05/2026)
# =========================================================================

class TestLichSuLanThi:
    """Test snapshot lich_su_thi day du + drill-down 1 lan thi cu the.

    Setup rieng (khong dung _setup_full_exam): seed truc tiep cau_hoi_dgnl bank
    qua POST /dgnl/ngan-hang vi bat_dau_thi() query bang `cau_hoi_dgnl`, khong
    phai `cau_hoi` (cua BKT).
    """

    async def _setup_dgnl_exam(self, client) -> dict:
        """Tao linh_vuc + vi_tri + ky_thi + 10 cau_hoi_dgnl + cau_truc_de."""
        uid = uuid.uuid4().hex[:6]
        lv = await _create_linh_vuc(client, f"LV-LSU-{uid}", "Lĩnh vực Lich Su Thi")
        vt = await _create_vi_tri(client, f"VT-LSU-{uid}", "Vị trí Lich Su Thi")
        kt = await _create_ky_thi(client, f"KT-LSU-{uid}")

        # Seed cau_hoi_dgnl bank: 5 DE + 3 TB + 2 KHO
        def _ch_body(noi_dung, do_kho):
            return {
                "linh_vuc_id": lv["id"],
                "noi_dung": noi_dung,
                "loai": "TRAC_NGHIEM_1",
                "do_kho": do_kho,
                "dap_an": {
                    "lua_chon": [
                        {"key": "A", "noi_dung": "Đáp A"},
                        {"key": "B", "noi_dung": "Đáp B"},
                    ],
                    "dap_an_dung": "A",
                },
            }

        for i in range(5):
            r = await client.post(f"{BASE}/dgnl/ngan-hang", json=_ch_body(f"DGNL Dễ {uid}-{i}", "DE"))
            assert r.status_code == 201, r.json()
        for i in range(3):
            r = await client.post(f"{BASE}/dgnl/ngan-hang", json=_ch_body(f"DGNL TB {uid}-{i}", "TRUNG_BINH"))
            assert r.status_code == 201, r.json()
        for i in range(2):
            r = await client.post(f"{BASE}/dgnl/ngan-hang", json=_ch_body(f"DGNL Khó {uid}-{i}", "KHO"))
            assert r.status_code == 201, r.json()

        # Cau truc de: 2 DE + 1 TB + 1 KHO = 4 cau
        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de", json={
            "vi_tri_id": vt["id"],
            "cau_truc": [
                {"linh_vuc_id": lv["id"], "so_cau_de": 2, "so_cau_trung_binh": 1, "so_cau_kho": 1},
            ],
        })
        assert resp.status_code == 201, resp.json()

        return {"linh_vuc": lv, "vi_tri": vt, "ky_thi": kt}

    async def _open_with_dap_an(self, client, admin_user) -> tuple[dict, str]:
        """Setup + bat hien_dap_an=True + giao thi sinh + mo ky thi."""
        data = await self._setup_dgnl_exam(client)
        kt = data["ky_thi"]
        vt = data["vi_tri"]
        # Bat hien_dap_an de chi_tiet co the duoc tra ve trong drill-down test
        r = await client.put(f"{BASE}/ky-thi/{kt['id']}", json={"hien_dap_an": True})
        assert r.status_code == 200, r.json()

        cc_id = "00327c43-c9a3-44d7-8306-7084e75cb2b5"  # admin_user idx=0
        await client.post(f"{BASE}/ky-thi/{kt['id']}/thi-sinh", json={
            "danh_sach": [{"cong_chuc_id": cc_id, "vi_tri_id": vt["id"]}],
        })
        await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "CHO_DUYET"})
        await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "DANG_MO"})
        return data, cc_id

    async def _thi_va_nop(self, client, kt_id: str):
        """Bat dau + nop bai 1 lan (chon A cho tat ca cau)."""
        start = await client.post(f"{BASE}/ky-thi/{kt_id}/bat-dau")
        assert start.status_code == 200, f"bat-dau failed: {start.status_code} {start.json()}"
        cau_hoi = start.json()["data"]["cau_hoi"]
        cau_tra_loi = [{"cau_hoi_id": c["id"], "tra_loi": {"dap_an": "A"}} for c in cau_hoi]
        nop = await client.post(f"{BASE}/ky-thi/{kt_id}/nop-bai", json={"cau_tra_loi": cau_tra_loi})
        assert nop.status_code == 200, f"nop-bai failed: {nop.status_code} {nop.json()}"

    async def test_nop_bai_snapshot_lich_su_voi_chi_tiet(self, client, admin_user):
        """Sau khi nop lan 1, lich_su_thi (summary) co 1 entry voi has_chi_tiet=True.

        Response /ket-qua tra ve summary projection -> KHONG co chi_tiet_tra_loi raw.
        Verify chi_tiet day du qua endpoint drill-down /ket-qua/{lan}.
        """
        data, cc_id = await self._open_with_dap_an(client, admin_user)
        kt = data["ky_thi"]
        await self._thi_va_nop(client, kt["id"])

        resp = await client.get(f"{BASE}/ky-thi/{kt['id']}/ket-qua")
        assert resp.status_code == 200
        lich_su = resp.json()["data"]["lich_su_thi"]
        assert lich_su is not None and len(lich_su) == 1
        entry = lich_su[0]
        assert entry["lan"] == 1
        assert entry["tong_so_cau"] == 4
        assert entry["thoi_gian_nop"] is not None
        # Summary KHONG kem chi_tiet raw, dung has_chi_tiet flag
        assert entry["has_chi_tiet"] is True
        assert "chi_tiet_tra_loi" not in entry

        # Drill-down lan 1 -> chi_tiet day du (4 cau)
        resp = await client.get(f"{BASE}/ky-thi/{kt['id']}/thi-sinh/{cc_id}/ket-qua/1")
        assert resp.status_code == 200
        assert len(resp.json()["data"]["chi_tiet"]) == 4

    async def test_thi_lai_upsert_idempotent(self, client, admin_user):
        """Retake + nop lan 2 -> lich_su_thi co 2 entry (lan 1, lan 2), khong trung."""
        data, cc_id = await self._open_with_dap_an(client, admin_user)
        kt = data["ky_thi"]
        await self._thi_va_nop(client, kt["id"])  # lan 1
        await self._thi_va_nop(client, kt["id"])  # lan 2 (retake)

        resp = await client.get(f"{BASE}/ky-thi/{kt['id']}/ket-qua")
        lich_su = resp.json()["data"]["lich_su_thi"]
        assert len(lich_su) == 2
        lans = sorted(e["lan"] for e in lich_su)
        assert lans == [1, 2]
        # Ca 2 entry summary deu co has_chi_tiet=True (upsert thay vi append duplicate)
        for e in lich_su:
            assert e["has_chi_tiet"] is True
        # Drill-down ca 2 lan -> chi_tiet day du
        for lan in [1, 2]:
            r = await client.get(f"{BASE}/ky-thi/{kt['id']}/thi-sinh/{cc_id}/ket-qua/{lan}")
            assert r.status_code == 200
            assert len(r.json()["data"]["chi_tiet"]) == 4

    async def test_danh_sach_thi_sinh_tra_ve_lich_su_summary(self, client, admin_user):
        """Endpoint /thi-sinh tra ve lich_su_thi summary (khong co chi_tiet_tra_loi)."""
        data, cc_id = await self._open_with_dap_an(client, admin_user)
        kt = data["ky_thi"]
        await self._thi_va_nop(client, kt["id"])

        resp = await client.get(f"{BASE}/ky-thi/{kt['id']}/thi-sinh")
        assert resp.status_code == 200
        items = resp.json()["data"]
        ts = next(i for i in items if i["cong_chuc_id"] == cc_id)
        assert ts["lich_su_thi"] is not None and len(ts["lich_su_thi"]) == 1
        entry = ts["lich_su_thi"][0]
        # Summary: co has_chi_tiet, KHONG co chi_tiet_tra_loi raw
        assert entry["has_chi_tiet"] is True
        assert "chi_tiet_tra_loi" not in entry
        assert entry["lan"] == 1
        assert entry["tong_so_cau"] == 4

    async def test_ket_qua_lan_thi_drill_down(self, client, admin_user):
        """GET /ky-thi/{id}/thi-sinh/{ccId}/ket-qua/{lan} tra dung data lan cu."""
        data, cc_id = await self._open_with_dap_an(client, admin_user)
        kt = data["ky_thi"]
        await self._thi_va_nop(client, kt["id"])  # lan 1
        await self._thi_va_nop(client, kt["id"])  # lan 2

        # Drill-down lan cu
        resp = await client.get(f"{BASE}/ky-thi/{kt['id']}/thi-sinh/{cc_id}/ket-qua/1")
        assert resp.status_code == 200, resp.json()
        result = resp.json()["data"]
        assert result["ket_qua"]["lan_thi"] == 1
        assert result["ket_qua"]["tong_so_cau"] == 4
        # chi_tiet co tu lich_su_thi entry (vi hien_dap_an=True)
        assert result["chi_tiet"] is not None
        assert len(result["chi_tiet"]) == 4

        # Drill-down lan hien tai
        resp = await client.get(f"{BASE}/ky-thi/{kt['id']}/thi-sinh/{cc_id}/ket-qua/2")
        assert resp.status_code == 200
        assert resp.json()["data"]["ket_qua"]["lan_thi"] == 2

        # Lan khong ton tai
        resp = await client.get(f"{BASE}/ky-thi/{kt['id']}/thi-sinh/{cc_id}/ket-qua/99")
        assert resp.status_code == 404


# =========================================================================
# XAC NHAN CA THI + DIEM CAO NHAT (B3 - 30/07/2026)
# =========================================================================

class TestXacNhanCaThi:
    """Xac nhan ca thi (chot ket qua), auto-confirm 10 phut, diem cao nhat."""

    # Tai dung helpers cua TestLichSuLanThi (khong ke thua de khoi re-run tests)
    _setup_dgnl_exam = TestLichSuLanThi._setup_dgnl_exam
    _open_with_dap_an = TestLichSuLanThi._open_with_dap_an
    _thi_va_nop = TestLichSuLanThi._thi_va_nop

    async def _thi_va_nop_dap_an(self, client, kt_id: str, dap_an: str):
        """Bat dau + nop bai, chon cung 1 dap an cho tat ca cau."""
        start = await client.post(f"{BASE}/ky-thi/{kt_id}/bat-dau")
        assert start.status_code == 200, start.json()
        cau_hoi = start.json()["data"]["cau_hoi"]
        cau_tra_loi = [{"cau_hoi_id": c["id"], "tra_loi": {"dap_an": dap_an}} for c in cau_hoi]
        nop = await client.post(f"{BASE}/ky-thi/{kt_id}/nop-bai", json={"cau_tra_loi": cau_tra_loi})
        assert nop.status_code == 200, nop.json()
        return nop.json()["data"]

    async def test_ket_qua_sau_nop_co_han_xac_nhan(self, client, admin_user):
        """Sau khi nop: da_xac_nhan=False, co han_xac_nhan (nop + 10'), so_lan_thi_toi_da."""
        data, _cc_id = await self._open_with_dap_an(client, admin_user)
        kq = await self._thi_va_nop_dap_an(client, data["ky_thi"]["id"], "A")
        assert kq["ket_qua"]["da_xac_nhan"] is False
        assert kq["ket_qua"]["han_xac_nhan"] is not None
        assert kq["ket_qua"]["so_lan_thi_toi_da"] == 2
        assert kq["ket_qua"]["thoi_gian_nop"] is not None

    async def test_xac_nhan_chan_thi_lai(self, client, admin_user):
        """Xac nhan xong -> bat-dau lai bi 400 DGNL_048 du con luot."""
        data, _cc_id = await self._open_with_dap_an(client, admin_user)
        kt_id = data["ky_thi"]["id"]
        await self._thi_va_nop(client, kt_id)  # lan 1 (con luot lan 2)

        resp = await client.post(f"{BASE}/ky-thi/{kt_id}/xac-nhan")
        assert resp.status_code == 200, resp.json()
        assert resp.json()["data"]["da_xac_nhan"] is True

        # Idempotent
        resp2 = await client.post(f"{BASE}/ky-thi/{kt_id}/xac-nhan")
        assert resp2.status_code == 200

        # Thi lai bi chan
        start = await client.post(f"{BASE}/ky-thi/{kt_id}/bat-dau")
        assert start.status_code == 400
        assert start.json()["detail"]["error"]["code"] == "DGNL_048"

        # Ket qua hien da_xac_nhan, khong con han
        kq = await client.get(f"{BASE}/ky-thi/{kt_id}/ket-qua")
        assert kq.json()["data"]["ket_qua"]["da_xac_nhan"] is True
        assert kq.json()["data"]["ket_qua"]["han_xac_nhan"] is None

    async def test_xac_nhan_chua_nop_400(self, client, admin_user):
        """Chua nop bai -> xac nhan bi 400 DGNL_047."""
        data, _cc_id = await self._open_with_dap_an(client, admin_user)
        resp = await client.post(f"{BASE}/ky-thi/{data['ky_thi']['id']}/xac-nhan")
        assert resp.status_code == 400
        assert resp.json()["detail"]["error"]["code"] == "DGNL_047"

    async def test_auto_xac_nhan_qua_10_phut(self, client, admin_user, db_session):
        """Qua 10 phut khong xac nhan -> thi lai bi 400 DGNL_049 + tu dong chot."""
        from sqlalchemy import text as sa_text

        data, cc_id = await self._open_with_dap_an(client, admin_user)
        kt_id = data["ky_thi"]["id"]
        await self._thi_va_nop(client, kt_id)

        # Gia lap da nop 11 phut truoc
        await db_session.execute(sa_text(
            "UPDATE lms.thi_sinh SET thoi_gian_nop = thoi_gian_nop - INTERVAL '11 minutes' "
            "WHERE ky_thi_id = :kt AND cong_chuc_id = :cc"
        ), {"kt": kt_id, "cc": cc_id})
        await db_session.commit()
        # Service dung chung session voi test -> expire de doc lai tu DB
        db_session.expire_all()

        start = await client.post(f"{BASE}/ky-thi/{kt_id}/bat-dau")
        assert start.status_code == 400
        assert start.json()["detail"]["error"]["code"] == "DGNL_049"

        # Lan sau bi chan voi ly do "da xac nhan" (flag da persist)
        start2 = await client.post(f"{BASE}/ky-thi/{kt_id}/bat-dau")
        assert start2.status_code == 400
        assert start2.json()["detail"]["error"]["code"] == "DGNL_048"

    async def test_diem_cao_nhat_giua_cac_lan(self, client, admin_user):
        """Lan 1 diem cao, lan 2 diem thap -> diem chinh thuc giu lan 1 (cao nhat)."""
        data, cc_id = await self._open_with_dap_an(client, admin_user)
        kt_id = data["ky_thi"]["id"]

        kq1 = await self._thi_va_nop_dap_an(client, kt_id, "A")  # dung het -> 100
        assert kq1["ket_qua"]["diem_tong"] == 100.0

        kq2 = await self._thi_va_nop_dap_an(client, kt_id, "B")  # sai het -> 0
        # Diem chinh thuc = cao nhat (lan 1), khong bi lan 2 ghi de
        assert kq2["ket_qua"]["diem_tong"] == 100.0
        assert kq2["ket_qua"]["xep_loai"] == "DAT"
        assert kq2["ket_qua"]["lan_tot_nhat"] == 1
        assert kq2["ket_qua"]["lan_thi"] == 2

        # Lich su van du 2 lan voi diem tung lan
        kq = await client.get(f"{BASE}/ky-thi/{kt_id}/ket-qua")
        lich_su = {e["lan"]: e["diem"] for e in kq.json()["data"]["lich_su_thi"]}
        assert lich_su[1] == 100.0
        assert lich_su[2] == 0

        # Drill-down lan 2 van xem duoc bai lam diem 0
        r = await client.get(f"{BASE}/ky-thi/{kt_id}/thi-sinh/{cc_id}/ket-qua/2")
        assert r.status_code == 200
        assert r.json()["data"]["ket_qua"]["diem_tong"] == 0


# =========================================================================
# MAU CAU TRUC DE (B4 - 30/07/2026)
# =========================================================================

class TestCauTrucDeTemplate:
    """Luu cau truc de lam mau + ap dung mau vao ky thi khac."""

    async def _setup(self, client) -> dict:
        uid = uuid.uuid4().hex[:6]
        lv = await _create_linh_vuc(client, f"LV-TPL-{uid}")
        vt = await _create_vi_tri(client, f"VT-TPL-{uid}")
        return {"lv": lv, "vt": vt, "uid": uid}

    async def test_tao_va_danh_sach_template(self, client, admin_user):
        data = await self._setup(client)
        resp = await client.post(f"{BASE}/cau-truc-de-template", json={
            "ten_template": f"Mẫu chuẩn {data['uid']}",
            "mo_ta": "Mẫu test",
            "cau_truc": [{
                "vi_tri_id": data["vt"]["id"], "linh_vuc_id": data["lv"]["id"],
                "so_cau_de": 2, "so_cau_trung_binh": 1, "so_cau_kho": 1,
            }],
        })
        assert resp.status_code == 201, resp.json()
        tpl = resp.json()["data"]
        assert tpl["ten_template"] == f"Mẫu chuẩn {data['uid']}"
        assert len(tpl["cau_truc"]) == 1

        ds = await client.get(f"{BASE}/cau-truc-de-template")
        assert ds.status_code == 200
        assert any(t["id"] == tpl["id"] for t in ds.json()["data"])

    async def test_ap_dung_template_vao_ky_thi(self, client, admin_user):
        """Lay cau_truc tu template roi upsert vao ky thi qua endpoint san co."""
        data = await self._setup(client)
        kt = await _create_ky_thi(client, f"KT-TPL-{data['uid']}")

        tpl_resp = await client.post(f"{BASE}/cau-truc-de-template", json={
            "ten_template": f"Mẫu áp dụng {data['uid']}",
            "cau_truc": [{
                "vi_tri_id": data["vt"]["id"], "linh_vuc_id": data["lv"]["id"],
                "so_cau_de": 3, "so_cau_trung_binh": 2, "so_cau_kho": 1,
            }],
        })
        tpl = tpl_resp.json()["data"]

        # FE flow: lay chi tiet template -> upsert cau-truc-de theo vi_tri
        chi_tiet = await client.get(f"{BASE}/cau-truc-de-template/{tpl['id']}")
        cau_truc = chi_tiet.json()["data"]["cau_truc"]
        upsert = await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de", json={
            "vi_tri_id": cau_truc[0]["vi_tri_id"],
            "cau_truc": [{
                "linh_vuc_id": c["linh_vuc_id"],
                "so_cau_de": c["so_cau_de"],
                "so_cau_trung_binh": c["so_cau_trung_binh"],
                "so_cau_kho": c["so_cau_kho"],
            } for c in cau_truc],
        })
        assert upsert.status_code == 201, upsert.json()

        ctd = await client.get(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de")
        assert ctd.status_code == 200

    async def test_xoa_mem_template(self, client, admin_user):
        data = await self._setup(client)
        tpl_resp = await client.post(f"{BASE}/cau-truc-de-template", json={
            "ten_template": f"Mẫu xóa {data['uid']}",
            "cau_truc": [{
                "vi_tri_id": data["vt"]["id"], "linh_vuc_id": data["lv"]["id"],
                "so_cau_de": 1, "so_cau_trung_binh": 0, "so_cau_kho": 0,
            }],
        })
        tpl_id = tpl_resp.json()["data"]["id"]

        del_resp = await client.delete(f"{BASE}/cau-truc-de-template/{tpl_id}")
        assert del_resp.status_code == 200

        # Da xoa mem -> chi tiet 404, khong con trong danh sach
        chi_tiet = await client.get(f"{BASE}/cau-truc-de-template/{tpl_id}")
        assert chi_tiet.status_code == 404
        ds = await client.get(f"{BASE}/cau-truc-de-template")
        assert not any(t["id"] == tpl_id for t in ds.json()["data"])

    async def test_cbcc_khong_duoc_xem_template(self, client, cbcc_user):
        resp = await client.get(f"{BASE}/cau-truc-de-template")
        assert resp.status_code == 403


# =========================================================================
# SUA MAU CAU TRUC DE TRUC TIEP (tab "Mau cau truc de")
# =========================================================================

class TestSuaMauCauTrucDe:
    """PUT / nhan-ban / chan trung ten — thay cho vong 'luu mau roi ap dung'."""

    async def _setup(self, client) -> dict:
        uid = uuid.uuid4().hex[:6]
        lv = await _create_linh_vuc(client, f"LV-EDT-{uid}")
        lv2 = await _create_linh_vuc(client, f"LV-EDT2-{uid}")
        vt = await _create_vi_tri(client, f"VT-EDT-{uid}")
        return {"lv": lv, "lv2": lv2, "vt": vt, "uid": uid}

    async def _tao_mau(self, client, data: dict, ten: str) -> dict:
        resp = await client.post(f"{BASE}/cau-truc-de-template", json={
            "ten_template": ten,
            "cau_truc": [{
                "vi_tri_id": data["vt"]["id"], "linh_vuc_id": data["lv"]["id"],
                "so_cau_de": 2, "so_cau_trung_binh": 1, "so_cau_kho": 1,
            }],
        })
        assert resp.status_code == 201, resp.json()
        return resp.json()["data"]

    async def test_sua_mau_thay_the_toan_bo_cau_truc(self, client, admin_user):
        data = await self._setup(client)
        tpl = await self._tao_mau(client, data, f"Mẫu sửa {data['uid']}")

        resp = await client.put(f"{BASE}/cau-truc-de-template/{tpl['id']}", json={
            "ten_template": f"Mẫu sửa {data['uid']} (v2)",
            "mo_ta": "Đã chỉnh",
            "cau_truc": [
                {"vi_tri_id": data["vt"]["id"], "linh_vuc_id": data["lv"]["id"],
                 "so_cau_de": 5, "so_cau_trung_binh": 0, "so_cau_kho": 0},
                {"vi_tri_id": data["vt"]["id"], "linh_vuc_id": data["lv2"]["id"],
                 "so_cau_de": 1, "so_cau_trung_binh": 2, "so_cau_kho": 3},
            ],
        })
        assert resp.status_code == 200, resp.json()
        moi = resp.json()["data"]
        assert moi["ten_template"] == f"Mẫu sửa {data['uid']} (v2)"
        assert moi["mo_ta"] == "Đã chỉnh"
        assert len(moi["cau_truc"]) == 2
        assert moi["cau_truc"][0]["so_cau_de"] == 5

    async def test_sua_mau_chi_doi_ten_giu_nguyen_cau_truc(self, client, admin_user):
        data = await self._setup(client)
        tpl = await self._tao_mau(client, data, f"Mẫu giữ {data['uid']}")

        resp = await client.put(f"{BASE}/cau-truc-de-template/{tpl['id']}", json={
            "ten_template": f"Mẫu giữ {data['uid']} đổi tên",
        })
        assert resp.status_code == 200
        assert len(resp.json()["data"]["cau_truc"]) == 1

    async def test_chan_trung_ten_khi_tao(self, client, admin_user):
        data = await self._setup(client)
        ten = f"Mẫu trùng {data['uid']}"
        await self._tao_mau(client, data, ten)

        resp = await client.post(f"{BASE}/cau-truc-de-template", json={
            "ten_template": f"  {ten.upper()}  ",  # khac hoa/thuong + khoang trang
            "cau_truc": [{
                "vi_tri_id": data["vt"]["id"], "linh_vuc_id": data["lv"]["id"],
                "so_cau_de": 1, "so_cau_trung_binh": 0, "so_cau_kho": 0,
            }],
        })
        assert resp.status_code == 400
        assert resp.json()["detail"]["error"]["code"] == "DGNL_071"

    async def test_chan_trung_ten_khi_sua(self, client, admin_user):
        data = await self._setup(client)
        a = await self._tao_mau(client, data, f"Mẫu A {data['uid']}")
        b = await self._tao_mau(client, data, f"Mẫu B {data['uid']}")

        resp = await client.put(f"{BASE}/cau-truc-de-template/{b['id']}", json={
            "ten_template": a["ten_template"],
        })
        assert resp.status_code == 400
        assert resp.json()["detail"]["error"]["code"] == "DGNL_071"

    async def test_sua_mau_giu_nguyen_ten_cua_chinh_no(self, client, admin_user):
        """Gui lai dung ten cu khong bi coi la trung."""
        data = await self._setup(client)
        tpl = await self._tao_mau(client, data, f"Mẫu tự {data['uid']}")

        resp = await client.put(f"{BASE}/cau-truc-de-template/{tpl['id']}", json={
            "ten_template": tpl["ten_template"], "mo_ta": "chỉ sửa mô tả",
        })
        assert resp.status_code == 200

    async def test_ten_da_xoa_mem_dung_lai_duoc(self, client, admin_user):
        data = await self._setup(client)
        ten = f"Mẫu tái dùng {data['uid']}"
        tpl = await self._tao_mau(client, data, ten)
        await client.delete(f"{BASE}/cau-truc-de-template/{tpl['id']}")

        lai = await self._tao_mau(client, data, ten)
        assert lai["id"] != tpl["id"]

    async def test_nhan_ban_mau(self, client, admin_user):
        data = await self._setup(client)
        tpl = await self._tao_mau(client, data, f"Mẫu gốc {data['uid']}")

        resp = await client.post(f"{BASE}/cau-truc-de-template/{tpl['id']}/nhan-ban", json={
            "ten_template": f"Mẫu bản sao {data['uid']}",
        })
        assert resp.status_code == 201, resp.json()
        ban_sao = resp.json()["data"]
        assert ban_sao["id"] != tpl["id"]
        assert ban_sao["cau_truc"] == tpl["cau_truc"]

    async def test_nhan_ban_chan_trung_ten(self, client, admin_user):
        data = await self._setup(client)
        tpl = await self._tao_mau(client, data, f"Mẫu NB {data['uid']}")

        resp = await client.post(f"{BASE}/cau-truc-de-template/{tpl['id']}/nhan-ban", json={
            "ten_template": tpl["ten_template"],
        })
        assert resp.status_code == 400
        assert resp.json()["detail"]["error"]["code"] == "DGNL_071"

    async def test_sua_mau_khong_ton_tai(self, client, admin_user):
        resp = await client.put(f"{BASE}/cau-truc-de-template/{uuid.uuid4()}", json={
            "ten_template": "Không tồn tại",
        })
        assert resp.status_code == 404

    async def test_cbcc_khong_duoc_sua_mau(self, client, cbcc_user):
        resp = await client.put(f"{BASE}/cau-truc-de-template/{uuid.uuid4()}", json={
            "ten_template": "X",
        })
        assert resp.status_code == 403


# =========================================================================
# AP DUNG MAU VAO KY THI — NGUYEN TU (1 transaction)
# =========================================================================

class TestApDungMauCauTruc:
    async def test_ap_dung_mau_nhieu_vi_tri(self, client, admin_user):
        uid = uuid.uuid4().hex[:6]
        lv = await _create_linh_vuc(client, f"LV-APD-{uid}")
        vt1 = await _create_vi_tri(client, f"VT-APD1-{uid}")
        vt2 = await _create_vi_tri(client, f"VT-APD2-{uid}")
        kt = await _create_ky_thi(client, f"KT-APD-{uid}")

        tpl = (await client.post(f"{BASE}/cau-truc-de-template", json={
            "ten_template": f"Mẫu áp {uid}",
            "cau_truc": [
                {"vi_tri_id": vt1["id"], "linh_vuc_id": lv["id"],
                 "so_cau_de": 3, "so_cau_trung_binh": 2, "so_cau_kho": 1},
                {"vi_tri_id": vt2["id"], "linh_vuc_id": lv["id"],
                 "so_cau_de": 4, "so_cau_trung_binh": 0, "so_cau_kho": 0},
            ],
        })).json()["data"]

        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de/ap-dung-mau", json={
            "template_id": tpl["id"],
        })
        assert resp.status_code == 201, resp.json()
        data = resp.json()["data"]
        assert len(data) == 2
        assert {d["vi_tri_id"] for d in data} == {vt1["id"], vt2["id"]}
        assert {d["tong_cau"] for d in data} == {6, 4}

    async def test_ap_dung_mau_giu_vi_tri_ngoai_mau(self, client, admin_user):
        """Mac dinh chi ghi de vi tri co trong mau — vi tri khac giu nguyen."""
        uid = uuid.uuid4().hex[:6]
        lv = await _create_linh_vuc(client, f"LV-KEEP-{uid}")
        vt_mau = await _create_vi_tri(client, f"VT-KEEP1-{uid}")
        vt_ngoai = await _create_vi_tri(client, f"VT-KEEP2-{uid}")
        kt = await _create_ky_thi(client, f"KT-KEEP-{uid}")

        await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de", json={
            "vi_tri_id": vt_ngoai["id"],
            "cau_truc": [{"linh_vuc_id": lv["id"], "so_cau_de": 7, "so_cau_trung_binh": 0, "so_cau_kho": 0}],
        })
        tpl = (await client.post(f"{BASE}/cau-truc-de-template", json={
            "ten_template": f"Mẫu keep {uid}",
            "cau_truc": [{"vi_tri_id": vt_mau["id"], "linh_vuc_id": lv["id"],
                          "so_cau_de": 1, "so_cau_trung_binh": 1, "so_cau_kho": 1}],
        })).json()["data"]

        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de/ap-dung-mau", json={
            "template_id": tpl["id"],
        })
        assert resp.status_code == 201
        theo_vt = {d["vi_tri_id"]: d["tong_cau"] for d in resp.json()["data"]}
        assert theo_vt[vt_ngoai["id"]] == 7
        assert theo_vt[vt_mau["id"]] == 3

    async def test_ap_dung_mau_ghi_de_toan_bo(self, client, admin_user):
        uid = uuid.uuid4().hex[:6]
        lv = await _create_linh_vuc(client, f"LV-WIPE-{uid}")
        vt_mau = await _create_vi_tri(client, f"VT-WIPE1-{uid}")
        vt_ngoai = await _create_vi_tri(client, f"VT-WIPE2-{uid}")
        kt = await _create_ky_thi(client, f"KT-WIPE-{uid}")

        await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de", json={
            "vi_tri_id": vt_ngoai["id"],
            "cau_truc": [{"linh_vuc_id": lv["id"], "so_cau_de": 7, "so_cau_trung_binh": 0, "so_cau_kho": 0}],
        })
        tpl = (await client.post(f"{BASE}/cau-truc-de-template", json={
            "ten_template": f"Mẫu wipe {uid}",
            "cau_truc": [{"vi_tri_id": vt_mau["id"], "linh_vuc_id": lv["id"],
                          "so_cau_de": 2, "so_cau_trung_binh": 0, "so_cau_kho": 0}],
        })).json()["data"]

        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de/ap-dung-mau", json={
            "template_id": tpl["id"], "ghi_de_toan_bo": True,
        })
        assert resp.status_code == 201
        data = resp.json()["data"]
        assert [d["vi_tri_id"] for d in data] == [vt_mau["id"]]

    async def test_ap_dung_mau_co_linh_vuc_da_xoa_khong_ghi_gi(self, client, admin_user):
        """Mau tro toi linh vuc da xoa -> 400 va KHONG dung vao cau truc dang co."""
        uid = uuid.uuid4().hex[:6]
        lv_ok = await _create_linh_vuc(client, f"LV-OK-{uid}")
        lv_xoa = await _create_linh_vuc(client, f"LV-DEL-{uid}")
        vt1 = await _create_vi_tri(client, f"VT-ATM1-{uid}")
        vt2 = await _create_vi_tri(client, f"VT-ATM2-{uid}")
        kt = await _create_ky_thi(client, f"KT-ATM-{uid}")

        # Cau truc dang co cua vt1 — phai con nguyen sau khi ap dung that bai
        await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de", json={
            "vi_tri_id": vt1["id"],
            "cau_truc": [{"linh_vuc_id": lv_ok["id"], "so_cau_de": 9, "so_cau_trung_binh": 0, "so_cau_kho": 0}],
        })

        tpl = (await client.post(f"{BASE}/cau-truc-de-template", json={
            "ten_template": f"Mẫu hỏng {uid}",
            "cau_truc": [
                {"vi_tri_id": vt1["id"], "linh_vuc_id": lv_ok["id"],
                 "so_cau_de": 1, "so_cau_trung_binh": 0, "so_cau_kho": 0},
                {"vi_tri_id": vt2["id"], "linh_vuc_id": lv_xoa["id"],
                 "so_cau_de": 1, "so_cau_trung_binh": 0, "so_cau_kho": 0},
            ],
        })).json()["data"]

        await client.delete(f"{BASE}/linh-vuc/{lv_xoa['id']}")

        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de/ap-dung-mau", json={
            "template_id": tpl["id"],
        })
        assert resp.status_code == 400
        assert resp.json()["detail"]["error"]["code"] == "DGNL_073"

        # KHONG co ghi nao xay ra — vt1 van 9 cau, vt2 chua ton tai
        ctd = (await client.get(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de")).json()["data"]
        theo_vt = {d["vi_tri_id"]: d["tong_cau"] for d in ctd}
        assert theo_vt == {vt1["id"]: 9}

    async def test_ap_dung_mau_trung_linh_vuc_bi_chan(self, client, admin_user):
        """Mau khai trung linh vuc trong 1 vi tri -> 400, khong de vo unique constraint."""
        uid = uuid.uuid4().hex[:6]
        lv = await _create_linh_vuc(client, f"LV-DUP-{uid}")
        vt = await _create_vi_tri(client, f"VT-DUP-{uid}")
        kt = await _create_ky_thi(client, f"KT-DUP-{uid}")

        tpl = (await client.post(f"{BASE}/cau-truc-de-template", json={
            "ten_template": f"Mẫu trùng lv {uid}",
            "cau_truc": [
                {"vi_tri_id": vt["id"], "linh_vuc_id": lv["id"],
                 "so_cau_de": 1, "so_cau_trung_binh": 0, "so_cau_kho": 0},
                {"vi_tri_id": vt["id"], "linh_vuc_id": lv["id"],
                 "so_cau_de": 2, "so_cau_trung_binh": 0, "so_cau_kho": 0},
            ],
        })).json()["data"]

        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de/ap-dung-mau", json={
            "template_id": tpl["id"],
        })
        assert resp.status_code == 400
        assert resp.json()["detail"]["error"]["code"] == "DGNL_073"

    async def test_upsert_trung_linh_vuc_bi_chan(self, client, admin_user):
        uid = uuid.uuid4().hex[:6]
        lv = await _create_linh_vuc(client, f"LV-DUP2-{uid}")
        vt = await _create_vi_tri(client, f"VT-DUP2-{uid}")
        kt = await _create_ky_thi(client, f"KT-DUP2-{uid}")

        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de", json={
            "vi_tri_id": vt["id"],
            "cau_truc": [
                {"linh_vuc_id": lv["id"], "so_cau_de": 1, "so_cau_trung_binh": 0, "so_cau_kho": 0},
                {"linh_vuc_id": lv["id"], "so_cau_de": 2, "so_cau_trung_binh": 0, "so_cau_kho": 0},
            ],
        })
        assert resp.status_code == 400
        assert resp.json()["detail"]["error"]["code"] == "DGNL_021"

    async def test_ap_dung_mau_khong_ton_tai(self, client, admin_user):
        kt = await _create_ky_thi(client)
        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de/ap-dung-mau", json={
            "template_id": str(uuid.uuid4()),
        })
        assert resp.status_code == 404

    async def test_cbcc_khong_duoc_ap_dung_mau(self, client, cbcc_user):
        resp = await client.post(f"{BASE}/ky-thi/{uuid.uuid4()}/cau-truc-de/ap-dung-mau", json={
            "template_id": str(uuid.uuid4()),
        })
        assert resp.status_code == 403


# =========================================================================
# NOI KHOA TRANG THAI KHI SUA CAU TRUC DE
# =========================================================================

class TestKhoaSuaCauTrucDe:
    """Cho sua o NHAP/CHO_DUYET; DANG_MO chi chan vi tri da co nguoi thi."""

    async def test_sua_duoc_khi_cho_duyet(self, client, admin_user):
        data = await _setup_full_exam(client, admin_user)
        kt, vt, lv = data["ky_thi"], data["vi_tri"], data["linh_vuc"]
        await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "CHO_DUYET"})

        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de", json={
            "vi_tri_id": vt["id"],
            "cau_truc": [{"linh_vuc_id": lv["id"], "so_cau_de": 3, "so_cau_trung_binh": 1, "so_cau_kho": 1}],
        })
        assert resp.status_code == 201, resp.json()

    async def test_sua_duoc_khi_dang_mo_va_chua_ai_thi(self, client, admin_user):
        data = await _setup_full_exam(client, admin_user)
        kt, vt, lv = data["ky_thi"], data["vi_tri"], data["linh_vuc"]
        await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "CHO_DUYET"})
        await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "DANG_MO"})

        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de", json={
            "vi_tri_id": vt["id"],
            "cau_truc": [{"linh_vuc_id": lv["id"], "so_cau_de": 1, "so_cau_trung_binh": 1, "so_cau_kho": 1}],
        })
        assert resp.status_code == 201, resp.json()

    async def test_chan_khi_dang_mo_va_da_co_nguoi_thi(self, client, admin_user):
        data = await _setup_full_exam(client, admin_user)
        kt, vt, lv = data["ky_thi"], data["vi_tri"], data["linh_vuc"]
        cc_id = "00327c43-c9a3-44d7-8306-7084e75cb2b5"  # admin_user idx=0

        await client.post(f"{BASE}/ky-thi/{kt['id']}/thi-sinh", json={
            "danh_sach": [{"cong_chuc_id": cc_id, "vi_tri_id": vt["id"]}],
        })
        await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "CHO_DUYET"})
        await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "DANG_MO"})

        bat_dau = await client.post(f"{BASE}/ky-thi/{kt['id']}/bat-dau")
        assert bat_dau.status_code == 200, bat_dau.json()

        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de", json={
            "vi_tri_id": vt["id"],
            "cau_truc": [{"linh_vuc_id": lv["id"], "so_cau_de": 1, "so_cau_trung_binh": 0, "so_cau_kho": 0}],
        })
        assert resp.status_code == 400
        assert resp.json()["detail"]["error"]["code"] == "DGNL_019"

        # Xoa cung bi chan
        xoa = await client.delete(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de/{vt['id']}")
        assert xoa.status_code == 400

    async def test_chan_khi_da_dong(self, client, admin_user):
        data = await _setup_full_exam(client, admin_user)
        kt, vt, lv = data["ky_thi"], data["vi_tri"], data["linh_vuc"]
        await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "CHO_DUYET"})
        await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "DANG_MO"})
        await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "DA_DONG"})

        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/cau-truc-de", json={
            "vi_tri_id": vt["id"],
            "cau_truc": [{"linh_vuc_id": lv["id"], "so_cau_de": 1, "so_cau_trung_binh": 0, "so_cau_kho": 0}],
        })
        assert resp.status_code == 400
        assert resp.json()["detail"]["error"]["code"] == "DGNL_019"


# =========================================================================
# IMPORT THI SINH TU EXCEL (B1 - 30/07/2026)
# =========================================================================

def _build_xlsx(ma_cc_list: list, header: str = "ma_cc") -> bytes:
    """Build file xlsx in-memory: 1 cot header + danh sach ma_cc."""
    import io as _io
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=header)
    for i, ma in enumerate(ma_cc_list, start=2):
        ws.cell(row=i, column=1, value=ma)
    out = _io.BytesIO()
    wb.save(out)
    return out.getvalue()


class TestImportThiSinh:
    """Import ma cong chuc tu Excel + tai file mau."""

    async def _setup_ky_thi_vi_tri(self, client) -> tuple[dict, dict]:
        uid = uuid.uuid4().hex[:6]
        vt = await _create_vi_tri(client, f"VT-IMP-{uid}")
        kt = await _create_ky_thi(client, f"KT-IMP-{uid}")
        return kt, vt

    async def _lay_ma_cc_that(self, db_session, n: int = 2) -> list[str]:
        """Lay n ma_cc that (is_active) tu public.cong_chuc de import."""
        from sqlalchemy import text as sa_text
        r = await db_session.execute(sa_text(
            "SELECT ma_cc FROM public.cong_chuc WHERE is_active = true ORDER BY ma_cc LIMIT :n"
        ), {"n": n})
        return [row[0] for row in r.all()]

    async def test_download_file_mau(self, client, admin_user):
        kt, _vt = await self._setup_ky_thi_vi_tri(client)
        resp = await client.get(f"{BASE}/ky-thi/{kt['id']}/thi-sinh/import/mau")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]

        import io as _io
        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(resp.content))
        ws = wb.active
        # File mau: dong 1 huong dan, dong 2 header ma_cc
        assert ws.cell(row=2, column=1).value == "ma_cc"

    async def test_import_thanh_cong(self, client, admin_user, db_session):
        kt, vt = await self._setup_ky_thi_vi_tri(client)
        ma_list = await self._lay_ma_cc_that(db_session, 2)
        content = _build_xlsx(ma_list)

        resp = await client.post(
            f"{BASE}/ky-thi/{kt['id']}/thi-sinh/import-excel?vi_tri_id={vt['id']}",
            files={"file": ("ds.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 201, resp.json()
        data = resp.json()["data"]
        assert data["thanh_cong"] == 2
        assert data["that_bai"] == 0

        # Import lai cung file -> tat ca bao trung
        resp2 = await client.post(
            f"{BASE}/ky-thi/{kt['id']}/thi-sinh/import-excel?vi_tri_id={vt['id']}",
            files={"file": ("ds.xlsx", content, "application/octet-stream")},
        )
        data2 = resp2.json()["data"]
        assert data2["thanh_cong"] == 0
        assert data2["that_bai"] == 2
        assert all("Đã được giao thi" in e["loi"] for e in data2["loi_chi_tiet"])

    async def test_import_loi_tung_dong(self, client, admin_user, db_session):
        """Ma khong ton tai + trung trong file -> loi per dong, dong hop le van vao."""
        kt, vt = await self._setup_ky_thi_vi_tri(client)
        ma_list = await self._lay_ma_cc_that(db_session, 1)
        content = _build_xlsx([ma_list[0], "MA-KHONG-TON-TAI", ma_list[0]])

        resp = await client.post(
            f"{BASE}/ky-thi/{kt['id']}/thi-sinh/import-excel?vi_tri_id={vt['id']}",
            files={"file": ("ds.xlsx", content, "application/octet-stream")},
        )
        assert resp.status_code == 201, resp.json()
        data = resp.json()["data"]
        assert data["thanh_cong"] == 1
        assert data["that_bai"] == 2
        loi_map = {e["ma_cc"]: e["loi"] for e in data["loi_chi_tiet"]}
        assert "không tồn tại" in loi_map["MA-KHONG-TON-TAI"]
        assert "Trùng lặp trong file" in loi_map[ma_list[0]]

    async def test_import_thieu_cot_ma_cc(self, client, admin_user):
        kt, vt = await self._setup_ky_thi_vi_tri(client)
        content = _build_xlsx(["X1"], header="ten_cot_sai")
        resp = await client.post(
            f"{BASE}/ky-thi/{kt['id']}/thi-sinh/import-excel?vi_tri_id={vt['id']}",
            files={"file": ("ds.xlsx", content, "application/octet-stream")},
        )
        assert resp.status_code == 400
        assert resp.json()["detail"]["error"]["code"] == "DGNL_061"


# =========================================================================
# VI PHAM — LOG CHI TIET (B5 - 30/07/2026)
# =========================================================================

class TestViPham:
    """Log vi pham chi tiet: ghi ngay kem gio, ly do giai trinh, xem danh sach."""

    _setup_dgnl_exam = TestLichSuLanThi._setup_dgnl_exam
    _open_with_dap_an = TestLichSuLanThi._open_with_dap_an

    async def _bat_dau(self, client, kt_id: str) -> dict:
        start = await client.post(f"{BASE}/ky-thi/{kt_id}/bat-dau")
        assert start.status_code == 200, start.json()
        return start.json()["data"]

    async def test_ghi_vi_pham_tang_counter_va_co_gio(self, client, admin_user):
        data, cc_id = await self._open_with_dap_an(client, admin_user)
        kt_id = data["ky_thi"]["id"]
        await self._bat_dau(client, kt_id)

        r1 = await client.post(f"{BASE}/ky-thi/{kt_id}/vi-pham", json={"loai_vi_pham": "EXIT_FULLSCREEN"})
        assert r1.status_code == 201, r1.json()
        assert r1.json()["data"]["so_lan_vi_pham"] == 1
        assert r1.json()["data"]["thoi_gian"] is not None
        vp_id = r1.json()["data"]["id"]

        r2 = await client.post(f"{BASE}/ky-thi/{kt_id}/vi-pham", json={"loai_vi_pham": "SWITCH_TAB"})
        assert r2.status_code == 201
        assert r2.json()["data"]["so_lan_vi_pham"] == 2

        # Admin xem danh sach chi tiet
        ds = await client.get(f"{BASE}/ky-thi/{kt_id}/thi-sinh/{cc_id}/vi-pham")
        assert ds.status_code == 200
        items = ds.json()["data"]
        assert len(items) == 2
        assert items[0]["id"] == vp_id
        assert items[0]["loai_vi_pham"] == "EXIT_FULLSCREEN"
        assert items[1]["loai_vi_pham"] == "SWITCH_TAB"
        assert all(i["thoi_gian"] for i in items)
        assert all(i["lan_thi"] == 1 for i in items)

    async def test_ghi_vi_pham_loai_khong_hop_le(self, client, admin_user):
        data, _cc = await self._open_with_dap_an(client, admin_user)
        kt_id = data["ky_thi"]["id"]
        await self._bat_dau(client, kt_id)
        r = await client.post(f"{BASE}/ky-thi/{kt_id}/vi-pham", json={"loai_vi_pham": "HACK"})
        assert r.status_code == 400
        assert r.json()["detail"]["error"]["code"] == "DGNL_050"

    async def test_ghi_vi_pham_khi_chua_thi(self, client, admin_user):
        data, _cc = await self._open_with_dap_an(client, admin_user)
        r = await client.post(f"{BASE}/ky-thi/{data['ky_thi']['id']}/vi-pham", json={"loai_vi_pham": "SWITCH_TAB"})
        assert r.status_code == 400
        assert r.json()["detail"]["error"]["code"] == "DGNL_051"

    async def test_cap_nhat_ly_do_giai_trinh(self, client, admin_user):
        data, cc_id = await self._open_with_dap_an(client, admin_user)
        kt_id = data["ky_thi"]["id"]
        await self._bat_dau(client, kt_id)
        r = await client.post(f"{BASE}/ky-thi/{kt_id}/vi-pham", json={"loai_vi_pham": "EXIT_FULLSCREEN"})
        vp_id = r.json()["data"]["id"]

        patch = await client.patch(
            f"{BASE}/ky-thi/{kt_id}/vi-pham/{vp_id}/ly-do",
            json={"ly_do": "Mất điện đột ngột, máy thoát fullscreen"},
        )
        assert patch.status_code == 200, patch.json()

        ds = await client.get(f"{BASE}/ky-thi/{kt_id}/thi-sinh/{cc_id}/vi-pham")
        assert ds.json()["data"][0]["ly_do"] == "Mất điện đột ngột, máy thoát fullscreen"

    async def test_luu_nhap_khong_ha_counter(self, client, admin_user):
        """luu-nhap gui so_lan_vi_pham thap hon -> counter server khong bi ha."""
        data, cc_id = await self._open_with_dap_an(client, admin_user)
        kt_id = data["ky_thi"]["id"]
        start = await self._bat_dau(client, kt_id)

        await client.post(f"{BASE}/ky-thi/{kt_id}/vi-pham", json={"loai_vi_pham": "SWITCH_TAB"})
        await client.post(f"{BASE}/ky-thi/{kt_id}/vi-pham", json={"loai_vi_pham": "SWITCH_TAB"})

        # Autosave voi counter cu (0) — khong duoc ghi de xuong
        cau_tra_loi = [{"cau_hoi_id": c["id"], "tra_loi": {"dap_an": "A"}} for c in start["cau_hoi"]]
        save = await client.post(f"{BASE}/ky-thi/{kt_id}/luu-nhap", json={
            "cau_tra_loi": cau_tra_loi, "so_lan_vi_pham": 0,
        })
        assert save.status_code == 200

        ds = await client.get(f"{BASE}/ky-thi/{kt_id}/thi-sinh")
        ts = next(i for i in ds.json()["data"] if i["cong_chuc_id"] == cc_id)
        assert ts["so_lan_vi_pham"] == 2

    async def test_export_excel_co_sheet_vi_pham(self, client, admin_user):
        import io as _io
        from openpyxl import load_workbook

        data, _cc = await self._open_with_dap_an(client, admin_user)
        kt_id = data["ky_thi"]["id"]
        await self._bat_dau(client, kt_id)
        await client.post(f"{BASE}/ky-thi/{kt_id}/vi-pham", json={"loai_vi_pham": "EXIT_FULLSCREEN"})

        resp = await client.get(f"{BASE}/ky-thi/{kt_id}/export")
        assert resp.status_code == 200
        wb = load_workbook(_io.BytesIO(resp.content), read_only=True)
        assert "Vi phạm chi tiết" in wb.sheetnames
        ws = wb["Vi phạm chi tiết"]
        rows = list(ws.iter_rows(min_row=4, values_only=True))
        assert len(rows) >= 1
        # Cot 6 = loai, cot 7 = thoi gian (dd/mm/yyyy HH:MM:SS)
        assert rows[0][5] == "Thoát toàn màn hình"
        assert rows[0][6] and "/" in str(rows[0][6])


# =========================================================================
# PHAN QUYEN
# =========================================================================

class TestPhanQuyen:
    async def test_cbcc_khong_tao_ky_thi(self, client, cbcc_user):
        now = now_vn()
        resp = await client.post(f"{BASE}/ky-thi", json={
            "ma_ky_thi": "KT-NOAUTH",
            "ten_ky_thi": "Test no auth",
            "ngay_bat_dau": now.isoformat(),
            "ngay_ket_thuc": (now + timedelta(days=1)).isoformat(),
        })
        assert resp.status_code == 403

    async def test_cbcc_khong_tao_linh_vuc(self, client, cbcc_user):
        resp = await client.post(f"{BASE}/linh-vuc", json={
            "ma_linh_vuc": "LV-NOAUTH",
            "ten_linh_vuc": "Test no auth",
        })
        assert resp.status_code == 403

    async def test_cbcc_xem_linh_vuc_ok(self, client, cbcc_user):
        """CBCC duoc xem danh sach linh vuc."""
        resp = await client.get(f"{BASE}/linh-vuc")
        assert resp.status_code == 200

    async def test_cbcc_xem_ky_thi_ok(self, client, cbcc_user):
        """CBCC duoc xem danh sach ky thi (chi thay duoc giao)."""
        resp = await client.get(f"{BASE}/ky-thi")
        assert resp.status_code == 200


# =========================================================================
# UNIT TEST — PHAN QUYEN XEM BAI LAM (siet chi CCT/PCCT, bo TDV/PDV)
# =========================================================================

class TestPhanQuyenXemBaiLam:
    """Test helper _is_lanh_dao: chi CCT/PCCT moi duoc coi la lanh dao xem bai lam.
    TDV/PDV (is_lanh_dao=True nhung vai_tro khac) phai bi tu choi."""

    def _make_service(self):
        # ThiSinhService can db nhung _is_lanh_dao khong su dung db -> truyen None.
        from lms_service.services.thi_sinh_service import ThiSinhService
        return ThiSinhService(db=None)  # type: ignore[arg-type]

    def _make_user(self, vai_tro: str, is_lanh_dao: bool = False, platform_roles=None):
        from shared.auth import TokenPayload
        return TokenPayload(
            sub="00000000-0000-0000-0000-000000000001",
            exp=int((now_vn() + timedelta(hours=1)).timestamp()),
            type="access",
            ma_cc="TEST",
            vai_tro=vai_tro,
            don_vi_id="a0000000-0000-0000-0000-000000000001",
            is_lanh_dao=is_lanh_dao,
            platform_roles=platform_roles or [],
        )

    def test_cct_la_lanh_dao(self):
        svc = self._make_service()
        assert svc._is_lanh_dao(self._make_user("CCT")) is True

    def test_pcct_la_lanh_dao(self):
        svc = self._make_service()
        assert svc._is_lanh_dao(self._make_user("PCCT")) is True

    def test_tdv_khong_la_lanh_dao(self):
        """TDV (TRUONG_PHONG, is_lanh_dao=True) PHAI bi tu choi."""
        svc = self._make_service()
        assert svc._is_lanh_dao(self._make_user("TRUONG_PHONG", is_lanh_dao=True)) is False

    def test_pdv_khong_la_lanh_dao(self):
        """PDV (PHO_PHONG, is_lanh_dao=True) PHAI bi tu choi."""
        svc = self._make_service()
        assert svc._is_lanh_dao(self._make_user("PHO_PHONG", is_lanh_dao=True)) is False

    def test_chuyen_vien_khong_la_lanh_dao(self):
        svc = self._make_service()
        assert svc._is_lanh_dao(self._make_user("CHUYEN_VIEN")) is False

    def test_super_admin_khong_la_lanh_dao(self):
        """SUPER_ADMIN duoc xem qua _is_manager, khong qua _is_lanh_dao."""
        svc = self._make_service()
        assert svc._is_lanh_dao(self._make_user("SUPER_ADMIN")) is False


# =========================================================================
# UNIT TEST — NORMALIZE DAP_AN_DUNG (bug fix [object Object] tren FE)
# =========================================================================

class TestNormalizeDapAnDung:
    """Test helper _normalize_dap_an_dung — chuan hoa shape dap_an_dung."""

    def test_trac_nghiem_1_unwrap_string(self):
        from lms_service.services.thi_sinh_service import ThiSinhService
        da = {"lua_chon": [{"key": "A", "noi_dung": "X"}], "dap_an_dung": "A"}
        assert ThiSinhService._normalize_dap_an_dung(da) == "A"

    def test_trac_nghiem_nhieu_unwrap_list(self):
        from lms_service.services.thi_sinh_service import ThiSinhService
        da = {"lua_chon": [{"key": "A"}, {"key": "B"}], "dap_an_dung": ["A", "C"]}
        assert ThiSinhService._normalize_dap_an_dung(da) == ["A", "C"]

    def test_dung_sai_unwrap_string(self):
        from lms_service.services.thi_sinh_service import ThiSinhService
        da = {"lua_chon": [{"key": "A"}, {"key": "B"}], "dap_an_dung": "A"}
        assert ThiSinhService._normalize_dap_an_dung(da) == "A"

    def test_tu_luan_unwrap_goi_y(self):
        from lms_service.services.thi_sinh_service import ThiSinhService
        da = {"goi_y": "Trinh bay cac buoc xu ly ho so"}
        assert ThiSinhService._normalize_dap_an_dung(da) == "Trinh bay cac buoc xu ly ho so"

    def test_none_passthrough(self):
        from lms_service.services.thi_sinh_service import ThiSinhService
        assert ThiSinhService._normalize_dap_an_dung(None) is None

    def test_string_passthrough(self):
        from lms_service.services.thi_sinh_service import ThiSinhService
        assert ThiSinhService._normalize_dap_an_dung("A") == "A"

    def test_list_passthrough(self):
        from lms_service.services.thi_sinh_service import ThiSinhService
        assert ThiSinhService._normalize_dap_an_dung(["A", "B"]) == ["A", "B"]

    def test_empty_dict_returns_none(self):
        from lms_service.services.thi_sinh_service import ThiSinhService
        assert ThiSinhService._normalize_dap_an_dung({}) is None

    def test_unknown_dict_returns_none(self):
        """Dict khong co dap_an_dung/goi_y → None de FE khong hien [object Object]."""
        from lms_service.services.thi_sinh_service import ThiSinhService
        assert ThiSinhService._normalize_dap_an_dung({"something_else": "x"}) is None

    def test_result_never_dict(self):
        """Tat ca code path KHONG bao gio tra ve dict (root cause bug)."""
        from lms_service.services.thi_sinh_service import ThiSinhService
        cases = [
            {"lua_chon": [], "dap_an_dung": "A"},
            {"lua_chon": [], "dap_an_dung": ["A", "B"]},
            {"goi_y": "x"},
            {},
            None,
            "A",
            ["A", "B"],
        ]
        for c in cases:
            assert not isinstance(ThiSinhService._normalize_dap_an_dung(c), dict), (
                f"input {c!r} cho ra dict — bug ['object Object'] van con!"
            )


# =========================================================================
# UNIT TEST — UPSERT LAN THI (idempotent theo lan)
# =========================================================================

class TestUpsertLanThi:
    """Test _upsert_lan_thi: append neu chua co, replace neu da co cung lan."""

    class _FakeTs:
        def __init__(self, lich_su=None):
            self.lich_su_thi = lich_su

    def test_append_khi_chua_co(self):
        from lms_service.services.thi_sinh_service import ThiSinhService
        ts = self._FakeTs(lich_su=[])
        ThiSinhService._upsert_lan_thi(ts, {"lan": 1, "diem": 60})
        assert len(ts.lich_su_thi) == 1
        assert ts.lich_su_thi[0]["lan"] == 1

    def test_append_khi_lich_su_None(self):
        from lms_service.services.thi_sinh_service import ThiSinhService
        ts = self._FakeTs(lich_su=None)
        ThiSinhService._upsert_lan_thi(ts, {"lan": 1, "diem": 60})
        assert len(ts.lich_su_thi) == 1

    def test_replace_khi_trung_lan(self):
        """Upsert lan 1 lan thu 2 -> replace, khong tao duplicate."""
        from lms_service.services.thi_sinh_service import ThiSinhService
        ts = self._FakeTs(lich_su=[{"lan": 1, "diem": 60}])
        ThiSinhService._upsert_lan_thi(ts, {"lan": 1, "diem": 75})
        assert len(ts.lich_su_thi) == 1
        assert ts.lich_su_thi[0]["diem"] == 75  # da bi replace

    def test_append_khi_lan_khac(self):
        """Append lan 2 vao lich su da co lan 1."""
        from lms_service.services.thi_sinh_service import ThiSinhService
        ts = self._FakeTs(lich_su=[{"lan": 1, "diem": 60}])
        ThiSinhService._upsert_lan_thi(ts, {"lan": 2, "diem": 80})
        assert len(ts.lich_su_thi) == 2
        lans = sorted(e["lan"] for e in ts.lich_su_thi)
        assert lans == [1, 2]


# =========================================================================
# SINGLE-SESSION (phien_thi) + GIAM SAT + SIET QUYEN ADMIN-ONLY (22/06/2026)
# =========================================================================

from lms_service.main import app as _app
from lms_service.dependencies import get_current_user as _get_current_user
from shared.auth import TokenPayload as _TokenPayload


def _override_current_user(user: _TokenPayload) -> None:
    """Swap user dang dang nhap giua chung 1 test (multi-user scenario)."""
    async def _o():
        return user
    _app.dependency_overrides[_get_current_user] = _o


def _make_token(vai_tro: str, is_lanh_dao: bool = False, platform_roles=None, sub=None) -> _TokenPayload:
    return _TokenPayload(
        sub=sub or "01014af6-1505-495b-95ab-8c1503cfa061",
        exp=int((now_vn() + timedelta(hours=1)).timestamp()),
        type="access",
        ma_cc=f"TEST-{vai_tro[:5]}",
        vai_tro=vai_tro,
        don_vi_id="a0000000-0000-0000-0000-000000000001",
        is_lanh_dao=is_lanh_dao,
        platform_roles=platform_roles or [],
    )


async def _setup_open_exam(client: AsyncClient, examinee_cc_id: str) -> dict:
    """Tao ky thi day du + giao 1 thi sinh + mo (DANG_MO). Tra ve {ky_thi, vi_tri}."""
    data = await _setup_full_exam(client, None)
    kt = data["ky_thi"]
    vt = data["vi_tri"]
    await client.post(f"{BASE}/ky-thi/{kt['id']}/thi-sinh", json={
        "danh_sach": [{"cong_chuc_id": examinee_cc_id, "vi_tri_id": vt["id"]}],
    })
    await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "CHO_DUYET"})
    await client.patch(f"{BASE}/ky-thi/{kt['id']}/trang-thai", json={"trang_thai": "DANG_MO"})
    return data


class TestPhienThiSingleSession:
    """Feature 1: 1 phien thi/tai khoan — thiet bi cu bi 409 khi co thiet bi moi."""

    # admin_user idx=0 -> cc_id nay
    _CC = "00327c43-c9a3-44d7-8306-7084e75cb2b5"

    async def test_bat_dau_tra_ve_phien_token(self, client, admin_user):
        data = await _setup_open_exam(client, self._CC)
        kt = data["ky_thi"]
        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/bat-dau")
        assert resp.status_code == 200
        assert resp.json()["data"].get("phien_token")

    async def test_nop_bai_token_cu_bi_409(self, client, admin_user):
        data = await _setup_open_exam(client, self._CC)
        kt = data["ky_thi"]

        # Thiet bi 1 bat dau -> token1
        r1 = await client.post(f"{BASE}/ky-thi/{kt['id']}/bat-dau")
        token1 = r1.json()["data"]["phien_token"]
        cau_hoi = r1.json()["data"]["cau_hoi"]

        # Thiet bi 2 bat dau (resume cung ky thi) -> token2, ghi de token1
        r2 = await client.post(f"{BASE}/ky-thi/{kt['id']}/bat-dau")
        token2 = r2.json()["data"]["phien_token"]
        assert token2 and token2 != token1

        cau_tra_loi = [{"cau_hoi_id": ch["id"], "tra_loi": {"dap_an": "A"}} for ch in cau_hoi]

        # Thiet bi 1 (token cu) nop -> 409
        resp_cu = await client.post(
            f"{BASE}/ky-thi/{kt['id']}/nop-bai",
            json={"cau_tra_loi": cau_tra_loi},
            headers={"X-Phien-Thi": token1},
        )
        assert resp_cu.status_code == 409
        assert resp_cu.json()["detail"]["error"]["code"] == "PHIEN_001"

        # Thiet bi 2 (token moi) nop -> 200
        resp_moi = await client.post(
            f"{BASE}/ky-thi/{kt['id']}/nop-bai",
            json={"cau_tra_loi": cau_tra_loi},
            headers={"X-Phien-Thi": token2},
        )
        assert resp_moi.status_code == 200

    async def test_luu_nhap_token_cu_bi_409(self, client, admin_user):
        data = await _setup_open_exam(client, self._CC)
        kt = data["ky_thi"]
        r1 = await client.post(f"{BASE}/ky-thi/{kt['id']}/bat-dau")
        token1 = r1.json()["data"]["phien_token"]
        await client.post(f"{BASE}/ky-thi/{kt['id']}/bat-dau")  # token2 ghi de

        resp = await client.post(
            f"{BASE}/ky-thi/{kt['id']}/luu-nhap",
            json={"cau_tra_loi": [], "so_lan_vi_pham": 1},
            headers={"X-Phien-Thi": token1},
        )
        assert resp.status_code == 409
        assert resp.json()["detail"]["error"]["code"] == "PHIEN_001"

    async def test_khong_co_token_van_chay(self, client, admin_user):
        """Client cu khong gui token -> khong enforce (tranh khoa khi trien khai)."""
        data = await _setup_open_exam(client, self._CC)
        kt = data["ky_thi"]
        r = await client.post(f"{BASE}/ky-thi/{kt['id']}/bat-dau")
        cau_hoi = r.json()["data"]["cau_hoi"]
        cau_tra_loi = [{"cau_hoi_id": ch["id"], "tra_loi": {"dap_an": "A"}} for ch in cau_hoi]
        resp = await client.post(f"{BASE}/ky-thi/{kt['id']}/nop-bai", json={"cau_tra_loi": cau_tra_loi})
        assert resp.status_code == 200


class TestGiamSat:
    """Feature 3: man hinh giam sat truc tiep — chi admin."""

    _CC = "00327c43-c9a3-44d7-8306-7084e75cb2b5"

    async def test_giam_sat_admin_ok(self, client, admin_user):
        data = await _setup_open_exam(client, self._CC)
        kt = data["ky_thi"]
        await client.post(f"{BASE}/ky-thi/{kt['id']}/bat-dau")

        resp = await client.get(f"{BASE}/ky-thi/{kt['id']}/giam-sat")
        assert resp.status_code == 200
        body = resp.json()["data"]
        assert body["tong_quan"]["tong_thi_sinh"] == 1
        assert body["tong_quan"]["dang_thi"] == 1
        assert len(body["thi_sinh"]) == 1
        ts = body["thi_sinh"][0]
        assert ts["trang_thai"] == "DANG_THI"
        assert "so_lan_vi_pham" in ts and "online" in ts

    async def test_giam_sat_cbcc_403(self, client, admin_user):
        data = await _setup_open_exam(client, self._CC)
        kt = data["ky_thi"]
        # Chuyen sang CBCC (khong co QT_DAO_TAO)
        _override_current_user(_make_token("CHUYEN_VIEN"))
        try:
            resp = await client.get(f"{BASE}/ky-thi/{kt['id']}/giam-sat")
            assert resp.status_code == 403
        finally:
            _override_current_user(admin_user)

    async def test_giam_sat_lanh_dao_403(self, client, admin_user):
        """Feature 4: lanh dao (CCT) KHONG con quyen giam sat DGNL."""
        data = await _setup_open_exam(client, self._CC)
        kt = data["ky_thi"]
        _override_current_user(_make_token("CCT", is_lanh_dao=True))
        try:
            resp = await client.get(f"{BASE}/ky-thi/{kt['id']}/giam-sat")
            assert resp.status_code == 403
        finally:
            _override_current_user(admin_user)


class TestSietQuyenAdminOnly:
    """Feature 4: chi admin (QT_DAO_TAO/SUPER_ADMIN) xem thi sinh/thong ke DGNL."""

    _CC = "00327c43-c9a3-44d7-8306-7084e75cb2b5"

    async def test_lanh_dao_khong_xem_thi_sinh(self, client, admin_user):
        data = await _setup_open_exam(client, self._CC)
        kt = data["ky_thi"]
        _override_current_user(_make_token("CCT", is_lanh_dao=True))
        try:
            resp = await client.get(f"{BASE}/ky-thi/{kt['id']}/thi-sinh")
            assert resp.status_code == 403
        finally:
            _override_current_user(admin_user)

    async def test_lanh_dao_khong_thong_ke(self, client, admin_user):
        data = await _setup_open_exam(client, self._CC)
        kt = data["ky_thi"]
        _override_current_user(_make_token("CCT", is_lanh_dao=True))
        try:
            resp = await client.get(f"{BASE}/ky-thi/{kt['id']}/thong-ke")
            assert resp.status_code == 403
        finally:
            _override_current_user(admin_user)

    async def test_admin_van_xem_thi_sinh_ok(self, client, admin_user):
        data = await _setup_open_exam(client, self._CC)
        kt = data["ky_thi"]
        resp = await client.get(f"{BASE}/ky-thi/{kt['id']}/thi-sinh")
        assert resp.status_code == 200


# =========================================================================
# XOA NHIEU CAU HOI DGNL (bulk soft-delete) — POST /dgnl/ngan-hang/xoa-nhieu
# =========================================================================

async def _tao_ch_dgnl(client, lv_id, noi_dung, do_kho="DE", da="A"):
    """Helper tao 1 cau hoi DGNL, tra ve id."""
    body = {
        "linh_vuc_id": lv_id, "do_kho": do_kho, "loai": "TRAC_NGHIEM_1",
        "noi_dung": noi_dung,
        "dap_an": {"lua_chon": [{"key": "A", "noi_dung": "A"}, {"key": "B", "noi_dung": "B"}],
                   "dap_an_dung": da},
    }
    r = await client.post(f"{BASE}/dgnl/ngan-hang", json=body)
    assert r.status_code == 201, r.json()
    return r.json()["data"]["id"]


class TestXoaNhieuCauHoiDgnl:
    async def test_xoa_theo_ids(self, client, admin_user):
        lv = await _create_linh_vuc(client, f"LV-BULK-{uuid.uuid4().hex[:6]}")
        ids = [await _tao_ch_dgnl(client, lv["id"], f"Câu {i}") for i in range(3)]

        resp = await client.post(f"{BASE}/dgnl/ngan-hang/xoa-nhieu", json={"ids": ids[:2]})
        assert resp.status_code == 200, resp.json()
        assert resp.json()["data"]["so_xoa"] == 2

        # Con lai 1 cau active trong linh vuc nay
        lst = await client.get(f"{BASE}/dgnl/ngan-hang", params={"linh_vuc_id": lv["id"]})
        assert lst.json()["pagination"]["total_items"] == 1

    async def test_xoa_tat_ca_theo_bo_loc(self, client, admin_user):
        lv = await _create_linh_vuc(client, f"LV-BULK2-{uuid.uuid4().hex[:6]}")
        for i in range(4):
            await _tao_ch_dgnl(client, lv["id"], f"Câu {i}", do_kho="DE")
        await _tao_ch_dgnl(client, lv["id"], "Câu khó", do_kho="KHO")

        # Xoa tat ca do_kho=DE trong linh vuc nay
        resp = await client.post(f"{BASE}/dgnl/ngan-hang/xoa-nhieu", json={
            "tat_ca_theo_bo_loc": True, "linh_vuc_id": lv["id"], "do_kho": "DE",
        })
        assert resp.status_code == 200
        assert resp.json()["data"]["so_xoa"] == 4

        lst = await client.get(f"{BASE}/dgnl/ngan-hang", params={"linh_vuc_id": lv["id"]})
        assert lst.json()["pagination"]["total_items"] == 1  # con lai cau KHO

    async def test_xoa_tat_ca_khong_bo_loc_bi_chan(self, client, admin_user):
        # tat_ca_theo_bo_loc nhung khong co dieu kien loc -> 400 (chan xoa toan bo)
        resp = await client.post(f"{BASE}/dgnl/ngan-hang/xoa-nhieu", json={
            "tat_ca_theo_bo_loc": True,
        })
        assert resp.status_code == 400

    async def test_xoa_khong_co_ids_bi_chan(self, client, admin_user):
        resp = await client.post(f"{BASE}/dgnl/ngan-hang/xoa-nhieu", json={"ids": []})
        assert resp.status_code == 400

    async def test_giang_vien_khong_duoc_xoa_nhieu(self, client, giang_vien_user):
        # GIANG_VIEN khong phai QT_DAO_TAO -> bi tu choi (403 do require_platform_role)
        resp = await client.post(f"{BASE}/dgnl/ngan-hang/xoa-nhieu",
                                 json={"ids": [str(uuid.uuid4())]})
        assert resp.status_code == 403
