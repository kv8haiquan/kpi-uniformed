"""
api/endpoints/diem_danh.py
============================
Module 4 — Điểm danh. 6 endpoint theo §6 HKG_API_SPECS.md, cộng 2 endpoint
bảng chi tiết + xuất Excel cho ban tổ chức (thêm 04/09/2026).
"""

from datetime import datetime
from uuid import UUID
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Response

from meeting_service.dependencies import (
    DatabaseDep,
    CurrentUserDep,
    require_can_edit_meeting,
    require_can_manage_diem_danh,
    require_can_view_meeting,
)
from meeting_service.models.cuoc_hop import CuocHop
from meeting_service.schemas.diem_danh import (
    BamTayBulk,
    DiemDanhResponse,
    DiemDanhSummary,
    QRSubmit,
    QRTokenResponse,
)
from meeting_service.services.audit_log_service import ghi_audit
from meeting_service.services.diem_danh_service import DiemDanhService


router = APIRouter(prefix="/diem-danh", tags=["Điểm danh"])
router_cuoc_hop = APIRouter(prefix="/cuoc-hop", tags=["Điểm danh"])


# ─── 1. Sinh QR token ─────────────────────────────────────────────────
@router.post(
    "/qr-token/{cuoc_hop_id}",
    response_model=None,
    summary="Sinh QR token cho cuộc họp (chu_toa/thu_ky)",
)
async def qr_token(
    db: DatabaseDep,
    user: CurrentUserDep,
    ch: CuocHop = Depends(require_can_edit_meeting),
):
    service = DiemDanhService(db)
    data = service.issue_qr_token(ch.id, ttl_seconds=3600)
    return {"success": True, "data": data}


# ─── 2. CBCC quét QR submit ───────────────────────────────────────────
@router.post("/quet", summary="CBCC quét QR submit token")
async def quet_qr(
    payload: QRSubmit,
    db: DatabaseDep,
    user: CurrentUserDep,
):
    service = DiemDanhService(db)
    dd = await service.submit_qr(payload.token, user)
    return {
        "success": True,
        "data": DiemDanhResponse.model_validate(dd).model_dump(mode="json"),
    }


# ─── 3. Bấm tay ───────────────────────────────────────────────────────
@router.post("/bam-tay", summary="Thư ký bấm tay điểm danh nhiều CBCC")
async def bam_tay(
    payload: BamTayBulk,
    db: DatabaseDep,
    user: CurrentUserDep,
):
    # Verify edit permission cho cuộc họp
    from sqlalchemy import select
    from meeting_service.models.cuoc_hop import CuocHop as CuocHopModel
    res = await db.execute(
        select(CuocHopModel).where(
            CuocHopModel.id == payload.cuoc_hop_id, CuocHopModel.is_deleted.is_(False)
        )
    )
    ch = res.scalar_one_or_none()
    if ch is None:
        raise HTTPException(
            status_code=404,
            detail={"success": False, "error": {"code": "MEETING_NOT_FOUND",
                    "message": "Không tìm thấy cuộc họp"}},
        )
    # G4-fix-5: chặn điểm danh tay cho cuộc họp đã hủy
    if ch.trang_thai == "HUY":
        raise HTTPException(
            status_code=409,
            detail={"success": False, "error": {"code": "MEETING_CANCELLED",
                    "message": "Cuộc họp đã hủy — không thể điểm danh"}},
        )
    user_id = UUID(user.sub)
    if not (
        user.is_admin or user.vai_tro in ("SUPER_ADMIN", "ADMIN")
        or ch.chu_toa_id == user_id or ch.thu_ky_id == user_id
        or "TRUONG_CNTT" in (user.platform_roles or [])
    ):
        raise HTTPException(
            status_code=403,
            detail={"success": False, "error": {"code": "NO_PERMISSION",
                    "message": "Chỉ chu_toa/thu_ky/admin mới được bấm tay"}},
        )

    service = DiemDanhService(db)
    results = await service.bam_tay(payload, user)
    return {
        "success": True,
        "data": {
            "so_diem_danh": len(results),
            "chi_tiet": [
                DiemDanhResponse.model_validate(r).model_dump(mode="json")
                for r in results
            ],
        },
    }


# ─── G4-fix-6.2: SELF CHECKIN (CBCC tự click trong app) ───────────────
@router_cuoc_hop.post(
    "/{cuoc_hop_id}/tu-diem-danh",
    summary="CBCC tự điểm danh (không cần QR)",
)
async def tu_diem_danh(
    cuoc_hop_id: UUID,
    db: DatabaseDep,
    user: CurrentUserDep,
):
    """CBCC click "Tôi có mặt" trong app trên máy tính."""
    service = DiemDanhService(db)
    dd = await service.tu_diem_danh(cuoc_hop_id, user)
    return {
        "success": True,
        "data": DiemDanhResponse.model_validate(dd).model_dump(mode="json"),
    }


@router_cuoc_hop.get(
    "/{cuoc_hop_id}/diem-danh-cua-toi",
    summary="Trạng thái điểm danh của user hiện tại",
)
async def diem_danh_cua_toi(
    cuoc_hop_id: UUID,
    db: DatabaseDep,
    user: CurrentUserDep,
    ch: CuocHop = Depends(require_can_view_meeting),
):
    """FE dùng để biết hiển thị nút 'Tôi có mặt' hay 'Đã điểm danh'.

    Fix 30/07/2026: bổ sung require_can_view_meeting — trước đây endpoint này
    không kiểm tra quyền nên user không thuộc cuộc họp vẫn gọi được (trả 200
    trong khi các endpoint cùng trang trả 403, gây log khó đọc).
    """
    service = DiemDanhService(db)
    return {"success": True, "data": await service.my_status(cuoc_hop_id, user)}


# ─── 4. Tổng hợp điểm danh ────────────────────────────────────────────
@router_cuoc_hop.get(
    "/{cuoc_hop_id}/diem-danh",
    summary="Tổng hợp điểm danh cuộc họp",
)
async def summary_diem_danh(
    db: DatabaseDep,
    user: CurrentUserDep,
    ch: CuocHop = Depends(require_can_view_meeting),
):
    service = DiemDanhService(db)
    data = await service.summary(ch.id)
    return {
        "success": True,
        "data": {
            **{k: v for k, v in data.items() if k != "chi_tiet"},
            "chi_tiet": [
                DiemDanhResponse.model_validate(d).model_dump(mode="json")
                for d in data["chi_tiet"]
            ],
        },
    }


# ─── 5. CHI TIẾT ĐIỂM DANH TỪNG THÀNH PHẦN ────────────────────────────
@router_cuoc_hop.get(
    "/{cuoc_hop_id}/diem-danh/chi-tiet",
    summary="Bảng điểm danh từng thành phần (ban tổ chức)",
)
async def chi_tiet_diem_danh(
    db: DatabaseDep,
    user: CurrentUserDep,
    ch: CuocHop = Depends(require_can_manage_diem_danh),
):
    """Endpoint RIÊNG, không nhồi vào `GET /{id}/diem-danh`.

    Endpoint tổng hợp cũ dùng `require_can_view_meeting` nên mọi người được
    mời đều gọi được; thêm danh sách từng người vào đó sẽ để công chức thường
    đọc được ai vắng, ai chưa điểm danh của cả cuộc họp.
    """
    service = DiemDanhService(db)
    return {"success": True, "data": await service.chi_tiet(ch, user)}


_MUI_GIO_VN = ZoneInfo("Asia/Ho_Chi_Minh")


def _gio_vn(iso: str | None) -> str:
    """ISO (UTC) → 'dd/mm/YYYY HH:MM' giờ Việt Nam. Rỗng nếu chưa điểm danh."""
    if not iso:
        return ""
    try:
        dt = datetime.fromisoformat(iso)
    except ValueError:
        return ""
    if dt.tzinfo is None:
        # Phòng trường hợp cột bị ghi giờ tường: coi như đã là giờ VN.
        return dt.strftime("%d/%m/%Y %H:%M")
    return dt.astimezone(_MUI_GIO_VN).strftime("%d/%m/%Y %H:%M")


# ─── 6. XUẤT EXCEL BẢNG ĐIỂM DANH ─────────────────────────────────────
@router_cuoc_hop.get(
    "/{cuoc_hop_id}/diem-danh/xuat-excel",
    summary="Xuất bảng điểm danh ra Excel",
)
async def xuat_excel_diem_danh(
    db: DatabaseDep,
    user: CurrentUserDep,
    ch: CuocHop = Depends(require_can_manage_diem_danh),
):
    service = DiemDanhService(db)
    kq = await service.chi_tiet(ch, user)

    # Nạp muộn để API JSON không phải trả giá cho openpyxl.
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    NHAN_TRANG_THAI = {
        "CO_MAT": "Có mặt",
        "DEN_MUON": "Đến muộn",
        "VANG_CO_PHEP": "Vắng có phép",
        "VANG_KHONG_PHEP": "Vắng không phép",
    }
    NHAN_HINH_THUC = {
        "QR": "Quét QR",
        "BAM_TAY": "Thư ký bấm tay",
        "TU_DIEM_DANH": "Tự điểm danh",
    }
    NHAN_LOAI_THAM_DU = {"BAT_BUOC": "Bắt buộc", "THAM_KHAO": "Tham khảo"}

    wb = Workbook()
    ws = wb.active
    ws.title = "Diem danh"

    cot = ["STT", "Họ tên", "Mã CC", "Đơn vị", "Chức vụ", "Loại tham dự",
           "Trạng thái", "Giờ điểm danh", "Hình thức", "Người chấm",
           "Lý do vắng / Ghi chú"]

    gio = ch.gio_bat_dau.strftime("%H:%M") if ch.gio_bat_dau else ""
    ws.append([f"BẢNG ĐIỂM DANH — {ch.tieu_de} ({ch.ngay_hop} {gio})".strip()])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cot))
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    th = kq["tong_hop"]
    ws.append([
        f"Tổng {th['tong_so']} người · có mặt {th['co_mat']}"
        f" · đến muộn {th['den_muon']}"
        f" · vắng có phép {th['vang_co_phep']}"
        f" · vắng không phép {th['vang_khong_phep']}"
        f" · chưa điểm danh {th['chua_diem_danh']}"
    ])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cot))
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.append([])
    ws.append(cot)

    dong_tieu_de = ws.max_row
    to_mau = PatternFill("solid", fgColor="D9E1F2")
    for c in range(1, len(cot) + 1):
        o = ws.cell(row=dong_tieu_de, column=c)
        o.font = Font(bold=True)
        o.fill = to_mau
        o.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(kq["danh_sach"], start=1):
        # gio_diem_danh là MỐC tuyệt đối, máy chủ trả ISO theo UTC. Phải đổi
        # về Asia/Ho_Chi_Minh trước khi in, nếu cắt thẳng chuỗi ISO thì báo
        # cáo hiện sai 7 tiếng (một cuộc họp 14:00 hoá ra "06:54").
        gio_dd = _gio_vn(r["gio_diem_danh"])
        ws.append([
            i,
            r["ho_ten"] or "",
            r["ma_cc"] or "",
            r["ten_don_vi"] or "",
            r["chuc_vu"] or "",
            NHAN_LOAI_THAM_DU.get(r["loai_tham_du"], r["loai_tham_du"] or ""),
            NHAN_TRANG_THAI.get(r["trang_thai"], "Chưa điểm danh"),
            gio_dd,
            NHAN_HINH_THUC.get(r["hinh_thuc"], ""),
            r["nguoi_diem_danh_ho_ten"] or "",
            r["ly_do_vang"] or "",
        ])

    for c, rong in enumerate([6, 26, 12, 26, 18, 13, 16, 17, 16, 22, 40], start=1):
        ws.column_dimensions[get_column_letter(c)].width = rong

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)

    await ghi_audit(
        db,
        hanh_dong="EXPORT_DIEM_DANH",
        nguoi_thuc_hien_id=UUID(user.sub),
        doi_tuong_loai="cuoc_hop",
        doi_tuong_id=ch.id,
        chi_tiet={"so_dong": len(kq["danh_sach"])},
    )
    await db.flush()

    # CuocHop KHÔNG có cột ma_hop; mã chỉ tồn tại với nguồn Lịch công tác
    # (ma_lich), nên lấy ngày họp làm tên file cho các cuộc họp HKG thuần.
    ten_file = f"diem-danh-{ch.ma_lich or ch.ngay_hop}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{ten_file}"'},
    )
