"""
api/endpoints/doi_soat.py
==========================
Đối soát tài liệu di trú — G4.9. Màn hình dùng MỘT LẦN rồi ẩn khỏi menu.

Chỉ Chánh Văn phòng và Quản trị viên vào được. Xuất Excel ở đây chính là biên
bản đối chiếu nộp khi nghiệm thu, nên mỗi dòng phải có người quyết định và
thời điểm.
"""

from datetime import date as date_type
from io import BytesIO
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, HTTPException, Query, Response
from pydantic import BaseModel

from meeting_service.dependencies import CurrentUserDep, DatabaseDep
from meeting_service.services.doi_soat_service import (
    NHAN_QUYET_DINH,
    DoiSoatService,
)
from meeting_service.services.lich_cong_tac_service import LoiNghiepVu

router = APIRouter(prefix="/doi-soat", tags=["Đối soát di trú"])


class YeuCauQuyetDinh(BaseModel):
    quyet_dinh: str
    cuoc_hop_id: Optional[UUID] = None
    ghi_chu: Optional[str] = None


def _loi(e: LoiNghiepVu) -> HTTPException:
    return HTTPException(
        e.http,
        detail={"success": False,
                "error": {"code": e.ma, "message": e.thong_diep}})


@router.get("/quyen", summary="Tôi có thấy màn hình đối soát không")
async def quyen(user: CurrentUserDep):
    """Để giao diện ẩn mục menu, không phải gọi rồi nhận 403."""
    return {"success": True,
            "data": {"duoc_xem": DoiSoatService.duoc_xem(user)}}


@router.get("/", summary="Danh sách thư mục cần đối soát")
async def danh_sach(
    db: DatabaseDep,
    user: CurrentUserDep,
    nhom: Optional[str] = Query(None, description="D hoặc E"),
    da_quyet_dinh: Optional[bool] = Query(None, alias="da-quyet-dinh"),
):
    try:
        kq = await DoiSoatService(db).danh_sach(
            user=user, nhom=nhom, da_quyet_dinh=da_quyet_dinh)
    except LoiNghiepVu as e:
        raise _loi(e) from e
    return {"success": True, "data": kq}


@router.get("/{doi_soat_id}/ung-vien", summary="Cuộc họp có thể là chủ thư mục")
async def ung_vien(
    doi_soat_id: UUID,
    db: DatabaseDep,
    user: CurrentUserDep,
    so_ngay: int = Query(3, ge=0, le=30, alias="so-ngay"),
):
    try:
        kq = await DoiSoatService(db).ung_vien(doi_soat_id, user=user,
                                                so_ngay=so_ngay)
    except LoiNghiepVu as e:
        raise _loi(e) from e
    return {"success": True, "data": kq}


@router.post("/{doi_soat_id}/quyet-dinh", summary="Ghi quyết định cho thư mục")
async def quyet_dinh(doi_soat_id: UUID, yc: YeuCauQuyetDinh,
                     db: DatabaseDep, user: CurrentUserDep):
    try:
        kq = await DoiSoatService(db).quyet_dinh(
            doi_soat_id, user=user, quyet_dinh=yc.quyet_dinh,
            cuoc_hop_id=yc.cuoc_hop_id, ghi_chu=yc.ghi_chu)
    except LoiNghiepVu as e:
        raise _loi(e) from e
    return {"success": True, "data": kq, "message": "Đã ghi quyết định"}


@router.delete("/{doi_soat_id}/quyet-dinh", summary="Bỏ quyết định để chọn lại")
async def huy_quyet_dinh(doi_soat_id: UUID, db: DatabaseDep,
                         user: CurrentUserDep):
    try:
        kq = await DoiSoatService(db).huy_quyet_dinh(doi_soat_id, user=user)
    except LoiNghiepVu as e:
        raise _loi(e) from e
    return {"success": True, "data": kq, "message": "Đã bỏ quyết định"}


@router.get("/xuat-excel", summary="Biên bản đối chiếu (Excel)")
async def xuat_excel(db: DatabaseDep, user: CurrentUserDep):
    try:
        kq = await DoiSoatService(db).danh_sach(user=user)
    except LoiNghiepVu as e:
        raise _loi(e) from e

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Bien ban doi chieu"

    cot = ["STT", "Nhóm", "Thư mục", "Số file", "Ngày suy ra", "Số GM",
           "Quyết định", "Cuộc họp được gắn", "Người quyết định",
           "Thời điểm", "Ghi chú"]

    th = kq["tong_hop"]
    ws.append(["BIÊN BẢN ĐỐI CHIẾU TÀI LIỆU DI TRÚ TỪ GOOGLE DRIVE"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cot))
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([f"{th['tong_thu_muc']} thư mục · {th['tong_file']} file · "
               f"đã quyết định {th['da_quyet_dinh']} · "
               f"còn lại {th['con_lai']} · lập ngày "
               f"{date_type.today():%d/%m/%Y}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cot))
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.append([])
    ws.append(cot)

    dong_tieu_de = ws.max_row
    for c in range(1, len(cot) + 1):
        o = ws.cell(row=dong_tieu_de, column=c)
        o.font = Font(bold=True)
        o.fill = PatternFill("solid", fgColor="D9E1F2")
        o.alignment = Alignment(horizontal="center", vertical="center")

    mau_chua_xu_ly = PatternFill("solid", fgColor="FFF2CC")

    for i, d in enumerate(kq["dong"], start=1):
        ws.append([
            i,
            d["nhom"],
            d["duong_dan_thu_muc"],
            d["so_file"],
            d["ngay_suy_ra"].strftime("%d/%m/%Y") if d["ngay_suy_ra"] else "",
            d["so_gm_suy_ra"] or "",
            d["quyet_dinh_nhan"] or "CHƯA QUYẾT ĐỊNH",
            str(d["cuoc_hop_id"] or ""),
            d["nguoi_quyet_dinh"] or "",
            (d["thoi_diem_quyet_dinh"].strftime("%d/%m/%Y %H:%M")
             if d["thoi_diem_quyet_dinh"] else ""),
            d["ghi_chu"] or "",
        ])
        if not d["quyet_dinh"]:
            for c in range(1, len(cot) + 1):
                ws.cell(row=ws.max_row, column=c).fill = mau_chua_xu_ly

    for c, w in enumerate([5, 7, 58, 9, 13, 10, 26, 38, 22, 18, 30], start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(row=dong_tieu_de + 1, column=1)

    # Sheet phụ: tên từng file, để bên nghiệm thu đối chiếu được tới từng file.
    ws2 = wb.create_sheet("Chi tiet file")
    ws2.append(["Thư mục", "Thư mục con", "Tên file", "Số byte", "Quyết định"])
    for c in range(1, 6):
        ws2.cell(row=1, column=c).font = Font(bold=True)
    for d in kq["dong"]:
        for f in d["danh_sach_file"]:
            ws2.append([d["duong_dan_thu_muc"], f.get("thu_muc_con", ""),
                        f.get("ten", ""), f.get("so_byte"),
                        d["quyet_dinh_nhan"] or "CHƯA QUYẾT ĐỊNH"])
    for c, w in enumerate([48, 34, 46, 12, 26], start=1):
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    ten_file = f"bien-ban-doi-chieu-di-tru-{date_type.today():%Y%m%d}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type=("application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"),
        headers={"Content-Disposition": f'attachment; filename="{ten_file}"'},
    )


@router.get("/danh-muc-quyet-dinh", summary="Danh mục 4 quyết định")
async def danh_muc(user: CurrentUserDep):
    return {"success": True,
            "data": [{"ma": k, "ten": v} for k, v in NHAN_QUYET_DINH.items()]}
