"""
api/endpoints/lich_cong_tac.py
===============================
Lịch công tác — di trú từ lichkv8.

Đọc trên cùng bảng `meeting.cuoc_hop` với Họp Không Giấy nên cuộc họp HKG tự
hiện lên lịch, không cần đồng bộ. Sự kiện nguồn HKG có cờ `co_the_mo_hkg` để
giao diện điều hướng thẳng sang màn hình chi tiết cuộc họp (tiêu chí 8.3).
"""

from datetime import date as date_type
from datetime import timedelta
from io import BytesIO
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, HTTPException, Query, Response, status

from meeting_service.dependencies import CurrentUserDep, DatabaseDep
from meeting_service.schemas.lich_cong_tac import (
    LOAI_LICH_VALUES,
    NHAN_LOAI_LICH,
    LichCongTacCreate,
    LichCongTacHuy,
    LichCongTacUpdate,
)
from meeting_service.services.lich_cong_tac_service import (
    LichCongTacService,
    LoiNghiepVu,
    la_quan_tri_lich,
)

router = APIRouter(prefix="/lich-cong-tac", tags=["Lịch công tác"])


def _loi(e: LoiNghiepVu) -> HTTPException:
    """Dịch lỗi nghiệp vụ sang HTTP theo đúng định dạng response chuẩn."""
    return HTTPException(
        e.http,
        detail={"success": False,
                "error": {"code": e.ma, "message": e.thong_diep}})


@router.get("/", summary="Danh sách sự kiện trên lịch")
async def danh_sach(
    db: DatabaseDep,
    user: CurrentUserDep,
    tu_ngay: Optional[date_type] = Query(None, alias="tu-ngay"),
    den_ngay: Optional[date_type] = Query(None, alias="den-ngay"),
    loai_lich: Optional[str] = Query(None, alias="loai-lich"),
    trang_thai: Optional[str] = Query(None, alias="trang-thai"),
    lanh_dao_id: Optional[UUID] = Query(None, alias="lanh-dao-id"),
    tim_kiem: Optional[str] = Query(None, alias="tim-kiem"),
    nguon: Optional[str] = Query(None),
    trang: int = Query(1, ge=1),
    so_dong: int = Query(50, ge=1, le=500, alias="so-dong"),
    moi_truoc: bool = Query(False, alias="moi-truoc",
                            description="Xếp ngày gần nhất lên đầu"),
):
    """Phân trang phía máy chủ.

    Hệ cũ tải toàn bộ mảng cuộc họp vào bộ nhớ trình duyệt và có ngưỡng cảnh
    báo hiệu năng 3000ms — không lặp lại cách đó.
    """
    if loai_lich and loai_lich not in LOAI_LICH_VALUES:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail={"success": False, "error": {
                "code": "LOAI_LICH_KHONG_HOP_LE",
                "message": f"loai_lich phải thuộc {LOAI_LICH_VALUES}"}})

    svc = LichCongTacService(db)
    items, tong = await svc.danh_sach(
        trang=trang, so_dong=so_dong,
        tu_ngay=tu_ngay, den_ngay=den_ngay, loai_lich=loai_lich,
        trang_thai=trang_thai, lanh_dao_id=lanh_dao_id, tim_kiem=tim_kiem,
        nguon=nguon, moi_truoc=moi_truoc,
    )
    return {
        "success": True,
        "data": items,
        "pagination": {
            "page": trang,
            "page_size": so_dong,
            "total_items": tong,
            "total_pages": (tong + so_dong - 1) // so_dong,
        },
    }


@router.get("/thang/{nam}/{thang}", summary="Lịch theo tháng")
async def theo_thang(
    nam: int,
    thang: int,
    db: DatabaseDep,
    user: CurrentUserDep,
    loai_lich: Optional[str] = Query(None, alias="loai-lich"),
    lanh_dao_id: Optional[UUID] = Query(None, alias="lanh-dao-id"),
):
    if not 1 <= thang <= 12:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail={"success": False, "error": {
                "code": "THANG_KHONG_HOP_LE", "message": "Tháng phải từ 1 đến 12"}})

    dau = date_type(nam, thang, 1)
    cuoi = (date_type(nam + (thang == 12), (thang % 12) + 1, 1)
            - timedelta(days=1))

    svc = LichCongTacService(db)
    items, tong = await svc.danh_sach(
        trang=1, so_dong=500, tu_ngay=dau, den_ngay=cuoi,
        loai_lich=loai_lich, lanh_dao_id=lanh_dao_id)

    # Gom theo ngày để giao diện lịch tháng render thẳng. Sự kiện nhiều ngày
    # xuất hiện ở mọi ngày nó kéo dài.
    theo_ngay: dict[str, list] = {}
    for it in items:
        bd = it["ngay_hien_thi"] or it["ngay_hop"]
        kt = it["ngay_ket_thuc"] or bd
        n = max(bd, dau)
        while n <= min(kt, cuoi):
            theo_ngay.setdefault(n.isoformat(), []).append(it)
            n += timedelta(days=1)

    return {"success": True,
            "data": {"nam": nam, "thang": thang, "tong": tong,
                     "theo_ngay": theo_ngay}}


@router.get("/tom-tat", summary="Tóm tắt lịch để gửi nhanh")
async def tom_tat(
    db: DatabaseDep,
    user: CurrentUserDep,
    tu_ngay: Optional[date_type] = Query(None, alias="tu-ngay"),
    so_ngay: int = Query(3, ge=1, le=31, alias="so-ngay"),
    chi_da_dang: bool = Query(True, alias="chi-da-dang"),
    kem_truc_ban: bool = Query(True, alias="kem-truc-ban"),
):
    """Mặc định 3 ngày — giữ đúng cấu hình của lichkv8."""
    bd = tu_ngay or date_type.today()
    svc = LichCongTacService(db)
    return {"success": True,
            "data": await svc.tom_tat(bd, bd + timedelta(days=so_ngay - 1),
                                      chi_da_dang=chi_da_dang,
                                      kem_truc_ban=kem_truc_ban)}


@router.get("/lanh-dao/{lanh_dao_id}", summary="Chương trình công tác của một lãnh đạo")
async def lich_lanh_dao(
    lanh_dao_id: UUID,
    db: DatabaseDep,
    user: CurrentUserDep,
    tu_ngay: Optional[date_type] = Query(None, alias="tu-ngay"),
    so_ngay: int = Query(7, ge=1, le=90, alias="so-ngay"),
):
    bd = tu_ngay or date_type.today()
    svc = LichCongTacService(db)
    return {"success": True,
            "data": await svc.lich_lanh_dao(lanh_dao_id, bd,
                                            bd + timedelta(days=so_ngay - 1))}


@router.get("/thong-ke", summary="Chỉ số tổng quan")
async def thong_ke(db: DatabaseDep, user: CurrentUserDep):
    svc = LichCongTacService(db)
    return {"success": True, "data": await svc.thong_ke()}


@router.get("/danh-muc-don-vi", summary="Danh mục đơn vị (để chọn đơn vị chuẩn bị)")
async def danh_muc_don_vi(db: DatabaseDep, user: CurrentUserDep):
    """Đọc thẳng public.don_vi — chỉ đọc, module này không sở hữu bảng đó."""
    from sqlalchemy import text as sa_text
    rows = (await db.execute(sa_text(
        "SELECT id, ma_don_vi, ten_don_vi FROM public.don_vi "
        " ORDER BY ma_don_vi"))).all()
    return {"success": True,
            "data": [{"id": i, "ma_don_vi": m, "ten_don_vi": t}
                     for i, m, t in rows]}


@router.get("/danh-muc", summary="Danh mục loại lịch")
async def danh_muc(user: CurrentUserDep):
    return {"success": True,
            "data": [{"ma": k, "ten": v} for k, v in NHAN_LOAI_LICH.items()]}


@router.get("/xuat-excel", summary="Xuất lịch ra Excel")
async def xuat_excel(
    db: DatabaseDep,
    user: CurrentUserDep,
    tu_ngay: Optional[date_type] = Query(None, alias="tu-ngay"),
    den_ngay: Optional[date_type] = Query(None, alias="den-ngay"),
    loai_lich: Optional[str] = Query(None, alias="loai-lich"),
    trang_thai: Optional[str] = Query(None, alias="trang-thai"),
    lanh_dao_id: Optional[UUID] = Query(None, alias="lanh-dao-id"),
    tim_kiem: Optional[str] = Query(None, alias="tim-kiem"),
    gioi_han: int = Query(2000, ge=1, le=5000, alias="gioi-han"),
):
    svc = LichCongTacService(db)
    items, _ = await svc.danh_sach(
        trang=1, so_dong=gioi_han, tu_ngay=tu_ngay, den_ngay=den_ngay,
        loai_lich=loai_lich, trang_thai=trang_thai, lanh_dao_id=lanh_dao_id,
        tim_kiem=tim_kiem)

    # Nạp muộn để API JSON không phải trả giá cho openpyxl.
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Lich cong tac"

    cot = ["STT", "Mã lịch", "Ngày", "Giờ", "Loại", "Nội dung", "Thành phần",
           "Địa điểm", "Chủ trì", "Lãnh đạo", "Đơn vị chuẩn bị", "Số văn bản",
           "Ghi chú", "Trạng thái", "Số file"]

    khoang = ""
    if tu_ngay or den_ngay:
        khoang = f" ({tu_ngay or '…'} → {den_ngay or '…'})"
    ws.append([f"LỊCH CÔNG TÁC{khoang}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cot))
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])
    ws.append(cot)

    dong_tieu_de = ws.max_row
    for c in range(1, len(cot) + 1):
        o = ws.cell(row=dong_tieu_de, column=c)
        o.font = Font(bold=True)
        o.fill = PatternFill("solid", fgColor="D9E1F2")
        o.alignment = Alignment(horizontal="center", vertical="center")

    mau_huy = PatternFill("solid", fgColor="F2F2F2")

    for i, d in enumerate(items, start=1):
        ngay = d["ngay_hien_thi"] or d["ngay_hop"]
        # Sự kiện nhiều ngày ghi thành khoảng, đọc bản in mới biết nó kéo dài.
        if d["ngay_ket_thuc"] and d["ngay_ket_thuc"] != ngay:
            ngay_txt = (f"{ngay:%d/%m/%Y} – "
                        f"{d['ngay_ket_thuc']:%d/%m/%Y}")
        else:
            ngay_txt = f"{ngay:%d/%m/%Y}" if ngay else ""

        gio = d["gio_bat_dau"].strftime("%H:%M") if d["gio_bat_dau"] else ""
        if d["gio_ket_thuc"]:
            gio += f" – {d['gio_ket_thuc']:%H:%M}"

        chu_tri = (d["chu_toa"]["ho_ten"] if d.get("chu_toa")
                   else (d.get("chu_tri_text") or ""))

        ws.append([
            i,
            d["ma_lich"] or "",
            ngay_txt,
            gio,
            d["loai_lich_nhan"] or d["loai_lich"] or "",
            d["tieu_de"],
            d.get("thanh_phan_text") or "",
            d["dia_diem"] or "",
            chu_tri,
            ", ".join(x["ho_ten"] for x in d["lanh_dao_lien_quan"]),
            d["don_vi_chuan_bi"] or "",
            d["so_van_ban"] or "",
            d.get("mo_ta") or "",
            d["trang_thai"],
            d["so_tai_lieu"],
        ])
        if d["trang_thai"] == "HUY":
            for c in range(1, len(cot) + 1):
                ws.cell(row=ws.max_row, column=c).fill = mau_huy

    for c, w in enumerate([5, 10, 22, 13, 11, 46, 30, 24, 22, 24, 22, 14, 30,
                           15, 8], start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(row=dong_tieu_de + 1, column=1)

    buf = BytesIO()
    wb.save(buf)
    ten_file = f"lich-cong-tac-{date_type.today():%Y%m%d}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type=("application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"),
        headers={"Content-Disposition": f'attachment; filename="{ten_file}"'},
    )


@router.post("/", status_code=status.HTTP_201_CREATED,
             summary="Tạo sự kiện lịch")
async def tao(du_lieu: LichCongTacCreate, db: DatabaseDep,
              user: CurrentUserDep):
    """Ai truy cập được module cũng tạo được, nhưng chỉ sửa được lịch mình tạo.

    Hệ cũ không phân quyền — 217 người từng tạo lịch — nên chặn tạo mới sẽ làm
    Văn phòng phải nhập hộ cả Chi cục.
    """
    svc = LichCongTacService(db)
    try:
        item = await svc.tao(du_lieu.model_dump(exclude_unset=True),
                             nguoi_id=UUID(user.sub))
    except LoiNghiepVu as e:
        raise _loi(e) from e
    return {"success": True, "data": item, "message": "Đã tạo lịch"}


@router.patch("/{cuoc_hop_id}", summary="Sửa sự kiện lịch")
async def cap_nhat(cuoc_hop_id: UUID, thay_doi: LichCongTacUpdate,
                   db: DatabaseDep, user: CurrentUserDep):
    svc = LichCongTacService(db)
    try:
        item = await svc.cap_nhat(
            cuoc_hop_id, thay_doi.model_dump(exclude_unset=True), user=user)
    except LoiNghiepVu as e:
        raise _loi(e) from e
    return {"success": True, "data": item, "message": "Đã lưu thay đổi"}


@router.post("/{cuoc_hop_id}/huy", summary="Huỷ lịch")
async def huy(cuoc_hop_id: UUID, du_lieu: LichCongTacHuy,
              db: DatabaseDep, user: CurrentUserDep):
    """Huỷ giữ lại bản ghi kèm lý do — không xoá, để còn tra lại được."""
    svc = LichCongTacService(db)
    try:
        item = await svc.huy(cuoc_hop_id, du_lieu.ly_do_huy.strip(), user=user)
    except LoiNghiepVu as e:
        raise _loi(e) from e
    return {"success": True, "data": item, "message": "Đã huỷ lịch"}


@router.delete("/{cuoc_hop_id}", summary="Xoá mềm sự kiện lịch")
async def xoa(cuoc_hop_id: UUID, db: DatabaseDep, user: CurrentUserDep):
    svc = LichCongTacService(db)
    try:
        await svc.xoa(cuoc_hop_id, user=user)
    except LoiNghiepVu as e:
        raise _loi(e) from e
    return {"success": True, "data": None, "message": "Đã xoá lịch"}


@router.get("/{cuoc_hop_id}/nhat-ky", summary="Nhật ký thay đổi của một lịch")
async def nhat_ky(cuoc_hop_id: UUID, db: DatabaseDep, user: CurrentUserDep,
                  gioi_han: int = Query(100, ge=1, le=500, alias="gioi-han")):
    svc = LichCongTacService(db)
    return {"success": True,
            "data": await svc.nhat_ky(cuoc_hop_id, gioi_han)}


@router.get("/quyen/cua-toi", summary="Quyền của tôi trên lịch công tác")
async def quyen_cua_toi(user: CurrentUserDep):
    """Để giao diện biết có hiện nút Sửa/Xoá hay không mà không phải đoán."""
    return {"success": True,
            "data": {"la_quan_tri_lich": la_quan_tri_lich(user),
                     "cong_chuc_id": user.sub}}


@router.get("/{cuoc_hop_id}", summary="Chi tiết một sự kiện")
async def chi_tiet(cuoc_hop_id: UUID, db: DatabaseDep, user: CurrentUserDep):
    svc = LichCongTacService(db)
    item = await svc.chi_tiet(cuoc_hop_id)
    if not item:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            detail={"success": False, "error": {
                "code": "KHONG_TIM_THAY",
                "message": "Không tìm thấy sự kiện trên lịch"}})
    return {"success": True, "data": item}
