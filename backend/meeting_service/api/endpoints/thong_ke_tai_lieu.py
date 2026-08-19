"""
api/endpoints/thong_ke_tai_lieu.py
==================================
Báo cáo Thống kê tài liệu họp — G4.6.

Theo dõi đơn vị được giao chuẩn bị đã nộp tài liệu chưa. Giấy mời thuần KHÔNG
tính là đã nộp; quy tắc phân loại port nguyên văn từ lichkv8, xem
`services/quy_tac_giay_moi.py`.
"""

from datetime import date as date_type
from io import BytesIO
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, HTTPException, Query, Response, status

from meeting_service.dependencies import CurrentUserDep, DatabaseDep
from meeting_service.services.thong_ke_tai_lieu_service import (
    NHAN_TINH_TRANG,
    TAT_CA,
    TINH_TRANG_VALUES,
    ThongKeTaiLieuService,
)

router = APIRouter(prefix="/thong-ke-tai-lieu", tags=["Thống kê tài liệu họp"])


def _tham_so(
    tu_ngay, den_ngay, tu_khoa, lanh_dao_id, trang_thai_lich,
    tinh_trang, tinh_lich_huy, gioi_han,
) -> dict:
    if tinh_trang not in TINH_TRANG_VALUES:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail={"success": False, "error": {
                "code": "TINH_TRANG_KHONG_HOP_LE",
                "message": f"tinh_trang phải thuộc {TINH_TRANG_VALUES}"}})
    return dict(
        tu_ngay=tu_ngay, den_ngay=den_ngay, tu_khoa=tu_khoa,
        lanh_dao_id=lanh_dao_id, trang_thai_lich=trang_thai_lich,
        tinh_trang=tinh_trang, tinh_lich_huy=tinh_lich_huy, gioi_han=gioi_han,
    )


@router.get("/", summary="Báo cáo tình trạng tài liệu theo cuộc họp")
async def bao_cao(
    db: DatabaseDep,
    user: CurrentUserDep,
    tu_ngay: Optional[date_type] = Query(None, alias="tu-ngay"),
    den_ngay: Optional[date_type] = Query(None, alias="den-ngay"),
    tu_khoa: Optional[str] = Query(None, alias="tu-khoa"),
    lanh_dao_id: Optional[UUID] = Query(None, alias="lanh-dao-id"),
    trang_thai_lich: Optional[str] = Query(None, alias="trang-thai-lich"),
    tinh_trang: str = Query(TAT_CA, alias="tinh-trang"),
    tinh_lich_huy: bool = Query(False, alias="tinh-lich-huy"),
    gioi_han: int = Query(500, ge=1, le=2000, alias="gioi-han"),
):
    svc = ThongKeTaiLieuService(db)
    kq = await svc.bao_cao(**_tham_so(
        tu_ngay, den_ngay, tu_khoa, lanh_dao_id, trang_thai_lich,
        tinh_trang, tinh_lich_huy, gioi_han))
    return {"success": True, "data": kq}


@router.get("/tinh-trang", summary="Danh mục tình trạng tài liệu")
async def danh_muc(user: CurrentUserDep):
    return {"success": True,
            "data": [{"ma": k, "ten": v} for k, v in NHAN_TINH_TRANG.items()]}


@router.get("/xuat-excel", summary="Xuất báo cáo ra Excel")
async def xuat_excel(
    db: DatabaseDep,
    user: CurrentUserDep,
    tu_ngay: Optional[date_type] = Query(None, alias="tu-ngay"),
    den_ngay: Optional[date_type] = Query(None, alias="den-ngay"),
    tu_khoa: Optional[str] = Query(None, alias="tu-khoa"),
    lanh_dao_id: Optional[UUID] = Query(None, alias="lanh-dao-id"),
    trang_thai_lich: Optional[str] = Query(None, alias="trang-thai-lich"),
    tinh_trang: str = Query(TAT_CA, alias="tinh-trang"),
    tinh_lich_huy: bool = Query(False, alias="tinh-lich-huy"),
    gioi_han: int = Query(2000, ge=1, le=5000, alias="gioi-han"),
):
    svc = ThongKeTaiLieuService(db)
    kq = await svc.bao_cao(**_tham_so(
        tu_ngay, den_ngay, tu_khoa, lanh_dao_id, trang_thai_lich,
        tinh_trang, tinh_lich_huy, gioi_han))

    # Nạp muộn để API JSON không phải trả giá cho openpyxl.
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Thong ke tai lieu hop"

    cot = ["STT", "Mã lịch", "Ngày", "Giờ", "Nội dung", "Lãnh đạo",
           "Đơn vị chuẩn bị", "Số văn bản", "Tài liệu", "Giấy mời",
           "Tình trạng"]

    khoang = ""
    if tu_ngay or den_ngay:
        khoang = f" ({tu_ngay or '…'} → {den_ngay or '…'})"
    ws.append([f"THỐNG KÊ TÀI LIỆU HỌP{khoang}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cot))
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    th = kq["tong_hop"]
    ws.append([
        f"Tổng {th['tong']} cuộc họp · đã gắn tài liệu {th['DA_GAN_TAI_LIEU']}"
        f" · thiếu tài liệu {th['THIEU_TAI_LIEU']}"
        f" · chưa giao chuẩn bị {th['CHUA_GIAO_CHUAN_BI']}"
    ])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cot))
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.append([])
    ws.append(cot)

    dong_tieu_de = ws.max_row
    to_mau = PatternFill("solid", fgColor="D9E1F2")
    for c in range(1, len(cot) + 1):
        o = ws.cell(row=dong_tieu_de, column=c)
        o.font = Font(bold=True)
        o.fill = to_mau
        o.alignment = Alignment(horizontal="center", vertical="center")

    # Đánh dấu dòng cần chú ý — Văn phòng dùng bản in này để nhắc đơn vị.
    mau_thieu = PatternFill("solid", fgColor="FCE4E4")
    mau_chua_giao = PatternFill("solid", fgColor="FFF2CC")

    for i, d in enumerate(kq["dong"], start=1):
        ws.append([
            i,
            d["ma_lich"] or "",
            d["ngay"].strftime("%d/%m/%Y") if d["ngay"] else "",
            d["gio_bat_dau"].strftime("%H:%M") if d["gio_bat_dau"] else "",
            d["tieu_de"],
            ", ".join(d["lanh_dao"]) or d["chu_tri"],
            d["don_vi_chuan_bi"] or "",
            d["so_van_ban"] or "",
            d["so_tai_lieu_chuan_bi"],
            d["so_giay_moi"],
            d["tinh_trang_nhan"],
        ])
        if d["tinh_trang"] == "THIEU_TAI_LIEU":
            for c in range(1, len(cot) + 1):
                ws.cell(row=ws.max_row, column=c).fill = mau_thieu
        elif d["tinh_trang"] == "CHUA_GIAO_CHUAN_BI":
            for c in range(1, len(cot) + 1):
                ws.cell(row=ws.max_row, column=c).fill = mau_chua_giao

    for c, w in enumerate([5, 10, 11, 7, 52, 26, 22, 16, 9, 9, 18], start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(row=dong_tieu_de + 1, column=1)

    buf = BytesIO()
    wb.save(buf)
    ten_file = f"thong-ke-tai-lieu-hop-{date_type.today():%Y%m%d}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type=("application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"),
        headers={"Content-Disposition": f'attachment; filename="{ten_file}"'},
    )
