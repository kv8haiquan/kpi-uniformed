"""
api/endpoints/truc_ban.py
==========================
Lịch trực ban — G4.7 (ma trận, nhập tay, nộp) và G4.8 (nhập từ Excel).

Nhập Excel đi hai bước: đọc file → **xem trước** → mới ghi. Hệ cũ ghi thẳng,
sai một dòng là hỏng cả bảng và không biết hỏng ở đâu.
"""

from datetime import date as date_type
from datetime import timedelta
from io import BytesIO
from typing import Optional
from uuid import UUID

from fastapi import (
    APIRouter,
    File,
    HTTPException,
    Query,
    Response,
    UploadFile,
    status,
)

from meeting_service.dependencies import CurrentUserDep, DatabaseDep
from meeting_service.schemas.truc_ban import (
    DongNhapExcel,
    KetQuaXemTruoc,
    TrucBanCreate,
    TrucBanNop,
    TrucBanUpdate,
    YeuCauGhiNhap,
)
from meeting_service.services.lich_cong_tac_service import LoiNghiepVu
from meeting_service.services.truc_ban_service import (
    CA_TRUC_VALUES,
    LOAI_TRUC_VALUES,
    TrucBanService,
    bac_chuc_vu,
    chuan_hoa_sdt,
    tuan_chua,
)

router = APIRouter(prefix="/truc-ban", tags=["Trực ban"])


def _loi(e: LoiNghiepVu) -> HTTPException:
    return HTTPException(
        e.http,
        detail={"success": False,
                "error": {"code": e.ma, "message": e.thong_diep}})


def _khoang(tu_ngay: Optional[date_type], den_ngay: Optional[date_type],
            tuan: Optional[str]) -> tuple[date_type, date_type]:
    """Ưu tiên khoảng ngày tự chọn; không có thì lấy theo tuần."""
    if tu_ngay and den_ngay:
        return tu_ngay, den_ngay
    lech = {"truoc": -1, "nay": 0, "sau": 1}.get(tuan or "nay", 0)
    return tuan_chua(date_type.today(), lech)


# ── xem ───────────────────────────────────────────────────────────────

@router.get("/tru-so", summary="Danh mục trụ sở trực ban")
async def danh_muc_tru_so(db: DatabaseDep, user: CurrentUserDep):
    return {"success": True,
            "data": await TrucBanService(db).danh_muc_tru_so()}


@router.get("/nguoi-goi-y/{tru_so_id}",
            summary="Công chức có thể phân trực ở trụ sở này")
async def nguoi_goi_y(
    tru_so_id: UUID,
    db: DatabaseDep,
    user: CurrentUserDep,
    tu_khoa: Optional[str] = Query(None, alias="tu-khoa"),
):
    try:
        ds = await TrucBanService(db).nguoi_goi_y(tru_so_id, tu_khoa=tu_khoa)
    except LoiNghiepVu as e:
        raise _loi(e) from e
    return {"success": True, "data": ds}


@router.get("/ma-tran", summary="Bảng ma trận trực ban")
async def ma_tran(
    db: DatabaseDep,
    user: CurrentUserDep,
    tu_ngay: Optional[date_type] = Query(None, alias="tu-ngay"),
    den_ngay: Optional[date_type] = Query(None, alias="den-ngay"),
    tuan: Optional[str] = Query(None, description="truoc | nay | sau"),
    don_vi_id: Optional[UUID] = Query(None, alias="don-vi-id"),
    chi_cuoi_tuan: bool = Query(True, alias="chi-cuoi-tuan"),
):
    bd, kt = _khoang(tu_ngay, den_ngay, tuan)
    try:
        kq = await TrucBanService(db).ma_tran(bd, kt, user=user,
                                              don_vi_id=don_vi_id,
                                              chi_cuoi_tuan=chi_cuoi_tuan)
    except LoiNghiepVu as e:
        raise _loi(e) from e
    return {"success": True, "data": kq}


@router.get("/danh-sach", summary="Dữ liệu chi tiết dạng bảng phẳng")
async def danh_sach(
    db: DatabaseDep,
    user: CurrentUserDep,
    tu_ngay: Optional[date_type] = Query(None, alias="tu-ngay"),
    den_ngay: Optional[date_type] = Query(None, alias="den-ngay"),
    tuan: Optional[str] = Query(None),
    tru_so_id: Optional[UUID] = Query(None, alias="tru-so-id"),
):
    bd, kt = _khoang(tu_ngay, den_ngay, tuan)
    return {"success": True,
            "data": await TrucBanService(db).danh_sach(bd, kt,
                                                       tru_so_id=tru_so_id)}


@router.get("/van-ban", summary="Bản text để sao chép sang Zalo")
async def van_ban(
    db: DatabaseDep,
    user: CurrentUserDep,
    tu_ngay: Optional[date_type] = Query(None, alias="tu-ngay"),
    den_ngay: Optional[date_type] = Query(None, alias="den-ngay"),
    tuan: Optional[str] = Query(None),
):
    bd, kt = _khoang(tu_ngay, den_ngay, tuan)
    return {"success": True,
            "data": {"van_ban": await TrucBanService(db).van_ban(bd, kt)}}


# ── ghi ───────────────────────────────────────────────────────────────

@router.post("/", status_code=status.HTTP_201_CREATED,
             summary="Thêm một người trực")
async def them(du_lieu: TrucBanCreate, db: DatabaseDep, user: CurrentUserDep):
    try:
        item = await TrucBanService(db).them(
            du_lieu.model_dump(exclude_unset=True), user=user)
    except LoiNghiepVu as e:
        raise _loi(e) from e
    return {"success": True, "data": item, "message": "Đã thêm người trực"}


@router.patch("/{truc_ban_id}", summary="Sửa một người trực")
async def sua(truc_ban_id: UUID, thay_doi: TrucBanUpdate,
              db: DatabaseDep, user: CurrentUserDep):
    try:
        item = await TrucBanService(db).sua(
            truc_ban_id, thay_doi.model_dump(exclude_unset=True), user=user)
    except LoiNghiepVu as e:
        raise _loi(e) from e
    return {"success": True, "data": item, "message": "Đã lưu"}


@router.delete("/{truc_ban_id}", summary="Xoá một người trực")
async def xoa(truc_ban_id: UUID, db: DatabaseDep, user: CurrentUserDep):
    try:
        await TrucBanService(db).xoa(truc_ban_id, user=user)
    except LoiNghiepVu as e:
        raise _loi(e) from e
    return {"success": True, "data": None, "message": "Đã xoá"}


@router.post("/nop", summary="Nộp chính thức một ô (khoá lại)")
async def nop(du_lieu: TrucBanNop, db: DatabaseDep, user: CurrentUserDep):
    try:
        kq = await TrucBanService(db).nop(du_lieu.ngay_truc,
                                          du_lieu.tru_so_id, user=user)
    except LoiNghiepVu as e:
        raise _loi(e) from e
    return {"success": True, "data": kq, "message": "Đã nộp lịch trực"}


@router.post("/mo-khoa", summary="Mở khoá ô đã nộp (chỉ quản trị lịch)")
async def mo_khoa(du_lieu: TrucBanNop, db: DatabaseDep, user: CurrentUserDep):
    try:
        kq = await TrucBanService(db).mo_khoa(du_lieu.ngay_truc,
                                              du_lieu.tru_so_id, user=user)
    except LoiNghiepVu as e:
        raise _loi(e) from e
    return {"success": True, "data": kq, "message": "Đã mở khoá"}


# ── xuất Excel ────────────────────────────────────────────────────────

@router.get("/xuat-excel", summary="Xuất bảng trực ban ra Excel")
async def xuat_excel(
    db: DatabaseDep,
    user: CurrentUserDep,
    tu_ngay: Optional[date_type] = Query(None, alias="tu-ngay"),
    den_ngay: Optional[date_type] = Query(None, alias="den-ngay"),
    tuan: Optional[str] = Query(None),
    chi_cuoi_tuan: bool = Query(True, alias="chi-cuoi-tuan"),
):
    bd, kt = _khoang(tu_ngay, den_ngay, tuan)
    svc = TrucBanService(db)
    try:
        mt = await svc.ma_tran(bd, kt, user=user,
                               chi_cuoi_tuan=chi_cuoi_tuan)
    except LoiNghiepVu as e:
        raise _loi(e) from e

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Truc ban"

    cot = ["Ngày", "Thứ"] + [t["ten_tru_so"] for t in mt["tru_so"]]

    ws.append([f"LỊCH TRỰC BAN ({bd:%d/%m/%Y} – {kt:%d/%m/%Y})"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cot))
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])
    ws.append(cot)

    dong_tieu_de = ws.max_row
    vien = Border(*[Side(style="thin")] * 4)
    for c in range(1, len(cot) + 1):
        o = ws.cell(row=dong_tieu_de, column=c)
        o.font = Font(bold=True)
        o.fill = PatternFill("solid", fgColor="D9E1F2")
        o.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        o.border = vien

    mau_cuoi_tuan = PatternFill("solid", fgColor="FFF7E6")

    for h in mt["hang"]:
        hang = [h["ngay"].strftime("%d/%m/%Y"), h["thu"]]
        for o in h["o"]:
            hang.append("\n".join(
                " · ".join(filter(None, [n["ho_ten"], n["chuc_vu"],
                                         n["so_dien_thoai"]]))
                for n in o["nguoi"]))
        ws.append(hang)
        for c in range(1, len(cot) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = vien
            if h["cuoi_tuan"]:
                cell.fill = mau_cuoi_tuan

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    for i in range(len(mt["tru_so"])):
        ws.column_dimensions[get_column_letter(3 + i)].width = 30
    ws.freeze_panes = ws.cell(row=dong_tieu_de + 1, column=3)

    buf = BytesIO()
    wb.save(buf)
    ten_file = f"truc-ban-{bd:%Y%m%d}-{kt:%Y%m%d}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type=("application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"),
        headers={"Content-Disposition": f'attachment; filename="{ten_file}"'},
    )


# ── nhập từ Excel (G4.8) ──────────────────────────────────────────────

# Tên cột chấp nhận được. Hệ cũ nhận nhiều biến thể vì mỗi đơn vị sửa file mẫu
# một kiểu; giữ đúng độ rộng đó, thêm chuẩn hoá bỏ dấu để khỏi liệt kê vô tận.
_COT = {
    "ngay_truc": ["ngay truc", "ngay", "ngay tr"],
    "ma_tru_so": ["unit code", "ma tru so", "ma don vi", "unitcode"],
    "ma_cc": ["ma cc", "ma cong chuc", "macc", "ma so cong chuc"],
    "ho_ten": ["ho ten", "ho va ten", "nguoi truc"],
    "chuc_vu": ["chuc vu", "chuc danh"],
    "so_dien_thoai": ["so dien thoai", "dien thoai", "sdt", "phone"],
    "ghi_chu": ["ghi chu", "note", "ghichu"],
    "ca_truc": ["ca truc", "ca"],
    "loai_truc": ["loai truc", "loai"],
}


def _chuan_ten_cot(s: object) -> str:
    import re
    import unicodedata
    t = unicodedata.normalize("NFD", str(s or "").lower())
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    t = t.replace("đ", "d")
    return re.sub(r"[^a-z0-9]+", " ", t).strip()


def _doc_ngay(v: object) -> Optional[date_type]:
    """Nhận cả ô ngày thật của Excel lẫn chuỗi dd/mm/yyyy."""
    from datetime import datetime
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date_type):
        return v
    s = str(v).strip()
    for khuon in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, khuon).date()
        except ValueError:
            continue
    return None


@router.get("/mau-import", summary="Tải file Excel mẫu")
async def mau_import(db: DatabaseDep, user: CurrentUserDep):
    """Sinh file mẫu ngay lúc tải, KHÔNG dùng file tĩnh mang từ lichkv8 sang.

    File tĩnh có hai chỗ hỏng theo thời gian: danh mục trụ sở chép cứng nên
    thêm/đổi trụ sở là file mẫu nói sai, và nó bắt gõ tay họ tên + chức vụ +
    số điện thoại cho từng lượt trực — đúng chỗ đẻ ra sai chính tả tên và số
    điện thoại lệch nhau giữa các tuần.

    Bản sinh động chỉ hỏi **mã công chức**; họ tên, chức vụ và số điện thoại
    do hệ thống điền. Kèm hai sheet tra cứu lấy thẳng từ cơ sở dữ liệu và ô
    chọn sẵn (data validation) để không gõ sai mã.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    svc = TrucBanService(db)
    tru_so = await svc.danh_muc_tru_so()
    danh_ba = await svc.danh_ba_cho_mau()

    wb = Workbook()
    dam = Font(bold=True)
    nen = PatternFill("solid", fgColor="D9E1F2")

    def _tieu_de(ws, cot: list[str], rong: list[int]) -> None:
        ws.append(cot)
        for i, r in enumerate(rong, start=1):
            ws.column_dimensions[get_column_letter(i)].width = r
            o = ws.cell(row=1, column=i)
            o.font, o.fill = dam, nen
            o.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"

    # ── sheet 1: nơi nhập ────────────────────────────────────────────
    ws = wb.active
    ws.title = "IMPORT_LICH_TRUC"
    _tieu_de(ws, ["NGAY_TRUC", "UNIT_CODE", "MA_CC", "SO_DIEN_THOAI",
                  "CA_TRUC", "LOAI_TRUC", "GHI_CHU"],
             [14, 14, 14, 16, 12, 14, 30])

    # Dòng mẫu dùng dữ liệu THẬT để đơn vị thấy ngay khuôn đúng: trụ sở đầu
    # danh mục và người ĐƯỢC PHÂN TRỰC ở chính trụ sở đó. Lấy bừa hai mã trong
    # danh bạ toàn Chi cục thì ra mẫu tự mâu thuẫn — mã trụ sở một nơi, người
    # một nơi — và đó là thứ người dùng sẽ chép theo.
    hom_nay = date_type.today()
    thu_bay = hom_nay + timedelta(days=(5 - hom_nay.weekday()) % 7 or 7)
    if tru_so:
        goi_y = await svc.nguoi_goi_y(tru_so[0]["id"], gioi_han=2)
        for ngay in (thu_bay, thu_bay + timedelta(days=1)):
            for g in goi_y:
                ws.append([ngay.strftime("%d/%m/%Y"), tru_so[0]["ma_tru_so"],
                           g["ma_cc"], None, "CA_NGAY", "CUOI_TUAN",
                           "DÒNG MẪU — xoá đi"])

    # ── sheet 2 + 3: tra cứu ─────────────────────────────────────────
    ws_dv = wb.create_sheet("DANH_MUC_DON_VI")
    _tieu_de(ws_dv, ["UNIT_CODE", "TEN_TRU_SO"], [16, 46])
    for t in tru_so:
        ws_dv.append([t["ma_tru_so"], t["ten_tru_so"]])

    ws_cc = wb.create_sheet("DANH_MUC_CONG_CHUC")
    _tieu_de(ws_cc, ["MA_CC", "HO_TEN", "CHUC_VU", "DON_VI",
                     "HE_THONG_DA_CO_SDT"], [14, 26, 22, 40, 20])
    for d in danh_ba:
        ws_cc.append([d["ma_cc"], d["ho_ten"], d["chuc_vu"] or "",
                      d["ten_don_vi"],
                      "Có" if d["he_thong_co_sdt"] else "CHƯA — phải tự điền"])

    # ── ô chọn sẵn ───────────────────────────────────────────────────
    # Excel chỉ nhận tham chiếu sang sheet khác khi tên sheet không có dấu
    # cách — ba sheet ở đây đều đặt tên không dấu cách nên dùng trực tiếp.
    def _gan_chon(cot: str, cong_thuc: str, thong_bao: str) -> None:
        dv = DataValidation(type="list", formula1=cong_thuc, allow_blank=True,
                            showDropDown=False, error=thong_bao,
                            errorTitle="Giá trị không có trong danh mục")
        ws.add_data_validation(dv)
        dv.add(f"{cot}2:{cot}1000")

    if tru_so:
        _gan_chon("B", f"=DANH_MUC_DON_VI!$A$2:$A${len(tru_so) + 1}",
                  "Chọn mã trụ sở trong sheet DANH_MUC_DON_VI")
    if danh_ba:
        _gan_chon("C", f"=DANH_MUC_CONG_CHUC!$A$2:$A${len(danh_ba) + 1}",
                  "Chọn mã công chức trong sheet DANH_MUC_CONG_CHUC")
    _gan_chon("E", '"' + ",".join(CA_TRUC_VALUES) + '"',
              "Ca trực chỉ nhận: " + ", ".join(CA_TRUC_VALUES))
    _gan_chon("F", '"' + ",".join(LOAI_TRUC_VALUES) + '"',
              "Loại trực chỉ nhận: " + ", ".join(LOAI_TRUC_VALUES))

    # ── sheet 4: hướng dẫn ───────────────────────────────────────────
    ws_hd = wb.create_sheet("HUONG_DAN")
    ws_hd.column_dimensions["A"].width = 110
    for dong in [
        "HƯỚNG DẪN NHẬP LỊCH TRỰC BAN — Chi cục Hải quan Khu vực VIII",
        "",
        "1. Nhập tại sheet IMPORT_LICH_TRUC. XOÁ các dòng mẫu trước khi nhập thật.",
        "2. NGAY_TRUC: dạng dd/mm/yyyy, ví dụ 06/06/2026.",
        "3. UNIT_CODE: chọn trong ô sổ xuống (danh mục ở sheet DANH_MUC_DON_VI).",
        "4. MA_CC: chọn trong ô sổ xuống (danh bạ ở sheet DANH_MUC_CONG_CHUC).",
        "   Hệ thống tự điền HỌ TÊN và CHỨC VỤ theo mã — không phải gõ tay nữa.",
        "5. SO_DIEN_THOAI: ĐỂ TRỐNG. Hệ thống lấy số của lượt trực gần nhất.",
        "   Chỉ điền khi cột HE_THONG_DA_CO_SDT ở sheet danh bạ ghi 'CHƯA',",
        "   hoặc khi người đó vừa đổi số. Điền vào là ghi đè số hệ thống biết.",
        "6. CA_TRUC: để trống = CA_NGAY. Các giá trị khác: "
        + ", ".join(CA_TRUC_VALUES[1:]) + ".",
        "7. LOAI_TRUC: để trống = CUOI_TUAN. Trực ngày lễ, Tết chọn LE_TET;",
        "   trực ngày thường chọn NGAY_THUONG.",
        "8. GHI_CHU: không bắt buộc.",
        "9. Không đổi tên cột ở dòng 1.",
        "",
        "Sau khi tải file lên, phần mềm hiện bảng XEM TRƯỚC: kiểm lại họ tên và",
        "số điện thoại hệ thống điền, đúng rồi mới bấm ghi. Dòng nào sai chỉ",
        "báo lỗi dòng đó, không làm hỏng cả file.",
        "",
        f"Danh mục trong file lấy từ hệ thống lúc {hom_nay:%d/%m/%Y} — "
        "tải lại file mẫu khi có người mới hoặc trụ sở mới.",
    ]:
        ws_hd.append([dong])
    ws_hd.cell(row=1, column=1).font = dam

    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type=("application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"),
        headers={"Content-Disposition":
                 'attachment; filename="Mau_import_lich_truc_ban.xlsx"'},
    )


@router.post("/nhap/xem-truoc", summary="Đọc file Excel và xem trước")
async def xem_truoc(db: DatabaseDep, user: CurrentUserDep,
                    file: UploadFile = File(...)):
    """Đọc và kiểm tra, KHÔNG ghi gì.

    Trả về từng dòng kèm lỗi của riêng dòng đó. Một dòng hỏng không làm hỏng
    cả file — người dùng thấy rõ dòng nào sai vì sao rồi tự quyết định.
    """
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail={"success": False, "error": {
                "code": "SAI_DINH_DANG",
                "message": "Chỉ nhận file .xlsx"}})

    noi_dung = await file.read()
    if len(noi_dung) > 10 * 1024 * 1024:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail={"success": False, "error": {
                "code": "FILE_QUA_LON", "message": "File tối đa 10MB"}})

    from openpyxl import load_workbook

    try:
        wb = load_workbook(BytesIO(noi_dung), data_only=True, read_only=True)
    except Exception as e:  # openpyxl ném nhiều loại lỗi khác nhau
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail={"success": False, "error": {
                "code": "KHONG_DOC_DUOC",
                "message": f"Không đọc được file Excel: {e}"}}) from e

    ws = (wb["IMPORT_LICH_TRUC"] if "IMPORT_LICH_TRUC" in wb.sheetnames
          else wb.worksheets[0])

    hang = list(ws.iter_rows(values_only=True))
    if not hang:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail={"success": False, "error": {
                "code": "FILE_RONG", "message": "File không có dữ liệu"}})

    tieu_de = [_chuan_ten_cot(x) for x in hang[0]]
    vi_tri: dict[str, int] = {}
    for khoa, bien_the in _COT.items():
        for i, t in enumerate(tieu_de):
            if t in bien_the:
                vi_tri[khoa] = i
                break

    # Người trực nhận diện bằng MA_CC (file mẫu mới) HOẶC HO_TEN (file cũ các
    # đơn vị còn giữ) — thiếu cả hai mới là thiếu cột bắt buộc.
    thieu = [k for k in ("ngay_truc", "ma_tru_so") if k not in vi_tri]
    if "ma_cc" not in vi_tri and "ho_ten" not in vi_tri:
        thieu.append("ma_cc (hoặc ho_ten)")
    if thieu:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail={"success": False, "error": {
                "code": "THIEU_COT",
                "message": f"File thiếu cột bắt buộc: {', '.join(thieu)}"}})

    svc = TrucBanService(db)
    tru_so = {t["ma_tru_so"].upper(): t for t in await svc.danh_muc_tru_so()}

    def _o(r: tuple, khoa: str) -> object:
        i = vi_tri.get(khoa)
        return r[i] if i is not None and i < len(r) else None

    def _chu(r: tuple, khoa: str) -> str:
        v = _o(r, khoa)
        return "" if v is None else str(v).strip()

    # Tra cả file bằng MỘT câu truy vấn thay vì mỗi dòng một lượt: file cả
    # tháng của 9 trụ sở là hơn trăm dòng, hỏi từng dòng thì chờ thấy rõ.
    ho_so = await svc.tra_cuu_ma_cc(
        [_chu(r, "ma_cc") for r in hang[1:] if _chu(r, "ma_cc")])

    ket_qua: list[DongNhapExcel] = []
    for so_dong, r in enumerate(hang[1:], start=2):
        def o(khoa: str) -> object:
            return _o(r, khoa)

        if all(x is None or str(x).strip() == ""
               for x in (o("ngay_truc"), o("ma_tru_so"),
                         o("ma_cc"), o("ho_ten"))):
            continue  # dòng trống ở cuối file

        loi: list[str] = []
        ngay = _doc_ngay(o("ngay_truc"))
        if ngay is None:
            loi.append("Ngày trực không đọc được (cần dd/mm/yyyy)")

        ma = str(o("ma_tru_so") or "").strip().upper()
        ts = tru_so.get(ma)
        if ts is None:
            loi.append(f"Mã trụ sở '{ma}' không có trong danh mục")

        # Excel hay đổi số điện thoại thành số thực rồi mất số 0 đứng đầu.
        sdt = o("so_dien_thoai")
        sdt_txt = ("" if sdt is None else
                   str(int(sdt)) if isinstance(sdt, float) and sdt.is_integer()
                   else str(sdt)).strip()

        # Mã công chức là đường chính; họ tên gõ tay chỉ còn để đọc được file
        # mẫu cũ các đơn vị đang giữ. Có mã thì hồ sơ hệ thống THẮNG giá trị
        # gõ tay — mã là thứ đã được chọn từ danh mục, tên gõ tay thì không.
        ma_cc = str(o("ma_cc") or "").strip().upper()
        cc = ho_so.get(ma_cc) if ma_cc else None
        cong_chuc_id = None
        ho_ten = str(o("ho_ten") or "").strip()
        chuc_vu = (str(o("chuc_vu")).strip() if o("chuc_vu") else None)

        if ma_cc and cc is None:
            loi.append(f"Mã công chức '{ma_cc}' không có trong danh bạ")
        elif cc is not None:
            if not cc["is_active"]:
                loi.append(f"Công chức '{ma_cc}' đã nghỉ/ngừng hoạt động")
            cong_chuc_id = cc["cong_chuc_id"]
            ho_ten = cc["ho_ten"]
            chuc_vu = cc["chuc_vu"]
            if not sdt_txt:
                sdt_txt = cc["so_dien_thoai"] or ""
                if not sdt_txt:
                    loi.append(
                        f"Hệ thống chưa biết số điện thoại của '{ma_cc}' "
                        "(chưa từng trực, chưa khai số) — điền cột "
                        "SO_DIEN_THOAI")
        elif not ho_ten:
            loi.append("Thiếu mã công chức")

        ca = str(o("ca_truc") or "CA_NGAY").strip().upper() or "CA_NGAY"
        if ca not in CA_TRUC_VALUES:
            # Không chặn ở đây thì dòng qua được xem trước rồi mới vỡ ở bước
            # ghi vì CHECK `ck_truc_ban_ca` — lúc đó lỗi hiện ra là 500.
            loi.append(f"Ca trực '{ca}' không hợp lệ "
                       f"({', '.join(CA_TRUC_VALUES)})")
        loai = str(o("loai_truc") or "CUOI_TUAN").strip().upper() or "CUOI_TUAN"
        if loai not in LOAI_TRUC_VALUES:
            loi.append(f"Loại trực '{loai}' không hợp lệ "
                       f"({', '.join(LOAI_TRUC_VALUES)})")

        ket_qua.append(DongNhapExcel(
            dong=so_dong,
            ngay_truc=ngay,
            ma_tru_so=ma or None,
            tru_so_id=ts["id"] if ts else None,
            ten_tru_so=ts["ten_tru_so"] if ts else None,
            ma_cc=ma_cc or None,
            cong_chuc_id=cong_chuc_id,
            ho_ten=ho_ten or None,
            chuc_vu=chuc_vu,
            so_dien_thoai=chuan_hoa_sdt(sdt_txt),
            ca_truc=ca,
            loai_truc=loai,
            ghi_chu=(str(o("ghi_chu")).strip() if o("ghi_chu") else None),
            hop_le=not loi,
            loi=loi,
        ))

    return {"success": True, "data": KetQuaXemTruoc(
        tong_dong=len(ket_qua),
        so_hop_le=sum(1 for d in ket_qua if d.hop_le),
        so_loi=sum(1 for d in ket_qua if not d.hop_le),
        dong=ket_qua,
    )}


@router.post("/nhap/ghi", summary="Ghi các dòng đã xem trước")
async def ghi_nhap(du_lieu: YeuCauGhiNhap, db: DatabaseDep,
                   user: CurrentUserDep):
    from sqlalchemy import update as sa_update

    from meeting_service.models.lich_cong_tac import TrucBan

    svc = TrucBanService(db)
    hop_le = [d for d in du_lieu.dong if d.hop_le and d.tru_so_id and d.ngay_truc]
    if not hop_le:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail={"success": False, "error": {
                "code": "KHONG_CO_DONG_HOP_LE",
                "message": "Không có dòng nào hợp lệ để ghi"}})

    # Kiểm quyền và khoá TRƯỚC khi ghi dòng nào — ghi được nửa file rồi mới
    # báo lỗi là tình huống tệ nhất, không biết phần nào đã vào.
    o_can_ghi = {(d.ngay_truc, d.tru_so_id) for d in hop_le}
    try:
        for ngay, ts in sorted(o_can_ghi, key=lambda x: (x[0], str(x[1]))):
            await svc._kiem_tra_sua(ts, user)          # noqa: SLF001
            await svc._chan_khi_da_khoa(ngay, ts)      # noqa: SLF001
    except LoiNghiepVu as e:
        raise _loi(e) from e

    if du_lieu.ghi_de:
        for ngay, ts in o_can_ghi:
            await db.execute(
                sa_update(TrucBan)
                .where(TrucBan.ngay_truc == ngay, TrucBan.tru_so_id == ts,
                       TrucBan.is_deleted.is_(False))
                .values(is_deleted=True, updated_by=UUID(user.sub)))

    for d in hop_le:
        db.add(TrucBan(
            ngay_truc=d.ngay_truc, tru_so_id=d.tru_so_id,
            # Trước đây nhập Excel để trống `cong_chuc_id` vì file chỉ có họ
            # tên gõ tay, không nối được về hồ sơ. Nay nhập bằng mã công chức
            # nên nối được — thiếu nó là số điện thoại của lượt sau lại không
            # tra ra được, đúng vòng luẩn quẩn phải gõ tay lại từ đầu.
            cong_chuc_id=d.cong_chuc_id,
            ho_ten=d.ho_ten or "", chuc_vu=d.chuc_vu,
            so_dien_thoai=d.so_dien_thoai,
            ca_truc=d.ca_truc or "CA_NGAY",
            loai_truc=d.loai_truc or "CUOI_TUAN",
            ghi_chu=d.ghi_chu, trang_thai="NHAP",
            created_by=UUID(user.sub), updated_by=UUID(user.sub)))

    from meeting_service.services.audit_log_service import ghi_audit
    await ghi_audit(
        db, hanh_dong="NHAP_EXCEL_TRUC_BAN", nguoi_thuc_hien_id=UUID(user.sub),
        doi_tuong_loai="TRUC_BAN",
        chi_tiet={"so_dong": len(hop_le), "ghi_de": du_lieu.ghi_de,
                  "so_o": len(o_can_ghi)})
    await db.commit()

    return {"success": True,
            "data": {"da_ghi": len(hop_le), "so_o": len(o_can_ghi)},
            "message": f"Đã ghi {len(hop_le)} dòng"}


@router.get("/thu-tu-chuc-vu", summary="Xem thứ tự sắp xếp theo chức vụ")
async def thu_tu_chuc_vu(user: CurrentUserDep,
                         chuc_vu: str = Query(..., alias="chuc-vu")):
    """Công cụ tra nhanh khi thấy thứ tự trong bảng có vẻ sai."""
    return {"success": True,
            "data": {"chuc_vu": chuc_vu, "bac": bac_chuc_vu(chuc_vu)}}
