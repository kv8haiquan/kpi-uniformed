"""
Test bảng điểm danh chi tiết từng thành phần + xuất Excel.

Đối tượng: GET /cuoc-hop/{id}/diem-danh/chi-tiet và .../xuat-excel
(thêm 04/09/2026 — trước đó tab Điểm danh chỉ có 6 ô số tổng hợp).

Điểm cần canh:
- Đi từ meeting.thanh_phan nên người CHƯA điểm danh vẫn phải có dòng
- LEFT JOIN diem_danh + xin_phep_vang KHÔNG được nhân bản dòng
- Cuộc họp đã HUY vẫn XEM được (khác require_can_edit_meeting vốn chặn 409)
- co_the_bam_tay phản chiếu đúng luật của POST /diem-danh/bam-tay
"""

from datetime import date, datetime, timedelta

import pytest
from httpx import AsyncClient
from sqlalchemy import text as sa_text
from sqlalchemy.ext.asyncio import AsyncSession


BASE_CH = "/api/v1/hop-khong-giay/cuoc-hop"
BASE_DD = "/api/v1/hop-khong-giay/diem-danh"

CC2 = "aaaaaaaa-0002-0000-0000-000000000002"
CC3 = "aaaaaaaa-0003-0000-0000-000000000003"
CC4 = "aaaaaaaa-0004-0000-0000-000000000004"


def _payload_hop(don_vi_id, chu_toa_id, thanh_phan):
    """Giờ bắt đầu = now-1m để cửa sổ điểm danh đang mở."""
    gio_now = (datetime.now() - timedelta(minutes=1)).time().strftime("%H:%M")
    return {
        "tieu_de": "Test G3 — Bảng điểm danh chi tiết",
        "khoi": "CHUYEN_MON",
        "hinh_thuc": "TRUC_TIEP",
        "ngay_hop": date.today().isoformat(),
        "gio_bat_dau": gio_now,
        "don_vi_to_chuc_id": str(don_vi_id),
        "chu_toa_id": str(chu_toa_id),
        "thanh_phan": thanh_phan,
    }


async def _tao_hop(client, seed, chu_toa_id, ma_ccs=(CC2, CC3, CC4)):
    resp = await client.post(BASE_CH + "/", json=_payload_hop(
        seed["don_vi_a"], chu_toa_id,
        thanh_phan=[
            {"cong_chuc_id": m, "loai_tham_du": "BAT_BUOC"} for m in ma_ccs
        ],
    ))
    assert resp.status_code in (200, 201), resp.text
    return resp.json()["data"]["id"]


def _url_chi_tiet(ch_id):
    return f"{BASE_CH}/{ch_id}/diem-danh/chi-tiet"


# ══════════════════════════════════════════════════════════════════════
# NỘI DUNG BẢNG
# ══════════════════════════════════════════════════════════════════════

@pytest.mark.asyncio
async def test_chi_tiet_gom_ca_nguoi_chua_diem_danh(
    client: AsyncClient, chu_toa_user, seed_test_users,
):
    """3 thành phần, chấm tay 1 người → vẫn phải trả đủ 3 dòng."""
    ch_id = await _tao_hop(client, seed_test_users, chu_toa_user.sub)

    await client.post(BASE_DD + "/bam-tay", json={
        "cuoc_hop_id": ch_id,
        "diem_danh": [{"cong_chuc_id": CC2, "trang_thai": "CO_MAT"}],
    })

    resp = await client.get(_url_chi_tiet(ch_id))
    assert resp.status_code == 200, resp.text
    data = resp.json()["data"]

    assert len(data["danh_sach"]) == 3
    theo_id = {r["cong_chuc_id"]: r for r in data["danh_sach"]}
    assert theo_id[CC2]["trang_thai"] == "CO_MAT"
    # Người chưa điểm danh: có dòng, trang_thai để trống
    assert theo_id[CC3]["trang_thai"] is None
    assert theo_id[CC3]["gio_diem_danh"] is None


@pytest.mark.asyncio
async def test_chi_tiet_join_ho_ten_don_vi(
    client: AsyncClient, chu_toa_user, seed_test_users,
):
    """Không được trả UUID trơ — phải JOIN sẵn tên người và tên đơn vị."""
    ch_id = await _tao_hop(client, seed_test_users, chu_toa_user.sub, (CC2,))

    resp = await client.get(_url_chi_tiet(ch_id))
    dong = resp.json()["data"]["danh_sach"][0]

    assert dong["ho_ten"] == "Test User TEST-G3-002"
    assert dong["ma_cc"] == "TEST-G3-002"
    assert dong["ten_don_vi"], "thiếu ten_don_vi — LEFT JOIN public.don_vi hỏng"
    assert dong["loai_tham_du"] == "BAT_BUOC"


@pytest.mark.asyncio
async def test_chi_tiet_khong_nhan_ban_dong(
    client: AsyncClient, chu_toa_user, seed_test_users,
    db_session: AsyncSession,
):
    """1 người + 1 bản điểm danh + 1 đơn xin phép → đúng 1 dòng.

    Canh đúng chỗ dễ vỡ nhất của câu SQL: hai LEFT JOIN cùng lúc.
    """
    ch_id = await _tao_hop(client, seed_test_users, chu_toa_user.sub, (CC2,))

    await client.post(BASE_DD + "/bam-tay", json={
        "cuoc_hop_id": ch_id,
        "diem_danh": [{"cong_chuc_id": CC2, "trang_thai": "VANG_CO_PHEP"}],
    })
    await db_session.execute(sa_text("""
        INSERT INTO meeting.xin_phep_vang
            (cuoc_hop_id, cong_chuc_id, ly_do, trang_thai)
        VALUES (:ch, :cc, 'Đi công tác Hà Nội', 'DA_DUYET')
    """), {"ch": ch_id, "cc": CC2})

    resp = await client.get(_url_chi_tiet(ch_id))
    danh_sach = resp.json()["data"]["danh_sach"]
    assert len(danh_sach) == 1, f"dòng bị nhân bản: {len(danh_sach)}"


@pytest.mark.asyncio
async def test_chi_tiet_ly_do_uu_tien_don_xin_phep(
    client: AsyncClient, chu_toa_user, seed_test_users,
    db_session: AsyncSession,
):
    """Có đơn → lấy lý do từ đơn; không đơn → rơi về ghi chú khi chấm tay."""
    ch_id = await _tao_hop(client, seed_test_users, chu_toa_user.sub, (CC2, CC3))

    await client.post(BASE_DD + "/bam-tay", json={
        "cuoc_hop_id": ch_id,
        "diem_danh": [
            {"cong_chuc_id": CC2, "trang_thai": "VANG_CO_PHEP",
             "ghi_chu": "ghi chú của thư ký"},
            {"cong_chuc_id": CC3, "trang_thai": "VANG_KHONG_PHEP",
             "ghi_chu": "không liên lạc được"},
        ],
    })
    # CC2 có thêm đơn xin phép → đơn phải thắng ghi chú
    await db_session.execute(sa_text("""
        INSERT INTO meeting.xin_phep_vang
            (cuoc_hop_id, cong_chuc_id, ly_do, trang_thai)
        VALUES (:ch, :cc, 'Nghỉ phép đã duyệt', 'DA_DUYET')
    """), {"ch": ch_id, "cc": CC2})

    resp = await client.get(_url_chi_tiet(ch_id))
    theo_id = {r["cong_chuc_id"]: r for r in resp.json()["data"]["danh_sach"]}

    assert theo_id[CC2]["ly_do_vang"] == "Nghỉ phép đã duyệt"
    assert theo_id[CC2]["nguon_ly_do"] == "DON_XIN_PHEP"
    assert theo_id[CC2]["xin_phep_trang_thai"] == "DA_DUYET"

    # Không có đơn → dùng ghi chú. Đây là đường THỰC TẾ đang dùng, vì luồng
    # đơn xin phép vắng chưa có ai dùng (prod 0 đơn).
    assert theo_id[CC3]["ly_do_vang"] == "không liên lạc được"
    assert theo_id[CC3]["nguon_ly_do"] == "GHI_CHU_DIEM_DANH"


@pytest.mark.asyncio
async def test_chi_tiet_tong_hop_khop_danh_sach(
    client: AsyncClient, chu_toa_user, seed_test_users,
):
    """6 con số tổng hợp phải đếm ra từ chính danh_sach, không lệch."""
    ch_id = await _tao_hop(client, seed_test_users, chu_toa_user.sub)

    await client.post(BASE_DD + "/bam-tay", json={
        "cuoc_hop_id": ch_id,
        "diem_danh": [
            {"cong_chuc_id": CC2, "trang_thai": "CO_MAT"},
            {"cong_chuc_id": CC3, "trang_thai": "VANG_CO_PHEP"},
        ],
    })

    data = (await client.get(_url_chi_tiet(ch_id))).json()["data"]
    th, ds = data["tong_hop"], data["danh_sach"]

    assert th["tong_so"] == len(ds) == 3
    assert th["co_mat"] == sum(1 for r in ds if r["trang_thai"] == "CO_MAT") == 1
    assert th["vang_co_phep"] == 1
    assert th["chua_diem_danh"] == sum(
        1 for r in ds if r["trang_thai"] is None) == 1
    assert th["den_muon"] == 0 and th["vang_khong_phep"] == 0


# ══════════════════════════════════════════════════════════════════════
# PHÂN QUYỀN
# ══════════════════════════════════════════════════════════════════════

@pytest.mark.asyncio
async def test_chi_tiet_cbcc_thuong_403(
    client: AsyncClient, chu_toa_user, seed_test_users,
):
    """CBCC được mời nhưng không phải ban tổ chức → không xem được cả bảng."""
    ch_id = await _tao_hop(client, seed_test_users, chu_toa_user.sub)

    from meeting_service.dependencies import get_current_user
    from meeting_service.main import app as fastapi_app
    from meeting_service.tests.conftest import _make_user

    cbcc = _make_user("TEST-G3-004", seed_test_users["don_vi_b"])

    async def _override():
        return cbcc
    fastapi_app.dependency_overrides[get_current_user] = _override

    resp = await client.get(_url_chi_tiet(ch_id))
    assert resp.status_code == 403, resp.text
    assert resp.json()["detail"]["error"]["code"] == "NO_PERMISSION"


@pytest.mark.asyncio
async def test_chi_tiet_hop_da_huy_van_xem_duoc(
    client: AsyncClient, chu_toa_user, seed_test_users,
    db_session: AsyncSession,
):
    """Regression: KHÔNG được dùng require_can_edit_meeting cho endpoint đọc.

    Dependency đó raise 409 MEETING_CANCELLED, làm bảng điểm danh của cuộc
    họp đã hủy không tra được nữa.
    """
    ch_id = await _tao_hop(client, seed_test_users, chu_toa_user.sub, (CC2,))
    await db_session.execute(sa_text(
        "UPDATE meeting.cuoc_hop SET trang_thai='HUY' WHERE id=:id"
    ), {"id": ch_id})

    resp = await client.get(_url_chi_tiet(ch_id))
    assert resp.status_code == 200, resp.text
    # Xem được, nhưng không cho chấm nữa
    assert resp.json()["data"]["co_the_bam_tay"] is False


@pytest.mark.asyncio
async def test_chi_tiet_co_the_bam_tay_theo_vai_tro(
    client: AsyncClient, chu_toa_user, seed_test_users,
):
    """Chủ tọa được chấm; CCT xem được bảng nhưng KHÔNG được chấm.

    Theo ma trận docs/HKG/HKG_PLATFORM_ROLES.md — cờ này giữ giao diện trung
    thực với luật của POST /diem-danh/bam-tay, tránh hiện nút rồi ăn 403.
    """
    ch_id = await _tao_hop(client, seed_test_users, chu_toa_user.sub, (CC2,))

    resp = await client.get(_url_chi_tiet(ch_id))
    assert resp.json()["data"]["co_the_bam_tay"] is True

    from meeting_service.dependencies import get_current_user
    from meeting_service.main import app as fastapi_app
    from meeting_service.tests.conftest import _make_user

    cct = _make_user("TEST-G3-004", seed_test_users["don_vi_b"], vai_tro="CCT")

    async def _override():
        return cct
    fastapi_app.dependency_overrides[get_current_user] = _override

    resp = await client.get(_url_chi_tiet(ch_id))
    assert resp.status_code == 200, resp.text
    assert resp.json()["data"]["co_the_bam_tay"] is False
    # nhưng vẫn thấy đủ dữ liệu
    assert len(resp.json()["data"]["danh_sach"]) == 1


# ══════════════════════════════════════════════════════════════════════
# XUẤT EXCEL
# ══════════════════════════════════════════════════════════════════════

@pytest.mark.asyncio
async def test_xuat_excel_tra_ve_xlsx(
    client: AsyncClient, chu_toa_user, seed_test_users,
    db_session: AsyncSession,
):
    ch_id = await _tao_hop(client, seed_test_users, chu_toa_user.sub, (CC2, CC3))
    await client.post(BASE_DD + "/bam-tay", json={
        "cuoc_hop_id": ch_id,
        "diem_danh": [{"cong_chuc_id": CC2, "trang_thai": "CO_MAT"}],
    })

    resp = await client.get(f"{BASE_CH}/{ch_id}/diem-danh/xuat-excel")
    assert resp.status_code == 200, resp.text
    assert "spreadsheetml" in resp.headers["content-type"]
    assert ".xlsx" in resp.headers["content-disposition"]
    # PK.. = chữ ký file zip/xlsx
    assert resp.content[:2] == b"PK"

    # Đọc lại file để chắc nội dung đúng, không chỉ đúng kiểu MIME
    from io import BytesIO
    from openpyxl import load_workbook
    ws = load_workbook(BytesIO(resp.content)).active
    assert "BẢNG ĐIỂM DANH" in str(ws["A1"].value)
    # 2 dòng tiêu đề + 1 dòng trống + 1 header + 2 người
    assert ws.max_row == 6
    ten_trong_file = [ws.cell(row=r, column=2).value for r in (5, 6)]
    assert "Test User TEST-G3-002" in ten_trong_file

    # HKG.txt §849: xuất dữ liệu BẮT BUỘC ghi log
    res = await db_session.execute(sa_text("""
        SELECT COUNT(*) FROM common.audit_log
         WHERE module='MEETING' AND hanh_dong='EXPORT_DIEM_DANH'
           AND doi_tuong_id=:ch_id
    """), {"ch_id": ch_id})
    assert res.scalar() == 1
