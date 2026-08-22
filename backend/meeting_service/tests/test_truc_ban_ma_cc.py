"""Nhập lịch trực bằng MÃ CÔNG CHỨC thay vì gõ tay họ tên.

File mẫu cũ mang từ lichkv8 bắt gõ tay họ tên + chức vụ + số điện thoại cho
TỪNG lượt trực. Hậu quả đo được trên dữ liệu thật: tên viết mỗi tuần một kiểu
(`Nguyễn Quang Chinh` / `Nguyễn Quang Chính`, `Đ/c Kiều Văn Ninh` /
`Kiều Văn Ninh`), và hai người KHÁC NHAU cùng mang một số điện thoại.

Nay chỉ hỏi mã công chức; họ tên, chức vụ, số điện thoại do hệ thống điền.

    DB_NAME=kpi_haiquan_test ALLOW_PROD_TEST=true \
    pytest meeting_service/tests/test_truc_ban_ma_cc.py -v
"""

from __future__ import annotations

from io import BytesIO

import pytest
from httpx import AsyncClient
from sqlalchemy import text as sa_text
from sqlalchemy.ext.asyncio import AsyncSession

pytestmark = pytest.mark.asyncio

BASE = "/api/v1/hop-khong-giay/truc-ban"
COT_MOI = ["NGAY_TRUC", "UNIT_CODE", "MA_CC", "SO_DIEN_THOAI",
           "CA_TRUC", "LOAI_TRUC", "GHI_CHU"]


def _xlsx(hang: list[list], tieu_de: list[str] | None = None) -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "IMPORT_LICH_TRUC"
    ws.append(tieu_de or COT_MOI)
    for h in hang:
        ws.append(h)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _gui(client: AsyncClient, noi_dung: bytes):
    return await client.post(
        f"{BASE}/nhap/xem-truoc",
        files={"file": ("truc.xlsx", noi_dung,
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet")})


@pytest.fixture
async def mot_nguoi(db_session: AsyncSession) -> dict:
    """Một công chức thật kèm số hệ thống tra được, để đối chiếu."""
    r = (await db_session.execute(sa_text("""
        SELECT cc.ma_cc, cc.ho_ten, cc.chuc_vu,
               COALESCE(NULLIF(btrim(z.so_goc), ''),
                        NULLIF(btrim(z.so_dien_thoai), '')) AS sdt
          FROM public.cong_chuc cc
          JOIN common.zalo_lien_ket z ON z.cong_chuc_id = cc.id
         WHERE cc.is_active AND z.so_goc ~ '^0[0-9]{9}$'
         ORDER BY cc.ma_cc LIMIT 1
    """))).first()
    if r is None:
        pytest.skip("chưa có công chức nào kèm số điện thoại")
    return {"ma_cc": r.ma_cc, "ho_ten": r.ho_ten,
            "chuc_vu": r.chuc_vu, "sdt": r.sdt}


# ── file mẫu ──────────────────────────────────────────────────────────

async def test_file_mau_sinh_tu_co_so_du_lieu(client: AsyncClient, admin_user):
    """Danh mục trong file mẫu phải lấy từ CSDL, không chép cứng.

    File tĩnh cũ chép cứng 9 trụ sở — thêm trụ sở là file mẫu nói sai mà
    không ai biết cho tới lúc đơn vị nhập lỗi.
    """
    from openpyxl import load_workbook

    resp = await client.get(f"{BASE}/mau-import")
    assert resp.status_code == 200, resp.text
    wb = load_workbook(BytesIO(resp.content))

    assert set(wb.sheetnames) == {"IMPORT_LICH_TRUC", "DANH_MUC_DON_VI",
                                  "DANH_MUC_CONG_CHUC", "HUONG_DAN"}

    ws = wb["IMPORT_LICH_TRUC"]
    tieu_de = [c.value for c in ws[1]]
    assert tieu_de == COT_MOI
    # Ba cột gõ tay cũ phải BIẾN MẤT khỏi chỗ nhập, nếu không đơn vị vẫn gõ.
    assert "HO_TEN" not in tieu_de and "CHUC_VU" not in tieu_de


async def test_file_mau_co_o_so_xuong(client: AsyncClient, admin_user):
    """Không có ô sổ xuống thì mã công chức lại thành thứ gõ tay — mất sạch
    tác dụng của việc bỏ cột họ tên."""
    from openpyxl import load_workbook

    resp = await client.get(f"{BASE}/mau-import")
    ws = load_workbook(BytesIO(resp.content))["IMPORT_LICH_TRUC"]
    cong_thuc = [dv.formula1 for dv in ws.data_validations.dataValidation]
    assert any("DANH_MUC_DON_VI" in f for f in cong_thuc)
    assert any("DANH_MUC_CONG_CHUC" in f for f in cong_thuc)
    assert any("CA_NGAY" in f for f in cong_thuc)
    assert any("CUOI_TUAN" in f for f in cong_thuc)


async def test_dong_mau_trong_file_mau_phai_hop_le(client: AsyncClient,
                                                   admin_user):
    """Nạp lại chính file mẫu phải ra 0 lỗi.

    Dòng mẫu là thứ người dùng chép theo; mẫu sai là dạy sai. Trước đây mẫu
    ghi mã trụ sở một nơi còn người một nơi.
    """
    resp = await client.get(f"{BASE}/mau-import")
    xt = (await _gui(client, resp.content)).json()["data"]
    assert xt["tong_dong"] > 0, "file mẫu phải có dòng mẫu"
    assert xt["so_loi"] == 0, [d["loi"] for d in xt["dong"] if d["loi"]]


# ── tra mã công chức ──────────────────────────────────────────────────

async def test_ma_cc_tu_dien_ho_ten_chuc_vu_so_dien_thoai(
    client: AsyncClient, admin_user, mot_nguoi,
):
    xt = (await _gui(client, _xlsx([
        ["05/12/2026", "CHICUC", mot_nguoi["ma_cc"], None, None, None, None],
    ]))).json()["data"]

    assert xt["so_loi"] == 0, xt["dong"][0]["loi"]
    d = xt["dong"][0]
    assert d["ho_ten"] == mot_nguoi["ho_ten"]
    assert d["chuc_vu"] == mot_nguoi["chuc_vu"]
    assert d["so_dien_thoai"] == mot_nguoi["sdt"]
    # Nối được về hồ sơ mới là điểm chính: thiếu nó thì lượt trực sau lại
    # không tra ra số, quay về gõ tay.
    assert d["cong_chuc_id"] is not None


async def test_ma_cc_khong_co_thi_bao_loi_dong_do(client: AsyncClient,
                                                  admin_user):
    xt = (await _gui(client, _xlsx([
        ["05/12/2026", "CHICUC", "KHONG-CO-MA-NAY", None, None, None, None],
    ]))).json()["data"]
    assert xt["so_loi"] == 1
    assert any("KHONG-CO-MA-NAY" in x for x in xt["dong"][0]["loi"])


async def test_ho_so_he_thong_thang_gia_tri_go_tay(
    client: AsyncClient, admin_user, mot_nguoi,
):
    """Có mã thì tên gõ tay bị bỏ qua — mã là thứ chọn từ danh mục, tên gõ
    tay thì không. Đây chính là chỗ đẻ ra 'Chinh'/'Chính' trên dữ liệu cũ."""
    xt = (await _gui(client, _xlsx(
        [["05/12/2026", "CHICUC", mot_nguoi["ma_cc"], "Tên Gõ Sai",
          "Chức vụ gõ sai", None, None]],
        tieu_de=["NGAY_TRUC", "UNIT_CODE", "MA_CC", "HO_TEN", "CHUC_VU",
                 "CA_TRUC", "GHI_CHU"]))).json()["data"]
    assert xt["dong"][0]["ho_ten"] == mot_nguoi["ho_ten"]
    assert xt["dong"][0]["chuc_vu"] == mot_nguoi["chuc_vu"]


async def test_so_dien_thoai_dien_tay_van_ghi_de(
    client: AsyncClient, admin_user, mot_nguoi,
):
    """Người vừa đổi số thì đơn vị phải sửa được, không bị hệ thống ép."""
    xt = (await _gui(client, _xlsx([
        ["05/12/2026", "CHICUC", mot_nguoi["ma_cc"], "0987654321",
         None, None, None],
    ]))).json()["data"]
    assert xt["dong"][0]["so_dien_thoai"] == "0987654321"


async def test_file_mau_cu_van_doc_duoc(client: AsyncClient, admin_user):
    """Các đơn vị còn giữ file cũ — đổi mẫu mà chặn file cũ là chặn công việc
    của họ giữa chừng."""
    xt = (await _gui(client, _xlsx(
        [["05/12/2026", "CHICUC", "Nguyễn Văn Cũ", "Công chức", "0912345678",
          None]],
        tieu_de=["NGAY_TRUC", "UNIT_CODE", "HO_TEN", "CHUC_VU",
                 "SO_DIEN_THOAI", "GHI_CHU"]))).json()["data"]
    assert xt["so_loi"] == 0, xt["dong"][0]["loi"]
    assert xt["dong"][0]["ho_ten"] == "Nguyễn Văn Cũ"


async def test_thieu_ca_ma_cc_lan_ho_ten_thi_tu_choi_file(
    client: AsyncClient, admin_user,
):
    resp = await _gui(client, _xlsx(
        [["05/12/2026", "CHICUC", None]],
        tieu_de=["NGAY_TRUC", "UNIT_CODE", "GHI_CHU"]))
    assert resp.status_code == 400
    assert resp.json()["detail"]["error"]["code"] == "THIEU_COT"


# ── ca trực / loại trực ───────────────────────────────────────────────

async def test_de_trong_thi_mac_dinh_ca_ngay_cuoi_tuan(
    client: AsyncClient, admin_user, mot_nguoi,
):
    """100% lịch trực đang chạy là CA_NGAY + CUOI_TUAN nên để trống phải ra
    đúng cặp đó — đơn vị không phải điền gì cho ca thường gặp."""
    xt = (await _gui(client, _xlsx([
        ["05/12/2026", "CHICUC", mot_nguoi["ma_cc"], None, None, None, None],
    ]))).json()["data"]
    assert xt["dong"][0]["ca_truc"] == "CA_NGAY"
    assert xt["dong"][0]["loai_truc"] == "CUOI_TUAN"


@pytest.mark.parametrize("ca,loai", [("SANG", "LE_TET"),
                                     ("DEM", "NGAY_THUONG")])
async def test_nhan_du_cac_gia_tri_hop_le(client: AsyncClient, admin_user,
                                          mot_nguoi, ca, loai):
    xt = (await _gui(client, _xlsx([
        ["05/12/2026", "CHICUC", mot_nguoi["ma_cc"], None, ca, loai, None],
    ]))).json()["data"]
    assert xt["so_loi"] == 0, xt["dong"][0]["loi"]
    assert xt["dong"][0]["ca_truc"] == ca
    assert xt["dong"][0]["loai_truc"] == loai


@pytest.mark.parametrize("ca,loai", [("Cả ngày", "CUOI_TUAN"),
                                     ("CA_NGAY", "Lễ Tết")])
async def test_gia_tri_sai_bi_chan_o_buoc_xem_truoc(
    client: AsyncClient, admin_user, mot_nguoi, ca, loai,
):
    """Không chặn ở xem trước thì dòng lọt tới bước ghi rồi mới vỡ vì CHECK
    `ck_truc_ban_ca` / `ck_truc_ban_loai` — người dùng nhận lỗi 500 sau khi
    đã bấm ghi, không biết phần nào đã vào."""
    xt = (await _gui(client, _xlsx([
        ["05/12/2026", "CHICUC", mot_nguoi["ma_cc"], None, ca, loai, None],
    ]))).json()["data"]
    assert xt["so_loi"] == 1
    assert xt["dong"][0]["hop_le"] is False
