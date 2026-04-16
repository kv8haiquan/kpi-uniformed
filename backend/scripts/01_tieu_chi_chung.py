#!/usr/bin/env python3
"""
01. THỐNG KÊ TIÊU CHÍ CHUNG
============================
Phần 1: Điểm Tiêu chí chung (30 điểm)
- Dưới 20 điểm: danh sách + lý do trừ
- Tròn 20 điểm: danh sách
- Trên 20 điểm: có/không minh chứng đổi mới sáng tạo

Chạy: python 01_tieu_chi_chung.py <thang> <nam>
Output: 01_TieuChiChung_01_2026.xlsx
"""

import sys
import asyncio
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, '/root/kpi-haiquan/backend')

from sqlalchemy import select
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
import os
from dotenv import load_dotenv

load_dotenv('/root/kpi-haiquan/backend/.env')

DB_HOST = os.getenv("DB_HOST", "localhost")
DB_PORT = os.getenv("DB_PORT", "5432")
DB_NAME = os.getenv("DB_NAME", "kpi_haiquan")
DB_USER = os.getenv("DB_USER", "kpi_user")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")

DATABASE_URL = f"postgresql+asyncpg://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"


async def get_data(thang: int, nam: int):
    from app.models.user_org import CongChuc, DonVi
    from app.models.kpi_assessment import DanhGiaThang, TieuChiChung
    
    engine = create_async_engine(DATABASE_URL)
    async with AsyncSession(engine) as db:
        # Lấy tất cả tiêu chí chung
        stmt_tc = select(TieuChiChung).where(TieuChiChung.is_active == True).order_by(TieuChiChung.ma_tieu_chi)
        result_tc = await db.execute(stmt_tc)
        all_tieu_chi = result_tc.scalars().all()
        
        # Lấy đánh giá tháng
        stmt = (
            select(DanhGiaThang)
            .options(
                selectinload(DanhGiaThang.cong_chuc).selectinload(CongChuc.don_vi),
                selectinload(DanhGiaThang.tieu_chi_chungs),
            )
            .where(DanhGiaThang.thang == thang, DanhGiaThang.nam == nam)
        )
        result = await db.execute(stmt)
        danh_gias = result.scalars().all()
        
        cong_chuc_list = []
        
        for dg in danh_gias:
            if not dg.cong_chuc:
                continue
            if hasattr(dg.cong_chuc, 'is_active') and dg.cong_chuc.is_active == False:
                continue
            if hasattr(dg.cong_chuc, 'deleted_at') and dg.cong_chuc.deleted_at is not None:
                continue
            
            tc_danh_gia_map = {str(tcdg.tieu_chi_id): tcdg for tcdg in (dg.tieu_chi_chungs or [])}
            
            tong_diem = 0
            diem_nhom1 = 0
            diem_nhom2 = 0
            diem_nhom3 = 0
            has_nhom3 = False
            has_ghi_chu_nhom3 = False
            ly_do_tru_diem = []
            ghi_chu_nhom3_list = []
            
            for tc in all_tieu_chi:
                tcdg = tc_danh_gia_map.get(str(tc.id))
                
                is_achieved_ld = tcdg.is_achieved_ld if tcdg else None
                is_achieved_cc = tcdg.is_achieved_cc if tcdg else False
                final_achieved = is_achieved_ld if is_achieved_ld is not None else is_achieved_cc
                
                diem_ld = float(tcdg.diem_phe_duyet) if tcdg and tcdg.diem_phe_duyet is not None else None
                diem_cc = float(tcdg.diem_tu_cham) if tcdg and tcdg.diem_tu_cham else 0
                diem = diem_ld if diem_ld is not None else diem_cc
                
                ghi_chu = tcdg.ghi_chu_cc if tcdg else ""
                ghi_chu_ld = tcdg.ghi_chu_ld if tcdg and hasattr(tcdg, 'ghi_chu_ld') else ""
                ly_do_dieu_chinh = tcdg.ly_do_dieu_chinh if tcdg and hasattr(tcdg, 'ly_do_dieu_chinh') else ""
                
                tong_diem += diem
                
                if tc.nhom_tieu_chi == 1:
                    diem_nhom1 += diem
                elif tc.nhom_tieu_chi == 2:
                    diem_nhom2 += diem
                elif tc.nhom_tieu_chi == 3:
                    diem_nhom3 += diem
                    if final_achieved:
                        has_nhom3 = True
                        if ghi_chu:
                            has_ghi_chu_nhom3 = True
                            ghi_chu_nhom3_list.append(f"[{tc.ma_tieu_chi}] {ghi_chu}")
                
                if not final_achieved and float(tc.diem_toi_da) > 0:
                    ly_do = ly_do_dieu_chinh or ghi_chu_ld or f"Không đạt tiêu chí {tc.ma_tieu_chi}"
                    ly_do_tru_diem.append({
                        "ma": tc.ma_tieu_chi,
                        "ten": tc.ten_tieu_chi,
                        "diem_tru": float(tc.diem_toi_da),
                        "ly_do": ly_do,
                        "nhom": tc.nhom_tieu_chi,
                    })
            
            cong_chuc_list.append({
                "ho_ten": dg.cong_chuc.ho_ten,
                "ma_cc": dg.cong_chuc.ma_cc,
                "don_vi": dg.cong_chuc.don_vi.ten_don_vi if dg.cong_chuc.don_vi else "",
                "tong_diem": tong_diem,
                "diem_nhom1": diem_nhom1,
                "diem_nhom2": diem_nhom2,
                "diem_nhom3": diem_nhom3,
                "has_nhom3": has_nhom3,
                "has_ghi_chu_nhom3": has_ghi_chu_nhom3,
                "minh_chung_nhom3": "\n".join(ghi_chu_nhom3_list) if ghi_chu_nhom3_list else "",
                "ly_do_tru_diem": ly_do_tru_diem,
            })
        
        return cong_chuc_list


def create_excel(data: list, thang: int, nam: int, output_path: str):
    wb = Workbook()
    
    # Styles
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    percent_font = Font(bold=True, color="0070C0")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    total = len(data)
    duoi_20 = [cc for cc in data if cc["tong_diem"] < 20]
    tron_20 = [cc for cc in data if cc["tong_diem"] == 20]
    tren_20 = [cc for cc in data if cc["tong_diem"] > 20]
    
    tren_20_co_nhom3 = [cc for cc in tren_20 if cc["has_nhom3"] and cc["has_ghi_chu_nhom3"]]
    tren_20_khong_nhom3 = [cc for cc in tren_20 if cc["has_nhom3"] and not cc["has_ghi_chu_nhom3"]]
    tren_20_30diem = [cc for cc in tren_20 if cc["tong_diem"] == 30]
    
    def pct(count):
        return f"{count/total*100:.1f}%" if total > 0 else "0%"
    
    # =========================================================================
    # SHEET 1: TỔNG HỢP
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Tổng hợp"
    
    ws1['A1'] = f"1. THỐNG KÊ TIÊU CHÍ CHUNG - THÁNG {thang}/{nam}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:E1')
    
    ws1['A2'] = f"Tổng số công chức: {total}"
    ws1['A2'].font = Font(bold=True)
    
    row = 4
    headers = ["Nhóm", "Số lượng", "Tỷ lệ", "Ghi chú"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    rows_data = [
        ("1. Dưới 20 điểm", len(duoi_20), pct(len(duoi_20)), "Xem Sheet 'Dưới 20 điểm'"),
        ("2. Tròn 20 điểm", len(tron_20), pct(len(tron_20)), "Hoàn thành tốt, không sai sót"),
        ("3. Trên 20 điểm", len(tren_20), pct(len(tren_20)), "Chi tiết bên dưới"),
        ("   3a. CÓ sản phẩm đổi mới", len(tren_20_co_nhom3), pct(len(tren_20_co_nhom3)), "Có minh chứng cụ thể"),
        ("   3b. KHÔNG CÓ sản phẩm đổi mới", len(tren_20_khong_nhom3), pct(len(tren_20_khong_nhom3)), "Chưa hiểu rõ quy định"),
        ("   3c. Đạt 30 điểm tối đa", len(tren_20_30diem), pct(len(tren_20_30diem)), "Xuất sắc"),
    ]
    
    for i, (nhom, sl, tl, gc) in enumerate(rows_data):
        r = row + 1 + i
        ws1.cell(row=r, column=1, value=nhom).border = border
        ws1.cell(row=r, column=2, value=sl).border = border
        ws1.cell(row=r, column=2).alignment = center_alignment
        ws1.cell(row=r, column=3, value=tl).border = border
        ws1.cell(row=r, column=3).alignment = center_alignment
        ws1.cell(row=r, column=3).font = percent_font
        ws1.cell(row=r, column=4, value=gc).border = border
        if i < 3:
            ws1.cell(row=r, column=1).font = Font(bold=True)
    
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 40
    
    # =========================================================================
    # SHEET 2: DƯỚI 20 ĐIỂM
    # =========================================================================
    ws2 = wb.create_sheet("Dưới 20 điểm")
    
    ws2['A1'] = f"DANH SÁCH CÔNG CHỨC DƯỚI 20 ĐIỂM - THÁNG {thang}/{nam}"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:G1')
    
    headers2 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "Tổng điểm", "Điểm trừ", "Lý do trừ điểm"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    for i, cc in enumerate(duoi_20, 1):
        r = 3 + i
        ws2.cell(row=r, column=1, value=i).border = border
        ws2.cell(row=r, column=2, value=cc["ho_ten"]).border = border
        ws2.cell(row=r, column=3, value=cc["ma_cc"]).border = border
        ws2.cell(row=r, column=4, value=cc["don_vi"]).border = border
        ws2.cell(row=r, column=5, value=cc["tong_diem"]).border = border
        
        diem_tru = sum(ly["diem_tru"] for ly in cc["ly_do_tru_diem"])
        ws2.cell(row=r, column=6, value=diem_tru).border = border
        
        ly_do_text = "; ".join([f"{ly['ma']}: {ly['ly_do']}" for ly in cc["ly_do_tru_diem"]])
        ws2.cell(row=r, column=7, value=ly_do_text).border = border
        ws2.cell(row=r, column=7).alignment = wrap_alignment
        ws2.row_dimensions[r].height = 40
    
    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 30
    ws2.column_dimensions['E'].width = 10
    ws2.column_dimensions['F'].width = 10
    ws2.column_dimensions['G'].width = 80
    
    # =========================================================================
    # SHEET 3: TRÊN 20 ĐIỂM
    # =========================================================================
    ws3 = wb.create_sheet("Trên 20 điểm")
    
    ws3['A1'] = f"DANH SÁCH CÔNG CHỨC TRÊN 20 ĐIỂM - THÁNG {thang}/{nam}"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:H1')
    
    headers3 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "Tổng điểm", "Điểm nhóm III", "Có MC", "Minh chứng đổi mới sáng tạo"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    tren_20_sorted = sorted(tren_20, key=lambda x: x["tong_diem"], reverse=True)
    
    for i, cc in enumerate(tren_20_sorted, 1):
        r = 3 + i
        ws3.cell(row=r, column=1, value=i).border = border
        ws3.cell(row=r, column=2, value=cc["ho_ten"]).border = border
        ws3.cell(row=r, column=3, value=cc["ma_cc"]).border = border
        ws3.cell(row=r, column=4, value=cc["don_vi"]).border = border
        ws3.cell(row=r, column=5, value=cc["tong_diem"]).border = border
        ws3.cell(row=r, column=6, value=cc["diem_nhom3"]).border = border
        
        if cc["has_ghi_chu_nhom3"]:
            co_mc = "Có"
            ws3.cell(row=r, column=7).fill = PatternFill("solid", fgColor="C6EFCE")
        else:
            co_mc = "Không"
            ws3.cell(row=r, column=7).fill = PatternFill("solid", fgColor="FFC7CE")
        ws3.cell(row=r, column=7, value=co_mc).border = border
        ws3.cell(row=r, column=7).alignment = center_alignment
        
        minh_chung = cc.get("minh_chung_nhom3", "")
        if not minh_chung:
            if cc["tong_diem"] == 30:
                minh_chung = "(Đạt điểm tối đa - chưa điền minh chứng)"
            else:
                minh_chung = "(Chưa điền minh chứng cụ thể)"
        ws3.cell(row=r, column=8, value=minh_chung).border = border
        ws3.cell(row=r, column=8).alignment = wrap_alignment
        ws3.row_dimensions[r].height = 60
    
    ws3.column_dimensions['A'].width = 5
    ws3.column_dimensions['B'].width = 25
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 25
    ws3.column_dimensions['E'].width = 10
    ws3.column_dimensions['F'].width = 12
    ws3.column_dimensions['G'].width = 8
    ws3.column_dimensions['H'].width = 100
    
    # =========================================================================
    # SHEET 4: TRÒN 20 ĐIỂM
    # =========================================================================
    ws4 = wb.create_sheet("Tròn 20 điểm")
    
    ws4['A1'] = f"DANH SÁCH CÔNG CHỨC TRÒN 20 ĐIỂM - THÁNG {thang}/{nam}"
    ws4['A1'].font = title_font
    ws4.merge_cells('A1:E1')
    
    ws4['A2'] = "Ghi chú: Hoàn thành tốt nhiệm vụ, không có sai sót."
    ws4['A2'].font = Font(italic=True, color="666666")
    
    headers4 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "Tổng điểm"]
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    for i, cc in enumerate(tron_20, 1):
        r = 4 + i
        ws4.cell(row=r, column=1, value=i).border = border
        ws4.cell(row=r, column=2, value=cc["ho_ten"]).border = border
        ws4.cell(row=r, column=3, value=cc["ma_cc"]).border = border
        ws4.cell(row=r, column=4, value=cc["don_vi"]).border = border
        ws4.cell(row=r, column=5, value=cc["tong_diem"]).border = border
    
    ws4.column_dimensions['A'].width = 5
    ws4.column_dimensions['B'].width = 25
    ws4.column_dimensions['C'].width = 12
    ws4.column_dimensions['D'].width = 35
    ws4.column_dimensions['E'].width = 12
    
    # Save
    wb.save(output_path)
    print(f"Đã tạo file: {output_path}")
    print(f"\n=== 1. TIÊU CHÍ CHUNG (30 điểm) ===")
    print(f"Tổng số công chức: {total}")
    print(f"- Dưới 20 điểm: {len(duoi_20)} ({pct(len(duoi_20))})")
    print(f"- Tròn 20 điểm: {len(tron_20)} ({pct(len(tron_20))})")
    print(f"- Trên 20 điểm: {len(tren_20)} ({pct(len(tren_20))})")
    print(f"  + Có minh chứng: {len(tren_20_co_nhom3)}")
    print(f"  + Không có minh chứng: {len(tren_20_khong_nhom3)}")
    print(f"  + Đạt 30 điểm: {len(tren_20_30diem)}")


async def main():
    if len(sys.argv) < 3:
        print("Usage: python 01_tieu_chi_chung.py <thang> <nam>")
        sys.exit(1)
    
    thang = int(sys.argv[1])
    nam = int(sys.argv[2])
    
    print(f"Đang lấy dữ liệu tháng {thang}/{nam}...")
    data = await get_data(thang, nam)
    
    if not data:
        print(f"Không có dữ liệu cho tháng {thang}/{nam}")
        sys.exit(1)
    
    output_path = f"01_TieuChiChung_{thang:02d}_{nam}.xlsx"
    create_excel(data, thang, nam, output_path)


if __name__ == "__main__":
    import warnings
    warnings.filterwarnings("ignore")
    
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(main())
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        try:
            loop.run_until_complete(loop.shutdown_asyncgens())
        except:
            pass
        loop.close()
