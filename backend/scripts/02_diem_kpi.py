#!/usr/bin/env python3
"""
02. THỐNG KÊ ĐIỂM KPI
======================
Phần 2: Điểm KPI (70 điểm)
- Đạt 70 điểm KPI: danh sách, trong đó vượt KPI bất thường
- Chưa đạt 70 điểm: do SP chưa đạt / do CL bị trừ / do TĐ bị trừ

Chạy: python 02_diem_kpi.py <thang> <nam>
Output: 02_DiemKPI_01_2026.xlsx
"""

import sys
import asyncio
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, '/root/kpi-haiquan/backend')

from sqlalchemy import select, text
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
import os
from dotenv import load_dotenv

load_dotenv('/root/kpi-haiquan/backend/.env')

DB_HOST = os.getenv("DB_HOST", "localhost")
DB_PORT = os.getenv("DB_PORT", "5432")
DB_NAME = os.getenv("DB_NAME", "kpi_haiquan")
DB_USER = os.getenv("DB_USER", "kpi_user")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")

DATABASE_URL = f"postgresql+asyncpg://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"


async def get_data(thang: int, nam: int):
    from app.models.user_org import CongChuc, DonVi
    
    engine = create_async_engine(DATABASE_URL)
    async with AsyncSession(engine) as db:
        so_ngay_trong_thang = calendar.monthrange(nam, thang)[1]
        
        # Load KPI từ chi_tiet_xep_loai
        kpi_result = await db.execute(text("""
            SELECT ct.cong_chuc_id::text, ct.diem_kpi, ct.diem_tong,
                   ct.is_lanh_dao, ct.xep_loai_he_thong,
                   ct.so_ngay_lam_viec, ct.so_ngay_nghi,
                   cc.ho_ten, cc.ma_cc, dv.ten_don_vi
            FROM chi_tiet_xep_loai ct
            JOIN bao_cao_xep_loai bc ON bc.id = ct.bao_cao_id
            JOIN cong_chuc cc ON cc.id = ct.cong_chuc_id
            LEFT JOIN don_vi dv ON dv.id = cc.don_vi_id
            WHERE bc.thang = :thang AND bc.nam = :nam
                  AND cc.is_active = true
        """), {"thang": thang, "nam": nam})
        
        kpi_list = []
        for row in kpi_result:
            kpi_list.append({
                "cong_chuc_id": row[0],
                "diem_kpi_70": float(row[1]) if row[1] is not None else None,
                "diem_tong_100": float(row[2]) if row[2] is not None else None,
                "is_lanh_dao": row[3],
                "xep_loai": row[4],
                "so_ngay_lv": float(row[5]) if row[5] is not None else None,
                "so_ngay_nghi": float(row[6]) if row[6] is not None else None,
                "ho_ten": row[7],
                "ma_cc": row[8],
                "don_vi": row[9] or "",
            })
        
        # Load nghỉ phép
        nghi_result = await db.execute(text("""
            SELECT cong_chuc_id::text, COALESCE(SUM(so_ngay), 0) as tong_nghi
            FROM dang_ky_nghi
            WHERE thang_ap_dung = :thang AND nam_ap_dung = :nam 
                  AND trang_thai = 'DA_PHE_DUYET' AND is_deleted = false
            GROUP BY cong_chuc_id
        """), {"thang": thang, "nam": nam})
        nghi_by_cc = {row[0]: float(row[1]) for row in nghi_result}
        
        # Load SP quy đổi CC
        sp_result = await db.execute(text("""
            SELECT cong_chuc_id::text,
                   COALESCE(SUM(so_sp_goc_quy_doi), 0) as tong_sp,
                   COALESCE(SUM(so_sp_goc_quy_doi * GREATEST(0, 1 - COALESCE(so_loi_chat_luong, 0) * 0.25)), 0) as sp_cl,
                   COALESCE(SUM(so_sp_goc_quy_doi * GREATEST(0, 1 - COALESCE(so_loi_tien_do, 0) * 0.25)), 0) as sp_td,
                   SUM(so_loi_chat_luong) as loi_cl,
                   SUM(so_loi_tien_do) as loi_td
            FROM ke_khai_cong_viec
            WHERE thang = :thang AND nam = :nam 
                  AND trang_thai = 'DA_PHE_DUYET' AND is_deleted = false
            GROUP BY cong_chuc_id
        """), {"thang": thang, "nam": nam})
        sp_by_cc = {}
        for row in sp_result:
            sp_by_cc[row[0]] = {
                "tong_sp": float(row[1]),
                "sp_cl": float(row[2]),
                "sp_td": float(row[3]),
                "loi_cl": int(row[4] or 0),
                "loi_td": int(row[5] or 0),
            }
        
        # Tính toán cho từng CC
        for cc in kpi_list:
            cc_id = cc["cong_chuc_id"]
            tong_nghi = nghi_by_cc.get(cc_id, 0)
            sp_data = sp_by_cc.get(cc_id, {"tong_sp": 0, "sp_cl": 0, "sp_td": 0, "loi_cl": 0, "loi_td": 0})
            
            # SP được giao = (ngày trong tháng - nghỉ) × 96
            sp_duoc_giao = (so_ngay_trong_thang - tong_nghi) * 96
            cc["sp_duoc_giao"] = sp_duoc_giao
            cc["sp_hoan_thanh"] = sp_data["tong_sp"]
            cc["sp_cl"] = sp_data["sp_cl"]
            cc["sp_td"] = sp_data["sp_td"]
            cc["loi_cl"] = sp_data["loi_cl"]
            cc["loi_td"] = sp_data["loi_td"]
            
            # Tỷ lệ vượt KPI
            if sp_duoc_giao > 0:
                cc["ty_le_vuot"] = (sp_data["tong_sp"] - sp_duoc_giao) / sp_duoc_giao * 100
            else:
                cc["ty_le_vuot"] = 0
        
        return kpi_list, so_ngay_trong_thang


def create_excel(data: list, thang: int, nam: int, so_ngay: int, output_path: str):
    wb = Workbook()
    
    # Styles
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    percent_font = Font(bold=True, color="0070C0")
    alert_fill = PatternFill("solid", fgColor="FFC7CE")
    good_fill = PatternFill("solid", fgColor="C6EFCE")
    warn_fill = PatternFill("solid", fgColor="FFEB9C")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Phân loại
    total = len(data)
    total_co_kpi = len([cc for cc in data if cc["diem_kpi_70"] is not None])
    
    # Chỉ tính CC không phải lãnh đạo (is_lanh_dao = False hoặc None)
    cc_list = [cc for cc in data if not cc.get("is_lanh_dao")]
    
    dat_kpi_70 = [cc for cc in cc_list if cc["diem_kpi_70"] is not None and cc["diem_kpi_70"] >= 70]
    chua_dat_kpi_70 = [cc for cc in cc_list if cc["diem_kpi_70"] is not None and cc["diem_kpi_70"] < 70]
    
    # Vượt KPI bất thường (vượt > 50%)
    vuot_kpi_bat_thuong = [cc for cc in dat_kpi_70 if cc["ty_le_vuot"] > 50]
    
    # Phân loại lý do chưa đạt
    chua_dat_do_so_luong = [cc for cc in chua_dat_kpi_70 if cc["sp_hoan_thanh"] < cc["sp_duoc_giao"]]
    chua_dat_do_chat_luong = [cc for cc in chua_dat_kpi_70 if cc["loi_cl"] > 0]
    chua_dat_do_tien_do = [cc for cc in chua_dat_kpi_70 if cc["loi_td"] > 0]
    
    def pct(count, base=total_co_kpi):
        return f"{count/base*100:.1f}%" if base > 0 else "0%"
    
    # =========================================================================
    # SHEET 1: TỔNG HỢP
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Tổng hợp"
    
    ws1['A1'] = f"2. THỐNG KÊ ĐIỂM KPI - THÁNG {thang}/{nam}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:E1')
    
    ws1['A2'] = f"Tổng CC có dữ liệu KPI: {total_co_kpi} | Số ngày trong tháng: {so_ngay}"
    ws1['A2'].font = Font(bold=True)
    
    row = 4
    headers = ["Nhóm", "Số lượng", "Tỷ lệ", "Ghi chú"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    rows_data = [
        ("I. Đạt KPI 70 điểm", len(dat_kpi_70), pct(len(dat_kpi_70)), ""),
        ("   - Vượt KPI bất thường (>50%)", len(vuot_kpi_bat_thuong), pct(len(vuot_kpi_bat_thuong)), "Cần xem xét cấp độ phức tạp"),
        ("II. Chưa đạt KPI 70 điểm", len(chua_dat_kpi_70), pct(len(chua_dat_kpi_70)), ""),
        ("   - Do SP chưa đạt", len(chua_dat_do_so_luong), pct(len(chua_dat_do_so_luong)), "SP hoàn thành < SP được giao"),
        ("   - Do CL bị trừ", len(chua_dat_do_chat_luong), pct(len(chua_dat_do_chat_luong)), "Có lỗi chất lượng"),
        ("   - Do TĐ bị trừ", len(chua_dat_do_tien_do), pct(len(chua_dat_do_tien_do)), "Có lỗi tiến độ"),
    ]
    
    for i, (nhom, sl, tl, gc) in enumerate(rows_data):
        r = row + 1 + i
        ws1.cell(row=r, column=1, value=nhom).border = border
        ws1.cell(row=r, column=2, value=sl).border = border
        ws1.cell(row=r, column=2).alignment = center_alignment
        ws1.cell(row=r, column=3, value=tl).border = border
        ws1.cell(row=r, column=3).alignment = center_alignment
        ws1.cell(row=r, column=3).font = percent_font
        ws1.cell(row=r, column=4, value=gc).border = border
        if i in [0, 2]:
            ws1.cell(row=r, column=1).font = Font(bold=True)
    
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 40
    
    # =========================================================================
    # SHEET 2: VƯỢT KPI BẤT THƯỜNG
    # =========================================================================
    ws2 = wb.create_sheet("Vượt KPI bất thường")
    
    ws2['A1'] = f"DANH SÁCH VƯỢT KPI BẤT THƯỜNG (>50%) - THÁNG {thang}/{nam}"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:I1')
    
    ws2['A2'] = "Ghi chú: Cần xem xét việc kê khai cấp độ phức tạp chưa chính xác."
    ws2['A2'].font = Font(italic=True, color="C00000")
    
    headers2 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "SP giao", "SP hoàn thành", "Tỷ lệ vượt", "Điểm KPI", "Ghi chú"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    vuot_sorted = sorted(vuot_kpi_bat_thuong, key=lambda x: x["ty_le_vuot"], reverse=True)
    
    for i, cc in enumerate(vuot_sorted, 1):
        r = 4 + i
        ws2.cell(row=r, column=1, value=i).border = border
        ws2.cell(row=r, column=2, value=cc["ho_ten"]).border = border
        ws2.cell(row=r, column=3, value=cc["ma_cc"]).border = border
        ws2.cell(row=r, column=4, value=cc["don_vi"]).border = border
        ws2.cell(row=r, column=5, value=f"{cc['sp_duoc_giao']:,.0f}").border = border
        ws2.cell(row=r, column=6, value=f"{cc['sp_hoan_thanh']:,.0f}").border = border
        
        ty_le_cell = ws2.cell(row=r, column=7, value=f"{cc['ty_le_vuot']:.1f}%")
        ty_le_cell.border = border
        ty_le_cell.alignment = center_alignment
        if cc["ty_le_vuot"] > 100:
            ty_le_cell.fill = alert_fill
            ty_le_cell.font = Font(bold=True, color="9C0006")
        elif cc["ty_le_vuot"] > 50:
            ty_le_cell.fill = warn_fill
        
        ws2.cell(row=r, column=8, value=cc["diem_kpi_70"]).border = border
        ws2.cell(row=r, column=9, value="Cần xác minh cấp độ").border = border
    
    for c, w in [('A', 5), ('B', 25), ('C', 12), ('D', 25), ('E', 12), ('F', 15), ('G', 12), ('H', 10), ('I', 25)]:
        ws2.column_dimensions[c].width = w
    
    # =========================================================================
    # SHEET 3: CHƯA ĐẠT KPI
    # =========================================================================
    ws3 = wb.create_sheet("Chưa đạt KPI")
    
    ws3['A1'] = f"DANH SÁCH CHƯA ĐẠT KPI 70 ĐIỂM - THÁNG {thang}/{nam}"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:K1')
    
    headers3 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "SP giao", "SP HT", "Lỗi CL", "Lỗi TĐ", "Điểm KPI", "Lý do chính"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    chua_dat_sorted = sorted(chua_dat_kpi_70, key=lambda x: x["diem_kpi_70"] or 0)
    
    for i, cc in enumerate(chua_dat_sorted, 1):
        r = 3 + i
        ws3.cell(row=r, column=1, value=i).border = border
        ws3.cell(row=r, column=2, value=cc["ho_ten"]).border = border
        ws3.cell(row=r, column=3, value=cc["ma_cc"]).border = border
        ws3.cell(row=r, column=4, value=cc["don_vi"]).border = border
        ws3.cell(row=r, column=5, value=f"{cc['sp_duoc_giao']:,.0f}").border = border
        ws3.cell(row=r, column=6, value=f"{cc['sp_hoan_thanh']:,.0f}").border = border
        
        loi_cl_cell = ws3.cell(row=r, column=7, value=cc["loi_cl"])
        loi_cl_cell.border = border
        if cc["loi_cl"] > 0:
            loi_cl_cell.fill = alert_fill
        
        loi_td_cell = ws3.cell(row=r, column=8, value=cc["loi_td"])
        loi_td_cell.border = border
        if cc["loi_td"] > 0:
            loi_td_cell.fill = warn_fill
        
        ws3.cell(row=r, column=9, value=cc["diem_kpi_70"]).border = border
        
        # Xác định lý do chính
        ly_do = []
        if cc["sp_hoan_thanh"] < cc["sp_duoc_giao"]:
            ly_do.append("SP chưa đạt")
        if cc["loi_cl"] > 0:
            ly_do.append(f"CL -{cc['loi_cl']} lỗi")
        if cc["loi_td"] > 0:
            ly_do.append(f"TĐ -{cc['loi_td']} lỗi")
        ws3.cell(row=r, column=10, value=", ".join(ly_do) if ly_do else "Khác").border = border
    
    for c, w in [('A', 5), ('B', 25), ('C', 12), ('D', 25), ('E', 10), ('F', 10), ('G', 8), ('H', 8), ('I', 10), ('J', 25)]:
        ws3.column_dimensions[c].width = w
    
    # Save
    wb.save(output_path)
    print(f"Đã tạo file: {output_path}")
    print(f"\n=== 2. ĐIỂM KPI (70 điểm) ===")
    print(f"Tổng CC có dữ liệu: {total_co_kpi}")
    print(f"I. Đạt KPI 70 điểm: {len(dat_kpi_70)} ({pct(len(dat_kpi_70))})")
    print(f"   - Vượt KPI bất thường: {len(vuot_kpi_bat_thuong)}")
    print(f"II. Chưa đạt KPI 70 điểm: {len(chua_dat_kpi_70)} ({pct(len(chua_dat_kpi_70))})")
    print(f"   - Do SP chưa đạt: {len(chua_dat_do_so_luong)}")
    print(f"   - Do CL bị trừ: {len(chua_dat_do_chat_luong)}")
    print(f"   - Do TĐ bị trừ: {len(chua_dat_do_tien_do)}")


async def main():
    if len(sys.argv) < 3:
        print("Usage: python 02_diem_kpi.py <thang> <nam>")
        sys.exit(1)
    
    thang = int(sys.argv[1])
    nam = int(sys.argv[2])
    
    print(f"Đang lấy dữ liệu tháng {thang}/{nam}...")
    data, so_ngay = await get_data(thang, nam)
    
    if not data:
        print(f"Không có dữ liệu cho tháng {thang}/{nam}")
        sys.exit(1)
    
    output_path = f"02_DiemKPI_{thang:02d}_{nam}.xlsx"
    create_excel(data, thang, nam, so_ngay, output_path)


if __name__ == "__main__":
    import warnings
    warnings.filterwarnings("ignore")
    
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(main())
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        try:
            loop.run_until_complete(loop.shutdown_asyncgens())
        except:
            pass
        loop.close()