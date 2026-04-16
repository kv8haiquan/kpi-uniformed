#!/usr/bin/env python3
"""
03. THỐNG KÊ LÃNH ĐẠO BỊ TRỪ ĐIỂM d, đ, e
==========================================
Phần 3: Lãnh đạo bị trừ điểm theo quy định
- d: Kết quả đơn vị
- đ: Tổ chức triển khai
- e: Đoàn kết nội bộ

Chạy: python 03_lanh_dao_dde.py <thang> <nam>
Output: 03_LanhDaoDDE_01_2026.xlsx
"""

import sys
import asyncio
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, '/root/kpi-haiquan/backend')

from sqlalchemy import text
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
import os
from dotenv import load_dotenv

load_dotenv('/root/kpi-haiquan/backend/.env')

DB_HOST = os.getenv("DB_HOST", "localhost")
DB_PORT = os.getenv("DB_PORT", "5432")
DB_NAME = os.getenv("DB_NAME", "kpi_haiquan")
DB_USER = os.getenv("DB_USER", "kpi_user")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")

DATABASE_URL = f"postgresql+asyncpg://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"


async def get_data(thang: int, nam: int):
    engine = create_async_engine(DATABASE_URL)
    async with AsyncSession(engine) as db:
        # Lấy danh sách lãnh đạo từ chi_tiet_xep_loai
        ld_result = await db.execute(text("""
            SELECT ct.cong_chuc_id::text, cc.ho_ten, cc.ma_cc, dv.ten_don_vi,
                   vt.ten_vai_tro, ct.diem_kpi, ct.diem_tong, ct.xep_loai_he_thong
            FROM chi_tiet_xep_loai ct
            JOIN bao_cao_xep_loai bc ON bc.id = ct.bao_cao_id
            JOIN cong_chuc cc ON cc.id = ct.cong_chuc_id
            LEFT JOIN don_vi dv ON dv.id = cc.don_vi_id
            LEFT JOIN vai_tro vt ON vt.id = cc.vai_tro_id
            WHERE bc.thang = :thang AND bc.nam = :nam
                  AND ct.is_lanh_dao = true
                  AND cc.is_active = true
        """), {"thang": thang, "nam": nam})
        
        lanh_dao_list = []
        for row in ld_result:
            lanh_dao_list.append({
                "cong_chuc_id": row[0],
                "ho_ten": row[1],
                "ma_cc": row[2],
                "don_vi": row[3] or "",
                "vai_tro": row[4] or "",
                "diem_kpi": float(row[5]) if row[5] is not None else None,
                "diem_tong": float(row[6]) if row[6] is not None else None,
                "xep_loai": row[7],
            })
        
        # Lấy đánh giá d, đ, e
        dde_result = await db.execute(text("""
            SELECT cong_chuc_id::text,
                   d_ket_qua_don_vi, d_ghi_chu, d_phe_duyet,
                   dd_to_chuc_trien_khai, dd_ghi_chu, dd_phe_duyet,
                   e_doan_ket_noi_bo, e_ghi_chu, e_phe_duyet,
                   trang_thai
            FROM danh_gia_dde
            WHERE thang = :thang AND nam = :nam
        """), {"thang": thang, "nam": nam})
        
        dde_by_cc = {}
        for row in dde_result:
            dde_by_cc[row[0]] = {
                "d_dat": row[1],  # True = Đạt, False = Không đạt
                "d_ghi_chu": row[2] or "",
                "d_phe_duyet": row[3],
                "dd_dat": row[4],
                "dd_ghi_chu": row[5] or "",
                "dd_phe_duyet": row[6],
                "e_dat": row[7],
                "e_ghi_chu": row[8] or "",
                "e_phe_duyet": row[9],
                "trang_thai": row[10],
            }
        
        # Gộp dữ liệu
        for ld in lanh_dao_list:
            cc_id = ld["cong_chuc_id"]
            dde = dde_by_cc.get(cc_id, {})
            ld["d_dat"] = dde.get("d_dat", True)
            ld["d_ghi_chu"] = dde.get("d_ghi_chu", "")
            ld["dd_dat"] = dde.get("dd_dat", True)
            ld["dd_ghi_chu"] = dde.get("dd_ghi_chu", "")
            ld["e_dat"] = dde.get("e_dat", True)
            ld["e_ghi_chu"] = dde.get("e_ghi_chu", "")
            
            # Tính số tiêu chí bị trừ
            ld["bi_tru_d"] = ld["d_dat"] == False
            ld["bi_tru_dd"] = ld["dd_dat"] == False
            ld["bi_tru_e"] = ld["e_dat"] == False
            ld["tong_bi_tru"] = sum([ld["bi_tru_d"], ld["bi_tru_dd"], ld["bi_tru_e"]])
        
        return lanh_dao_list


def create_excel(data: list, thang: int, nam: int, output_path: str):
    wb = Workbook()
    
    # Styles
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    percent_font = Font(bold=True, color="0070C0")
    alert_fill = PatternFill("solid", fgColor="FFC7CE")
    good_fill = PatternFill("solid", fgColor="C6EFCE")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Phân loại
    total_ld = len(data)
    bi_tru_d = [ld for ld in data if ld["bi_tru_d"]]
    bi_tru_dd = [ld for ld in data if ld["bi_tru_dd"]]
    bi_tru_e = [ld for ld in data if ld["bi_tru_e"]]
    bi_tru_any = [ld for ld in data if ld["tong_bi_tru"] > 0]
    khong_bi_tru = [ld for ld in data if ld["tong_bi_tru"] == 0]
    
    def pct(count):
        return f"{count/total_ld*100:.1f}%" if total_ld > 0 else "0%"
    
    # =========================================================================
    # SHEET 1: TỔNG HỢP
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Tổng hợp"
    
    ws1['A1'] = f"3. THỐNG KÊ LÃNH ĐẠO BỊ TRỪ ĐIỂM d, đ, e - THÁNG {thang}/{nam}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:E1')
    
    ws1['A2'] = f"Tổng số lãnh đạo: {total_ld}"
    ws1['A2'].font = Font(bold=True)
    
    row = 4
    headers = ["Tiêu chí", "Số lượng bị trừ", "Tỷ lệ", "Ghi chú"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    rows_data = [
        ("d - Kết quả đơn vị", len(bi_tru_d), pct(len(bi_tru_d)), "Đơn vị không hoàn thành nhiệm vụ"),
        ("đ - Tổ chức triển khai", len(bi_tru_dd), pct(len(bi_tru_dd)), "Triển khai không đạt yêu cầu"),
        ("e - Đoàn kết nội bộ", len(bi_tru_e), pct(len(bi_tru_e)), "Có vấn đề đoàn kết"),
        ("TỔNG BỊ TRỪ (ít nhất 1 tiêu chí)", len(bi_tru_any), pct(len(bi_tru_any)), ""),
        ("Không bị trừ", len(khong_bi_tru), pct(len(khong_bi_tru)), ""),
    ]
    
    for i, (nhom, sl, tl, gc) in enumerate(rows_data):
        r = row + 1 + i
        ws1.cell(row=r, column=1, value=nhom).border = border
        ws1.cell(row=r, column=2, value=sl).border = border
        ws1.cell(row=r, column=2).alignment = center_alignment
        ws1.cell(row=r, column=3, value=tl).border = border
        ws1.cell(row=r, column=3).alignment = center_alignment
        ws1.cell(row=r, column=3).font = percent_font
        ws1.cell(row=r, column=4, value=gc).border = border
        
        if i == 3:  # Tổng bị trừ
            ws1.cell(row=r, column=1).font = Font(bold=True)
            if sl > 0:
                ws1.cell(row=r, column=2).fill = alert_fill
    
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 40
    
    # =========================================================================
    # SHEET 2: DANH SÁCH BỊ TRỪ
    # =========================================================================
    ws2 = wb.create_sheet("Danh sách bị trừ")
    
    ws2['A1'] = f"DANH SÁCH LÃNH ĐẠO BỊ TRỪ ĐIỂM d, đ, e - THÁNG {thang}/{nam}"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:K1')
    
    headers2 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "Chức vụ", "d", "đ", "e", "Tổng trừ", "Lý do", "Điểm tổng"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    # Sắp xếp theo số tiêu chí bị trừ giảm dần
    bi_tru_sorted = sorted(bi_tru_any, key=lambda x: x["tong_bi_tru"], reverse=True)
    
    for i, ld in enumerate(bi_tru_sorted, 1):
        r = 3 + i
        ws2.cell(row=r, column=1, value=i).border = border
        ws2.cell(row=r, column=2, value=ld["ho_ten"]).border = border
        ws2.cell(row=r, column=3, value=ld["ma_cc"]).border = border
        ws2.cell(row=r, column=4, value=ld["don_vi"]).border = border
        ws2.cell(row=r, column=5, value=ld["vai_tro"]).border = border
        
        # d
        d_cell = ws2.cell(row=r, column=6, value="✗" if ld["bi_tru_d"] else "✓")
        d_cell.border = border
        d_cell.alignment = center_alignment
        if ld["bi_tru_d"]:
            d_cell.fill = alert_fill
            d_cell.font = Font(bold=True, color="9C0006")
        else:
            d_cell.fill = good_fill
        
        # đ
        dd_cell = ws2.cell(row=r, column=7, value="✗" if ld["bi_tru_dd"] else "✓")
        dd_cell.border = border
        dd_cell.alignment = center_alignment
        if ld["bi_tru_dd"]:
            dd_cell.fill = alert_fill
            dd_cell.font = Font(bold=True, color="9C0006")
        else:
            dd_cell.fill = good_fill
        
        # e
        e_cell = ws2.cell(row=r, column=8, value="✗" if ld["bi_tru_e"] else "✓")
        e_cell.border = border
        e_cell.alignment = center_alignment
        if ld["bi_tru_e"]:
            e_cell.fill = alert_fill
            e_cell.font = Font(bold=True, color="9C0006")
        else:
            e_cell.fill = good_fill
        
        # Tổng trừ
        ws2.cell(row=r, column=9, value=ld["tong_bi_tru"]).border = border
        ws2.cell(row=r, column=9).alignment = center_alignment
        ws2.cell(row=r, column=9).font = Font(bold=True)
        
        # Lý do
        ly_do = []
        if ld["bi_tru_d"]:
            ly_do.append(f"d: {ld['d_ghi_chu']}" if ld['d_ghi_chu'] else "d")
        if ld["bi_tru_dd"]:
            ly_do.append(f"đ: {ld['dd_ghi_chu']}" if ld['dd_ghi_chu'] else "đ")
        if ld["bi_tru_e"]:
            ly_do.append(f"e: {ld['e_ghi_chu']}" if ld['e_ghi_chu'] else "e")
        ws2.cell(row=r, column=10, value="; ".join(ly_do)).border = border
        ws2.cell(row=r, column=10).alignment = wrap_alignment
        
        ws2.cell(row=r, column=11, value=ld["diem_tong"]).border = border
        
        ws2.row_dimensions[r].height = 30
    
    for c, w in [('A', 5), ('B', 25), ('C', 12), ('D', 25), ('E', 15), 
                 ('F', 6), ('G', 6), ('H', 6), ('I', 10), ('J', 50), ('K', 10)]:
        ws2.column_dimensions[c].width = w
    
    # =========================================================================
    # SHEET 3: TẤT CẢ LÃNH ĐẠO
    # =========================================================================
    ws3 = wb.create_sheet("Tất cả lãnh đạo")
    
    ws3['A1'] = f"DANH SÁCH TẤT CẢ LÃNH ĐẠO - THÁNG {thang}/{nam}"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:J1')
    
    headers3 = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "Chức vụ", "d", "đ", "e", "Điểm KPI", "Điểm tổng"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    for i, ld in enumerate(data, 1):
        r = 3 + i
        ws3.cell(row=r, column=1, value=i).border = border
        ws3.cell(row=r, column=2, value=ld["ho_ten"]).border = border
        ws3.cell(row=r, column=3, value=ld["ma_cc"]).border = border
        ws3.cell(row=r, column=4, value=ld["don_vi"]).border = border
        ws3.cell(row=r, column=5, value=ld["vai_tro"]).border = border
        
        for col_idx, key in [(6, "bi_tru_d"), (7, "bi_tru_dd"), (8, "bi_tru_e")]:
            cell = ws3.cell(row=r, column=col_idx, value="✗" if ld[key] else "✓")
            cell.border = border
            cell.alignment = center_alignment
            if ld[key]:
                cell.fill = alert_fill
            else:
                cell.fill = good_fill
        
        ws3.cell(row=r, column=9, value=ld["diem_kpi"]).border = border
        ws3.cell(row=r, column=10, value=ld["diem_tong"]).border = border
    
    for c, w in [('A', 5), ('B', 25), ('C', 12), ('D', 25), ('E', 15), 
                 ('F', 6), ('G', 6), ('H', 6), ('I', 10), ('J', 10)]:
        ws3.column_dimensions[c].width = w
    
    # Save
    wb.save(output_path)
    print(f"Đã tạo file: {output_path}")
    print(f"\n=== 3. LÃNH ĐẠO BỊ TRỪ d, đ, e ===")
    print(f"Tổng số lãnh đạo: {total_ld}")
    print(f"- d bị trừ: {len(bi_tru_d)} ({pct(len(bi_tru_d))})")
    print(f"- đ bị trừ: {len(bi_tru_dd)} ({pct(len(bi_tru_dd))})")
    print(f"- e bị trừ: {len(bi_tru_e)} ({pct(len(bi_tru_e))})")
    print(f"TỔNG BỊ TRỪ: {len(bi_tru_any)}/{total_ld}")


async def main():
    if len(sys.argv) < 3:
        print("Usage: python 03_lanh_dao_dde.py <thang> <nam>")
        sys.exit(1)
    
    thang = int(sys.argv[1])
    nam = int(sys.argv[2])
    
    print(f"Đang lấy dữ liệu tháng {thang}/{nam}...")
    data = await get_data(thang, nam)
    
    if not data:
        print(f"Không có dữ liệu lãnh đạo cho tháng {thang}/{nam}")
        sys.exit(1)
    
    output_path = f"03_LanhDaoDDE_{thang:02d}_{nam}.xlsx"
    create_excel(data, thang, nam, output_path)


if __name__ == "__main__":
    import warnings
    warnings.filterwarnings("ignore")
    
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(main())
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        try:
            loop.run_until_complete(loop.shutdown_asyncgens())
        except:
            pass
        loop.close()