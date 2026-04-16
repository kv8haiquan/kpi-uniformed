#!/usr/bin/env python3
"""
04. THỐNG KÊ KHỐI LƯỢNG CÔNG VIỆC
==================================
Phần 4: Báo cáo thống kê tình hình thực hiện nhiệm vụ

4.1. Khối lượng công việc:
- SP1: Tờ khai HQ
- SP2: Văn bản hành chính
- SP3: Giờ trực làm việc
- SP4: Giờ tuần tra kiểm soát

4.2. Mức độ phức tạp:
- C1: Dễ - Đơn giản
- C2: Trung bình - Thông thường
- C3: Khó - Nâng cao
- C4: Rất khó - Phức tạp
- C5: Đặc biệt khó - Đặc thù

Chạy: python 04_khoi_luong_cong_viec.py <thang> <nam>
Output: 04_KhoiLuongCongViec_01_2026.xlsx
"""

import sys
import asyncio
import calendar
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, '/root/kpi-haiquan/backend')

from sqlalchemy import text
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
import os
from dotenv import load_dotenv

load_dotenv('/root/kpi-haiquan/backend/.env')

DB_HOST = os.getenv("DB_HOST", "localhost")
DB_PORT = os.getenv("DB_PORT", "5432")
DB_NAME = os.getenv("DB_NAME", "kpi_haiquan")
DB_USER = os.getenv("DB_USER", "kpi_user")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")

DATABASE_URL = f"postgresql+asyncpg://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"

SP_NAMES = {
    "SP1": "Tờ khai HQ (kiểm tra chi tiết hồ sơ)",
    "SP2": "Văn bản hành chính",
    "SP3": "Giờ trực làm việc",
    "SP4": "Giờ tuần tra kiểm soát",
}

CAP_DO_NAMES = {
    "C1": "Dễ - Đơn giản",
    "C2": "Trung bình - Thông thường",
    "C3": "Khó - Nâng cao",
    "C4": "Rất khó - Phức tạp",
    "C5": "Đặc biệt khó - Đặc thù",
}


async def get_data(thang: int, nam: int):
    engine = create_async_engine(DATABASE_URL)
    async with AsyncSession(engine) as db:
        so_ngay_trong_thang = calendar.monthrange(nam, thang)[1]
        
        # =================================================================
        # 1. Thống kê theo loại SP (SP1, SP2, SP3, SP4)
        # =================================================================
        sp_result = await db.execute(text("""
            SELECT sp.ma_sp, sp.ten_sp,
                   COUNT(*) as so_khai,
                   COALESCE(SUM(kk.so_sp_goc_quy_doi), 0) as tong_sp_quy_doi
            FROM ke_khai_cong_viec kk
            JOIN danh_muc_sp_cong_viec dm ON dm.id = kk.danh_muc_sp_id
            JOIN sp_cong_viec_chuan sp ON sp.id = dm.sp_chuan_id
            WHERE kk.thang = :thang AND kk.nam = :nam 
                  AND kk.trang_thai = 'DA_PHE_DUYET' AND kk.is_deleted = false
            GROUP BY sp.ma_sp, sp.ten_sp
            ORDER BY sp.ma_sp
        """), {"thang": thang, "nam": nam})
        
        sp_data = []
        for row in sp_result:
            sp_data.append({
                "ma_sp": row[0],
                "ten_sp": row[1],
                "so_khai": int(row[2]),
                "tong_sp": float(row[3]),
            })
        
        # =================================================================
        # 2. Thống kê theo cấp độ phức tạp (C1-C5)
        # =================================================================
        cap_do_result = await db.execute(text("""
            SELECT cd.ma_cap_do, cd.ten_cap_do,
                   COUNT(*) as so_khai,
                   COALESCE(SUM(kk.so_sp_goc_quy_doi), 0) as tong_sp_quy_doi
            FROM ke_khai_cong_viec kk
            JOIN cap_do_phuc_tap cd ON cd.id = kk.cap_do_id
            WHERE kk.thang = :thang AND kk.nam = :nam 
                  AND kk.trang_thai = 'DA_PHE_DUYET' AND kk.is_deleted = false
            GROUP BY cd.ma_cap_do, cd.ten_cap_do
            ORDER BY cd.ma_cap_do
        """), {"thang": thang, "nam": nam})
        
        cap_do_data = []
        for row in cap_do_result:
            cap_do_data.append({
                "ma_cap_do": row[0],
                "ten_cap_do": row[1],
                "so_khai": int(row[2]),
                "tong_sp": float(row[3]),
            })
        
        # =================================================================
        # 3. Chi tiết theo đơn vị
        # =================================================================
        don_vi_result = await db.execute(text("""
            SELECT dv.ten_don_vi, sp.ma_sp, cd.ma_cap_do,
                   COUNT(*) as so_khai,
                   COALESCE(SUM(kk.so_sp_goc_quy_doi), 0) as tong_sp
            FROM ke_khai_cong_viec kk
            JOIN danh_muc_sp_cong_viec dm ON dm.id = kk.danh_muc_sp_id
            JOIN sp_cong_viec_chuan sp ON sp.id = dm.sp_chuan_id
            JOIN cap_do_phuc_tap cd ON cd.id = kk.cap_do_id
            JOIN cong_chuc cc ON cc.id = kk.cong_chuc_id
            JOIN don_vi dv ON dv.id = cc.don_vi_id
            WHERE kk.thang = :thang AND kk.nam = :nam 
                  AND kk.trang_thai = 'DA_PHE_DUYET' AND kk.is_deleted = false
            GROUP BY dv.ten_don_vi, sp.ma_sp, cd.ma_cap_do
            ORDER BY dv.ten_don_vi, sp.ma_sp, cd.ma_cap_do
        """), {"thang": thang, "nam": nam})
        
        don_vi_data = []
        for row in don_vi_result:
            don_vi_data.append({
                "don_vi": row[0],
                "ma_sp": row[1],
                "ma_cap_do": row[2],
                "so_khai": int(row[3]),
                "tong_sp": float(row[4]),
            })
        
        # Tổng SP
        tong_sp_all = sum(item["tong_sp"] for item in sp_data)
        
        return sp_data, cap_do_data, don_vi_data, tong_sp_all, so_ngay_trong_thang


def create_excel(sp_data, cap_do_data, don_vi_data, tong_sp_all, so_ngay, thang, nam, output_path):
    wb = Workbook()
    
    # Styles
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    percent_font = Font(bold=True, color="0070C0")
    sp1_fill = PatternFill("solid", fgColor="DAEEF3")
    sp2_fill = PatternFill("solid", fgColor="E2EFDA")
    sp3_fill = PatternFill("solid", fgColor="FDE9D9")
    sp4_fill = PatternFill("solid", fgColor="E4DFEC")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    SP_FILLS = {"SP1": sp1_fill, "SP2": sp2_fill, "SP3": sp3_fill, "SP4": sp4_fill}
    
    def pct(val):
        return f"{val/tong_sp_all*100:.1f}%" if tong_sp_all > 0 else "0%"
    
    # =========================================================================
    # SHEET 1: TỔNG HỢP
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Tổng hợp"
    
    ws1['A1'] = f"4. BÁO CÁO THỐNG KÊ TÌNH HÌNH THỰC HIỆN NHIỆM VỤ - THÁNG {thang}/{nam}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:E1')
    
    ws1['A2'] = f"Chi cục Hải quan Khu vực VIII | Tổng SP: {tong_sp_all:,.0f}"
    ws1['A2'].font = Font(bold=True)
    
    # 4.1. Khối lượng công việc
    row = 4
    ws1.cell(row=row, column=1, value="4.1. KHỐI LƯỢNG CÔNG VIỆC")
    ws1.cell(row=row, column=1).font = Font(bold=True, size=13, color="2F5496")
    ws1.merge_cells(f'A{row}:E{row}')
    
    row += 1
    headers = ["Loại SP", "Tên", "Số lượng SP", "Tỷ lệ"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    for sp in sp_data:
        row += 1
        ws1.cell(row=row, column=1, value=sp["ma_sp"]).border = border
        ws1.cell(row=row, column=1).fill = SP_FILLS.get(sp["ma_sp"], sp1_fill)
        ws1.cell(row=row, column=1).font = Font(bold=True)
        
        ten_sp = SP_NAMES.get(sp["ma_sp"], sp["ten_sp"])
        ws1.cell(row=row, column=2, value=ten_sp).border = border
        
        ws1.cell(row=row, column=3, value=f"{sp['tong_sp']:,.0f}").border = border
        ws1.cell(row=row, column=3).alignment = center_alignment
        
        ws1.cell(row=row, column=4, value=pct(sp["tong_sp"])).border = border
        ws1.cell(row=row, column=4).alignment = center_alignment
        ws1.cell(row=row, column=4).font = percent_font
    
    # Tổng
    row += 1
    ws1.cell(row=row, column=1, value="TỔNG").border = border
    ws1.cell(row=row, column=1).font = Font(bold=True)
    ws1.cell(row=row, column=2).border = border
    ws1.cell(row=row, column=3, value=f"{tong_sp_all:,.0f}").border = border
    ws1.cell(row=row, column=3).font = Font(bold=True)
    ws1.cell(row=row, column=3).alignment = center_alignment
    ws1.cell(row=row, column=4, value="100%").border = border
    ws1.cell(row=row, column=4).font = Font(bold=True, color="0070C0")
    ws1.cell(row=row, column=4).alignment = center_alignment
    
    # 4.2. Mức độ phức tạp
    row += 3
    ws1.cell(row=row, column=1, value="4.2. MỨC ĐỘ PHỨC TẠP CỦA CÔNG VIỆC")
    ws1.cell(row=row, column=1).font = Font(bold=True, size=13, color="2F5496")
    ws1.merge_cells(f'A{row}:E{row}')
    
    row += 1
    headers2 = ["Cấp độ", "Tên", "Số lượng SP", "Tỷ lệ"]
    for col, h in enumerate(headers2, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    cap_do_fills = {
        "C1": PatternFill("solid", fgColor="C6EFCE"),
        "C2": PatternFill("solid", fgColor="D9EAD3"),
        "C3": PatternFill("solid", fgColor="FFEB9C"),
        "C4": PatternFill("solid", fgColor="FFC7CE"),
        "C5": PatternFill("solid", fgColor="E6B8AF"),
    }
    
    for cd in cap_do_data:
        row += 1
        ws1.cell(row=row, column=1, value=cd["ma_cap_do"]).border = border
        ws1.cell(row=row, column=1).fill = cap_do_fills.get(cd["ma_cap_do"], sp1_fill)
        ws1.cell(row=row, column=1).font = Font(bold=True)
        ws1.cell(row=row, column=1).alignment = center_alignment
        
        ten_cd = CAP_DO_NAMES.get(cd["ma_cap_do"], cd["ten_cap_do"])
        ws1.cell(row=row, column=2, value=ten_cd).border = border
        
        ws1.cell(row=row, column=3, value=f"{cd['tong_sp']:,.0f}").border = border
        ws1.cell(row=row, column=3).alignment = center_alignment
        
        ws1.cell(row=row, column=4, value=pct(cd["tong_sp"])).border = border
        ws1.cell(row=row, column=4).alignment = center_alignment
        ws1.cell(row=row, column=4).font = percent_font
    
    # Tổng
    row += 1
    tong_cap_do = sum(cd["tong_sp"] for cd in cap_do_data)
    ws1.cell(row=row, column=1, value="TỔNG").border = border
    ws1.cell(row=row, column=1).font = Font(bold=True)
    ws1.cell(row=row, column=2).border = border
    ws1.cell(row=row, column=3, value=f"{tong_cap_do:,.0f}").border = border
    ws1.cell(row=row, column=3).font = Font(bold=True)
    ws1.cell(row=row, column=3).alignment = center_alignment
    ws1.cell(row=row, column=4, value="100%").border = border
    ws1.cell(row=row, column=4).font = Font(bold=True, color="0070C0")
    ws1.cell(row=row, column=4).alignment = center_alignment
    
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 40
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 12
    
    # =========================================================================
    # SHEET 2: CHI TIẾT THEO ĐƠN VỊ - LOẠI SP
    # =========================================================================
    ws2 = wb.create_sheet("Theo đơn vị - Loại SP")
    
    ws2['A1'] = f"CHI TIẾT KHỐI LƯỢNG CÔNG VIỆC THEO ĐƠN VỊ - THÁNG {thang}/{nam}"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:G1')
    
    # Pivot: đơn vị -> ma_sp -> tổng SP
    pivot_dv_sp = defaultdict(lambda: defaultdict(float))
    for item in don_vi_data:
        pivot_dv_sp[item["don_vi"]][item["ma_sp"]] += item["tong_sp"]
    
    headers2 = ["STT", "Đơn vị", "SP1", "SP2", "SP3", "SP4", "Tổng"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    row = 3
    for i, (dv, sp_map) in enumerate(sorted(pivot_dv_sp.items()), 1):
        row += 1
        ws2.cell(row=row, column=1, value=i).border = border
        ws2.cell(row=row, column=2, value=dv).border = border
        
        tong_dv = 0
        for col_idx, ma_sp in enumerate(["SP1", "SP2", "SP3", "SP4"], 3):
            val = sp_map.get(ma_sp, 0)
            tong_dv += val
            cell = ws2.cell(row=row, column=col_idx, value=f"{val:,.0f}" if val > 0 else "")
            cell.border = border
            cell.alignment = center_alignment
            if val > 0:
                cell.fill = SP_FILLS.get(ma_sp, sp1_fill)
        
        ws2.cell(row=row, column=7, value=f"{tong_dv:,.0f}").border = border
        ws2.cell(row=row, column=7).font = Font(bold=True)
        ws2.cell(row=row, column=7).alignment = center_alignment
    
    # Tổng hàng
    row += 1
    ws2.cell(row=row, column=1).border = border
    ws2.cell(row=row, column=2, value="TỔNG").border = border
    ws2.cell(row=row, column=2).font = Font(bold=True)
    
    for col_idx, ma_sp in enumerate(["SP1", "SP2", "SP3", "SP4"], 3):
        tong = sum(sp_map.get(ma_sp, 0) for sp_map in pivot_dv_sp.values())
        cell = ws2.cell(row=row, column=col_idx, value=f"{tong:,.0f}")
        cell.border = border
        cell.font = Font(bold=True)
        cell.alignment = center_alignment
    
    ws2.cell(row=row, column=7, value=f"{tong_sp_all:,.0f}").border = border
    ws2.cell(row=row, column=7).font = Font(bold=True)
    ws2.cell(row=row, column=7).alignment = center_alignment
    
    for c, w in [('A', 5), ('B', 35), ('C', 12), ('D', 12), ('E', 12), ('F', 12), ('G', 12)]:
        ws2.column_dimensions[c].width = w
    
    # =========================================================================
    # SHEET 3: CHI TIẾT THEO ĐƠN VỊ - CẤP ĐỘ
    # =========================================================================
    ws3 = wb.create_sheet("Theo đơn vị - Cấp độ")
    
    ws3['A1'] = f"CHI TIẾT MỨC ĐỘ PHỨC TẠP THEO ĐƠN VỊ - THÁNG {thang}/{nam}"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:H1')
    
    # Pivot: đơn vị -> ma_cap_do -> tổng SP
    pivot_dv_cd = defaultdict(lambda: defaultdict(float))
    for item in don_vi_data:
        pivot_dv_cd[item["don_vi"]][item["ma_cap_do"]] += item["tong_sp"]
    
    headers3 = ["STT", "Đơn vị", "C1", "C2", "C3", "C4", "C5", "Tổng"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    row = 3
    for i, (dv, cd_map) in enumerate(sorted(pivot_dv_cd.items()), 1):
        row += 1
        ws3.cell(row=row, column=1, value=i).border = border
        ws3.cell(row=row, column=2, value=dv).border = border
        
        tong_dv = 0
        for col_idx, ma_cd in enumerate(["C1", "C2", "C3", "C4", "C5"], 3):
            val = cd_map.get(ma_cd, 0)
            tong_dv += val
            cell = ws3.cell(row=row, column=col_idx, value=f"{val:,.0f}" if val > 0 else "")
            cell.border = border
            cell.alignment = center_alignment
            if val > 0:
                cell.fill = cap_do_fills.get(ma_cd)
        
        ws3.cell(row=row, column=8, value=f"{tong_dv:,.0f}").border = border
        ws3.cell(row=row, column=8).font = Font(bold=True)
        ws3.cell(row=row, column=8).alignment = center_alignment
    
    # Tổng hàng
    row += 1
    ws3.cell(row=row, column=1).border = border
    ws3.cell(row=row, column=2, value="TỔNG").border = border
    ws3.cell(row=row, column=2).font = Font(bold=True)
    
    for col_idx, ma_cd in enumerate(["C1", "C2", "C3", "C4", "C5"], 3):
        tong = sum(cd_map.get(ma_cd, 0) for cd_map in pivot_dv_cd.values())
        cell = ws3.cell(row=row, column=col_idx, value=f"{tong:,.0f}")
        cell.border = border
        cell.font = Font(bold=True)
        cell.alignment = center_alignment
    
    ws3.cell(row=row, column=8, value=f"{tong_sp_all:,.0f}").border = border
    ws3.cell(row=row, column=8).font = Font(bold=True)
    ws3.cell(row=row, column=8).alignment = center_alignment
    
    for c, w in [('A', 5), ('B', 35), ('C', 10), ('D', 10), ('E', 10), ('F', 10), ('G', 10), ('H', 12)]:
        ws3.column_dimensions[c].width = w
    
    # Save
    wb.save(output_path)
    
    # Console output
    print(f"Đã tạo file: {output_path}")
    print(f"\n=== 4.1. KHỐI LƯỢNG CÔNG VIỆC ===")
    for sp in sp_data:
        print(f"  {sp['ma_sp']} ({SP_NAMES.get(sp['ma_sp'], '')}): {sp['tong_sp']:,.0f} SP ({pct(sp['tong_sp'])})")
    print(f"\n=== 4.2. MỨC ĐỘ PHỨC TẠP ===")
    for cd in cap_do_data:
        print(f"  {cd['ma_cap_do']} ({CAP_DO_NAMES.get(cd['ma_cap_do'], '')}): {cd['tong_sp']:,.0f} SP ({pct(cd['tong_sp'])})")


async def main():
    if len(sys.argv) < 3:
        print("Usage: python 04_khoi_luong_cong_viec.py <thang> <nam>")
        sys.exit(1)
    
    thang = int(sys.argv[1])
    nam = int(sys.argv[2])
    
    print(f"Đang lấy dữ liệu tháng {thang}/{nam}...")
    sp_data, cap_do_data, don_vi_data, tong_sp_all, so_ngay = await get_data(thang, nam)
    
    if not sp_data:
        print(f"Không có dữ liệu cho tháng {thang}/{nam}")
        sys.exit(1)
    
    output_path = f"04_KhoiLuongCongViec_{thang:02d}_{nam}.xlsx"
    create_excel(sp_data, cap_do_data, don_vi_data, tong_sp_all, so_ngay, thang, nam, output_path)


if __name__ == "__main__":
    import warnings
    warnings.filterwarnings("ignore")
    
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(main())
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        try:
            loop.run_until_complete(loop.shutdown_asyncgens())
        except:
            pass
        loop.close()
