#!/usr/bin/env python3
"""
05. THỐNG KÊ DANH MỤC CÔNG VIỆC
================================
Báo cáo chi tiết từng đầu mục công việc và danh sách user kê khai

Chạy: python 05_danh_muc_cong_viec.py <thang> <nam>
Output: 05_DanhMucCongViec_02_2026.xlsx
"""

import sys
import asyncio
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, '/root/kpi-haiquan/backend')

from sqlalchemy import text
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
import os
from dotenv import load_dotenv

load_dotenv('/root/kpi-haiquan/backend/.env')

DB_HOST = os.getenv("DB_HOST", "localhost")
DB_PORT = os.getenv("DB_PORT", "5432")
DB_NAME = os.getenv("DB_NAME", "kpi_haiquan")
DB_USER = os.getenv("DB_USER", "kpi_user")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")

DATABASE_URL = f"postgresql+asyncpg://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"


async def get_data(thang: int, nam: int):
    engine = create_async_engine(DATABASE_URL)
    async with AsyncSession(engine) as db:
        # =================================================================
        # 1. Lấy tất cả kê khai công việc với thông tin chi tiết
        # =================================================================
        result = await db.execute(text("""
            SELECT 
                dm.id as danh_muc_id,
                dm.ten_cong_viec,
                sp.ma_sp,
                sp.ten_sp,
                cc.id as cong_chuc_id,
                cc.ho_ten,
                cc.ma_cc,
                dv.ten_don_vi,
                cd.ma_cap_do,
                cd.ten_cap_do,
                COUNT(*) as so_lan_khai,
                COALESCE(SUM(kk.so_sp_goc_quy_doi), 0) as tong_sp_quy_doi,
                COALESCE(SUM(kk.so_luong), 0) as tong_so_luong
            FROM ke_khai_cong_viec kk
            JOIN danh_muc_sp_cong_viec dm ON dm.id = kk.danh_muc_sp_id
            JOIN sp_cong_viec_chuan sp ON sp.id = dm.sp_chuan_id
            JOIN cong_chuc cc ON cc.id = kk.cong_chuc_id
            LEFT JOIN don_vi dv ON dv.id = cc.don_vi_id
            JOIN cap_do_phuc_tap cd ON cd.id = kk.cap_do_id
            WHERE kk.thang = :thang AND kk.nam = :nam 
                  AND kk.trang_thai = 'DA_PHE_DUYET' AND kk.is_deleted = false
            GROUP BY dm.id, dm.ten_cong_viec,
                     sp.ma_sp, sp.ten_sp,
                     cc.id, cc.ho_ten, cc.ma_cc, dv.ten_don_vi,
                     cd.ma_cap_do, cd.ten_cap_do
            ORDER BY sp.ma_sp, dm.ten_cong_viec, dv.ten_don_vi, cc.ho_ten
        """), {"thang": thang, "nam": nam})
        
        raw_data = []
        for row in result:
            raw_data.append({
                "danh_muc_id": row[0],
                "ten_cong_viec": row[1],
                "ma_sp": row[2],
                "ten_sp": row[3],
                "cong_chuc_id": row[4],
                "ho_ten": row[5],
                "ma_cc": row[6],
                "don_vi": row[7] or "",
                "ma_cap_do": row[8],
                "ten_cap_do": row[9],
                "so_lan_khai": int(row[10]),
                "tong_sp_quy_doi": float(row[11]),
                "tong_so_luong": float(row[12]),
            })
        
        # =================================================================
        # 2. Tổng hợp theo danh mục công việc
        # =================================================================
        danh_muc_map = defaultdict(lambda: {
            "ten_cong_viec": "",
            "ma_sp": "",
            "ten_sp": "",
            "users": [],
            "tong_sp": 0,
            "tong_lan_khai": 0,
            "cap_do_stats": defaultdict(int),
        })
        
        for item in raw_data:
            dm_id = item["danh_muc_id"]
            dm = danh_muc_map[dm_id]
            dm["ten_cong_viec"] = item["ten_cong_viec"]
            dm["ma_sp"] = item["ma_sp"]
            dm["ten_sp"] = item["ten_sp"]
            dm["tong_sp"] += item["tong_sp_quy_doi"]
            dm["tong_lan_khai"] += item["so_lan_khai"]
            dm["cap_do_stats"][item["ma_cap_do"]] += item["so_lan_khai"]
            
            # Tìm user đã có chưa
            user_found = False
            for u in dm["users"]:
                if u["cong_chuc_id"] == item["cong_chuc_id"]:
                    u["so_lan_khai"] += item["so_lan_khai"]
                    u["tong_sp"] += item["tong_sp_quy_doi"]
                    u["cap_do_list"].add(item["ma_cap_do"])
                    user_found = True
                    break
            
            if not user_found:
                dm["users"].append({
                    "cong_chuc_id": item["cong_chuc_id"],
                    "ho_ten": item["ho_ten"],
                    "ma_cc": item["ma_cc"],
                    "don_vi": item["don_vi"],
                    "so_lan_khai": item["so_lan_khai"],
                    "tong_sp": item["tong_sp_quy_doi"],
                    "cap_do_list": {item["ma_cap_do"]},
                })
        
        # Convert to list và sort
        danh_muc_list = []
        for dm_id, dm in danh_muc_map.items():
            dm["danh_muc_id"] = dm_id
            dm["so_user"] = len(dm["users"])
            # Sort users theo đơn vị, tên
            dm["users"] = sorted(dm["users"], key=lambda x: (x["don_vi"], x["ho_ten"]))
            danh_muc_list.append(dm)
        
        # Sort theo ma_sp, tên công việc
        danh_muc_list = sorted(danh_muc_list, key=lambda x: (x["ma_sp"], x["ten_cong_viec"]))
        
        return danh_muc_list


def create_excel(data: list, thang: int, nam: int, output_path: str):
    wb = Workbook()
    
    # Styles
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    dm_header_fill = PatternFill("solid", fgColor="2F5496")
    sub_header_fill = PatternFill("solid", fgColor="D6DCE4")
    sp1_fill = PatternFill("solid", fgColor="DAEEF3")
    sp2_fill = PatternFill("solid", fgColor="E2EFDA")
    sp3_fill = PatternFill("solid", fgColor="FDE9D9")
    sp4_fill = PatternFill("solid", fgColor="E4DFEC")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    SP_FILLS = {"SP1": sp1_fill, "SP2": sp2_fill, "SP3": sp3_fill, "SP4": sp4_fill}
    
    # =========================================================================
    # SHEET 1: TỔNG HỢP DANH MỤC
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Tổng hợp"
    
    ws1['A1'] = f"5. THỐNG KÊ DANH MỤC CÔNG VIỆC - THÁNG {thang}/{nam}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:H1')
    
    ws1['A2'] = f"Tổng số đầu mục công việc: {len(data)}"
    ws1['A2'].font = Font(bold=True)
    
    headers = ["STT", "Loại SP", "Tên công việc", "Số user kê khai", "Số lần kê khai", "Tổng SP quy đổi", "Cấp độ phổ biến", "Số cấp độ"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    row = 4
    for i, dm in enumerate(data, 1):
        row += 1
        ws1.cell(row=row, column=1, value=i).border = border
        ws1.cell(row=row, column=1).alignment = center_alignment
        
        ma_sp_cell = ws1.cell(row=row, column=2, value=dm["ma_sp"])
        ma_sp_cell.border = border
        ma_sp_cell.alignment = center_alignment
        ma_sp_cell.fill = SP_FILLS.get(dm["ma_sp"], sp1_fill)
        ma_sp_cell.font = Font(bold=True)
        
        ws1.cell(row=row, column=3, value=dm["ten_cong_viec"]).border = border
        
        ws1.cell(row=row, column=4, value=dm["so_user"]).border = border
        ws1.cell(row=row, column=4).alignment = center_alignment
        
        ws1.cell(row=row, column=5, value=dm["tong_lan_khai"]).border = border
        ws1.cell(row=row, column=5).alignment = center_alignment
        
        ws1.cell(row=row, column=6, value=f"{dm['tong_sp']:,.0f}").border = border
        ws1.cell(row=row, column=6).alignment = center_alignment
        
        # Cấp độ phổ biến nhất
        cap_do_pho_bien = max(dm["cap_do_stats"], key=dm["cap_do_stats"].get) if dm["cap_do_stats"] else "-"
        ws1.cell(row=row, column=7, value=cap_do_pho_bien).border = border
        ws1.cell(row=row, column=7).alignment = center_alignment
        
        so_cap_do = len(dm["cap_do_stats"])
        cap_do_cell = ws1.cell(row=row, column=8, value=so_cap_do)
        cap_do_cell.border = border
        cap_do_cell.alignment = center_alignment
        if so_cap_do >= 4:
            cap_do_cell.fill = PatternFill("solid", fgColor="FFC7CE")
            cap_do_cell.font = Font(bold=True, color="9C0006")
        elif so_cap_do == 3:
            cap_do_cell.fill = PatternFill("solid", fgColor="FFEB9C")
    
    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 10
    ws1.column_dimensions['C'].width = 50
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 18
    ws1.column_dimensions['G'].width = 15
    ws1.column_dimensions['H'].width = 12
    
    # =========================================================================
    # SHEET 2: CHI TIẾT TỪNG DANH MỤC VÀ USER
    # =========================================================================
    ws2 = wb.create_sheet("Chi tiết theo công việc")
    
    ws2['A1'] = f"CHI TIẾT DANH MỤC CÔNG VIỆC VÀ USER KÊ KHAI - THÁNG {thang}/{nam}"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:H1')
    
    row = 3
    for dm_idx, dm in enumerate(data, 1):
        # Header cho mỗi danh mục
        ws2.merge_cells(f'A{row}:H{row}')
        header_text = f"{dm_idx}. [{dm['ma_sp']}] {dm['ten_cong_viec']} ({dm['so_user']} user | {dm['tong_lan_khai']} lần | {dm['tong_sp']:,.0f} SP)"
        cell = ws2.cell(row=row, column=1, value=header_text)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = dm_header_fill
        for col in range(2, 9):
            ws2.cell(row=row, column=col).fill = dm_header_fill
        row += 1
        
        # Sub-header
        sub_headers = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "Số lần khai", "Tổng SP", "Cấp độ sử dụng"]
        for col, h in enumerate(sub_headers, 1):
            cell = ws2.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True, size=10)
            cell.fill = sub_header_fill
            cell.border = border
            cell.alignment = center_alignment
        row += 1
        
        # Users
        for u_idx, user in enumerate(dm["users"], 1):
            ws2.cell(row=row, column=1, value=u_idx).border = border
            ws2.cell(row=row, column=1).alignment = center_alignment
            
            ws2.cell(row=row, column=2, value=user["ho_ten"]).border = border
            ws2.cell(row=row, column=3, value=user["ma_cc"]).border = border
            ws2.cell(row=row, column=4, value=user["don_vi"]).border = border
            
            ws2.cell(row=row, column=5, value=user["so_lan_khai"]).border = border
            ws2.cell(row=row, column=5).alignment = center_alignment
            
            ws2.cell(row=row, column=6, value=f"{user['tong_sp']:,.0f}").border = border
            ws2.cell(row=row, column=6).alignment = center_alignment
            
            cap_do_str = ", ".join(sorted(user["cap_do_list"]))
            ws2.cell(row=row, column=7, value=cap_do_str).border = border
            ws2.cell(row=row, column=7).alignment = center_alignment
            
            row += 1
        
        row += 1  # Khoảng cách giữa các danh mục
    
    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 30
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 12
    ws2.column_dimensions['G'].width = 18
    
    # =========================================================================
    # SHEET 3: PIVOT - CÔNG VIỆC × ĐƠN VỊ
    # =========================================================================
    ws3 = wb.create_sheet("Pivot công việc - đơn vị")
    
    ws3['A1'] = f"PIVOT: CÔNG VIỆC × ĐƠN VỊ - THÁNG {thang}/{nam}"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:K1')
    
    # Lấy danh sách đơn vị
    all_don_vi = set()
    for dm in data:
        for user in dm["users"]:
            if user["don_vi"]:
                all_don_vi.add(user["don_vi"])
    all_don_vi = sorted(all_don_vi)
    
    # Header
    headers3 = ["STT", "Tên công việc"] + all_don_vi + ["Tổng"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws3.row_dimensions[3].height = 40
    
    row = 3
    for i, dm in enumerate(data, 1):
        row += 1
        ws3.cell(row=row, column=1, value=i).border = border
        ws3.cell(row=row, column=1).alignment = center_alignment
        
        ws3.cell(row=row, column=2, value=dm["ten_cong_viec"]).border = border
        
        # Tính số user theo đơn vị
        dv_user_count = defaultdict(int)
        for user in dm["users"]:
            if user["don_vi"]:
                dv_user_count[user["don_vi"]] += 1
        
        tong_user = 0
        for col_idx, dv in enumerate(all_don_vi, 3):
            count = dv_user_count.get(dv, 0)
            tong_user += count
            cell = ws3.cell(row=row, column=col_idx, value=count if count > 0 else "")
            cell.border = border
            cell.alignment = center_alignment
            if count > 0:
                cell.fill = PatternFill("solid", fgColor="E2EFDA")
        
        ws3.cell(row=row, column=len(all_don_vi) + 3, value=tong_user).border = border
        ws3.cell(row=row, column=len(all_don_vi) + 3).font = Font(bold=True)
        ws3.cell(row=row, column=len(all_don_vi) + 3).alignment = center_alignment
    
    ws3.column_dimensions['A'].width = 5
    ws3.column_dimensions['B'].width = 45
    for col in range(3, len(all_don_vi) + 4):
        ws3.column_dimensions[get_column_letter(col)].width = 12
    
    # Save
    wb.save(output_path)
    
    # Console output
    print(f"Đã tạo file: {output_path}")
    print(f"\n=== 5. THỐNG KÊ DANH MỤC CÔNG VIỆC ===")
    print(f"Tổng số đầu mục công việc: {len(data)}")
    
    # Top 10 công việc nhiều user nhất
    top_10 = sorted(data, key=lambda x: x["so_user"], reverse=True)[:10]
    print(f"\nTop 10 công việc nhiều user kê khai nhất:")
    for i, dm in enumerate(top_10, 1):
        print(f"  {i}. [{dm['ma_sp']}] {dm['ten_cong_viec'][:40]}... ({dm['so_user']} user)")
    
    # Thống kê theo loại SP
    sp_stats = defaultdict(lambda: {"count": 0, "users": set()})
    for dm in data:
        sp_stats[dm["ma_sp"]]["count"] += 1
        for user in dm["users"]:
            sp_stats[dm["ma_sp"]]["users"].add(user["cong_chuc_id"])
    
    print(f"\nThống kê theo loại SP:")
    for ma_sp in sorted(sp_stats.keys()):
        stats = sp_stats[ma_sp]
        print(f"  {ma_sp}: {stats['count']} đầu mục | {len(stats['users'])} user")


async def main():
    if len(sys.argv) < 3:
        print("Usage: python 05_danh_muc_cong_viec.py <thang> <nam>")
        sys.exit(1)
    
    thang = int(sys.argv[1])
    nam = int(sys.argv[2])
    
    print(f"Đang lấy dữ liệu tháng {thang}/{nam}...")
    data = await get_data(thang, nam)
    
    if not data:
        print(f"Không có dữ liệu cho tháng {thang}/{nam}")
        sys.exit(1)
    
    output_path = f"05_DanhMucCongViec_{thang:02d}_{nam}.xlsx"
    create_excel(data, thang, nam, output_path)


if __name__ == "__main__":
    import warnings
    warnings.filterwarnings("ignore")
    
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(main())
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        try:
            loop.run_until_complete(loop.shutdown_asyncgens())
        except:
            pass
        loop.close()