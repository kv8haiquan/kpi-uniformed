#!/usr/bin/env python3
"""
05b. THỐNG KÊ DANH MỤC CÔNG VIỆC THEO CẤP ĐỘ
================================================
Báo cáo chi tiết số lần kê khai và tổng SP của từng cấp độ cho mỗi đầu mục công việc

Chạy: python 05b_danh_muc_theo_cap_do.py <thang> <nam>
Output: 05b_DanhMucTheoCapDo_02_2026.xlsx
"""

import sys
import asyncio
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, '/root/kpi-haiquan/backend')

from sqlalchemy import text
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
import os
from dotenv import load_dotenv

load_dotenv('/root/kpi-haiquan/backend/.env')

DB_HOST = os.getenv("DB_HOST", "localhost")
DB_PORT = os.getenv("DB_PORT", "5432")
DB_NAME = os.getenv("DB_NAME", "kpi_haiquan")
DB_USER = os.getenv("DB_USER", "kpi_user")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")

DATABASE_URL = f"postgresql+asyncpg://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"


async def get_data(thang: int, nam: int):
    engine = create_async_engine(DATABASE_URL)
    async with AsyncSession(engine) as db:
        # =================================================================
        # 1. Lấy danh sách cấp độ
        # =================================================================
        cap_do_result = await db.execute(text("""
            SELECT ma_cap_do, ten_cap_do, he_so
            FROM cap_do_phuc_tap
            ORDER BY he_so
        """))
        cap_do_list = []
        for row in cap_do_result:
            cap_do_list.append({
                "ma_cap_do": row[0],
                "ten_cap_do": row[1],
                "he_so": float(row[2]),
            })

        # =================================================================
        # 2. Lấy tất cả kê khai công việc chi tiết theo cấp độ
        # =================================================================
        result = await db.execute(text("""
            SELECT
                dm.id as danh_muc_id,
                dm.ten_cong_viec,
                sp.ma_sp,
                sp.ten_sp,
                cd.ma_cap_do,
                cd.ten_cap_do,
                cd.he_so,
                COUNT(*) as so_lan_khai,
                COUNT(DISTINCT kk.cong_chuc_id) as so_user,
                COALESCE(SUM(kk.so_luong), 0) as tong_so_luong,
                COALESCE(SUM(kk.so_sp_goc_quy_doi), 0) as tong_sp_quy_doi
            FROM ke_khai_cong_viec kk
            JOIN danh_muc_sp_cong_viec dm ON dm.id = kk.danh_muc_sp_id
            JOIN sp_cong_viec_chuan sp ON sp.id = dm.sp_chuan_id
            JOIN cap_do_phuc_tap cd ON cd.id = kk.cap_do_id
            WHERE kk.thang = :thang AND kk.nam = :nam
                  AND kk.trang_thai = 'DA_PHE_DUYET' AND kk.is_deleted = false
            GROUP BY dm.id, dm.ten_cong_viec,
                     sp.ma_sp, sp.ten_sp,
                     cd.ma_cap_do, cd.ten_cap_do, cd.he_so
            ORDER BY sp.ma_sp, dm.ten_cong_viec, cd.he_so
        """), {"thang": thang, "nam": nam})

        raw_data = []
        for row in result:
            raw_data.append({
                "danh_muc_id": row[0],
                "ten_cong_viec": row[1],
                "ma_sp": row[2],
                "ten_sp": row[3],
                "ma_cap_do": row[4],
                "ten_cap_do": row[5],
                "he_so": float(row[6]),
                "so_lan_khai": int(row[7]),
                "so_user": int(row[8]),
                "tong_so_luong": float(row[9]),
                "tong_sp_quy_doi": float(row[10]),
            })

        # =================================================================
        # 3. Tổng hợp theo danh mục
        # =================================================================
        danh_muc_map = defaultdict(lambda: {
            "ten_cong_viec": "",
            "ma_sp": "",
            "ten_sp": "",
            "cap_do_details": {},  # ma_cap_do -> {so_lan, so_user, tong_sp, tong_sl}
            "tong_sp": 0,
            "tong_lan_khai": 0,
            "tong_user": 0,
        })

        for item in raw_data:
            dm_id = item["danh_muc_id"]
            dm = danh_muc_map[dm_id]
            dm["ten_cong_viec"] = item["ten_cong_viec"]
            dm["ma_sp"] = item["ma_sp"]
            dm["ten_sp"] = item["ten_sp"]
            dm["tong_sp"] += item["tong_sp_quy_doi"]
            dm["tong_lan_khai"] += item["so_lan_khai"]

            dm["cap_do_details"][item["ma_cap_do"]] = {
                "ten_cap_do": item["ten_cap_do"],
                "he_so": item["he_so"],
                "so_lan_khai": item["so_lan_khai"],
                "so_user": item["so_user"],
                "tong_so_luong": item["tong_so_luong"],
                "tong_sp_quy_doi": item["tong_sp_quy_doi"],
            }

        # Tính tổng user (cần query riêng vì 1 user có thể khai nhiều cấp độ)
        for dm_id in danh_muc_map:
            pass  # Sẽ query bổ sung

        user_count_result = await db.execute(text("""
            SELECT dm.id, COUNT(DISTINCT kk.cong_chuc_id) as so_user
            FROM ke_khai_cong_viec kk
            JOIN danh_muc_sp_cong_viec dm ON dm.id = kk.danh_muc_sp_id
            WHERE kk.thang = :thang AND kk.nam = :nam
                  AND kk.trang_thai = 'DA_PHE_DUYET' AND kk.is_deleted = false
            GROUP BY dm.id
        """), {"thang": thang, "nam": nam})
        for row in user_count_result:
            if row[0] in danh_muc_map:
                danh_muc_map[row[0]]["tong_user"] = int(row[1])

        # =================================================================
        # 4. Lấy chi tiết user theo từng công việc + cấp độ
        # =================================================================
        user_detail_result = await db.execute(text("""
            SELECT
                dm.id as danh_muc_id,
                cc.ho_ten,
                cc.ma_cc,
                dv.ten_don_vi,
                cd.ma_cap_do,
                COUNT(*) as so_lan_khai,
                COALESCE(SUM(kk.so_luong), 0) as tong_so_luong,
                COALESCE(SUM(kk.so_sp_goc_quy_doi), 0) as tong_sp_quy_doi
            FROM ke_khai_cong_viec kk
            JOIN danh_muc_sp_cong_viec dm ON dm.id = kk.danh_muc_sp_id
            JOIN cong_chuc cc ON cc.id = kk.cong_chuc_id
            LEFT JOIN don_vi dv ON dv.id = cc.don_vi_id
            JOIN cap_do_phuc_tap cd ON cd.id = kk.cap_do_id
            WHERE kk.thang = :thang AND kk.nam = :nam
                  AND kk.trang_thai = 'DA_PHE_DUYET' AND kk.is_deleted = false
            GROUP BY dm.id, cc.ho_ten, cc.ma_cc, dv.ten_don_vi, cd.ma_cap_do
            ORDER BY dm.id, dv.ten_don_vi, cc.ho_ten, cd.ma_cap_do
        """), {"thang": thang, "nam": nam})

        user_details = defaultdict(list)  # dm_id -> list of user+cap_do rows
        for row in user_detail_result:
            user_details[row[0]].append({
                "ho_ten": row[1],
                "ma_cc": row[2],
                "don_vi": row[3] or "",
                "ma_cap_do": row[4],
                "so_lan_khai": int(row[5]),
                "tong_so_luong": float(row[6]),
                "tong_sp_quy_doi": float(row[7]),
            })

        # Convert to list
        danh_muc_list = []
        for dm_id, dm in danh_muc_map.items():
            dm["danh_muc_id"] = dm_id
            dm["user_details"] = user_details.get(dm_id, [])
            danh_muc_list.append(dm)

        danh_muc_list = sorted(danh_muc_list, key=lambda x: (x["ma_sp"], x["ten_cong_viec"]))

    await engine.dispose()
    return danh_muc_list, cap_do_list


def create_excel(data: list, cap_do_list: list, thang: int, nam: int, output_path: str):
    wb = Workbook()

    # Styles
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    dm_header_fill = PatternFill("solid", fgColor="2F5496")
    sub_header_fill = PatternFill("solid", fgColor="D6DCE4")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    wrap_align = Alignment(wrap_text=True, vertical='center', horizontal='center')

    SP_FILLS = {
        "SP1": PatternFill("solid", fgColor="DAEEF3"),
        "SP2": PatternFill("solid", fgColor="E2EFDA"),
        "SP3": PatternFill("solid", fgColor="FDE9D9"),
        "SP4": PatternFill("solid", fgColor="E4DFEC"),
    }

    # Màu cho từng cấp độ
    CD_FILLS = {
        0: PatternFill("solid", fgColor="E2EFDA"),  # xanh nhạt
        1: PatternFill("solid", fgColor="DAEEF3"),  # xanh dương nhạt
        2: PatternFill("solid", fgColor="FDE9D9"),  # cam nhạt
        3: PatternFill("solid", fgColor="E4DFEC"),  # tím nhạt
        4: PatternFill("solid", fgColor="FFC7CE"),  # đỏ nhạt
        5: PatternFill("solid", fgColor="FFEB9C"),  # vàng nhạt
    }

    ma_cap_do_list = [cd["ma_cap_do"] for cd in cap_do_list]

    # =========================================================================
    # SHEET 1: TỔNG HỢP - SỐ LẦN KÊ KHAI THEO CẤP ĐỘ
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Số lần kê khai theo CĐ"

    ws1['A1'] = f"THỐNG KÊ SỐ LẦN KÊ KHAI THEO CẤP ĐỘ - THÁNG {thang}/{nam}"
    ws1['A1'].font = title_font
    ws1.merge_cells(f'A1:{get_column_letter(5 + len(cap_do_list) * 2)}1')

    ws1['A2'] = f"Tổng số đầu mục công việc: {len(data)}"
    ws1['A2'].font = Font(bold=True)

    # Header row 1: merge cột cho từng cấp độ
    row = 4
    fixed_headers = ["STT", "Loại SP", "Tên công việc", "Tổng user", "Tổng lần khai"]
    for col, h in enumerate(fixed_headers, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = wrap_align
        # Merge 2 rows cho fixed headers
        ws1.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col)

    col_start = len(fixed_headers) + 1
    for i, cd in enumerate(cap_do_list):
        # Merge 2 cột cho mỗi cấp độ (Số lần | SP)
        merge_start = col_start + i * 2
        merge_end = merge_start + 1
        ws1.merge_cells(start_row=row, start_column=merge_start, end_row=row, end_column=merge_end)
        cell = ws1.cell(row=row, column=merge_start, value=f"{cd['ma_cap_do']}\n({cd['ten_cap_do']})")
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = wrap_align
        # Border cho merged cell
        ws1.cell(row=row, column=merge_end).border = border
        ws1.cell(row=row, column=merge_end).fill = header_fill

    # Cột tổng SP cuối
    tong_sp_col = col_start + len(cap_do_list) * 2
    ws1.merge_cells(start_row=row, start_column=tong_sp_col, end_row=row + 1, end_column=tong_sp_col)
    cell = ws1.cell(row=row, column=tong_sp_col, value="Tổng SP\nquy đổi")
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = border
    cell.alignment = wrap_align

    # Header row 2: sub-headers cho từng cấp độ
    row = 5
    for i in range(len(cap_do_list)):
        col_lan = col_start + i * 2
        col_sp = col_lan + 1
        fill = CD_FILLS.get(i, CD_FILLS[0])

        cell_lan = ws1.cell(row=row, column=col_lan, value="Lần")
        cell_lan.font = Font(bold=True, size=9)
        cell_lan.fill = fill
        cell_lan.border = border
        cell_lan.alignment = center_align

        cell_sp = ws1.cell(row=row, column=col_sp, value="SP")
        cell_sp.font = Font(bold=True, size=9)
        cell_sp.fill = fill
        cell_sp.border = border
        cell_sp.alignment = center_align

    ws1.cell(row=row, column=tong_sp_col).border = border

    # Data rows
    row = 6
    tong_cot = defaultdict(float)  # Để tính tổng cuối bảng
    for i, dm in enumerate(data, 1):
        ws1.cell(row=row, column=1, value=i).border = border
        ws1.cell(row=row, column=1).alignment = center_align

        ma_sp_cell = ws1.cell(row=row, column=2, value=dm["ma_sp"])
        ma_sp_cell.border = border
        ma_sp_cell.alignment = center_align
        ma_sp_cell.fill = SP_FILLS.get(dm["ma_sp"], SP_FILLS["SP1"])
        ma_sp_cell.font = Font(bold=True)

        ws1.cell(row=row, column=3, value=dm["ten_cong_viec"]).border = border

        ws1.cell(row=row, column=4, value=dm["tong_user"]).border = border
        ws1.cell(row=row, column=4).alignment = center_align

        ws1.cell(row=row, column=5, value=dm["tong_lan_khai"]).border = border
        ws1.cell(row=row, column=5).alignment = center_align

        for j, cd in enumerate(cap_do_list):
            col_lan = col_start + j * 2
            col_sp = col_lan + 1
            fill = CD_FILLS.get(j, CD_FILLS[0])

            detail = dm["cap_do_details"].get(cd["ma_cap_do"])
            if detail:
                so_lan = detail["so_lan_khai"]
                sp = detail["tong_sp_quy_doi"]

                cell_lan = ws1.cell(row=row, column=col_lan, value=so_lan)
                cell_lan.fill = fill
                cell_sp_val = ws1.cell(row=row, column=col_sp, value=round(sp, 1))
                cell_sp_val.fill = fill

                tong_cot[f"lan_{j}"] += so_lan
                tong_cot[f"sp_{j}"] += sp
            else:
                ws1.cell(row=row, column=col_lan, value="")
                ws1.cell(row=row, column=col_sp, value="")

            ws1.cell(row=row, column=col_lan).border = border
            ws1.cell(row=row, column=col_lan).alignment = center_align
            ws1.cell(row=row, column=col_sp).border = border
            ws1.cell(row=row, column=col_sp).alignment = center_align

        tong_sp_cell = ws1.cell(row=row, column=tong_sp_col, value=round(dm["tong_sp"], 1))
        tong_sp_cell.border = border
        tong_sp_cell.alignment = center_align
        tong_sp_cell.font = Font(bold=True)

        row += 1

    # Dòng tổng cộng
    tong_fill = PatternFill("solid", fgColor="FFC000")
    ws1.cell(row=row, column=1).border = border
    ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    cell_tong = ws1.cell(row=row, column=2, value="TỔNG CỘNG")
    cell_tong.font = Font(bold=True, size=11)
    cell_tong.fill = tong_fill
    cell_tong.border = border
    ws1.cell(row=row, column=3).border = border
    ws1.cell(row=row, column=3).fill = tong_fill

    tong_user_all = sum(dm["tong_user"] for dm in data)
    tong_lan_all = sum(dm["tong_lan_khai"] for dm in data)
    ws1.cell(row=row, column=4, value=tong_user_all).border = border
    ws1.cell(row=row, column=4).alignment = center_align
    ws1.cell(row=row, column=4).font = Font(bold=True)
    ws1.cell(row=row, column=4).fill = tong_fill

    ws1.cell(row=row, column=5, value=tong_lan_all).border = border
    ws1.cell(row=row, column=5).alignment = center_align
    ws1.cell(row=row, column=5).font = Font(bold=True)
    ws1.cell(row=row, column=5).fill = tong_fill

    for j in range(len(cap_do_list)):
        col_lan = col_start + j * 2
        col_sp = col_lan + 1

        cell_lan = ws1.cell(row=row, column=col_lan, value=int(tong_cot.get(f"lan_{j}", 0)))
        cell_lan.border = border
        cell_lan.alignment = center_align
        cell_lan.font = Font(bold=True)
        cell_lan.fill = tong_fill

        cell_sp = ws1.cell(row=row, column=col_sp, value=round(tong_cot.get(f"sp_{j}", 0), 1))
        cell_sp.border = border
        cell_sp.alignment = center_align
        cell_sp.font = Font(bold=True)
        cell_sp.fill = tong_fill

    tong_sp_all = sum(dm["tong_sp"] for dm in data)
    cell_tsp = ws1.cell(row=row, column=tong_sp_col, value=round(tong_sp_all, 1))
    cell_tsp.border = border
    cell_tsp.alignment = center_align
    cell_tsp.font = Font(bold=True, size=12)
    cell_tsp.fill = tong_fill

    # Column widths
    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 10
    ws1.column_dimensions['C'].width = 50
    ws1.column_dimensions['D'].width = 10
    ws1.column_dimensions['E'].width = 12
    for j in range(len(cap_do_list)):
        ws1.column_dimensions[get_column_letter(col_start + j * 2)].width = 8
        ws1.column_dimensions[get_column_letter(col_start + j * 2 + 1)].width = 10
    ws1.column_dimensions[get_column_letter(tong_sp_col)].width = 12
    ws1.row_dimensions[4].height = 45

    # =========================================================================
    # SHEET 2: CHI TIẾT USER THEO TỪNG CÔNG VIỆC + CẤP ĐỘ
    # =========================================================================
    ws2 = wb.create_sheet("Chi tiết user theo CĐ")

    ws2['A1'] = f"CHI TIẾT USER KÊ KHAI THEO TỪNG CẤP ĐỘ - THÁNG {thang}/{nam}"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:I1')

    row = 3
    for dm_idx, dm in enumerate(data, 1):
        # Header cho mỗi danh mục
        cap_do_summary_parts = []
        for cd in cap_do_list:
            detail = dm["cap_do_details"].get(cd["ma_cap_do"])
            if detail:
                cap_do_summary_parts.append(
                    f"{cd['ma_cap_do']}: {detail['so_lan_khai']} lần/{detail['tong_sp_quy_doi']:,.1f} SP"
                )

        ws2.merge_cells(f'A{row}:I{row}')
        header_text = f"{dm_idx}. [{dm['ma_sp']}] {dm['ten_cong_viec']}"
        cell = ws2.cell(row=row, column=1, value=header_text)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = dm_header_fill
        for col in range(2, 10):
            ws2.cell(row=row, column=col).fill = dm_header_fill
        row += 1

        # Dòng tổng hợp cấp độ
        ws2.merge_cells(f'A{row}:I{row}')
        summary_text = f"   Tổng: {dm['tong_user']} user | {dm['tong_lan_khai']} lần | {dm['tong_sp']:,.1f} SP  ——  " + " | ".join(cap_do_summary_parts)
        cell = ws2.cell(row=row, column=1, value=summary_text)
        cell.font = Font(italic=True, size=10, color="333333")
        cell.fill = PatternFill("solid", fgColor="F2F2F2")
        for col in range(2, 10):
            ws2.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F2F2F2")
        row += 1

        # Sub-header
        sub_headers = ["STT", "Họ và tên", "Mã CC", "Đơn vị", "Cấp độ", "Số lần khai", "Số lượng", "SP quy đổi"]
        for col, h in enumerate(sub_headers, 1):
            cell = ws2.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True, size=10)
            cell.fill = sub_header_fill
            cell.border = border
            cell.alignment = center_align
        row += 1

        # User details - group theo user
        user_groups = defaultdict(list)
        for ud in dm["user_details"]:
            key = (ud["ho_ten"], ud["ma_cc"], ud["don_vi"])
            user_groups[key].append(ud)

        # Sort theo đơn vị, tên
        sorted_users = sorted(user_groups.items(), key=lambda x: (x[0][2], x[0][0]))

        u_idx = 0
        for (ho_ten, ma_cc, don_vi), details in sorted_users:
            u_idx += 1
            # Sort cap do theo thứ tự
            details = sorted(details, key=lambda x: ma_cap_do_list.index(x["ma_cap_do"]) if x["ma_cap_do"] in ma_cap_do_list else 99)

            first_row = row
            for d_idx, d in enumerate(details):
                if d_idx == 0:
                    ws2.cell(row=row, column=1, value=u_idx).border = border
                    ws2.cell(row=row, column=1).alignment = center_align
                    ws2.cell(row=row, column=2, value=ho_ten).border = border
                    ws2.cell(row=row, column=3, value=ma_cc).border = border
                    ws2.cell(row=row, column=4, value=don_vi).border = border
                else:
                    # Dòng tiếp theo của cùng user - để trống tên
                    ws2.cell(row=row, column=1).border = border
                    ws2.cell(row=row, column=2).border = border
                    ws2.cell(row=row, column=3).border = border
                    ws2.cell(row=row, column=4).border = border

                cd_idx = ma_cap_do_list.index(d["ma_cap_do"]) if d["ma_cap_do"] in ma_cap_do_list else 0
                cd_fill = CD_FILLS.get(cd_idx, CD_FILLS[0])

                cap_do_cell = ws2.cell(row=row, column=5, value=d["ma_cap_do"])
                cap_do_cell.border = border
                cap_do_cell.alignment = center_align
                cap_do_cell.fill = cd_fill
                cap_do_cell.font = Font(bold=True)

                ws2.cell(row=row, column=6, value=d["so_lan_khai"]).border = border
                ws2.cell(row=row, column=6).alignment = center_align

                ws2.cell(row=row, column=7, value=d["tong_so_luong"]).border = border
                ws2.cell(row=row, column=7).alignment = center_align

                ws2.cell(row=row, column=8, value=round(d["tong_sp_quy_doi"], 1)).border = border
                ws2.cell(row=row, column=8).alignment = center_align

                row += 1

            # Merge STT, tên, mã CC, đơn vị nếu user có nhiều cấp độ
            if len(details) > 1:
                for merge_col in [1, 2, 3, 4]:
                    ws2.merge_cells(start_row=first_row, start_column=merge_col,
                                    end_row=first_row + len(details) - 1, end_column=merge_col)
                    ws2.cell(row=first_row, column=merge_col).alignment = Alignment(vertical='center')

        row += 1  # Khoảng cách giữa các danh mục

    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 30
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 12
    ws2.column_dimensions['G'].width = 12
    ws2.column_dimensions['H'].width = 12

    # =========================================================================
    # SHEET 3: TỔNG HỢP CẤP ĐỘ TOÀN CHI CỤC
    # =========================================================================
    ws3 = wb.create_sheet("Tổng hợp cấp độ")

    ws3['A1'] = f"TỔNG HỢP SỬ DỤNG CẤP ĐỘ PHỨC TẠP - THÁNG {thang}/{nam}"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:G1')

    # Header
    headers3 = ["Cấp độ", "Tên cấp độ", "Hệ số", "Số đầu mục CV", "Tổng lần kê khai", "Tổng SP quy đổi", "% SP"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    # Tổng hợp theo cấp độ
    cd_totals = defaultdict(lambda: {"ten": "", "he_so": 0, "so_dm": 0, "so_lan": 0, "sp": 0})
    for dm in data:
        for ma_cd, detail in dm["cap_do_details"].items():
            cd_totals[ma_cd]["ten"] = detail["ten_cap_do"]
            cd_totals[ma_cd]["he_so"] = detail["he_so"]
            cd_totals[ma_cd]["so_dm"] += 1
            cd_totals[ma_cd]["so_lan"] += detail["so_lan_khai"]
            cd_totals[ma_cd]["sp"] += detail["tong_sp_quy_doi"]

    tong_sp_all = sum(v["sp"] for v in cd_totals.values())

    row = 4
    for i, cd in enumerate(cap_do_list):
        totals = cd_totals.get(cd["ma_cap_do"])
        if not totals:
            continue

        fill = CD_FILLS.get(i, CD_FILLS[0])

        cell = ws3.cell(row=row, column=1, value=cd["ma_cap_do"])
        cell.border = border
        cell.alignment = center_align
        cell.fill = fill
        cell.font = Font(bold=True)

        ws3.cell(row=row, column=2, value=totals["ten"]).border = border
        ws3.cell(row=row, column=2).fill = fill

        ws3.cell(row=row, column=3, value=totals["he_so"]).border = border
        ws3.cell(row=row, column=3).alignment = center_align

        ws3.cell(row=row, column=4, value=totals["so_dm"]).border = border
        ws3.cell(row=row, column=4).alignment = center_align

        ws3.cell(row=row, column=5, value=totals["so_lan"]).border = border
        ws3.cell(row=row, column=5).alignment = center_align

        ws3.cell(row=row, column=6, value=round(totals["sp"], 1)).border = border
        ws3.cell(row=row, column=6).alignment = center_align

        pct = (totals["sp"] / tong_sp_all * 100) if tong_sp_all > 0 else 0
        pct_cell = ws3.cell(row=row, column=7, value=f"{pct:.1f}%")
        pct_cell.border = border
        pct_cell.alignment = center_align
        if pct >= 30:
            pct_cell.font = Font(bold=True, color="C00000")

        row += 1

    # Dòng tổng
    tong_fill = PatternFill("solid", fgColor="FFC000")
    ws3.cell(row=row, column=1, value="TỔNG").font = Font(bold=True)
    ws3.cell(row=row, column=1).fill = tong_fill
    ws3.cell(row=row, column=1).border = border

    for col in [2, 3]:
        ws3.cell(row=row, column=col).fill = tong_fill
        ws3.cell(row=row, column=col).border = border

    ws3.cell(row=row, column=4, value=len(data)).border = border
    ws3.cell(row=row, column=4).alignment = center_align
    ws3.cell(row=row, column=4).font = Font(bold=True)
    ws3.cell(row=row, column=4).fill = tong_fill

    ws3.cell(row=row, column=5, value=sum(v["so_lan"] for v in cd_totals.values())).border = border
    ws3.cell(row=row, column=5).alignment = center_align
    ws3.cell(row=row, column=5).font = Font(bold=True)
    ws3.cell(row=row, column=5).fill = tong_fill

    ws3.cell(row=row, column=6, value=round(tong_sp_all, 1)).border = border
    ws3.cell(row=row, column=6).alignment = center_align
    ws3.cell(row=row, column=6).font = Font(bold=True)
    ws3.cell(row=row, column=6).fill = tong_fill

    ws3.cell(row=row, column=7, value="100%").border = border
    ws3.cell(row=row, column=7).alignment = center_align
    ws3.cell(row=row, column=7).font = Font(bold=True)
    ws3.cell(row=row, column=7).fill = tong_fill

    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 25
    ws3.column_dimensions['C'].width = 10
    ws3.column_dimensions['D'].width = 15
    ws3.column_dimensions['E'].width = 18
    ws3.column_dimensions['F'].width = 18
    ws3.column_dimensions['G'].width = 10

    # Save
    wb.save(output_path)

    # Console output
    print(f"Đã tạo file: {output_path}")
    print(f"\n=== THỐNG KÊ DANH MỤC CÔNG VIỆC THEO CẤP ĐỘ ===")
    print(f"Tổng số đầu mục công việc: {len(data)}")

    print(f"\nTổng hợp theo cấp độ:")
    print(f"{'Cấp độ':<10} {'Tên':<25} {'Hệ số':<8} {'Số lần khai':>12} {'Tổng SP':>12} {'%':>8}")
    print("-" * 80)
    for cd in cap_do_list:
        totals = cd_totals.get(cd["ma_cap_do"])
        if totals:
            pct = (totals["sp"] / tong_sp_all * 100) if tong_sp_all > 0 else 0
            print(f"{cd['ma_cap_do']:<10} {totals['ten']:<25} {totals['he_so']:<8.1f} {totals['so_lan']:>12} {totals['sp']:>12,.1f} {pct:>7.1f}%")
    print("-" * 80)
    print(f"{'TỔNG':<10} {'':<25} {'':<8} {sum(v['so_lan'] for v in cd_totals.values()):>12} {tong_sp_all:>12,.1f} {'100.0%':>8}")


async def main():
    if len(sys.argv) < 3:
        print("Usage: python 05b_danh_muc_theo_cap_do.py <thang> <nam>")
        sys.exit(1)

    thang = int(sys.argv[1])
    nam = int(sys.argv[2])

    print(f"Đang lấy dữ liệu tháng {thang}/{nam}...")
    data, cap_do_list = await get_data(thang, nam)

    if not data:
        print(f"Không có dữ liệu cho tháng {thang}/{nam}")
        sys.exit(1)

    output_path = f"05b_DanhMucTheoCapDo_{thang:02d}_{nam}.xlsx"
    create_excel(data, cap_do_list, thang, nam, output_path)


if __name__ == "__main__":
    import warnings
    warnings.filterwarnings("ignore")

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(main())
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        try:
            loop.run_until_complete(loop.shutdown_asyncgens())
        except:
            pass
        loop.close()
