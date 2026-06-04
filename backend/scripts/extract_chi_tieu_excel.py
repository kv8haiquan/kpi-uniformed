#!/usr/bin/env python3
"""
scripts/extract_chi_tieu_excel.py
=================================
Trich xuat "chi tieu giao nam" tu file Rà soát chỉ tiêu (Excel) ra CSV preview.
KHONG ghi DB — chi doc Excel + xuat CSV de nguoi dung ra soat.

Chay:
    cd backend && python scripts/extract_chi_tieu_excel.py

Output: scripts/dumps/chi_tieu_giao_nam_preview.csv

Moi sheet 1 lĩnh vực, layout khac nhau -> cau hinh cot rieng (xem SHEETS).
Voi sheet "kho" (QLRR ti le, CBL theo to/tram) -> trich best-effort + co canh bao.
"""

import csv
import sys
import unicodedata
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
REPO = SCRIPT_DIR.parent.parent
EXCEL = REPO / "docs" / "Chi Tieu" / "Chỉ_tiêu_KPI_T4_2026_TatCaDonVi_highlight.xlsx"
OUT = SCRIPT_DIR / "dumps" / "chi_tieu_giao_nam_preview.csv"
NAM = 2026

# ---- Map ten don vi (Excel) -> ma_don_vi (public.don_vi) ----
DON_VI_MAP = {
    # ten day du + viet tat
    "mong cai": "HQCK-MC", "mc": "HQCK-MC",
    "hon gai": "HQCK-HG", "hg": "HQCK-HG",
    "cam pha": "HQCK-CP", "cp": "HQCK-CP",
    "van gia": "HQCK-VG", "vg": "HQCK-VG",
    "hoanh mo": "HQCK-HM", "hm": "HQCK-HM",
    "bac phong sinh": "HQCK-BPS", "bps": "HQCK-BPS",
    # don vi STQ / KSHQ
    "pt&ktstq": "PTSTQ", "pt va ktstq": "PTSTQ", "ptstq": "PTSTQ", "ktstq": "PTSTQ",
    "doi pt&ktstq": "PTSTQ", "doi pt va ktstq": "PTSTQ",
    "doi kshq": "KSHQ", "kshq": "KSHQ", "doi kiem soat hq": "KSHQ", "doi ks hq": "KSHQ",
    # phong
    "nvhq": "NVHQ", "phong nvhq": "NVHQ", "phong nghiep vu": "NVHQ",
    "qlrr": "QLRR", "phong qlrr": "QLRR",
    "cntt": "CNTT", "phong cntt": "CNTT",
    "van phong": "VP", "vp": "VP",
    "tccb": "TCCB", "phong tccb": "TCCB", "phong tcb": "TCCB", "phong tcbc": "TCCB",
}

# Gia tri trong cot giao nam co nghia "khong co dong giao nam" -> bo qua
KHONG_GIAO = {"khong giao chi tieu", "khong giao", "khong co", "-"}


def khong_dau(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.replace("đ", "d").replace("Đ", "D").lower().strip()


def map_don_vi(ten: str):
    k = khong_dau(ten)
    return DON_VI_MAP.get(k)


def parse_num(v):
    """Parse so co the co dau , thap phan hoac . phan cach nghin."""
    if v is None:
        return None
    s = str(v).strip().replace("%", "")
    if not s or "/" in s:
        return None
    if "." in s and "," in s:           # 1.234,5 -> nghin . + thap phan ,
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:                       # 13,4 -> thap phan
        s = s.replace(",", ".")
    elif s.count(".") == 1 and len(s.split(".")[1]) == 3:  # 4.700 -> phan cach nghin
        s = s.replace(".", "")
    try:
        return float(s)
    except ValueError:
        return None


# ---- Cau hinh tung BANG. Cot target = cot "Chỉ tiêu giao năm 2026" (da ghi ro trong file). ----
SHEETS = [
    {   # Giám sát quản lý — bảng 1 (chỉ tiêu 1-3)
        "sheet": "Giám sát, Thuế", "linh_vuc": "Giám sát quản lý",
        "data_start": 5, "data_end": 12, "donvi_col": 1,
        "targets": [
            (5, "Tăng kim ngạch XNK (không gồm KNQ, TNTX)", "PHAP_LENH"),
            (11, "Tăng doanh nghiệp làm thủ tục", "PHAP_LENH"),
            (17, "Hội nghị đối thoại, tham vấn", "PHAP_LENH"),
        ],
    },
    {   # Giám sát quản lý — bảng 2 (chỉ tiêu 4-6)
        "sheet": "Giám sát, Thuế", "linh_vuc": "Giám sát quản lý",
        "data_start": 17, "data_end": 24, "donvi_col": 1,
        "targets": [
            (5, "Rà soát, đề xuất cắt giảm TTHC", "PHAP_LENH"),
            (11, "Phát hiện vi phạm SHTT/hàng giả", "PHAP_LENH"),
            (17, "Đề xuất cung cấp thông tin KTSTQ", "PHAP_LENH"),
        ],
    },
    {   # Thuế XNK — bảng 3
        "sheet": "Giám sát, Thuế", "linh_vuc": "Thuế XNK",
        "data_start": 31, "data_end": 37, "donvi_col": 1,
        "targets": [
            (5, "Số thu thuế XNK (pháp lệnh 18.300 tỷ)", "PHAP_LENH"),
            (11, "Số thu thuế XNK (phấn đấu 25.000 tỷ)", "PHAN_DAU"),
        ],
    },
    {   # KTSTQ
        "sheet": " STQ, Đào tạo", "linh_vuc": "Kiểm tra sau thông quan",
        "data_start": 3, "data_end": 12, "donvi_col": 1,
        "targets": [
            (5, "Số cuộc kiểm tra STQ", "PHAP_LENH"),
            (11, "Số thu NSNN (KTSTQ)", "PHAP_LENH"),
            (12, "Số thu NSNN (KTSTQ)", "PHAN_DAU"),
            (15, "Tỷ lệ phát hiện vi phạm KTSTQ", "PHAP_LENH"),
            (21, "Phiếu cung cấp thông tin KTSTQ", "PHAP_LENH"),
        ],
    },
    {   # Đào tạo, tập huấn
        "sheet": " STQ, Đào tạo", "linh_vuc": "Đào tạo, tập huấn",
        "data_start": 16, "data_end": 23, "donvi_col": 1,
        "targets": [
            (5, "Số lớp tập huấn", "PHAP_LENH"),
            (11, "Số lượt người tham gia tập huấn", "PHAP_LENH"),
        ],
    },
    {   # QLRR (cấu trúc cột phức tạp — best-effort theo cột "Chỉ tiêu giao năm 2026")
        "sheet": "V. QLRR", "linh_vuc": "Quản lý rủi ro",
        "data_start": 6, "data_end": 15, "donvi_col": 1,
        "targets": [
            (5, "Số hồ sơ DN trọng điểm", "PHAP_LENH"),
            (11, "Số tiêu chí phân tích (doanh nghiệp)", "PHAP_LENH"),
            (17, "Số tiêu chí phân tích (mặt hàng)", "PHAP_LENH"),
            (23, "Số lượng DN được giao thu (Cục HQ giao)", "PHAP_LENH"),
            (29, "Số lượng DN giao thu (Chi cục giao)", "PHAP_LENH"),
        ],
    },
    {   # CBL — chỉ block 1 (r3..r10); KHÔNG còn TT MC/TT HL trong file này
        "sheet": "VI. CBL", "linh_vuc": "Kiểm soát chống buôn lậu",
        "data_start": 3, "data_end": 11, "donvi_col": 1,
        "targets": [
            (6, "Số vụ chủ trì bắt giữ", "PHAP_LENH"),
            (12, "Số vụ phối hợp bắt giữ", "PHAP_LENH"),
            (18, "Tổng trị giá chủ trì bắt giữ", "PHAP_LENH"),
            (24, "Số thu NSNN từ xử phạt VPHC", "PHAP_LENH"),
            (30, "HQ khởi tố (vụ)", "PHAP_LENH"),
        ],
    },
    {   # Truyền thông — 4 kênh (cột "CHỈ TIÊU NĂM"), KHÔNG lấy "TỶ LỆ SO CHỈ TIÊU NĂM"
        "sheet": "Truyền thông", "linh_vuc": "Truyền thông",
        "data_start": 4, "data_end": 17, "donvi_col": 1,
        "targets": [
            (5, "Tin/bài - Cổng TTĐT Chi cục", "PHAP_LENH"),
            (11, "Tin/bài - Fanpage DDCI Quảng Ninh", "PHAP_LENH"),
            (17, "Tin/bài - Fanpage DDCI Chi cục", "PHAP_LENH"),
            (23, "Tin/bài - ISEC", "PHAP_LENH"),
        ],
    },
]


def la_dong_nhom(stt, ten):
    """Bo qua dong nhom (I, II, III...) hoac dong tong."""
    s = khong_dau(ten)
    if s == "" or s.startswith("tong") or s in ("don vi", "dvi", "stt"):
        return True
    st = str(stt or "").strip().upper()
    return st in ("I", "II", "III", "IV", "V", "VI", "VII")


def main():
    if not EXCEL.exists():
        print(f"Khong tim thay file Excel: {EXCEL}"); sys.exit(1)
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    OUT.parent.mkdir(parents=True, exist_ok=True)

    out_rows = []
    canh_bao = 0
    so_khong_giao = 0
    for cfg in SHEETS:
        if cfg["sheet"] not in wb.sheetnames:
            print(f"  ! Bo qua sheet thieu: {cfg['sheet']}"); continue
        ws = wb[cfg["sheet"]]
        rows = list(ws.iter_rows(values_only=True))
        data_end = cfg.get("data_end", len(rows))
        trong_lien_tiep = 0
        for ri in range(cfg["data_start"], min(data_end, len(rows))):
            row = rows[ri]
            ten_dv = row[cfg["donvi_col"]] if len(row) > cfg["donvi_col"] else None
            stt = row[0] if row else None
            if ten_dv is None or str(ten_dv).strip() == "":
                trong_lien_tiep += 1
                if trong_lien_tiep >= 3:
                    break          # het bang
                continue
            trong_lien_tiep = 0
            if la_dong_nhom(stt, ten_dv):
                continue
            ma_dv = map_don_vi(str(ten_dv))
            for col, ten_ct, loai in cfg["targets"]:
                raw = row[col] if len(row) > col else None
                if raw is None or str(raw).strip() == "":
                    continue
                if khong_dau(raw).strip() in KHONG_GIAO:
                    so_khong_giao += 1
                    continue           # don vi khong duoc giao chi tieu nay
                num = parse_num(raw)
                ghi_chu = []
                if ma_dv is None:
                    ghi_chu.append("DON VI chua map")
                if num is None:
                    ghi_chu.append("gia tri khong phai so")
                if ghi_chu:
                    canh_bao += 1
                out_rows.append({
                    "linh_vuc": cfg["linh_vuc"], "sheet": cfg["sheet"],
                    "chi_tieu": ten_ct, "loai_muc": loai,
                    "don_vi_excel": str(ten_dv).strip(), "ma_don_vi": ma_dv or "",
                    "nam": NAM,
                    "gia_tri_giao_raw": "" if raw is None else str(raw),
                    "gia_tri_giao_num": "" if num is None else num,
                    "ghi_chu": "; ".join(ghi_chu),
                })

    with open(OUT, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=[
            "linh_vuc", "sheet", "chi_tieu", "loai_muc", "don_vi_excel",
            "ma_don_vi", "nam", "gia_tri_giao_raw", "gia_tri_giao_num", "ghi_chu",
        ])
        w.writeheader()
        w.writerows(out_rows)

    print(f"✓ Trich xuat {len(out_rows)} dong giao nam -> {OUT}")
    print(f"  Trong do {canh_bao} dong co canh bao; {so_khong_giao} o 'Khong giao chi tieu' (da bo qua).")
    # Thong ke nhanh theo linh vuc
    from collections import Counter
    c = Counter(r["linh_vuc"] for r in out_rows)
    for lv, n in c.items():
        print(f"    - {lv}: {n} dong")


if __name__ == "__main__":
    main()
