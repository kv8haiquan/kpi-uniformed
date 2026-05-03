"""
scripts/seed_pl3_catalog.py
===========================
Phase A.7 — Seed ~2.812 mục PL3 từ file Excel vào danh_muc_sp_cong_viec.

Cách chạy:
    cd backend && source venv/bin/activate
    python scripts/seed_pl3_catalog.py

Input:
    docs/Danh mục công việc.xlsx (sheet PL3, ~3.455 rows × 28 cols)

Logic (đã tinh giản 28/04/2026):
    1. Đọc sheet PL3 từ row 8 trở đi (row 1-7 là header).
    2. Detect section headers bằng regex `^([IVXLCDM]+)\\.\\s*(.+)$` ở cột A.
    3. Lấy 5 cột "quan trọng":
         - col E: Phân nhóm (text "Nhóm X" → parse số 1-5)
         - col K: diem_cham
         - col L: he_so_quy_doi
         - col C: cong_viec_chi_tiet
         - col D: san_pham_dau_ra
    4. Derive khung_diem_toi_da từ phan_nhom: {1:100, 2:200, 3:300, 4:400, 5:500}.
    5. KHÔNG validate sum 4 cột chấm hay he_so == diem_cham/25 (trust Excel).
    6. Skip:
         - Section header rows (đã pre-scan)
         - Row trước section header đầu tiên (row 8 "đơn vị SPCV chuẩn")
         - Row thiếu cong_viec_chi_tiet/san_pham_dau_ra/diem_cham/he_so/phan_nhom
    7. Idempotent: ON CONFLICT (ma_danh_muc) DO UPDATE.

Output: "X inserted, Y updated, Z skipped".
"""

from __future__ import annotations

import re
import sys
from decimal import Decimal
from pathlib import Path
from typing import Optional

# Cho phép import app.config từ scripts/
PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl
from sqlalchemy import create_engine, text

from app.config import settings


# =============================================================================
# CONFIG
# =============================================================================

EXCEL_PATH = PROJECT_ROOT.parent / "docs" / "Danh mục công việc.xlsx"
SHEET_NAME = "PL3"
DATA_START_ROW = 8

# Mapping nhom_pl3 → khung_diem_toi_da (Quyết định nghiệp vụ)
NHOM_KHUNG_MAP = {1: 100, 2: 200, 3: 300, 4: 400, 5: 500}

# Cột Excel (1-based, openpyxl)
COL_STT = 1
COL_NHIEM_VU = 2
COL_CONG_VIEC = 3
COL_SAN_PHAM = 4
COL_PHAN_NHOM = 5
COL_KHUNG_DIEM = 6  # Dùng làm fallback khi phan_nhom (col E) trống
COL_DIEM_CHAM = 11
COL_HE_SO = 12
COL_GHI_CHU = 13

# Fallback: khung_diem → nhom_pl3
KHUNG_TO_NHOM = {100: 1, 200: 2, 300: 3, 400: 4, 500: 5}

SECTION_REGEX = re.compile(r"^([IVXLCDM]+)\.\s*(.+)$")


# =============================================================================
# HELPERS
# =============================================================================

def to_int(val) -> Optional[int]:
    if val is None or val == "":
        return None
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return None


def to_decimal(val) -> Optional[Decimal]:
    if val is None or val == "":
        return None
    try:
        return Decimal(str(val))
    except (TypeError, ValueError):
        return None


def detect_nhom_from_phan_nhom(s: Optional[str]) -> Optional[int]:
    """'Nhóm 3' → 3."""
    if not s:
        return None
    m = re.search(r"(\d+)", str(s))
    if not m:
        return None
    n = int(m.group(1))
    return n if 1 <= n <= 5 else None


def is_section_header(val) -> Optional[tuple[str, str]]:
    if val is None:
        return None
    s = str(val).strip()
    m = SECTION_REGEX.match(s)
    if m:
        return m.group(1).upper(), m.group(2).strip()
    return None


# =============================================================================
# MAIN
# =============================================================================

def main() -> int:
    if not EXCEL_PATH.exists():
        print(f"[ERROR] Không tìm thấy file Excel: {EXCEL_PATH}", file=sys.stderr)
        return 1

    print(f"[INFO] Reading {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"[ERROR] Sheet '{SHEET_NAME}' không tồn tại", file=sys.stderr)
        return 1
    ws = wb[SHEET_NAME]
    print(f"[INFO] Sheet '{SHEET_NAME}': {ws.max_row} rows × {ws.max_column} cols")

    # =========================================================================
    # 1) Pre-scan section headers
    # =========================================================================
    sections: list[tuple[int, str, str]] = []
    for r in range(1, ws.max_row + 1):
        info = is_section_header(ws.cell(row=r, column=COL_STT).value)
        if info:
            sections.append((r, info[0], info[1]))

    print(f"[INFO] {len(sections)} section headers:")
    for r, lv, ten in sections:
        print(f"       row {r}: {lv} - {ten[:60]}")

    if len(sections) != 15:
        print(f"[ERROR] Kỳ vọng 15 sections (I-XV), tìm thấy {len(sections)}",
              file=sys.stderr)
        return 1

    def linh_vuc_at(row: int) -> tuple[Optional[str], Optional[str]]:
        current_lv, current_ten = None, None
        for sr, lv, ten in sections:
            if sr <= row:
                current_lv, current_ten = lv, ten
            else:
                break
        return current_lv, current_ten

    # =========================================================================
    # 2) Parse data rows (trust Excel, không validate sum)
    # =========================================================================
    rows_to_insert: list[dict] = []
    skipped_no_data = 0
    skipped_pre_section = 0
    skipped_invalid_nhom = 0
    seen_ma: set[str] = set()  # Detect duplicate ma_danh_muc (lỗi gõ stt trong Excel)
    duplicate_count = 0

    for r in range(DATA_START_ROW, ws.max_row + 1):
        stt_raw = ws.cell(row=r, column=COL_STT).value

        # Skip section headers
        if is_section_header(stt_raw):
            continue

        nhiem_vu = ws.cell(row=r, column=COL_NHIEM_VU).value
        cong_viec = ws.cell(row=r, column=COL_CONG_VIEC).value
        san_pham = ws.cell(row=r, column=COL_SAN_PHAM).value
        phan_nhom_raw = ws.cell(row=r, column=COL_PHAN_NHOM).value
        khung_diem_excel = to_int(ws.cell(row=r, column=COL_KHUNG_DIEM).value)
        diem_cham = to_int(ws.cell(row=r, column=COL_DIEM_CHAM).value)
        he_so = to_decimal(ws.cell(row=r, column=COL_HE_SO).value)
        ghi_chu = ws.cell(row=r, column=COL_GHI_CHU).value

        # Skip dòng thiếu data (nhiệm vụ header / dòng trống)
        if not cong_viec or not san_pham or diem_cham is None or he_so is None:
            skipped_no_data += 1
            continue

        # Lĩnh vực
        linh_vuc, ten_linh_vuc = linh_vuc_at(r)
        if not linh_vuc:
            # Row trước section header đầu tiên (row 8 "Đơn vị SPCV chuẩn")
            skipped_pre_section += 1
            continue

        # Nhóm — ưu tiên parse từ phan_nhom, fallback derive từ khung_diem Excel
        nhom = detect_nhom_from_phan_nhom(phan_nhom_raw)
        if nhom is None and khung_diem_excel in KHUNG_TO_NHOM:
            nhom = KHUNG_TO_NHOM[khung_diem_excel]
            print(f"[INFO] Row {r}: phan_nhom trống, fallback nhom={nhom} từ khung_diem={khung_diem_excel}")

        if nhom is None:
            skipped_invalid_nhom += 1
            print(f"[WARN] Skip row {r}: phan_nhom='{phan_nhom_raw}' và khung_diem='{khung_diem_excel}' đều không xác định được nhóm")
            continue

        # khung_diem_toi_da derive cứng từ nhóm (đảm bảo consistency)
        khung_diem = NHOM_KHUNG_MAP[nhom]

        # ma_danh_muc
        if isinstance(stt_raw, (int, float)):
            stt_str = str(int(stt_raw)) if float(stt_raw).is_integer() else str(stt_raw)
        else:
            stt_str = str(stt_raw).strip() if stt_raw is not None else f"r{r}"

        ma_danh_muc = f"PL3-{linh_vuc}-{stt_str}"

        # Excel có một số stt duplicate trong cùng lĩnh vực (lỗi gõ tay) →
        # append -r{row} cho occurrence thứ 2 trở đi để giữ unique.
        if ma_danh_muc in seen_ma:
            ma_danh_muc = f"{ma_danh_muc}-r{r}"
            duplicate_count += 1
        seen_ma.add(ma_danh_muc)

        if len(ma_danh_muc) > 30:
            print(f"[WARN] Skip row {r}: ma_danh_muc='{ma_danh_muc}' > 30 chars")
            skipped_no_data += 1
            continue

        rows_to_insert.append({
            "ma_danh_muc": ma_danh_muc,
            "ten_cong_viec": str(cong_viec).strip()[:500],
            "mo_ta": str(ghi_chu).strip() if ghi_chu else None,
            "nguon_du_lieu": "PL3",
            "linh_vuc": linh_vuc,
            "ten_linh_vuc": ten_linh_vuc,
            "nhiem_vu": str(nhiem_vu).strip()[:500] if nhiem_vu else None,
            "cong_viec_chi_tiet": str(cong_viec).strip(),
            "san_pham_dau_ra": str(san_pham).strip(),
            "nhom_pl3": nhom,
            "khung_diem_toi_da": khung_diem,
            "diem_cham": diem_cham,
            "he_so_quy_doi": he_so,
            "is_active": True,
        })

    # =========================================================================
    # 3) Báo cáo parse
    # =========================================================================
    print()
    print(f"[INFO] Parse summary:")
    print(f"       - Rows OK to insert: {len(rows_to_insert)}")
    print(f"       - Skipped (no full data / nhiệm vụ header / trống): {skipped_no_data}")
    print(f"       - Skipped (trước section header đầu tiên — row 8): {skipped_pre_section}")
    print(f"       - Skipped (phan_nhom không hợp lệ): {skipped_invalid_nhom}")
    print(f"       - Duplicate stt (renamed -r{{row}}): {duplicate_count}")
    print(f"       - Total processed = {len(rows_to_insert) + skipped_no_data + skipped_pre_section + skipped_invalid_nhom}")

    if not rows_to_insert:
        print("[STOP] Không có row nào để insert.")
        return 1

    # =========================================================================
    # 4) Insert vào DB (idempotent UPSERT)
    # =========================================================================
    print()
    print(f"[INFO] Connecting DB: {settings.database_url_sync.split('@')[-1]}")
    engine = create_engine(settings.database_url_sync)

    inserted = 0
    updated = 0

    upsert_sql = text("""
        INSERT INTO danh_muc_sp_cong_viec (
            ma_danh_muc, ten_cong_viec, mo_ta,
            nguon_du_lieu, linh_vuc, ten_linh_vuc,
            nhiem_vu, cong_viec_chi_tiet, san_pham_dau_ra,
            nhom_pl3, khung_diem_toi_da,
            diem_cham, he_so_quy_doi,
            is_active, sp_chuan_id, don_vi_ap_dung_id
        ) VALUES (
            :ma_danh_muc, :ten_cong_viec, :mo_ta,
            :nguon_du_lieu, :linh_vuc, :ten_linh_vuc,
            :nhiem_vu, :cong_viec_chi_tiet, :san_pham_dau_ra,
            :nhom_pl3, :khung_diem_toi_da,
            :diem_cham, :he_so_quy_doi,
            :is_active, NULL, NULL
        )
        ON CONFLICT (ma_danh_muc) DO UPDATE SET
            ten_cong_viec = EXCLUDED.ten_cong_viec,
            mo_ta = EXCLUDED.mo_ta,
            nguon_du_lieu = EXCLUDED.nguon_du_lieu,
            linh_vuc = EXCLUDED.linh_vuc,
            ten_linh_vuc = EXCLUDED.ten_linh_vuc,
            nhiem_vu = EXCLUDED.nhiem_vu,
            cong_viec_chi_tiet = EXCLUDED.cong_viec_chi_tiet,
            san_pham_dau_ra = EXCLUDED.san_pham_dau_ra,
            nhom_pl3 = EXCLUDED.nhom_pl3,
            khung_diem_toi_da = EXCLUDED.khung_diem_toi_da,
            diem_cham = EXCLUDED.diem_cham,
            he_so_quy_doi = EXCLUDED.he_so_quy_doi,
            is_active = EXCLUDED.is_active,
            updated_at = NOW()
        RETURNING (xmax = 0) AS inserted
    """)

    with engine.begin() as conn:
        for row in rows_to_insert:
            result = conn.execute(upsert_sql, row).first()
            if result and result[0]:
                inserted += 1
            else:
                updated += 1

    print()
    print(f"[DONE] inserted={inserted}, updated={updated}, total_db={inserted + updated}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
