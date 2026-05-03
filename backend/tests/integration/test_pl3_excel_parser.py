"""
tests/integration/test_pl3_excel_parser.py
==========================================
Test module parse Excel PL3 (dùng cho seed CLI + admin import endpoint).

Test scope:
- Parse file thật → 15 sections, 2.812 rows, 0 errors.
- Validate he_so vs diem_cham/25 với tolerance.
- Detect duplicate ma_danh_muc → suffix -r{row}.
- Reject sheet sai tên.
"""

from __future__ import annotations

import io
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from app.core.pl3_excel_parser import (
    HE_SO_TOLERANCE,
    NHOM_KHUNG_MAP,
    SHEET_NAME,
    parse_pl3_excel,
)


REAL_EXCEL = Path("/root/kpi-haiquan/docs/Danh mục công việc.xlsx")


# =============================================================================
# Tests với file Excel thật
# =============================================================================

class TestRealExcel:
    @pytest.mark.skipif(not REAL_EXCEL.exists(), reason="File Excel thật không tồn tại")
    def test_parse_real_file(self):
        result = parse_pl3_excel(REAL_EXCEL)
        assert len(result.sections) == 15, f"Sections != 15: {len(result.sections)}"
        # 2.812 rows OK to insert (đã verify Phase A)
        assert len(result.rows) == 2812, f"Rows != 2812: {len(result.rows)}"
        assert len(result.errors) == 0, f"Có {len(result.errors)} errors"
        assert result.duplicate_count == 18, f"Duplicate != 18: {result.duplicate_count}"

    @pytest.mark.skipif(not REAL_EXCEL.exists(), reason="File Excel thật không tồn tại")
    def test_section_headers_order(self):
        result = parse_pl3_excel(REAL_EXCEL)
        ma_lvs = [s[1] for s in result.sections]
        expected = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII",
                    "IX", "X", "XI", "XII", "XIII", "XIV", "XV"]
        assert ma_lvs == expected

    @pytest.mark.skipif(not REAL_EXCEL.exists(), reason="File Excel thật không tồn tại")
    def test_he_so_consistency(self):
        result = parse_pl3_excel(REAL_EXCEL)
        # Mọi row đều phải pass validate he_so ≈ diem_cham/25
        for r in result.rows:
            expected_he_so = Decimal(r.diem_cham) / Decimal("25")
            assert abs(r.he_so_quy_doi - expected_he_so) <= HE_SO_TOLERANCE


# =============================================================================
# Tests với synthetic Excel (in-memory)
# =============================================================================

def _build_synthetic_xlsx(rows_data: list[tuple], with_section: bool = True) -> bytes:
    """
    Tạo file xlsx in-memory với:
    - Row 1-7: header (skip)
    - Row 8: section "I" (if with_section)
    - Row 9+: data rows
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Header rows (1-7)
    for r in range(1, 8):
        ws.cell(row=r, column=1, value=f"header_r{r}")
    # Stt | Nhiệm vụ | CV | SP | PN | Khung | G | H | I | J | K | L | Note
    ws.cell(row=7, column=1, value="Stt")
    ws.cell(row=7, column=11, value="Điểm chấm")
    ws.cell(row=7, column=12, value="Hệ số quy đổi")

    cur_row = 8
    if with_section:
        ws.cell(row=cur_row, column=1, value="I. LĨNH VỰC TEST")
        cur_row += 1

    for stt, cv, sp, phan_nhom, khung, diem_cham, he_so in rows_data:
        ws.cell(row=cur_row, column=1, value=stt)
        ws.cell(row=cur_row, column=3, value=cv)
        ws.cell(row=cur_row, column=4, value=sp)
        ws.cell(row=cur_row, column=5, value=phan_nhom)
        ws.cell(row=cur_row, column=6, value=khung)
        ws.cell(row=cur_row, column=11, value=diem_cham)
        ws.cell(row=cur_row, column=12, value=he_so)
        cur_row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestSyntheticExcel:
    def test_simple_valid(self):
        xlsx = _build_synthetic_xlsx([
            ("1.1", "CV 1", "SP 1", "Nhóm 1", 100, 75, 3.0),
            ("1.2", "CV 2", "SP 2", "Nhóm 2", 200, 150, 6.0),
        ])
        # Note: synthetic chỉ có 1 section ('I'). Parser yêu cầu 15.
        # → expect errors về section count
        result = parse_pl3_excel(xlsx)
        assert len(result.sections) == 1
        assert len(result.errors) >= 1
        assert "15 section" in result.errors[0].error

    def test_he_so_mismatch_rejected(self):
        # Tạo file synthetic với 15 sections fake để pass section check
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.cell(row=7, column=1, value="Stt")

        cur = 8
        # Tạo 15 sections, mỗi section 1 row data hợp lệ
        roman = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII",
                 "IX", "X", "XI", "XII", "XIII", "XIV", "XV"]
        for lv in roman:
            ws.cell(row=cur, column=1, value=f"{lv}. Section {lv}")
            cur += 1
            ws.cell(row=cur, column=1, value=f"{roman.index(lv)+1}.1")
            ws.cell(row=cur, column=3, value="CV valid")
            ws.cell(row=cur, column=4, value="SP")
            ws.cell(row=cur, column=5, value="Nhóm 1")
            ws.cell(row=cur, column=6, value=100)
            ws.cell(row=cur, column=11, value=75)
            ws.cell(row=cur, column=12, value=3.0)
            cur += 1

        # Thêm 1 row có he_so SAI (3.0 thay vì should be 5.0)
        ws.cell(row=cur, column=1, value="99.99")
        ws.cell(row=cur, column=3, value="CV mismatch")
        ws.cell(row=cur, column=4, value="SP")
        ws.cell(row=cur, column=5, value="Nhóm 1")
        ws.cell(row=cur, column=6, value=100)
        ws.cell(row=cur, column=11, value=125)  # diem/25 = 5.0
        ws.cell(row=cur, column=12, value=3.0)  # SAI

        buf = io.BytesIO()
        wb.save(buf)
        result = parse_pl3_excel(buf.getvalue())

        # Section check pass; he_so mismatch → error
        assert len(result.sections) == 15
        assert len(result.errors) >= 1
        assert any("he_so" in e.error.lower() for e in result.errors)

    def test_wrong_sheet_name_raises(self):
        wb = openpyxl.Workbook()
        wb.active.title = "WrongName"
        buf = io.BytesIO()
        wb.save(buf)

        with pytest.raises(ValueError, match="Sheet 'PL3' không tồn tại"):
            parse_pl3_excel(buf.getvalue())


class TestKhungNhomMap:
    def test_map_consistency(self):
        assert NHOM_KHUNG_MAP == {1: 100, 2: 200, 3: 300, 4: 400, 5: 500}
